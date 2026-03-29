#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工商业储能市场开拓分析Excel生成脚本
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ========== 样式定义 ==========
# 深蓝色表头背景
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
# 绿色背景（达标）
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# 浅绿色（绿色标注的门槛）
LIGHT_GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
# 红色背景（不达标/警示）
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# 浅红色
LIGHT_RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
# 黄色警示
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
# 白色背景
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
# 浅灰背景
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
# 浅蓝背景
LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
# 浅绿背景
LIGHT_GREEN_BG = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# 字体
WHITE_BOLD_FONT = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
DARK_BOLD_FONT = Font(name='微软雅黑', size=11, bold=True, color="1F3864")
NORMAL_FONT = Font(name='微软雅黑', size=10)
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
LARGE_BOLD_FONT = Font(name='微软雅黑', size=14, bold=True, color="1F3864")
RED_BOLD_FONT = Font(name='微软雅黑', size=11, bold=True, color="9C0006")
GREEN_BOLD_FONT = Font(name='微软雅黑', size=11, bold=True, color="006100")
ORANGE_BOLD_FONT = Font(name='微软雅黑', size=10, bold=True, color="9C6500")

# 对齐
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

# 边框
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ========== 辅助函数 ==========
def set_header_style(cell):
    """设置表头样式"""
    cell.font = WHITE_BOLD_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

def set_cell_style(cell, fill=None, font=None, alignment=None, border=True):
    """设置单元格样式"""
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = THIN_BORDER

def set_column_width(ws, col_widths):
    """设置列宽"""
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

def write_cell(ws, row, col, value, fill=None, font=None, alignment=None, 
               start_col=None, end_col=None, start_row=None, end_row=None):
    """写入单元格（支持合并）"""
    if start_col and end_col:
        ws.merge_cells(start_row=start_row or row, start_column=start_col,
                       end_row=end_row or row, end_column=end_col)
        cell = ws.cell(row=start_row or row, column=start_col)
    elif start_row and end_row:
        ws.merge_cells(start_row=start_row, start_column=col,
                       end_row=end_row, end_column=col)
        cell = ws.cell(row=start_row, column=col)
    else:
        cell = ws.cell(row=row, column=col)
    
    cell.value = value
    set_cell_style(cell, fill=fill, font=font, alignment=alignment)
    return cell

# ========== Sheet1：模式一经济账 ==========
def create_sheet1(wb):
    ws = wb.active
    ws.title = "模式一经济账"
    
    # 标题
    write_cell(ws, 1, 1, "模式一：纯储能投资经济账",
               fill=HEADER_FILL,
               font=Font(name='微软雅黑', size=16, bold=True, color="FFFFFF"),
               alignment=CENTER_ALIGN,
               start_col=1, end_col=6, start_row=1, end_row=1)
    ws.row_dimensions[1].height = 40
    
    # 门槛结论区域标题
    write_cell(ws, 3, 1, "📊 门槛结论",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=3, end_row=3)
    ws.row_dimensions[3].height = 25
    
    # 门槛条件表头
    write_cell(ws, 4, 1, "指标", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=4, end_row=4)
    write_cell(ws, 4, 3, "条件", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=4, end_row=4)
    write_cell(ws, 4, 5, "结论", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=6, start_row=4, end_row=4)
    
    # IRR行
    write_cell(ws, 5, 1, "IRR = 10%", fill=LIGHT_GRAY_FILL, font=BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=5, end_row=5)
    write_cell(ws, 5, 3, "电价差 ≥ 0.674 元/kWh", fill=LIGHT_GRAY_FILL, font=BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=5, end_row=5)
    write_cell(ws, 5, 5, "✅ 绿色可达", fill=GREEN_FILL, font=GREEN_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=6, start_row=5, end_row=5)
    
    # 回收期行
    write_cell(ws, 6, 1, "回收期 ≤ 7年", fill=WHITE_FILL, font=BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=6, end_row=6)
    write_cell(ws, 6, 3, "电价差 ≥ 0.588 元/kWh", fill=WHITE_FILL, font=BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=6, end_row=6)
    write_cell(ws, 6, 5, "✅ 绿色可达", fill=GREEN_FILL, font=GREEN_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=6, start_row=6, end_row=6)
    
    # 基础参数表标题
    write_cell(ws, 8, 1, "📋 基础参数表",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=8, end_row=8)
    ws.row_dimensions[8].height = 25
    
    # 参数表表头
    write_cell(ws, 9, 1, "参数名称", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=1, start_row=9, end_row=9)
    write_cell(ws, 9, 2, "参数值", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=2, end_col=3, start_row=9, end_row=9)
    write_cell(ws, 9, 4, "单位", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=4, end_col=4, start_row=9, end_row=9)
    write_cell(ws, 9, 5, "说明", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=6, start_row=9, end_row=9)
    
    # 参数数据
    params = [
        ("储能容量", "10", "MWh", "电池组总容量"),
        ("总投资", "1000", "万元", "含设备及施工"),
        ("设备单价", "0.8", "元/Wh", "电池组成本"),
        ("施工单价", "0.2", "元/Wh", "建设施工成本"),
        ("运营年限", "10", "年", "项目运营周期"),
        ("年运行次数", "330", "次/年", "充放电循环"),
        ("系统效率", "88", "%", "AC-AC转换效率"),
        ("年衰减率", "2", "%", "电池每年衰减"),
        ("DOD深度", "95", "%", "放电深度"),
        ("年运营成本", "10", "万元/年", "运维人员及杂费"),
    ]
    
    for idx, (name, value, unit, desc) in enumerate(params):
        row = 10 + idx
        row_fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, name, fill=LIGHT_GRAY_FILL, font=BOLD_FONT,
                   alignment=CENTER_ALIGN, start_col=1, end_col=1, start_row=row, end_row=row)
        write_cell(ws, row, 2, value, fill=row_fill,
                   font=Font(name='微软雅黑', size=11, bold=True),
                   alignment=CENTER_ALIGN, start_col=2, end_col=3, start_row=row, end_row=row)
        write_cell(ws, row, 4, unit, fill=row_fill, font=NORMAL_FONT,
                   alignment=CENTER_ALIGN, start_col=4, end_col=4, start_row=row, end_row=row)
        write_cell(ws, row, 5, desc, fill=row_fill, font=NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=5, end_col=6, start_row=row, end_row=row)
    
    # 设置列宽
    set_column_width(ws, {'A': 15, 'B': 12, 'C': 10, 'D': 10, 'E': 15, 'F': 15})
    
    return ws


# ========== Sheet2：模式二经济账 ==========
def create_sheet2(wb):
    ws = wb.create_sheet("模式二经济账")
    
    # 标题
    write_cell(ws, 1, 1, "模式二：光储一体化投资经济账（不换电芯版）",
               fill=HEADER_FILL,
               font=Font(name='微软雅黑', size=16, bold=True, color="FFFFFF"),
               alignment=CENTER_ALIGN,
               start_col=1, end_col=7, start_row=1, end_row=1)
    ws.row_dimensions[1].height = 40
    
    # 警示信息
    write_cell(ws, 2, 1, "⚠️ 结论标注：模式二经济性较差，建议调整参数（配比/规模）",
               fill=LIGHT_RED_FILL,
               font=RED_BOLD_FONT,
               alignment=CENTER_ALIGN,
               start_col=1, end_col=7, start_row=2, end_row=2)
    ws.row_dimensions[2].height = 30
    
    # 门槛结论区域标题
    write_cell(ws, 4, 1, "📊 门槛结论",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=7, start_row=4, end_row=4)
    ws.row_dimensions[4].height = 25
    
    # 门槛结论表头
    write_cell(ws, 5, 1, "指标", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=5, end_row=5)
    write_cell(ws, 5, 3, "条件", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=5, end_row=5)
    write_cell(ws, 5, 5, "计算结果", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=5, start_row=5, end_row=5)
    write_cell(ws, 5, 6, "状态", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=6, end_col=7, start_row=5, end_row=5)
    
    # 门槛结论数据
    conclusions = [
        ("LCOE", "储能度电成本", "待计算", "⚠️ 待计算"),
        ("IRR=10%", "储能卖电价门槛", "待计算", "⚠️ 待计算"),
        ("回收期≤7年", "投资回收期", "待计算", "⚠️ 待计算"),
    ]
    
    for idx, (name, cond, result, status) in enumerate(conclusions):
        row = 6 + idx
        row_fill = LIGHT_GRAY_FILL if idx % 2 == 0 else WHITE_FILL
        
        write_cell(ws, row, 1, name, fill=row_fill, font=BOLD_FONT,
                   alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=row, end_row=row)
        write_cell(ws, row, 3, cond, fill=row_fill, font=NORMAL_FONT,
                   alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=row, end_row=row)
        write_cell(ws, row, 5, result, fill=YELLOW_FILL, font=BOLD_FONT,
                   alignment=CENTER_ALIGN, start_col=5, end_col=5, start_row=row, end_row=row)
        write_cell(ws, row, 6, status, fill=YELLOW_FILL, font=ORANGE_BOLD_FONT,
                   alignment=CENTER_ALIGN, start_col=6, end_col=7, start_row=row, end_row=row)
    
    # 基础参数表标题
    write_cell(ws, 10, 1, "📋 基础参数表",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=7, start_row=10, end_row=10)
    ws.row_dimensions[10].height = 25
    
    # 参数表表头
    write_cell(ws, 11, 1, "参数名称", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=1, start_row=11, end_row=11)
    write_cell(ws, 11, 2, "参数值", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=2, end_col=3, start_row=11, end_row=11)
    write_cell(ws, 11, 4, "单位", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=4, end_col=4, start_row=11, end_row=11)
    write_cell(ws, 11, 5, "说明", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=7, start_row=11, end_row=11)
    
    # 参数数据
    params = [
        ("光伏装机", "1", "MW", "峰值功率"),
        ("储能容量", "2", "MWh", "电池组总容量"),
        ("光伏成本", "2", "元/W", "含组件及安装"),
        ("储能成本", "1", "元/Wh", "电池组成本"),
        ("总投资", "500", "万元", "光储系统总投资"),
        ("年日照时数", "1200", "h", "有效发电小时"),
        ("系统效率", "80", "%", "综合效率"),
        ("年衰减率", "1.5", "%", "电池每年衰减（优于模式一）"),
        ("年运行次数", "330", "次/年", "充放电循环"),
        ("运营年限", "20", "年", "不换电芯长期运行"),
    ]
    
    for idx, (name, value, unit, desc) in enumerate(params):
        row = 12 + idx
        row_fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, name, fill=LIGHT_GRAY_FILL, font=BOLD_FONT,
                   alignment=CENTER_ALIGN, start_col=1, end_col=1, start_row=row, end_row=row)
        write_cell(ws, row, 2, value, fill=row_fill,
                   font=Font(name='微软雅黑', size=11, bold=True),
                   alignment=CENTER_ALIGN, start_col=2, end_col=3, start_row=row, end_row=row)
        write_cell(ws, row, 4, unit, fill=row_fill, font=NORMAL_FONT,
                   alignment=CENTER_ALIGN, start_col=4, end_col=4, start_row=row, end_row=row)
        write_cell(ws, row, 5, desc, fill=row_fill, font=NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=5, end_col=7, start_row=row, end_row=row)
    
    # 收入结构标题
    write_cell(ws, 23, 1, "💰 收入结构（待计算）",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=7, start_row=23, end_row=23)
    ws.row_dimensions[23].height = 25
    
    # 收入结构表头
    write_cell(ws, 24, 1, "收入来源", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=24, end_row=24)
    write_cell(ws, 24, 3, "比例", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=24, end_row=24)
    write_cell(ws, 24, 5, "说明", fill=HEADER_FILL, font=WHITE_BOLD_FONT,
               alignment=CENTER_ALIGN, start_col=5, end_col=7, start_row=24, end_row=24)
    
    # 收入结构数据
    incomes = [
        ("光伏直售电", "31.8%", "自发自用余电上网"),
        ("储能峰谷套利", "64%", "低充高放收益分成"),
        ("其他收益", "4.2%", "需求响应/辅助服务"),
    ]
    
    for idx, (source, ratio, desc) in enumerate(incomes):
        row = 25 + idx
        row_fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, source, fill=row_fill, font=BOLD_FONT,
                   alignment=CENTER_ALIGN, start_col=1, end_col=2, start_row=row, end_row=row)
        write_cell(ws, row, 3, ratio, fill=YELLOW_FILL,
                   font=Font(name='微软雅黑', size=11, bold=True),
                   alignment=CENTER_ALIGN, start_col=3, end_col=4, start_row=row, end_row=row)
        write_cell(ws, row, 5, desc, fill=row_fill, font=NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=5, end_col=7, start_row=row, end_row=row)
    
    # 设置列宽
    set_column_width(ws, {'A': 15, 'B': 12, 'C': 10, 'D': 10, 'E': 15, 'F': 12, 'G': 12})
    
    return ws


# ========== Sheet3：现货市场价格数据 ==========
def create_sheet3(wb):
    ws = wb.create_sheet("现货市场价格数据")
    
    # 标题
    write_cell(ws, 1, 1, "各省电力现货市场价格数据（2025-2026年）",
               fill=HEADER_FILL,
               font=Font(name='微软雅黑', size=16, bold=True, color="FFFFFF"),
               alignment=CENTER_ALIGN,
               start_col=1, end_col=8, start_row=1, end_row=1)
    ws.row_dimensions[1].height = 40
    
    # 表头
    headers = ["省份", "午间低电价\n(元/kWh)", "晚高峰\n(元/kWh)", "峰谷价差\n(元/kWh)", 
               "负电价情况", "负电价\n(元/kWh)", "IRR≥10%", "回收期≤7年"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        set_header_style(cell)
        cell.alignment = CENTER_ALIGN
    
    ws.row_dimensions[3].height = 35
    
    # 数据
    data = [
        ("山东", "0.02~0.05", "0.40~0.50", "0.35~0.57", "频繁", "-0.07", "❌不满足", "⚠️勉强"),
        ("山西", "0~0.01", "0.40~0.80", "0.40~1.50", "频繁", "-", "✅满足", "✅满足"),
        ("广东", "0.02~0.05", "0.45~0.60", "0.40~1.29", "极少", "-", "✅满足", "✅满足"),
        ("甘肃", "0.04", "0.30~0.40", "0.25~0.46", "地板频繁", "-", "❌不满足", "❌不满足"),
        ("浙江", "0.02~0.05", "0.40~0.50", "0.35~1.27", "频繁", "-", "✅满足", "✅满足"),
    ]
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
            # 省份列
            if col_idx == 1:
                cell.font = BOLD_FONT
                cell.fill = LIGHT_GRAY_FILL
            # 负电价列
            elif col_idx == 6:
                if value == "-0.07":
                    cell.fill = LIGHT_RED_FILL
                else:
                    cell.fill = WHITE_FILL
            # IRR列
            elif col_idx == 7:
                if "✅满足" in value:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_BOLD_FONT
                elif "❌" in value:
                    cell.fill = RED_FILL
                    cell.font = RED_BOLD_FONT
                else:
                    cell.fill = YELLOW_FILL
                    cell.font = ORANGE_BOLD_FONT
            # 回收期列
            elif col_idx == 8:
                if "✅满足" in value:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_BOLD_FONT
                elif "❌" in value:
                    cell.fill = RED_FILL
                    cell.font = RED_BOLD_FONT
                elif "⚠️勉强" in value:
                    cell.fill = YELLOW_FILL
                    cell.font = ORANGE_BOLD_FONT
                else:
                    cell.fill = WHITE_FILL
            else:
                cell.fill = WHITE_FILL if row_idx % 2 == 0 else LIGHT_GRAY_FILL
    
    # 说明
    write_cell(ws, 10, 1, "📌 说明：负电价表示午间光伏发电高峰期电力供过于求，储能可在负电价时段充电，高电价时段放电套利",
               fill=WHITE_FILL,
               font=Font(name='微软雅黑', size=9, italic=True, color="666666"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=8, start_row=10, end_row=10)
    
    # 设置列宽
    set_column_width(ws, {'A': 10, 'B': 14, 'C': 14, 'D': 14, 'E': 12, 'F': 10, 'G': 12, 'H': 14})
    
    return ws


# ========== Sheet4：省份梯队划分 ==========
def create_sheet4(wb):
    ws = wb.create_sheet("省份梯队划分")
    
    # 标题
    write_cell(ws, 1, 1, "省份梯队划分与开拓建议",
               fill=HEADER_FILL,
               font=Font(name='微软雅黑', size=16, bold=True, color="FFFFFF"),
               alignment=CENTER_ALIGN,
               start_col=1, end_col=5, start_row=1, end_row=1)
    ws.row_dimensions[1].height = 40
    
    # 模式一梯队标题
    write_cell(ws, 3, 1, "📊 模式一梯队（纯储能）",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=13, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=5, start_row=3, end_row=3)
    ws.row_dimensions[3].height = 28
    
    # 模式一表头
    tier1_headers = ["梯队", "省份", "准入条件", "市场状态", "开拓建议"]
    for col, header in enumerate(tier1_headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        set_header_style(cell)
    
    # 模式一数据
    tier1_data = [
        ("第一梯队", "山西/广东/浙江", "峰谷价差 ≥ 0.674元/kWh", "✅ 现货市场已开放", "⭐ 优先开拓"),
        ("第二梯队", "山东/江苏/河南", "光伏装机大\n现货即将全覆盖", "⚠️ 现货建设中", "📋 可储备项目"),
        ("第三梯队", "其他省份", "持续观察", "📌 政策待明确", "🔍 持续跟踪"),
    ]
    
    tier_fills = [GREEN_FILL, YELLOW_FILL, LIGHT_GRAY_FILL]
    
    for row_idx, (row_data, fill) in enumerate(zip(tier1_data, tier_fills), 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = CENTER_ALIGN if col_idx in [1, 2] else LEFT_ALIGN
            cell.border = THIN_BORDER
            
            if col_idx in [1, 4, 5]:
                cell.fill = fill
                if col_idx == 1:
                    cell.font = DARK_BOLD_FONT
                elif col_idx == 4:
                    if "✅" in value:
                        cell.font = GREEN_BOLD_FONT
                    elif "⚠️" in value:
                        cell.font = ORANGE_BOLD_FONT
                    else:
                        cell.font = NORMAL_FONT
                elif col_idx == 5:
                    cell.font = DARK_BOLD_FONT
            else:
                cell.fill = WHITE_FILL if row_idx % 2 == 0 else LIGHT_GRAY_FILL
    
    # 模式二梯队标题
    write_cell(ws, 9, 1, "📊 模式二梯队（光储一体化）",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=13, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=5, start_row=9, end_row=9)
    ws.row_dimensions[9].height = 28
    
    # 模式二表头
    for col, header in enumerate(tier1_headers, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        set_header_style(cell)
    
    # 模式二数据
    tier2_data = [
        ("第一梯队", "河北/山东/河南/安徽", "光伏装机大\n午间低电价明确", "✅ 午间电价优势明显", "⭐ 优先开拓"),
        ("第二梯队", "江苏/浙江/广东", "光伏装机大\n现货市场成熟", "✅ 市场机制完善", "📋 重点跟进"),
        ("第三梯队", "其他省份", "持续观察", "📌 资源待开发", "🔍 持续跟踪"),
    ]
    
    for row_idx, (row_data, fill) in enumerate(zip(tier2_data, tier_fills), 11):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = CENTER_ALIGN if col_idx in [1, 2] else LEFT_ALIGN
            cell.border = THIN_BORDER
            
            if col_idx in [1, 4, 5]:
                cell.fill = fill
                if col_idx == 1:
                    cell.font = DARK_BOLD_FONT
                elif col_idx == 4:
                    if "✅" in value:
                        cell.font = GREEN_BOLD_FONT
                    elif "⚠️" in value:
                        cell.font = ORANGE_BOLD_FONT
                    else:
                        cell.font = NORMAL_FONT
                elif col_idx == 5:
                    cell.font = DARK_BOLD_FONT
            else:
                cell.fill = WHITE_FILL if row_idx % 2 == 0 else LIGHT_GRAY_FILL
    
    # 设置列宽
    set_column_width(ws, {'A': 12, 'B': 22, 'C': 22, 'D': 20, 'E': 15})
    
    return ws


# ========== Sheet5：综合结论与建议 ==========
def create_sheet5(wb):
    ws = wb.create_sheet("综合结论与建议")
    
    # 标题
    write_cell(ws, 1, 1, "综合结论与开拓建议",
               fill=HEADER_FILL,
               font=Font(name='微软雅黑', size=16, bold=True, color="FFFFFF"),
               alignment=CENTER_ALIGN,
               start_col=1, end_col=6, start_row=1, end_row=1)
    ws.row_dimensions[1].height = 40
    
    # 政策背景标题
    write_cell(ws, 3, 1, "📌 政策背景",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=3, end_row=3)
    ws.row_dimensions[3].height = 25
    
    write_cell(ws, 4, 1, "2026年3月起，全国取消行政性峰谷电价，改为市场化交易。电力现货市场建设加速推进，工商业储能迎来新机遇。",
               fill=WHITE_FILL,
               font=NORMAL_FONT,
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=4, end_row=4)
    ws.row_dimensions[4].height = 30
    
    # 模式一结论标题
    write_cell(ws, 6, 1, "✅ 模式一结论（纯储能）",
               fill=LIGHT_GREEN_BG,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=6, end_row=6)
    ws.row_dimensions[6].height = 25
    
    mode1_items = [
        ("主推省份：", "山西 / 广东 / 浙江"),
        ("门槛电价差：", "≥ 0.674 元/kWh"),
        ("IRR条件：", "IRR ≥ 10%，需峰谷价差 ≥ 0.674元/kWh"),
        ("回收期条件：", "≤ 7年，需峰谷价差 ≥ 0.588元/kWh"),
    ]
    
    for idx, (label, value) in enumerate(mode1_items):
        row = 7 + idx
        fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, label, fill=fill, font=BOLD_FONT,
                   alignment=LEFT_ALIGN, start_col=1, end_col=2, start_row=row, end_row=row)
        write_cell(ws, row, 3, value, fill=fill, font=GREEN_BOLD_FONT if "山西" in value or "0.674" in value or "0.588" in value else NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=3, end_col=6, start_row=row, end_row=row)
        ws.row_dimensions[row].height = 22
    
    # 模式二结论标题
    write_cell(ws, 12, 1, "⚠️ 模式二结论（光储一体化）",
               fill=PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=12, end_row=12)
    ws.row_dimensions[12].height = 25
    
    mode2_items = [
        ("经济性评价：", "经济性不足，建议调整参数"),
        ("调整方向：", "优化光储配比/扩大规模"),
        ("适用场景：", "光伏装机大、午间低电价明确省份"),
    ]
    
    for idx, (label, value) in enumerate        row = 13 + idx
        fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, label, fill=fill, font=BOLD_FONT,
                   alignment=LEFT_ALIGN, start_col=1, end_col=2, start_row=row, end_row=row)
        write_cell(ws, row, 3, value, fill=fill, font=RED_BOLD_FONT if "不足" in value else NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=3, end_col=6, start_row=row, end_row=row)
        ws.row_dimensions[row].height = 22
    
    # 开拓建议标题
    write_cell(ws, 17, 1, "🎯 开拓建议",
               fill=LIGHT_BLUE_FILL,
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=17, end_row=17)
    ws.row_dimensions[17].height = 25
    
    suggestions = [
        ("优先级：", "第一梯队省份（山西/广东/浙江）优先开展工商业储能项目"),
        ("时机：", "现货市场已开放省份立即启动，其他省份密切跟踪政策"),
        ("模式选择：", "优先选择模式一（纯储能），模式二需优化参数后决策"),
    ]
    
    for idx, (label, value) in enumerate(suggestions):
        row = 18 + idx
        fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, label, fill=fill, font=BOLD_FONT,
                   alignment=LEFT_ALIGN, start_col=1, end_col=2, start_row=row, end_row=row)
        write_cell(ws, row, 3, value, fill=fill, font=GREEN_BOLD_FONT if "优先" in value else NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=3, end_col=6, start_row=row, end_row=row)
        ws.row_dimensions[row].height = 22
    
    # 风险提示标题
    write_cell(ws, 22, 1, "⚠️ 风险提示",
               fill=PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
               font=Font(name='微软雅黑', size=12, bold=True, color="1F3864"),
               alignment=LEFT_ALIGN,
               start_col=1, end_col=6, start_row=22, end_row=22)
    ws.row_dimensions[22].height = 25
    
    risks = [
        ("电价差收窄风险：", "随着新能源装机增长，峰谷价差可能收窄，影响项目收益"),
        ("市场价格波动风险：", "电力现货市场价格波动大，需做好收益预测和风险对控"),
        ("政策变化风险：", "电力市场化改革推进，补贴政策和市场规则可能调整"),
        ("技术迭代风险：", "电池技术进步可能影响现有项目经济性"),
    ]
    
    for idx, (label, value) in enumerate(risks):
        row = 23 + idx
        fill = WHITE_FILL if idx % 2 == 0 else LIGHT_GRAY_FILL
        
        write_cell(ws, row, 1, label, fill=fill, font=RED_BOLD_FONT,
                   alignment=LEFT_ALIGN, start_col=1, end_col=2, start_row=row, end_row=row)
        write_cell(ws, row, 3, value, fill=fill, font=NORMAL_FONT,
                   alignment=LEFT_ALIGN, start_col=3, end_col=6, start_row=row, end_row=row)
        ws.row_dimensions[row].height = 22
    
    # 设置列宽
    set_column_width(ws, {'A': 15, 'B': 12, 'C': 15, 'D': 15, 'E': 15, 'F': 15})
    
    return ws


# ========== 主函数 ==========
def main():
    """生成工商业储能市场开拓分析Excel文件"""
    # 创建工作簿
    wb = Workbook()
    
    # 创建各个Sheet
    print("正在创建 Sheet1：模式一经济账...")
    create_sheet1(wb)
    
    print("正在创建 Sheet2：模式二经济账...")
    create_sheet2(wb)
    
    print("正在创建 Sheet3：现货市场价格数据...")
    create_sheet3(wb)
    
    print("正在创建 Sheet4：省份梯队划分...")
    create_sheet4(wb)
    
    print("正在创建 Sheet5：综合结论与建议...")
    create_sheet5(wb)
    
    # 保存文件
    output_path = "/Users/zhaoruicn/.openclaw/workspace/工商业储能市场开拓分析.xlsx"
    wb.save(output_path)
    print(f"\n✅ Excel文件已生成：{output_path}")
    
    return output_path


if __name__ == "__main__":
    main()
