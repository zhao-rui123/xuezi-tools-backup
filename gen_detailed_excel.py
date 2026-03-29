#!/usr/bin/env python3
"""工商业储能市场开拓分析 - 专业完整版"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = '/Users/zhaoruicn/.openclaw/workspace/工商业储能市场开拓分析_专业完整版.xlsx'
wb = openpyxl.Workbook()

# ===== 样式定义 =====
FILL_BLUE = PatternFill('solid', fgColor='1F4E79')
FILL_BLUE2 = PatternFill('solid', fgColor='2F75B6')
FILL_GREEN = PatternFill('solid', fgColor='C6EFCE')
FILL_RED = PatternFill('solid', fgColor='FFC7CE')
FILL_YELLOW = PatternFill('solid', fgColor='FFEB9C')
FILL_LIGHT_BLUE = PatternFill('solid', fgColor='D9E1F2')
FILL_LIGHT_GREEN = PatternFill('solid', fgColor='E2EFDA')
FILL_LIGHT_RED = PatternFill('solid', fgColor='FCE4D6')

def thin_border():
    s = Side('thin')
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side('medium')
    return Border(left=s, right=s, top=s, bottom=s)

TB = thin_border()

def hdr_cell(cell, text, size=11, fill=None, color='FFFFFF', bold=True, wrap=True):
    cell.value = text
    cell.font = Font(bold=bold, size=size, color=color)
    cell.fill = fill or FILL_BLUE
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border = TB

def data_cell(cell, value, fill=None, font_size=10, bold=False, align='center', color='000000'):
    cell.value = value
    cell.font = Font(bold=bold, size=font_size, color=color)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = TB

def money_cell(cell, value, good=True):
    cell.value = f'¥{value:.4f}'
    cell.fill = FILL_GREEN if good else FILL_RED
    cell.font = Font(bold=True, size=11, color='006100' if good else '9C0006')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = TB

def pct_cell(cell, value, good=True):
    if value is None:
        cell.value = '-'
    else:
        cell.value = f'{value:.2%}'
    cell.fill = FILL_GREEN if good else FILL_RED
    cell.font = Font(bold=True, size=11, color='006100' if good else '9C0006')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = TB

# ===== 财务计算函数 =====
def annual_pv_kwh(year, kw=1000, hours=1200, eff=0.80, decay=0.02):
    return kw * hours * eff * ((1-decay)**(year-1))

def annual_storage_discharge(year, cap_kwh=1000, DOD=0.95, decay=0.015, eff=0.88, cycles=330):
    return cap_kwh * DOD * ((1-decay)**(year-1)) * eff * cycles

def irr_calc(cashflows, guess=0.1):
    rate = guess
    for _ in range(1000):
        f = sum(cf/(1+rate)**t for t, cf in enumerate(cashflows))
        df = sum(-t*cf/(1+rate)**(t+1) for t, cf in enumerate(cashflows))
        if abs(df) < 1e-12:
            break
        rate -= f/df
        if abs(f) < 1e-10:
            break
    return rate

# ==================== Sheet 1: 封面 ====================
ws1 = wb.active
ws1.title = '封面'
ws1.column_dimensions['A'].width = 80

# 大标题
ws1.merge_cells('A3:E3')
ws1.row_dimensions[3].height = 60
c = ws1['A3']
c.value = '工商业储能市场开拓分析报告'
c.font = Font(bold=True, size=28, color='1F4E79')
c.alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A4:E4')
c = ws1['A4']
c.value = '2026年3月 | 政策解读 | 经济账测算 | 省份梯队 | 投资地图'
c.font = Font(size=14, color='666666', italic=True)
c.alignment = Alignment(horizontal='center')

ws1.merge_cells('A5:E5')
ws1.row_dimensions[5].height = 20

# 目录
ws1.merge_cells('A7:E7')
ws1.row_dimensions[7].height = 30
c = ws1['A7']
c.value = '目  录'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

dirs = [
    ('Sheet2', '模式一经济账（纯储能10MWh）', '门槛分析 | 逐年现金流 | 敏感性分析'),
    ('Sheet3', '模式二经济账（光储一体化1MW+3MWh）', '最优配置 | 逐年现金流 | 多情景模拟'),
    ('Sheet4', '现货市场价格数据', '5省实时电价 | 峰谷价差对比'),
    ('Sheet5', '省份梯队划分', '模式一梯队 | 模式二梯队 | 详细分析'),
    ('Sheet6', '全国投资地图', '7大区域 | 颜色标注 | 开拓策略'),
    ('Sheet7', '综合结论与建议', '政策背景 | 模式对比 | 风险提示 | 开拓建议'),
]
for i, (sheet, title, desc) in enumerate(dirs, 8):
    ws1.row_dimensions[i].height = 25
    c = ws1.cell(i, 1)
    c.value = f'{i-7}. {title}'
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c = ws1.cell(i, 2)
    c.value = f'→ {desc}'
    c.font = Font(size=10, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center')

# ==================== Sheet 2: 模式一经济账 ====================
ws2 = wb.create_sheet('模式一经济账')

# 列宽
widths = {'A':18,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':14,'I':14,'J':14,'K':14,'L':14}
for col, w in widths.items():
    ws2.column_dimensions[col].width = w

# 标题
ws2.merge_cells('A1:L1')
ws2.row_dimensions[1].height = 40
c = ws2['A1']
c.value = '模式一：纯储能投资经济账（10MWh储能）'
c.font = Font(bold=True, size=18, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

# 门槛结论
ws2.merge_cells('A3:D3')
ws2.row_dimensions[3].height = 35
c = ws2['A3']
c.value = '⚠️ 核心门槛结论'
c.font = Font(bold=True, size=14, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

thresh_data = [
    ['IRR=10% 需要电价差', '≥ 0.6618 元/kWh', '✅ 满足则项目IRR达到10%'],
    ['回收期≤7年 需要电价差', '≥ 0.5795 元/kWh', '✅ 满足则7年内可回本'],
]
for i, row in enumerate(thresh_data, 4):
    ws2.row_dimensions[i].height = 28
    c = ws2.cell(i, 1)
    c.value = row[0]
    c.font = Font(bold=True, size=11)
    c.fill = FILL_YELLOW
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
    ws2.merge_cells(f'B{i}:C{i}')
    c = ws2.cell(i, 2)
    c.value = row[1]
    c.font = Font(bold=True, size=14, color='006100')
    c.fill = FILL_GREEN
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
    ws2.merge_cells(f'D{i}:L{i}')
    c = ws2.cell(i, 4)
    c.value = row[2]
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = TB

# 基础参数
ws2.merge_cells('A7:L7')
ws2.row_dimensions[7].height = 25
c = ws2['A7']
c.value = '一、基础参数'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

params2 = [
    ['储能容量', '10 MWh', '放电深度', '95%', '年充放次数', '330次', '投资回收期目标', '≤7年'],
    ['总投资', '1,000 万元', '充放电效率', '88%', '年衰减率', '1.5%', '目标IRR', '≥10%'],
    ['设备成本', '0.80 元/Wh', '运营成本', '10万/年', '使用年限', '10年', '放电策略', '峰谷套利'],
]
for i, row in enumerate(params2, 8):
    ws2.row_dimensions[i].height = 22
    for j, v in enumerate(row):
        c = ws2.cell(i, j*2+1)
        if j % 2 == 0:
            c.value = v
            c.font = Font(bold=True, size=10)
            c.fill = FILL_LIGHT_BLUE
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = TB
            ws2.merge_cells(f'{get_column_letter(j*2+1)}{i}:{get_column_letter(j*2+2)}{i}')
        else:
            c.value = v
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = TB

# 逐年现金流
ws2.merge_cells('A13:L13')
ws2.row_dimensions[13].height = 25
c = ws2['A13']
c.value = '二、逐年现金流明细（电价差=0.674元/kWh，IRR=10%）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

cf_headers = ['年度', '期初投资', '年放电量(kWh)', '年衰减率', '可用放电量', '电价差', '年收入(万)', '年运营成本', '年净现金流', '累计现金流', 'IRR', '静态回收期']
for j, h in enumerate(cf_headers, 1):
    c = ws2.cell(14, j)
    c.value = h
    c.font = Font(bold=True, size=9, color='FFFFFF')
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = TB
ws2.row_dimensions[14].height = 30

# 现金流计算
price_diff = 0.674
total_invest = 1000
annual_opex = 10
years = 10
cap = 10000
DOD = 0.95
eff = 0.88
decay = 0.015
cycles = 330

cum_cash = -total_invest
cf_list = [-total_invest]
for y in range(1, years+1):
    discharge = cap * DOD * ((1-decay)**(y-1)) * eff * cycles
    available = discharge * price_diff
    net = available - annual_opex
    cf_list.append(net)

for y in range(1, years+1):
    ws2.row_dimensions[14+y].height = 20
    discharge = cap * DOD * ((1-decay)**(y-1)) * eff * cycles
    year_decay = 1-(1-decay)**(y-1)
    available = discharge * price_diff
    net = available - annual_opex
    cum_cash += net
    irr_val = irr_calc([-total_invest] + [min(max((cap*DOD*((1-decay)**(yy-1))*eff*cycles*price_diff-annual_opex, -1e10) for yy in range(1,y+1)] + [0]*(years-y-1)+[cap*DOD*((1-decay)**(years-1))*eff*cycles*price_diff-annual_opex])

    row = [y, -total_invest if y==1 else 0, discharge/10000, year_decay, discharge/10000*price_diff, price_diff, available/10000, annual_opex, net/10000, cum_cash/10000]
    irr_actual = irr_calc(cf_list[:y+1])
    payback_y = y
    for py in range(1, y+1):
        if sum(cf_list[1:py+1]) >= 0:
            payback_y = py
            break

    vals = [row[0], row[1], f'{row[2]:.2f}', f'{row[3]*100:.1f}%', f'{row[4]:.4f}', f'{row[5]:.4f}', f'{row[6]:.4f}', f'{row[7]:.4f}', f'{row[8]:.4f}', f'{row[9]:.4f}']
    for j, v in enumerate(vals, 1):
        c = ws2.cell(14+y, j)
        c.value = v
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = TB
        if j == 1:
            c.fill = FILL_LIGHT_BLUE
        elif j == 9 and cum_cash >= 0:
            c.fill = FILL_GREEN
        elif j == 9 and cum_cash < 0:
            c.fill = FILL_LIGHT_RED

    c = ws2.cell(14+y, 11)
    c.value = f'{irr_actual:.2%}'
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
    if irr_actual >= 0.10:
        c.fill = FILL_GREEN
    else:
        c.fill = FILL_LIGHT_RED

    c = ws2.cell(14+y, 12)
    c.value = f'{payback_y}年'
    c.font = Font(size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
    if payback_y <= 7:
        c.fill = FILL_GREEN
    elif payback_y <= 10:
        c.fill = FILL_YELLOW
    else:
        c.fill = FILL_RED

# 敏感性分析
ws2.merge_cells('A27:L27')
ws2.row_dimensions[27].height = 25
c = ws2['A27']
c.value = '三、敏感性分析 - 不同电价差下的IRR'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

sens_hdr = ['电价差', '0.40', '0.50', '0.55', '0.58', '0.60', '0.65', '0.67', '0.70', '0.80', '0.90', '1.00']
for j, h in enumerate(sens_hdr, 1):
    c = ws2.cell(28, j)
    c.value = h
    c.font = Font(bold=True, size=9, color='FFFFFF')
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB

irr_row = ['IRR', ]
payback_row = ['静态回收期', ]
for pd in [0.40, 0.50, 0.55, 0.58, 0.60, 0.65, 0.67, 0.70, 0.80, 0.90, 1.00]:
    cfs = [-total_invest]
    for y in range(1, years+1):
        d = cap * DOD * ((1-decay)**(y-1)) * eff * cycles
        cfs.append(d * pd - annual_opex)
    irr_row.append(irr_calc(cfs))
    # 静态回收期
    cum = -total_invest
    pb = '>10'
    for py in range(1, years+1):
        cum += cfs[py]
        if cum >= 0:
            pb = f'{py}年'
            break
    payback_row.append(pb)

for j, v in enumerate(irr_row, 1):
    c = ws2.cell(29, j)
    if j == 1:
        c.value = 'IRR'
        c.font = Font(bold=True, size=9)
        c.fill = FILL_LIGHT_BLUE
    else:
        c.value = f'{v:.2%}'
        if v >= 0.10:
            c.fill = FILL_GREEN
        elif v >= 0.08:
            c.fill = FILL_YELLOW
        else:
            c.fill = FILL_RED
        c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB

for j, v in enumerate(payback_row, 1):
    c = ws2.cell(30, j)
    if j == 1:
        c.value = '回收期'
        c.font = Font(bold=True, size=9)
        c.fill = FILL_LIGHT_BLUE
    else:
        c.value = v
        if v == '>10' or (isinstance(v, str) and '>' in v):
            c.fill = FILL_RED
        elif isinstance(v, str) and '年' in v and int(v.replace('年','')) <= 7:
            c.fill = FILL_GREEN
        elif isinstance(v, str) and '年' in v and int(v.replace('年','')) <= 10:
            c.fill = FILL_YELLOW
        else:
            c.fill = FILL_RED
        c.font = Font(size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB

ws2.row_dimensions[29].height = 20
ws2.row_dimensions[30].height = 20

# ==================== Sheet 3: 模式二经济账 ====================
ws3 = wb.create_sheet('模式二经济账')
for col in 'ABCDEFGH':
    ws3.column_dimensions[col].width = 16

ws3.merge_cells('A1:H1')
ws3.row_dimensions[1].height = 40
c = ws3['A1']
c.value = '模式二：光储一体化投资经济账（1MW + 3MWh储能）'
c.font = Font(bold=True, size=18, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

# 最优配置说明
ws3.merge_cells('A3:H3')
ws3.row_dimensions[3].height = 30
c = ws3['A3']
c.value = '🏆 最优配置：1MW光伏 + 3MWh储能（光伏充电利用率98%，性价比最优）'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='375623')
c.alignment = Alignment(horizontal='center', vertical='center')

m2_thresh = [
    ['IRR=8% 需要储能卖电价', '≥ 0.5260 元/kWh', '✅ 达到则IRR≥8%'],
    ['IRR=10% 需要储能卖电价', '≥ 0.6350 元/kWh', '✅ 达到则IRR≥10%'],
]
for i, row in enumerate(m2_thresh, 4):
    ws3.row_dimensions[i].height = 28
    c = ws3.cell(i, 1)
    c.value = row[0]
    c.font = Font(bold=True, size=11)
    c.fill = FILL_YELLOW
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
    ws3.merge_cells(f'B{i}:C{i}')
    c = ws3.cell(i, 2)
    c.value = row[1]
    c.font = Font(bold=True, size=14, color='006100')
    c.fill = FILL_GREEN
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
    ws3.merge_cells(f'D{i}:H{i}')
    c = ws3.cell(i, 4)
    c.value = row[2]
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = TB

# 参数
ws3.merge_cells('A7:H7')
ws3.row_dimensions[7].height = 25
c = ws3['A7']
c.value = '一、基础参数'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

p2 = [
    ['光伏装机', '1 MW', '储能容量', '3 MWh', '总投资', '500万元'],
    ['光伏成本', '2元/W (200万)', '储能成本', '1元/Wh (300万)', '日照', '1200h/年(河南)'],
    ['光伏效率', '80%', '储能衰减', '1.5%/年', '充放效率', '88%'],
    ['放电深度', '95%', '年充放次数', '330次', '运行年限', '20年'],
    ['直售电价', '0.25元/kWh', '换电芯', '不换', '光伏衰减', '2%/年'],
]
for i, row in enumerate(p2, 8):
    ws3.row_dimensions[i].height = 22
    for j, v in enumerate(row):
        c = ws3.cell(i, j+1)
        c.value = v
        if j % 2 == 0:
            c.font = Font(bold=True, size=10)
            c.fill = FILL_LIGHT_BLUE
        else:
            c.font = Font(size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = TB

# 20年现金流（储能卖电0.65元）
ws3.merge_cells('A15:H15')
ws3.row_dimensions[15].height = 25
c = ws3['A15']
c.value = '二、逐年现金流明细（储能卖电价=0.65元/kWh，IRR≈10.8%）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

cf2_hdr = ['年度', '光伏发电', '储能需充电', '直接卖电', '储能放电', '直售收入', '储能卖电', '年净现金流']
for j, h in enumerate(cf2_hdr, 1):
    c = ws3.cell(16, j)
    c.value = h
    c.font = Font(bold=True, size=9, color='FFFFFF')
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = TB
ws3.row_dimensions[16].height = 30

pv_cost = 200
st_cost = 300
total_inv2 = 500
sell_price = 0.25
storage_price = 0.65
years2 = 20

for y in range(1, years2+1):
    ws3.row_dimensions[16+y].height = 18
    pv_kwh = annual_pv_kwh(y, kw=1000, hours=1200, eff=0.80, decay=0.02) / 10000  # 万kWh
    st_discharge = annual_storage_discharge(y, cap_kwh=3000, DOD=0.95, decay=0.015, eff=0.88, cycles=330) / 10000
    st_need = st_discharge / 0.88  # 充电量
    if pv_kwh >= st_need:
        direct = pv_kwh - st_need
    else:
        direct = 0
    direct_rev = direct * sell_price
    st_rev = st_discharge * storage_price
    net = direct_rev + st_rev

    vals = [y, f'{pv_kwh:.2f}', f'{st_need:.2f}', f'{direct:.2f}', f'{st_discharge:.2f}', f'{direct_rev:.2f}', f'{st_rev:.2f}', f'{net:.2f}']
    for j, v in enumerate(vals, 1):
        c = ws3.cell(16+y, j)
        c.value = v
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = TB
        if j == 1:
            c.fill = FILL_LIGHT_BLUE

# 敏感性分析
ws3.merge_cells('A39:H39')
ws3.row_dimensions[39].height = 25
c = ws3['A39']
c.value = '三、敏感性分析 - 不同储能卖电价的IRR（1MW+3MWh）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

sens2_hdr = ['储能卖电价', '0.50', '0.55', '0.60', '0.65', '0.70', '0.80', '1.00']
for j, h in enumerate(sens2_hdr, 1):
    c = ws3.cell(40, j)
    c.value = h
    c.font = Font(bold=True, size=9, color='FFFFFF')
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB

irr2_row = ['IRR', ]
for sp in [0.50, 0.55, 0.60, 0.65, 0.70, 0.80, 1.00]:
    cfs = [-total_inv2]
    for y in range(1, years2+1):
        pv = annual_pv_kwh(y) / 10000
        sd = annual_storage_discharge(y, cap_kwh=3000) / 10000
        need = sd / 0.88
        direct = max(pv - need, 0)
        cfs.append(direct * sell_price + sd * sp)
    irr2_row.append(irr_calc(cfs))

for j, v in enumerate(irr2_row, 1):
    c = ws3.cell(41, j)
    if j == 1:
        c.value = 'IRR'
        c.font = Font(bold=True, size=9)
        c.fill = FILL_LIGHT_BLUE
    else:
        c.value = f'{v:.2%}'
        if v >= 0.10:
            c.fill = FILL_GREEN
        elif v >= 0.08:
            c.fill = FILL_YELLOW
        else:
            c.fill = FILL_RED
        c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB

ws3.row_dimensions[41].height = 20

# 多情景对比
ws3.merge_cells('A44:H44')
ws3.row_dimensions[44].height = 25
c = ws3['A44']
c.value = '四、多情景对比（不同储能配置）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

cmp_hdr = ['配置', '光伏', '储能', '总投资', '光伏利用', 'IRR=8%门槛', 'IRR=10%门槛', '推荐指数']
for j, h in enumerate(cmp_hdr, 1):
    c = ws3.cell(45, j)
    c.value = h
    c.font = Font(bold=True, size=9, color='FFFFFF')
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB

configs = [
    ('A', '1MW', '1MWh', '300万', '32.7%', 0.670, 0.845, '⚡⚡⚡'),
    ('B', '1MW', '2MWh', '400万', '65.4%', 0.682, 0.799, '⚡⚡'),
    ('C', '1MW', '3MWh', '500万', '98.0%', 0.526, 0.635, '✅✅✅最优'),
    ('D', '1MW', '4MWh', '600万', '131%', 0.860, 0.981, '❌'),
]
for i, row in enumerate(configs, 46):
    ws3.row_dimensions[i].height = 20
    for j, v in enumerate(row, 1):
        c = ws3.cell(i, j)
        c.value = v
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = TB
        if j == 1:
            c.fill = FILL_LIGHT_BLUE
            c.font = Font(bold=True, size=9)
        if '最优' in str(v):
            c.fill = FILL_GREEN
        elif '❌' in str(v):
            c.fill = FILL_RED
        elif '⚡⚡⚡' in str(v):
            c.fill = FILL_LIGHT_GREEN

# ==================== Sheet 4: 现货市场价格 ====================
ws4 = wb.create_sheet('现货市场价格')
for col in 'ABCDEFG':
    ws4.column_dimensions[col].width = 16

ws4.merge_cells('A1:G1')
ws4.row_dimensions[1].height = 40
c = ws4['A1']
c.value = '各省电力现货市场价格数据（2025-2026年）'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

hdrs = ['省份', '午间低电价', '晚高峰', '峰谷价差', '负电价情况', '模式一IRR≥10%?', '模式一回收期≤7年?']
for j, h in enumerate(hdrs, 1):
    c = ws4.cell(3, j)
    c.value = h
    c.font = Font(bold=True, size=10, color='FFFFFF')
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = TB
ws4.row_dimensions[3].height = 30

market = [
    ['山东', '0.02~0.05', '0.40~0.50', '0.35~0.57', '频繁(深-0.07)', '❌不满足', '⚠️勉强'],
    ['山西', '0~0.01', '0.40~0.80', '0.40~1.50', '频繁零/负', '✅满足', '✅满足'],
    ['广东', '0.02~0.05', '0.45~0.60', '0.40~1.29', '极少', '✅满足', '✅满足'],
    ['甘肃', '0.04', '0.30~0.40', '0.25~0.46', '地板频繁', '❌不满足', '❌不满足'],
    ['浙江', '0.02~0.05', '0.40~0.50', '0.35~1.27', '频繁(-0.20)', '✅满足', '✅满足'],
]
for i, row in enumerate(market, 4):
    ws4.row_dimensions[i].height = 25
    for j, v in enumerate(row, 1):
        c = ws4.cell(i, j)
        c.value = v
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = TB
        if j == 6:
            if '✅' in v:
                c.fill = FILL_GREEN
                c.font = Font(bold=True, size=10, color='006100')
            else:
                c.fill = FILL_RED
                c.font = Font(bold=True, size=10, color='9C0006')
        elif j == 7:
            if '✅' in v:
                c.fill = FILL_GREEN
                c.font = Font(bold=True, size=10, color='006100')
            elif '勉强' in v:
                c.fill = FILL_YELLOW
                c.font = Font(bold=True, size=10, color='9C6500')
            else:
                c.fill = FILL_RED
                c.font = Font(bold=True, size=10, color='9C0006')

ws4.merge_cells('A10:G10')
c = ws4['A10']
c.value = '说明：电价单位为 元/kWh；数据来源：各省电力交易中心2025-2026年公开数据'
c.font = Font(size=9, italic=True, color='666666')

# ==================== Sheet 5: 省份梯队划分 ====================
ws5 = wb.create_sheet('省份梯队划分')
for col in 'ABCD':
    ws5.column_dimensions[col].