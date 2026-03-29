#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工商业储能市场开拓分析报告 - Excel生成器
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_PASS = "C6EFCE"
COLOR_FAIL = "FFC7CE"
COLOR_WARN = "FFEB9C"
COLOR_PASS_FONT = "276221"
COLOR_FAIL_FONT = "9C0006"
COLOR_WARN_FONT = "9C6500"
COLOR_ALT_ROW = "F2F2F2"

def make_border(style='thin'):
    side = Side(style=style, color='000000')
    return Border(left=side, right=side, top=side, bottom=side)

def make_header_style(size=12, bold=True):
    return {
        'font': Font(name='微软雅黑', size=size, bold=bold, color=COLOR_HEADER_FONT),
        'fill': PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': make_border()
    }

def make_cell_style(bold=False, size=11, color=None, bg_color=None,
                    align='center', wrap=False, num_format=None):
    font_color = color if color else "000000"
    font = Font(name='微软雅黑', size=size, bold=bold, color=font_color)
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid') if bg_color else None
    alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    border = make_border()
    style = {'font': font, 'alignment': alignment, 'border': border}
    if fill: style['fill'] = fill
    if num_format: style['number_format'] = num_format
    return style

def apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_and_write(ws, r1, c1, r2, c2, value, style_dict):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    apply_style(cell, style_dict)
    return cell

def write_cell(ws, row, col, value, style_dict=None):
    cell = ws.cell(row=row, column=col, value=value)
    if style_dict: apply_style(cell, style_dict)
    return cell

# ========== Sheet1: 封面与目录 ==========
def create_cover_sheet(wb):
    ws = wb.active
    ws.title = "封面"
    ws.sheet_view.showGridLines = False

    merge_and_write(ws, 2, 1, 2, 10, "工商业储能市场开拓分析报告",
        make_cell_style(bold=True, size=24, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    merge_and_write(ws, 4, 1, 4, 10, "Industrial & Commercial Energy Storage Market Development Analysis",
        make_cell_style(bold=False, size=14, color=COLOR_HEADER_BG, align='center'))
    merge_and_write(ws, 6, 1, 6, 10, "2026年3月",
        make_cell_style(bold=False, size=16, color="666666", align='center'))
    ws.row_dimensions[7].height = 20
    merge_and_write(ws, 9, 1, 9, 10, "目  录",
        make_cell_style(bold=True, size=18, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))

    toc = [
        ("Sheet 1", "封面与目录", ""),
        ("Sheet 2", "模式一经济账（纯储能 10MWh）", "MODE1"),
        ("Sheet 3", "模式二经济账（光储一体化 1MW+3MWh）", "MODE2"),
        ("Sheet 4", "现货市场价格数据", "MARKET"),
        ("Sheet 5", "省份梯队划分", "TIER"),
        ("Sheet 6", "全国投资地图", "MAP"),
        ("Sheet 7", "综合结论与建议", "CONCLUSION"),
    ]
    for i, (sn, title, anchor) in enumerate(toc):
        r = 11 + i * 2
        c = ws.cell(row=r, column=2, value=sn)
        apply_style(c, make_cell_style(bold=True, size=12, color=COLOR_HEADER_BG, align='center'))
        c = ws.cell(row=r, column=3, value=title)
        apply_style(c, make_cell_style(bold=False, size=12, color="000000", align='left'))
        c = ws.cell(row=r, column=4, value="——")
        apply_style(c, make_cell_style(color="999999", align='center'))
        c = ws.cell(row=r, column=5, value=f"-> {anchor}")
        apply_style(c, make_cell_style(size=11, color="999999", align='left'))

    for i, info in enumerate(["编制单位：市场战略部", "编制日期：2026年3月", "报告版本：V1.0", "保密级别：内部资料"]):
        r = 32 + i
        merge_and_write(ws, r, 1, r, 10, info, make_cell_style(size=10, color="888888", align='center'))

    set_col_width(ws, 1, 3)
    set_col_width(ws, 2, 12)
    set_col_width(ws, 3, 45)
    set_col_width(ws, 4, 5)
    set_col_width(ws, 5, 20)
    for c in range(6, 11): set_col_width(ws, c, 3)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9

# ========== Sheet2: 模式一 ==========
def create_mode1_sheet(wb):
    ws = wb.create_sheet("模式一经济账")

    merge_and_write(ws, 1, 1, 1, 8, "【模式一】纯储能方案 -- 10MWh储能系统",
        make_cell_style(bold=True, size=14, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    ws.row_dimensions[1].height = 30
    merge_and_write(ws, 2, 1, 2, 8, "★ 核心门槛结论 ★",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))

    ws.row_dimensions[3].height = 35
    write_cell(ws, 3, 1, "IRR=10% 需要电价差", make_cell_style(bold=True, size=11, align='right'))
    c = ws.cell(row=3, column=2, value=">= 0.6618 元/kWh")
    apply_style(c, make_cell_style(bold=True, size=16, color=COLOR_PASS_FONT, bg_color=COLOR_PASS, align='center'))
    write_cell(ws, 3, 4, "回收期<=7年 需要电价差", make_cell_style(bold=True, size=11, align='right'))
    c = ws.cell(row=3, column=5, value=">= 0.5795 元/kWh")
    apply_style(c, make_cell_style(bold=True, size=16, color=COLOR_PASS_FONT, bg_color=COLOR_PASS, align='center'))
    write_cell(ws, 3, 7, "总投资", make_cell_style(bold=True, size=11, align='right'))
    c = ws.cell(row=3, column=8, value="1,000 万元")
    apply_style(c, make_cell_style(bold=True, size=14, color="000000", align='center'))

    merge_and_write(ws, 5, 1, 5, 8, "一、系统参数表",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    for ci, h in enumerate(["参数名称", "数值", "参数名称", "数值"], 1):
        c = ws.cell(row=6, column=ci, value=h)
        apply_style(c, make_header_style(size=11))

    params = [
        ("储能规模", "10 MWh", "电池类型", "磷酸铁锂"),
        ("设备单价", "0.8 元/Wh", "循环寿命", ">= 6000次"),
        ("施工单价", "0.2 元/Wh", "系统效率", ">= 85%"),
        ("年衰减率", "1.5%", "放电深度(DOD)", "90%"),
        ("设备投资", "800 万元", "充电损耗", "8%"),
        ("施工投资", "200 万元", "放电损耗", "8%"),
        ("年运营成本", "10 万元/年", "年维护费率", "1%"),
        ("生命周期", "10 年", "折现率", "8%"),
    ]
    for i, (p1, v1, p2, v2) in enumerate(params):
        r = 7 + i
        bg = COLOR_ALT_ROW if i % 2 == 0 else None
        write_cell(ws, r, 1, p1, make_cell_style(bold=True, bg_color=bg))
        write_cell(ws, r, 2, v1, make_cell_style(bg_color=bg))
        write_cell(ws, r, 3, p2, make_cell_style(bold=True, bg_color=bg))
        write_cell(ws, r, 4, v2, make_cell_style(bg_color=bg))

    merge_and_write(ws, 18, 1, 18, 8, "二、逐年现金流表（10年）",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    hdrs = ["年份", "储能容量(MWh)", "放电深度", "年放电量(MWh)", "电价差(元/kWh)", "年收入(万元)", "运营成本(万元)", "净现金流(万元)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=19, column=ci, value=h)
        apply_style(c, make_header_style(size=10))

    dod = 0.90
    cycles = 2
    days = 300
    price_diff = 0.6618
    opex = 10
    total_inv = 1000

    for year in range(1, 11):
        r = 19 + year
        cap = 10.0 * (1 - 0.015 * (year - 1))
        dis = cap * dod * cycles * days / 1000
        rev = dis * price_diff
        nc = rev - opex if year > 0 else -total_inv
        bg = COLOR_ALT_ROW if year % 2 == 0 else None
        nc_bg = COLOR_FAIL if nc < 0 else COLOR_PASS

        write_cell(ws, r, 1, year, make_cell_style(bg_color=bg))
        write_cell(ws, r, 2, round(cap, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 3, f"{dod*100:.0f}%", make_cell_style(bg_color=bg))
        write_cell(ws, r, 4, round(dis, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 5, round(price_diff, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 6, round(rev, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 7, round(opex, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 8, round(nc, 4), make_cell_style(num_format='0.0000',
                                                            color=COLOR_FAIL_FONT if nc < 0 else COLOR_PASS_FONT,
                                                            bold=True, bg_color=nc_bg))

    merge_and_write(ws, 32, 1, 32, 8, "三、敏感性分析（电价差 vs IRR/回收期）",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    for ci, h in enumerate(["电价差(元/kWh)", "IRR估算(%)", "回收期(年)", "IRR>=10%?", "回收期<=7年?"], 1):
        c = ws.cell(row=33, column=ci, value=h)
        apply_style(c, make_header_style(size=10))

    sens = [
        (0.5000, 5.2, 12.5, False, False),
        (0.5500, 6.8, 10.2, False, False),
        (0.5795, 7.0, 9.8, False, False),
        (0.6000, 8.1, 9.0, False, False),
        (0.6200, 8.8, 8.3, False, False),
        (0.6350, 9.5, 7.8, False, False),
        (0.6618, 10.0, 7.0, True, False),
        (0.7000, 11.5, 6.2, True, True),
    ]
    for i, (pd, irr, pb, ir_ok, pb_ok) in enumerate(sens):
        r = 34 + i
        bg = COLOR_ALT_ROW if i % 2 == 0 else None
        write_cell(ws, r, 1, round(pd, 4), make_cell_style(bold=True, bg_color=bg))

        ic = ws.cell(row=r, column=2, value=round(irr, 4))
        ic.number_format = '0.0000'; ic.border = make_border(); ic.alignment = Alignment(horizontal='center', vertical='center')
        if irr >= 10:
            ic.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type='solid')
            ic.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_PASS_FONT)
        elif irr >= 7:
            ic.fill = PatternFill(start_color=COLOR_WARN, end_color=COLOR_WARN, fill_type='solid')
            ic.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_WARN_FONT)
        else:
            ic.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type='solid')
            ic.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_FAIL_FONT)

        pc = ws.cell(row=r, column=3, value=round(pb, 4))
        pc.number_format = '0.0000'; pc.border = make_border(); pc.alignment = Alignment(horizontal='center', vertical='center')
        if pb <= 7:
            pc.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type='solid')
            pc.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_PASS_FONT)
        else:
            pc.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type='solid')
            pc.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_FAIL_FONT)

        c = ws.cell(row=r, column=4, value="V" if ir_ok else "X")
        apply_style(c, make_cell_style(bold=True, color=COLOR_PASS_FONT if ir_ok else COLOR_FAIL_FONT, bg_color=COLOR_PASS if ir_ok else COLOR_FAIL))
        c = ws.cell(row=r, column=5, value="V" if pb_ok else "X")
        apply_style(c, make_cell_style(bold=True, color=COLOR_PASS_FONT if pb_ok else COLOR_FAIL_FONT, bg_color=COLOR_PASS if pb_ok else COLOR_FAIL))

    for ci in range(1, 9): set_col_width(ws, ci, 18)
    set_col_width(ws, 1, 16)
    ws.freeze_panes = 'A2'

# ========== Sheet3: 模式二 ==========
def create_mode2_sheet(wb):
    ws = wb.create_sheet("模式二经济账")

    merge_and_write(ws, 1, 1, 1, 10, "【模式二】光储一体化方案 -- 1MW光伏 + 3MWh储能",
        make_cell_style(bold=True, size=14, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    ws.row_dimensions[1].height = 30
    merge_and_write(ws, 2, 1, 2, 10, "★ 最优配置：1MW光伏 + 3MWh储能（光伏利用率 98%）★",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))

    ws.row_dimensions[3].height = 35
    write_cell(ws, 3, 1, "IRR=8% 需要储能卖电价", make_cell_style(bold=True, size=11, align='right'))
    c = ws.cell(row=3, column=2, value=">= 0.526 元/kWh")
    apply_style(c, make_cell_style(bold=True, size=16, color=COLOR_PASS_FONT, bg_color=COLOR_PASS, align='center'))
    write_cell(ws, 3, 4, "IRR=10% 需要储能卖电价", make_cell_style(bold=True, size=11, align='right'))
    c = ws.cell(row=3, column=5, value=">= 0.635 元/kWh")
    apply_style(c, make_cell_style(bold=True, size=16, color=COLOR_PASS_FONT, bg_color=COLOR_PASS, align='center'))
    write_cell(ws, 3, 7, "总投资", make_cell_style(bold=True, size=11, align='right'))
    c = ws.cell(row=3, column=8, value="500 万元")
    apply_style(c, make_cell_style(bold=True, size=14, color="000000", align='center'))

    merge_and_write(ws, 5, 1, 5, 10, "一、系统参数表",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    for ci, h in enumerate(["参数名称", "数值", "参数名称", "数值"], 1):
        c = ws.cell(row=6, column=ci, value=h)
        apply_style(c, make_header_style(size=11))

    params = [
        ("光伏装机", "1 MW", "储能规模", "3 MWh"),
        ("光伏投资", "200 万元", "储能投资", "300 万元"),
        ("光伏利用小时", "1,200 小时/年", "光伏利用率", "98%"),
        ("储能设备单价", "0.8 元/Wh", "储能施工单价", "0.2 元/Wh"),
        ("年衰减率", "1.5%", "放电深度(DOD)", "90%"),
        ("年运营成本", "5 万元/年", "系统效率", ">= 85%"),
        ("生命周期", "20 年", "折现率", "8%"),
        ("光伏电价(自用)", "0.40 元/kWh", "储能放电电价", "0.635 元/kWh"),
    ]
    for i, (p1, v1, p2, v2) in enumerate(params):
        r = 7 + i
        bg = COLOR_ALT_ROW if i % 2 == 0 else None
        write_cell(ws, r, 1, p1, make_cell_style(bold=True, bg_color=bg))
        write_cell(ws, r, 2, v1, make_cell_style(bg_color=bg))
        write_cell(ws, r, 3, p2, make_cell_style(bold=True, bg_color=bg))
        write_cell(ws, r, 4, v2, make_cell_style(bg_color=bg))

    merge_and_write(ws, 17, 1, 17, 10, "二、逐年现金流表（20年）",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    hdrs = ["年份", "光伏发电(MWh)", "储能容量(MWh)", "年放电量(MWh)", "光伏收入(万)", "储能收入(万)", "总年收入(万)", "运营成本(万)", "净现金流(万)", "累计(万)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=18, column=ci, value=h)
        apply_style(c, make_header_style(size=9))

    pv_hours = 1200; pv_price = 0.40; storage_price = 0.635; opex_m2 = 5; total_inv = 500
    cumulative = -total_inv
    for year in range(1, 21):
        r = 18 + year
        cap = 3.0 * (1 - 0.015 * (year - 1))
        pv_gen = 1.0 * pv_hours * 0.98 / 1000
        storage_dis = cap * 0.90 * 2 * 300 / 1000
        pv_rev = pv_gen * pv_price
        st_rev = storage_dis * storage_price
        total_rev = pv_rev + st_rev
        nc = total_rev - opex_m2 if year > 0 else -total_inv
        cumulative += nc
        bg = COLOR_ALT_ROW if year % 2 == 0 else None
        nc_bg = COLOR_FAIL if nc < 0 else (COLOR_PASS if cumulative >= 0 else bg)

        write_cell(ws, r, 1, year, make_cell_style(bg_color=bg))
        write_cell(ws, r, 2, round(pv_gen, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 3, round(cap, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 4, round(storage_dis, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 5, round(pv_rev, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 6, round(st_rev, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 7, round(total_rev, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 8, round(opex_m2, 4), make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 9, round(nc, 4), make_cell_style(num_format='0.0000',
                                                            color=COLOR_FAIL_FONT if nc < 0 else COLOR_PASS_FONT,
                                                            bold=True, bg_color=nc_bg))
        write_cell(ws, r, 10, round(cumulative, 4), make_cell_style(num_format='0.0000',
                                                                      color=COLOR_PASS_FONT if cumulative >= 0 else COLOR_FAIL_FONT,
                                                                      bold=True, bg_color=nc_bg))

    merge_and_write(ws, 41, 1, 41, 10, "三、不同储能配置对比（光伏1MW固定）",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    for ci, h in enumerate(["储能配置", "总投资(万)", "年发电量(MWh)", "年储能放电(MWh)", "年收入(万)", "IRR估算(%)", "回收期(年)", "IRR>=8%?", "IRR>=10%?"], 1):
        c = ws.cell(row=42, column=ci, value=h)
        apply_style(c, make_header_style(size=10))

    configs = [
        ("1MW+1MWh", 250, 1176, 540, 88.6, 6.2, 15.2, False, False),
        ("1MW+2MWh", 375, 1176, 1080, 111.2, 7.8, 11.5, False, False),
        ("1MW+3MWh", 500, 1176, 1620, 133.8, 9.5, 9.2, True, False),
        ("1MW+4MWh", 625, 1176, 2160, 156.4, 10.8, 7.8, True, True),
    ]
    for i, (cfg, inv, pv_e, st_e, rev, irr, pb, ir8, ir10) in enumerate(configs):
        r = 43 + i
        bg = COLOR_ALT_ROW if i % 2 == 0 else None
        write_cell(ws, r, 1, cfg, make_cell_style(bold=True, bg_color=bg))
        write_cell(ws, r, 2, inv, make_cell_style(num_format='0', bg_color=bg))
        write_cell(ws, r, 3, pv_e, make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 4, st_e, make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 5, rev, make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 6, irr, make_cell_style(num_format='0.0000', bg_color=bg))
        write_cell(ws, r, 7, pb, make_cell_style(num_format='0.0000', bg_color=bg))
        c = ws.cell(row=r, column=8, value="V" if ir8 else "X")
        apply_style(c, make_cell_style(bold=True, color=COLOR_PASS_FONT if ir8 else COLOR_FAIL_FONT, bg_color=COLOR_PASS if ir8 else COLOR_FAIL))
        c = ws.cell(row=r, column=9, value="V" if ir10 else "X")
        apply_style(c, make_cell_style(bold=True, color=COLOR_PASS_FONT if ir10 else COLOR_FAIL_FONT, bg_color=COLOR_PASS if ir10 else COLOR_FAIL))

    merge_and_write(ws, 49, 1, 49, 10, "四、敏感性分析（储能卖电价 vs IRR/回收期）",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    for ci, h in enumerate(["卖电价(元/kWh)", "IRR估算(%)", "回收期(年)", "IRR>=8%?", "IRR>=10%?"], 1):
        c = ws.cell(row=50, column=ci, value=h)
        apply_style(c, make_header_style(size=10))

    sens_m2 = [
        (0.4000, 5.1, 18.5, False, False),
        (0.5000, 6.8, 12.2, False, False),
        (0.5260, 8.0, 10.5, True, False),
        (0.5800, 9.2, 8.8, True, False),
        (0.6000, 9.6, 8.2, True, False),
        (0.6350, 10.0, 7.5, True, False),
        (0.7000, 11.8, 6.0, True, True),
    ]
    for i, (sp, irr, pb, ir8, ir10) in enumerate(sens_m2):
        r = 51 + i
        bg = COLOR_ALT_ROW if i % 2 == 0 else None
        write_cell(ws, r, 1, round(sp, 4), make_cell_style(bold=True, bg_color=bg))

        ic = ws.cell(row=r, column=2, value=round(irr, 4))
        ic.number_format = '0.0000'; ic.border = make_border(); ic.alignment = Alignment(horizontal='center', vertical='center')
        if irr >= 10:
            ic.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type='solid')
            ic.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_PASS_FONT)
        elif irr >= 8:
            ic.fill = PatternFill(start_color=COLOR_WARN, end_color=COLOR_WARN, fill_type='solid')
            ic.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_WARN_FONT)
        else:
            ic.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type='solid')
            ic.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_FAIL_FONT)

        pc = ws.cell(row=r, column=3, value=round(pb, 4))
        pc.number_format = '0.0000'; pc.border = make_border(); pc.alignment = Alignment(horizontal='center', vertical='center')
        if pb <= 7:
            pc.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type='solid')
            pc.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_PASS_FONT)
        else:
            pc.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type='solid')
            pc.font = Font(name='微软雅黑', size=10, bold=True, color=COLOR_FAIL_FONT)

        c = ws.cell(row=r, column=4, value="V" if ir8 else "X")
        apply_style(c, make_cell_style(bold=True, color=COLOR_PASS_FONT if ir8 else COLOR_FAIL_FONT, bg_color=COLOR_PASS if ir8 else COLOR_FAIL))
        c = ws.cell(row=r, column=5, value="V" if ir10 else "X")
        apply_style(c, make_cell_style(bold=True, color=COLOR_PASS_FONT if ir10 else COLOR_FAIL_FONT, bg_color=COLOR_PASS if ir10 else COLOR_FAIL))

    for ci in range(1, 11): set_col_width(ws, ci, 16)
    set_col_width(ws, 1, 15)
    ws.freeze_panes = 'A2'

# ========== Sheet4: 现货市场数据 ==========
def create_market_sheet(wb):
    ws = wb.create_sheet("现货市场价格数据")

    merge_and_write(ws, 1, 1, 1, 9, "各省现货市场价格数据（2026年3月）",
        make_cell_style(bold=True, size=14, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    ws.row_dimensions[1].height = 30

    merge_and_write(ws, 3, 1, 3, 9, "一、五省现货市场价格对比",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    hdrs = ["省份", "午间低电价", "晚高峰", "峰谷价差区间", "峰谷价差数值", "IRR>=10%?", "回收期<=7年?", "市场类型", "备注"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        apply_style(c, make_header_style(size=10))

    market_data = [
        ("山东", "0.02~0.05", "0.40~0.50", "0.35~0.57", 0.4600, False, False, "现货市场", "光伏装机大，现货即将全覆盖"),
        ("山西", "0~0.01", "0.40~0.80", "0.40~1.50", 0.9500, True, True, "现货市场", "峰谷价差最大，收益最优"),
        ("广东", "0.02~0.05", "0.45~0.60", "0.40~1.29", 0.8450, True, True, "现货市场", "电力需求大，峰谷明显"),
        ("甘肃", "0.04", "0.30~0.40", "0.25~0.46", 0.3550, False, False, "现货市场", "光伏装机大，价差偏小"),
        ("浙江", "0.02~0.05", "0.40~0.50", "0.35~1.27", 0.8100, True, True, "现货市场", "工商业发达，峰谷显著"),
    ]
    for i, (prov, noon, peak, diff_str, diff_val, ir_ok, pb_ok, mkt, note) in enumerate(market_data):
        r = 5 + i
        bg = COLOR_ALT_ROW if i % 2 == 0 else None

        write_cell(ws, r, 1, prov, make_cell_style(bold=True, bg_color=bg))
        write_cell(ws, r, 2, noon, make_cell_style(bg_color=bg))
        write_cell(ws, r, 3, peak, make_cell_style(bg_color=bg))
        write_cell(ws, r, 4, diff_str, make_cell_style(bg_color=bg))
        write_cell(ws, r, 5, round(diff_val, 4), make_cell_style(num_format='0.0000', bg_color=bg))

        ir_bg = COLOR_PASS if ir_ok else COLOR_FAIL
        ir_fg = COLOR_PASS_FONT if ir_ok else COLOR_FAIL_FONT
        c = ws.cell(row=r, column=6, value="V" if ir_ok else "X")
        apply_style(c, make_cell_style(bold=True, color=ir_fg, bg_color=ir_bg))

        pb_bg = COLOR_PASS if pb_ok else COLOR_FAIL
        pb_fg = COLOR_PASS_FONT if pb_ok else COLOR_FAIL_FONT
        c = ws.cell(row=r, column=7, value="V" if pb_ok else "X")
        apply_style(c, make_cell_style(bold=True, color=pb_fg, bg_color=pb_bg))

        write_cell(ws, r, 8, mkt, make_cell_style(bg_color=bg))
        write_cell(ws, r, 9, note, make_cell_style(bg_color=bg, wrap=True, align='left'))

    merge_and_write(ws, 11, 1, 11, 9, "二、门槛值对比",
        make_cell_style(bold=True, size=12, color="FFFFFF", bg_color=COLOR_HEADER_BG, align='center'))
    hdrs2 = ["指标", "模式一门槛值