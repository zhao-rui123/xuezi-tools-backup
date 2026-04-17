#!/usr/bin/env python3
"""
储能电价循环优化算法
====================
输入：电价时间序列（1小时间隔）
输出：最优的充放电循环组合，总收益最大化

规则：
1. 先充后放（充电时段在放电时段之前）
2. 不允许时间重叠（一个循环的放电结束后，下一个循环的充电才能开始）
3. 允许跨天（今天充电，明天放电）
4. 目标：总收益（放电价-充电价之和）最大的不重叠循环组合

算法设计（加权区间调度 Weighted Interval Scheduling）：
=========================================================
核心思路：
- 每个"循环"= 连续充电时段 + 连续放电时段，中间可以有间隔
- 循环的"收益" = 放电时段平均电价 - 充电时段平均电价
- 循环占用的时间 = [充电开始, 放电结束]
- 问题转化为：从所有可能的循环中，选出不重叠且总收益最大的子集

由于枚举所有可能的(充电段, 放电段)组合数量巨大，我们采用更高效的方法：

**贪心+动态规划混合策略：**
1. 将时间序列按天分组，识别每天的"谷时段"和"峰时段"
2. 生成候选循环：谷时段充电 → 峰时段放电（同天或跨天）
3. 用加权区间调度DP选出最优不重叠组合

**候选循环生成策略：**
- 对每天识别连续的低价时段（充电候选）和高价时段（放电候选）
- 充电候选：价格低于当天均值的连续时段
- 放电候选：价格高于当天均值的连续时段
- 生成所有合法的(充电段i, 放电段j)对，其中充电段结束 <= 放电段开始
- 只保留收益为正的循环
"""

import openpyxl
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import List, Tuple, Optional
import bisect
import sys
import os


@dataclass
class TimeSlot:
    """一个小时的时间槽"""
    index: int          # 全局索引（0-based）
    date: datetime      # 日期
    hour: int           # 小时 (0-23)
    price_da: float     # 日前电价
    price_rt: float     # 实时电价

    @property
    def timestamp(self) -> datetime:
        return self.date.replace(hour=self.hour)

    def __repr__(self):
        return f"[{self.index}] {self.date.strftime('%Y-%m-%d')} {self.hour:02d}:00 DA={self.price_da:.2f} RT={self.price_rt:.2f}"


@dataclass
class CycleCandidate:
    """一个充放电循环候选"""
    charge_slots: List[TimeSlot]      # 充电时段
    discharge_slots: List[TimeSlot]   # 放电时段
    profit: float                      # 收益 (放电均价 - 充电均价) * 时长
    start_index: int                   # 占用的起始全局索引
    end_index: int                     # 占用的结束全局索引

    @property
    def charge_avg_price(self) -> float:
        return sum(s.price_rt for s in self.charge_slots) / len(self.charge_slots)

    @property
    def discharge_avg_price(self) -> float:
        return sum(s.price_rt for s in self.discharge_slots) / len(self.discharge_slots)

    @property
    def charge_hours(self) -> int:
        return len(self.charge_slots)

    @property
    def discharge_hours(self) -> int:
        return len(self.discharge_slots)

    def __repr__(self):
        cs = self.charge_slots[0]
        ce = self.charge_slots[-1]
        ds = self.discharge_slots[0]
        de = self.discharge_slots[-1]
        return (f"充电: {cs.date.strftime('%m-%d')} {cs.hour:02d}:00-{ce.hour+1:02d}:00 "
                f"({self.charge_hours}h, 均价{self.charge_avg_price:.1f}) → "
                f"放电: {ds.date.strftime('%m-%d')} {ds.hour:02d}:00-{de.hour+1:02d}:00 "
                f"({self.discharge_hours}h, 均价{self.discharge_avg_price:.1f}) "
                f"收益={self.profit:.1f}")


def read_excel_prices(filepath: str, price_type: str = "实时") -> List[TimeSlot]:
    """
    读取Excel电价数据，返回按时间排序的TimeSlot列表。

    Excel结构：
    - 多个月份块，每块有header行(日期) + 24行小时数据
    - 每天2列：日前、实时
    - 块之间有空行和星期/类型标题行

    Args:
        filepath: Excel文件路径
        price_type: "日前" 或 "实时"，决定用哪列价格做主价格
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    slots = []
    global_index = 0

    # 找到所有月份块的起始行（通过col A = "时段"来定位）
    block_headers = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "时段":
            block_headers.append(r)

    for hdr_row in block_headers:
        # 解析日期列：从col 2开始，每2列一天（日前、实时）
        col = 2
        day_columns = []  # [(date, da_col, rt_col), ...]

        while col <= ws.max_column:
            date_val = ws.cell(hdr_row, col).value
            if not isinstance(date_val, datetime):
                break
            type_val = ws.cell(hdr_row - 1, col).value  # 上一行是"日前"/"实时"
            if type_val == "日前":
                da_col = col
                rt_col = col + 1
                day_columns.append((date_val, da_col, rt_col))
                col += 2
            elif type_val == "实时":
                # 如果第一列就是实时，跳过（不应该发生）
                col += 1
            else:
                # 可能是合计列
                break

        # 读取24行小时数据
        data_start = hdr_row + 1
        for hour in range(24):
            row = data_start + hour
            time_label = ws.cell(row, 1).value
            if time_label is None:
                continue

            for date_val, da_col, rt_col in day_columns:
                price_da = ws.cell(row, da_col).value
                price_rt = ws.cell(row, rt_col).value

                # 跳过空值
                if price_da is None and price_rt is None:
                    continue

                price_da = float(price_da) if price_da is not None else 0.0
                price_rt = float(price_rt) if price_rt is not None else 0.0

                slot = TimeSlot(
                    index=global_index,
                    date=date_val.replace(hour=0, minute=0, second=0),
                    hour=hour,
                    price_da=price_da,
                    price_rt=price_rt,
                )
                slots.append(slot)
                global_index += 1

    # 按时间排序（日期+小时）
    slots.sort(key=lambda s: (s.date, s.hour))

    # 重新分配全局索引
    for i, s in enumerate(slots):
        s.index = i

    wb.close()
    return slots


def find_contiguous_segments(slots: List[TimeSlot], threshold: float,
                              mode: str = "below") -> List[List[TimeSlot]]:
    """
    找出连续的低价/高价时段。

    Args:
        slots: 一天内的24个时间槽
        threshold: 价格阈值
        mode: "below" = 低于阈值（充电候选），"above" = 高于阈值（放电候选）

    Returns:
        连续时段的列表
    """
    segments = []
    current = []

    for s in slots:
        price = s.price_rt
        if (mode == "below" and price < threshold) or \
           (mode == "above" and price > threshold):
            current.append(s)
        else:
            if current:
                segments.append(current)
                current = []
    if current:
        segments.append(current)

    return segments


def generate_candidates(slots: List[TimeSlot],
                        max_charge_hours: int = 8,
                        max_discharge_hours: int = 8,
                        min_profit_per_hour: float = 0.0,
                        lookforward_days: int = 3) -> List[CycleCandidate]:
    """
    生成所有候选充放电循环。

    策略：
    1. 按天分组
    2. 每天计算均价，识别低价段（充电候选）和高价段（放电候选）
    3. 对每个充电段，在其后的lookforward_days天内寻找放电段配对
    4. 只保留正收益的循环

    Args:
        slots: 全部时间槽
        max_charge_hours: 最大充电时长
        max_discharge_hours: 最大放电时长
        min_profit_per_hour: 每小时最低收益阈值
        lookforward_days: 向前看几天寻找放电机会
    """
    # 按天分组
    days = {}
    for s in slots:
        key = s.date.strftime('%Y-%m-%d')
        if key not in days:
            days[key] = []
        days[key].append(s)

    day_keys = sorted(days.keys())

    # 为每天计算统计量并识别充放电段
    day_charge_segs = {}    # date_key -> [segment, ...]
    day_discharge_segs = {} # date_key -> [segment, ...]

    for dk in day_keys:
        day_slots = sorted(days[dk], key=lambda s: s.hour)
        prices = [s.price_rt for s in day_slots]
        if not prices:
            continue

        avg_price = sum(prices) / len(prices)
        # 使用更激进的阈值：充电用 P25，放电用 P75
        sorted_prices = sorted(prices)
        p25 = sorted_prices[len(sorted_prices) // 4]
        p75 = sorted_prices[3 * len(sorted_prices) // 4]

        charge_segs = find_contiguous_segments(day_slots, p25, "below")
        discharge_segs = find_contiguous_segments(day_slots, p75, "above")

        # 也用均值生成一批，增加候选多样性
        charge_segs_avg = find_contiguous_segments(day_slots, avg_price * 0.85, "below")
        discharge_segs_avg = find_contiguous_segments(day_slots, avg_price * 1.15, "above")

        # 合并去重（按起始index）
        seen_starts = set()
        all_charge = []
        for seg in charge_segs + charge_segs_avg:
            key = seg[0].index
            if key not in seen_starts:
                seen_starts.add(key)
                all_charge.append(seg)

        seen_starts = set()
        all_discharge = []
        for seg in discharge_segs + discharge_segs_avg:
            key = seg[0].index
            if key not in seen_starts:
                seen_starts.add(key)
                all_discharge.append(seg)

        day_charge_segs[dk] = all_charge
        day_discharge_segs[dk] = all_discharge

    # 生成候选循环
    candidates = []

    for ci, charge_dk in enumerate(day_keys):
        for charge_seg in day_charge_segs.get(charge_dk, []):
            # 限制充电时长
            for ch_len in range(1, min(len(charge_seg), max_charge_hours) + 1):
                # 取最便宜的ch_len个连续小时
                # 滑动窗口找最低均价
                best_charge = None
                best_charge_avg = float('inf')
                for start in range(len(charge_seg) - ch_len + 1):
                    window = charge_seg[start:start + ch_len]
                    avg = sum(s.price_rt for s in window) / ch_len
                    if avg < best_charge_avg:
                        best_charge_avg = avg
                        best_charge = window

                if best_charge is None:
                    continue

                charge_end_index = best_charge[-1].index

                # 在当天及后续lookforward_days天内寻找放电段
                for di in range(ci, min(ci + lookforward_days + 1, len(day_keys))):
                    discharge_dk = day_keys[di]
                    for discharge_seg in day_discharge_segs.get(discharge_dk, []):
                        # 放电必须在充电之后
                        valid_discharge = [s for s in discharge_seg if s.index > charge_end_index]
                        if not valid_discharge:
                            continue

                        # 限制放电时长
                        for dh_len in range(1, min(len(valid_discharge), max_discharge_hours) + 1):
                            # 滑动窗口找最高均价
                            best_discharge = None
                            best_discharge_avg = -float('inf')
                            for start in range(len(valid_discharge) - dh_len + 1):
                                window = valid_discharge[start:start + dh_len]
                                avg = sum(s.price_rt for s in window) / dh_len
                                if avg > best_discharge_avg:
                                    best_discharge_avg = avg
                                    best_discharge = window

                            if best_discharge is None:
                                continue

                            # 计算收益：取充放电时长的较小值作为有效时长
                            effective_hours = min(ch_len, dh_len)
                            profit = (best_discharge_avg - best_charge_avg) * effective_hours

                            if profit <= min_profit_per_hour * effective_hours:
                                continue

                            candidate = CycleCandidate(
                                charge_slots=best_charge,
                                discharge_slots=best_discharge,
                                profit=profit,
                                start_index=best_charge[0].index,
                                end_index=best_discharge[-1].index,
                            )
                            candidates.append(candidate)

    # 去重：相同start_index和end_index的只保留收益最高的
    best_by_range = {}
    for c in candidates:
        key = (c.start_index, c.end_index)
        if key not in best_by_range or c.profit > best_by_range[key].profit:
            best_by_range[key] = c

    candidates = list(best_by_range.values())

    # 按结束时间排序（DP需要）
    candidates.sort(key=lambda c: c.end_index)

    return candidates


def weighted_interval_scheduling(candidates: List[CycleCandidate]) -> List[CycleCandidate]:
    """
    加权区间调度算法（动态规划）。

    选出不重叠且总收益最大的循环子集。

    时间复杂度：O(n log n)，n = 候选数量

    算法：
    1. 按end_index排序（已排序）
    2. 对每个候选i，用二分查找找到最后一个不与i冲突的候选p(i)
    3. dp[i] = max(dp[i-1], profit[i] + dp[p(i)])
    """
    if not candidates:
        return []

    n = len(candidates)

    # 提取结束索引用于二分查找
    end_indices = [c.end_index for c in candidates]

    # p[i] = 最后一个end_index < candidates[i].start_index的候选索引
    # 即不与candidates[i]冲突的最近前驱
    def find_last_compatible(i: int) -> int:
        """二分查找：找最大的j < i，使得candidates[j].end_index < candidates[i].start_index"""
        target = candidates[i].start_index
        lo, hi = 0, i - 1
        result = -1
        while lo <= hi:
            mid = (lo + hi) // 2
            if end_indices[mid] < target:
                result = mid
                lo = mid + 1
            else:
                hi = mid - 1
        return result

    # DP
    dp = [0.0] * n
    choice = [False] * n  # 是否选择了第i个候选

    dp[0] = max(0, candidates[0].profit)
    choice[0] = candidates[0].profit > 0

    for i in range(1, n):
        # 不选i
        exclude = dp[i - 1]

        # 选i
        p = find_last_compatible(i)
        include = candidates[i].profit + (dp[p] if p >= 0 else 0)

        if include > exclude:
            dp[i] = include
            choice[i] = True
        else:
            dp[i] = exclude
            choice[i] = False

    # 回溯找出选中的候选
    selected = []
    i = n - 1
    while i >= 0:
        if choice[i]:
            selected.append(candidates[i])
            p = find_last_compatible(i)
            i = p
        else:
            i -= 1

    selected.reverse()
    return selected


def validate_solution(cycles: List[CycleCandidate]) -> bool:
    """验证解的合法性：不重叠、先充后放"""
    for i, c in enumerate(cycles):
        # 充电在放电之前
        if c.charge_slots[-1].index >= c.discharge_slots[0].index:
            print(f"❌ 循环 {i+1}: 充电未在放电之前！")
            return False

        # 与下一个循环不重叠
        if i + 1 < len(cycles):
            next_c = cycles[i + 1]
            if c.end_index >= next_c.start_index:
                print(f"❌ 循环 {i+1} 和 {i+2} 时间重叠！")
                print(f"   循环{i+1} 结束: index={c.end_index}")
                print(f"   循环{i+2} 开始: index={next_c.start_index}")
                return False

    print("✅ 所有循环验证通过：不重叠、先充后放")
    return True


def print_results(cycles: List[CycleCandidate], slots: List[TimeSlot]):
    """打印优化结果"""
    print("=" * 80)
    print("储能电价循环优化结果")
    print("=" * 80)
    print(f"数据范围: {slots[0].date.strftime('%Y-%m-%d')} ~ {slots[-1].date.strftime('%Y-%m-%d')}")
    print(f"总时间槽: {len(slots)} 小时")
    print(f"优化循环数: {len(cycles)}")
    print()

    total_profit = 0
    total_charge_hours = 0
    total_discharge_hours = 0

    for i, c in enumerate(cycles):
        cs = c.charge_slots[0]
        ce = c.charge_slots[-1]
        ds = c.discharge_slots[0]
        de = c.discharge_slots[-1]

        charge_start = f"{cs.date.strftime('%m-%d')} {cs.hour:02d}:00"
        charge_end = f"{ce.date.strftime('%m-%d')} {ce.hour+1:02d}:00"
        discharge_start = f"{ds.date.strftime('%m-%d')} {ds.hour:02d}:00"
        discharge_end = f"{de.date.strftime('%m-%d')} {de.hour+1:02d}:00"

        print(f"循环 {i+1:3d}: "
              f"充电 {charge_start}~{charge_end} ({c.charge_hours}h, "
              f"均价 {c.charge_avg_price:7.1f} 元/MWh) → "
              f"放电 {discharge_start}~{discharge_end} ({c.discharge_hours}h, "
              f"均价 {c.discharge_avg_price:7.1f} 元/MWh) "
              f"| 收益 {c.profit:8.1f} 元/MWh")

        total_profit += c.profit
        total_charge_hours += c.charge_hours
        total_discharge_hours += c.discharge_hours

    print()
    print("-" * 80)
    print(f"总收益:       {total_profit:,.1f} 元/MWh")
    print(f"总充电时长:   {total_charge_hours} 小时")
    print(f"总放电时长:   {total_discharge_hours} 小时")
    print(f"平均每循环收益: {total_profit/len(cycles):,.1f} 元/MWh" if cycles else "")
    print(f"循环数量:     {len(cycles)}")
    print("=" * 80)


def optimize(filepath: str,
             price_type: str = "实时",
             max_charge_hours: int = 4,
             max_discharge_hours: int = 4,
             min_profit_per_hour: float = 50.0,
             lookforward_days: int = 2) -> Tuple[List[CycleCandidate], List[TimeSlot]]:
    """
    主优化函数。

    Args:
        filepath: Excel文件路径
        price_type: "日前" 或 "实时"
        max_charge_hours: 单次最大充电时长
        max_discharge_hours: 单次最大放电时长
        min_profit_per_hour: 每小时最低收益（过滤低价值循环）
        lookforward_days: 跨天寻找放电机会的天数

    Returns:
        (选中的循环列表, 全部时间槽)
    """
    print(f"📖 读取电价数据: {os.path.basename(filepath)}")
    slots = read_excel_prices(filepath, price_type)
    print(f"   共 {len(slots)} 个时间槽")

    # 基本统计
    prices = [s.price_rt for s in slots]
    print(f"   电价范围: {min(prices):.1f} ~ {max(prices):.1f} 元/MWh")
    print(f"   电价均值: {sum(prices)/len(prices):.1f} 元/MWh")
    print()

    print(f"🔍 生成候选循环 (充电≤{max_charge_hours}h, 放电≤{max_discharge_hours}h, "
          f"最低收益≥{min_profit_per_hour}/h, 跨天≤{lookforward_days}天)...")
    candidates = generate_candidates(
        slots,
        max_charge_hours=max_charge_hours,
        max_discharge_hours=max_discharge_hours,
        min_profit_per_hour=min_profit_per_hour,
        lookforward_days=lookforward_days,
    )
    print(f"   候选循环数: {len(candidates)}")
    print()

    print("🧮 运行加权区间调度DP...")
    selected = weighted_interval_scheduling(candidates)
    print(f"   选中循环数: {len(selected)}")
    print()

    # 验证
    validate_solution(selected)
    print()

    # 打印结果
    print_results(selected, slots)

    return selected, slots


if __name__ == "__main__":
    filepath = "/Users/zhaoruicn/.openclaw/media/inbound/2026ç_è---ce6c191b-e89c-43af-ac82-de759b1f9813.xlsx"

    if len(sys.argv) > 1:
        filepath = sys.argv[1]

    # 运行优化
    cycles, slots = optimize(
        filepath,
        price_type="实时",
        max_charge_hours=4,       # 单次最多充4小时
        max_discharge_hours=4,    # 单次最多放4小时
        min_profit_per_hour=30.0, # 每小时至少30元/MWh收益
        lookforward_days=2,       # 最多跨2天
    )
