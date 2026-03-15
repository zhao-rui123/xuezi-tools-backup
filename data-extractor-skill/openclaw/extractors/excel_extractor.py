"""
Excel数据提取模块

提供全面的Excel数据提取功能，包括:
- 多工作表数据提取
- 公式计算结果提取
- 合并单元格处理
- 数据透视表提取
- 图表数据提取
- 元数据提取
"""

import json
import logging
from typing import Any, Dict, List, Optional, Tuple, Union, Iterator
from pathlib import Path
from dataclasses import dataclass, field
from datetime import datetime, date
import warnings

import numpy as np

# 设置日志
logger = logging.getLogger("openclaw.extractors.excel")


@dataclass
class ExcelCell:
    """Excel单元格"""
    row: int
    col: int
    value: Any
    formula: Optional[str] = None
    data_type: str = "unknown"  # string, number, date, boolean, empty
    style: Optional[Dict[str, Any]] = None
    is_merged: bool = False
    merge_range: Optional[str] = None


@dataclass
class ExcelSheet:
    """Excel工作表"""
    name: str
    index: int
    data: List[List[Any]]
    headers: List[str] = field(default_factory=list)
    cells: List[List[ExcelCell]] = field(default_factory=list)
    merged_cells: List[str] = field(default_factory=list)
    max_row: int = 0
    max_col: int = 0
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "name": self.name,
            "index": self.index,
            "headers": self.headers,
            "data": self.data,
            "dimensions": {"rows": self.max_row, "cols": self.max_col},
            "merged_cells": self.merged_cells,
        }


@dataclass
class ExcelChart:
    """Excel图表"""
    name: str
    sheet: str
    chart_type: str
    title: Optional[str] = None
    data_range: Optional[str] = None


@dataclass
class ExcelPivotTable:
    """Excel数据透视表"""
    name: str
    sheet: str
    source_range: Optional[str] = None
    row_fields: List[str] = field(default_factory=list)
    col_fields: List[str] = field(default_factory=list)
    data_fields: List[str] = field(default_factory=list)


@dataclass
class ExcelMetadata:
    """Excel元数据"""
    title: Optional[str] = None
    subject: Optional[str] = None
    creator: Optional[str] = None
    keywords: Optional[str] = None
    description: Optional[str] = None
    last_modified_by: Optional[str] = None
    created: Optional[datetime] = None
    modified: Optional[datetime] = None
    sheet_count: int = 0


@dataclass
class ExcelExtractionResult:
    """Excel提取结果"""
    metadata: ExcelMetadata = field(default_factory=ExcelMetadata)
    sheets: List[ExcelSheet] = field(default_factory=list)
    charts: List[ExcelChart] = field(default_factory=list)
    pivot_tables: List[ExcelPivotTable] = field(default_factory=list)
    defined_names: Dict[str, str] = field(default_factory=dict)
    
    def get_sheet(self, name: str) -> Optional[ExcelSheet]:
        """根据名称获取工作表"""
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        return None
    
    def get_sheet_by_index(self, index: int) -> Optional[ExcelSheet]:
        """根据索引获取工作表"""
        for sheet in self.sheets:
            if sheet.index == index:
                return sheet
        return None
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "metadata": {
                "title": self.metadata.title,
                "creator": self.metadata.creator,
                "sheet_count": self.metadata.sheet_count,
            },
            "sheets": [s.to_dict() for s in self.sheets],
            "charts_count": len(self.charts),
            "pivot_tables_count": len(self.pivot_tables),
            "defined_names": self.defined_names,
        }


class ExcelExtractor:
    """
    Excel数据提取器
    
    提供全面的Excel数据提取功能，支持.xlsx, .xls, .xlsm格式。
    
    Example:
        >>> from openclaw import ExcelExtractor, Config
        >>> extractor = ExcelExtractor(Config().excel)
        >>> 
        >>> # 提取所有数据
        >>> result = extractor.extract("data.xlsx")
        >>> 
        >>> # 提取指定工作表
        >>> sheet = extractor.extract_sheet("data.xlsx", sheet_name="Sheet1")
        >>> 
        >>> # 提取为DataFrame
        >>> df = extractor.to_dataframe("data.xlsx", sheet_name="Sheet1")
        
        >>> # 批量提取多个文件
        >>> for result in extractor.extract_batch(["1.xlsx", "2.xlsx"]):
        ...     print(result.metadata.title)
    """
    
    def __init__(self, config=None):
        """
        初始化Excel提取器
        
        Args:
            config: ExcelConfig配置对象
        """
        self.config = config
        self.logger = logging.getLogger("openclaw.extractors.excel")
        self._check_dependencies()
    
    def _check_dependencies(self) -> None:
        """检查依赖库"""
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            self.logger.warning("openpyxl未安装，Excel提取功能受限")
            self.openpyxl = None
        
        try:
            import pandas as pd
            self.pandas = pd
        except ImportError:
            self.logger.warning("pandas未安装，DataFrame转换功能受限")
            self.pandas = None
        
        try:
            import xlrd
            self.xlrd = xlrd
        except ImportError:
            self.logger.warning("xlrd未安装，.xls格式支持受限")
            self.xlrd = None
    
    def extract(
        self,
        file_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None,
        sheet_index: Optional[Union[int, List[int]]] = None,
    ) -> ExcelExtractionResult:
        """
        提取Excel中的所有数据
        
        Args:
            file_path: Excel文件路径
            sheet_names: 指定工作表名称列表
            sheet_index: 指定工作表索引
        
        Returns:
            ExcelExtractionResult: 提取结果
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        self.logger.info(f"开始提取Excel: {file_path}")
        
        result = ExcelExtractionResult()
        
        # 提取元数据
        result.metadata = self.extract_metadata(file_path)
        
        # 提取工作表
        result.sheets = self.extract_sheets(file_path, sheet_names, sheet_index)
        result.metadata.sheet_count = len(result.sheets)
        
        # 提取图表
        result.charts = self.extract_charts(file_path)
        
        # 提取数据透视表
        result.pivot_tables = self.extract_pivot_tables(file_path)
        
        # 提取定义的名称
        result.defined_names = self.extract_defined_names(file_path)
        
        self.logger.info(
            f"Excel提取完成: {len(result.sheets)}个工作表, "
            f"{len(result.charts)}个图表, {len(result.pivot_tables)}个数据透视表"
        )
        
        return result
    
    def extract_metadata(self, file_path: Union[str, Path]) -> ExcelMetadata:
        """
        提取Excel元数据
        
        Args:
            file_path: Excel文件路径
        
        Returns:
            ExcelMetadata: 元数据对象
        """
        file_path = Path(file_path)
        metadata = ExcelMetadata()
        
        if not self.openpyxl:
            return metadata
        
        try:
            wb = self.openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            props = wb.properties
            
            metadata.title = props.title
            metadata.subject = props.subject
            metadata.creator = props.creator
            metadata.keywords = props.keywords
            metadata.description = props.description
            metadata.last_modified_by = props.lastModifiedBy
            metadata.created = props.created
            metadata.modified = props.modified
            metadata.sheet_count = len(wb.sheetnames)
            
            wb.close()
        except Exception as e:
            self.logger.warning(f"提取元数据失败: {e}")
        
        return metadata
    
    def extract_sheets(
        self,
        file_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None,
        sheet_index: Optional[Union[int, List[int]]] = None,
    ) -> List[ExcelSheet]:
        """
        提取工作表数据
        
        Args:
            file_path: Excel文件路径
            sheet_names: 指定工作表名称列表
            sheet_index: 指定工作表索引
        
        Returns:
            List[ExcelSheet]: 工作表列表
        """
        file_path = Path(file_path)
        sheets = []
        
        if not self.openpyxl:
            self.logger.error("openpyxl未安装，无法提取工作表")
            return sheets
        
        # 确定是否计算公式
        data_only = True
        if self.config and not self.config.evaluate_formulas:
            data_only = False
        
        try:
            wb = self.openpyxl.load_workbook(file_path, data_only=data_only)
            
            # 确定要处理的工作表
            target_sheets = []
            if sheet_names:
                target_sheets = [(name, wb[name]) for name in sheet_names if name in wb]
            elif sheet_index is not None:
                if isinstance(sheet_index, int):
                    if 0 <= sheet_index < len(wb.sheetnames):
                        name = wb.sheetnames[sheet_index]
                        target_sheets = [(name, wb[name])]
                else:
                    for idx in sheet_index:
                        if 0 <= idx < len(wb.sheetnames):
                            name = wb.sheetnames[idx]
                            target_sheets.append((name, wb[name]))
            else:
                target_sheets = [(name, wb[name]) for name in wb.sheetnames]
            
            # 提取每个工作表
            for idx, (name, ws) in enumerate(target_sheets):
                try:
                    sheet = self._extract_sheet_data(ws, name, idx)
                    sheets.append(sheet)
                except Exception as e:
                    self.logger.warning(f"提取工作表 '{name}' 失败: {e}")
            
            wb.close()
        except Exception as e:
            self.logger.error(f"提取工作表失败: {e}")
        
        return sheets
    
    def _extract_sheet_data(
        self,
        ws,
        name: str,
        index: int,
    ) -> ExcelSheet:
        """提取单个工作表的数据"""
        # 获取数据范围
        header_row = self.config.header_row if self.config else 0
        start_row = self.config.start_row if self.config else 1
        end_row = self.config.end_row if self.config else ws.max_row
        start_col = self.config.start_col if self.config else 1
        end_col = self.config.end_col if self.config else ws.max_column
        
        # 处理列索引
        if isinstance(start_col, str):
            start_col = self.openpyxl.utils.column_index_from_string(start_col)
        if isinstance(end_col, str):
            end_col = self.openpyxl.utils.column_index_from_string(end_col)
        
        # 提取合并单元格信息
        merged_cells = []
        if self.config and self.config.detect_merged_cells:
            merged_cells = [str(mc) for mc in ws.merged_cells.ranges]
        
        # 提取数据
        data = []
        cells = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
        ), start_row):
            row_data = []
            row_cells = []
            
            for col_idx, cell in enumerate(row, start_col):
                value = self._convert_cell_value(cell.value)
                
                # 检测数据类型
                data_type = self._detect_data_type(value)
                
                # 创建单元格对象
                excel_cell = ExcelCell(
                    row=row_idx,
                    col=col_idx,
                    value=value,
                    formula=cell.value if isinstance(cell.value, str) and str(cell.value).startswith("=") else None,
                    data_type=data_type,
                    is_merged=cell.coordinate in merged_cells,
                )
                
                row_data.append(value)
                row_cells.append(excel_cell)
            
            data.append(row_data)
            cells.append(row_cells)
        
        # 提取表头
        if header_row is not None and 0 <= header_row < len(data):
            headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(data[header_row])]
        
        return ExcelSheet(
            name=name,
            index=index,
            data=data,
            headers=headers,
            cells=cells,
            merged_cells=merged_cells,
            max_row=len(data),
            max_col=len(data[0]) if data else 0,
        )
    
    def _convert_cell_value(self, value: Any) -> Any:
        """转换单元格值"""
        if value is None:
            return None
        
        # 处理日期时间
        if isinstance(value, (datetime, date)):
            if self.config:
                if isinstance(value, datetime):
                    return value.strftime(self.config.datetime_format)
                else:
                    return value.strftime(self.config.date_format)
            return str(value)
        
        # 处理数值
        if isinstance(value, (int, float)):
            return value
        
        # 处理字符串
        if isinstance(value, str):
            return value.strip()
        
        return str(value)
    
    def _detect_data_type(self, value: Any) -> str:
        """检测数据类型"""
        if value is None:
            return "empty"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, (datetime, date)):
            return "date"
        return "string"
    
    def extract_sheet(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        sheet_index: int = 0,
    ) -> Optional[ExcelSheet]:
        """
        提取单个工作表
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（优先使用）
            sheet_index: 工作表索引
        
        Returns:
            ExcelSheet: 工作表对象
        """
        if sheet_name:
            sheets = self.extract_sheets(file_path, sheet_names=[sheet_name])
        else:
            sheets = self.extract_sheets(file_path, sheet_index=sheet_index)
        
        return sheets[0] if sheets else None
    
    def extract_charts(self, file_path: Union[str, Path]) -> List[ExcelChart]:
        """
        提取图表信息
        
        Args:
            file_path: Excel文件路径
        
        Returns:
            List[ExcelChart]: 图表列表
        """
        file_path = Path(file_path)
        charts = []
        
        if not self.openpyxl:
            return charts
        
        try:
            wb = self.openpyxl.load_workbook(file_path, read_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 检查是否有图表
                if hasattr(ws, '_charts') and ws._charts:
                    for idx, chart in enumerate(ws._charts):
                        chart_info = ExcelChart(
                            name=f"Chart_{idx+1}",
                            sheet=sheet_name,
                            chart_type=str(chart.__class__.__name__),
                            title=chart.title.text if hasattr(chart, 'title') and chart.title else None,
                        )
                        charts.append(chart_info)
            
            wb.close()
        except Exception as e:
            self.logger.warning(f"提取图表失败: {e}")
        
        return charts
    
    def extract_pivot_tables(self, file_path: Union[str, Path]) -> List[ExcelPivotTable]:
        """
        提取数据透视表信息
        
        Args:
            file_path: Excel文件路径
        
        Returns:
            List[ExcelPivotTable]: 数据透视表列表
        """
        file_path = Path(file_path)
        pivot_tables = []
        
        if not self.openpyxl:
            return pivot_tables
        
        try:
            wb = self.openpyxl.load_workbook(file_path, read_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 检查是否有数据透视表
                if hasattr(ws, '_pivots') and ws._pivots:
                    for idx, pivot in enumerate(ws._pivots):
                        pivot_info = ExcelPivotTable(
                            name=pivot.name if hasattr(pivot, 'name') else f"PivotTable_{idx+1}",
                            sheet=sheet_name,
                        )
                        pivot_tables.append(pivot_info)
            
            wb.close()
        except Exception as e:
            self.logger.warning(f"提取数据透视表失败: {e}")
        
        return pivot_tables
    
    def extract_defined_names(self, file_path: Union[str, Path]) -> Dict[str, str]:
        """
        提取定义的名称
        
        Args:
            file_path: Excel文件路径
        
        Returns:
            Dict[str, str]: 名称到范围的映射
        """
        file_path = Path(file_path)
        defined_names = {}
        
        if not self.openpyxl:
            return defined_names
        
        try:
            wb = self.openpyxl.load_workbook(file_path, read_only=True)
            
            for name in wb.defined_names:
                defined_names[name] = str(wb.defined_names[name].attr_text)
            
            wb.close()
        except Exception as e:
            self.logger.warning(f"提取定义名称失败: {e}")
        
        return defined_names
    
    def to_dataframe(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None,
        sheet_index: int = 0,
        header: Optional[int] = 0,
    ) -> Any:
        """
        提取为pandas DataFrame
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            sheet_index: 工作表索引
            header: 表头行
        
        Returns:
            DataFrame: pandas DataFrame对象
        """
        if not self.pandas:
            raise RuntimeError("pandas未安装，无法转换为DataFrame")
        
        file_path = Path(file_path)
        
        try:
            if sheet_name:
                return self.pandas.read_excel(file_path, sheet_name=sheet_name, header=header)
            else:
                return self.pandas.read_excel(file_path, sheet_name=sheet_index, header=header)
        except Exception as e:
            self.logger.error(f"转换为DataFrame失败: {e}")
            raise
    
    def extract_batch(
        self,
        file_paths: List[Union[str, Path]],
    ) -> Iterator[ExcelExtractionResult]:
        """
        批量提取多个Excel文件
        
        Args:
            file_paths: 文件路径列表
        
        Yields:
            ExcelExtractionResult: 每个文件的提取结果
        """
        for file_path in file_paths:
            try:
                yield self.extract(file_path)
            except Exception as e:
                self.logger.error(f"提取文件 {file_path} 失败: {e}")
                yield ExcelExtractionResult()
    
    def compare_sheets(
        self,
        file_path1: Union[str, Path],
        file_path2: Union[str, Path],
        sheet_name: str,
    ) -> Dict[str, Any]:
        """
        比较两个Excel文件的指定工作表
        
        Args:
            file_path1: 第一个文件路径
            file_path2: 第二个文件路径
            sheet_name: 工作表名称
        
        Returns:
            Dict: 比较结果
        """
        sheet1 = self.extract_sheet(file_path1, sheet_name=sheet_name)
        sheet2 = self.extract_sheet(file_path2, sheet_name=sheet_name)
        
        if not sheet1 or not sheet2:
            return {"error": "无法提取工作表"}
        
        differences = []
        
        # 比较数据
        max_rows = max(len(sheet1.data), len(sheet2.data))
        max_cols = max(
            len(sheet1.data[0]) if sheet1.data else 0,
            len(sheet2.data[0]) if sheet2.data else 0,
        )
        
        for row in range(max_rows):
            for col in range(max_cols):
                val1 = sheet1.data[row][col] if row < len(sheet1.data) and col < len(sheet1.data[row]) else None
                val2 = sheet2.data[row][col] if row < len(sheet2.data) and col < len(sheet2.data[row]) else None
                
                if val1 != val2:
                    differences.append({
                        "row": row,
                        "col": col,
                        "file1_value": val1,
                        "file2_value": val2,
                    })
        
        return {
            "sheet_name": sheet_name,
            "rows_file1": len(sheet1.data),
            "rows_file2": len(sheet2.data),
            "differences_count": len(differences),
            "differences": differences[:100],  # 限制差异数量
        }
