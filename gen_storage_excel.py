#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = '/Users/zhaoruicn/.openclaw/workspace/工商业储能市场开拓分析_专业版.xlsx'

wb = openpyxl.Workbook()

# Colors
BLUE_HDR = PatternFill('solid', fgColor='1F4E79')
BLUE_SUB = PatternFill('solid', fgColor='2F75B6')
GREEN = PatternFill('solid', fgColor='C6EFCE')
GREEN_TXT = Font(color='006100', bold=True, size=12)
RED = PatternFill('solid', fgColor='FFC7CE')
RED_TXT = Font(color='9C0006', bold=True, size=12)
YELLOW = PatternFill('solid', fgColor='FFEB9C')
YELLOW_TXT = Font(color='9C6500', bold=True, size=12)
WHITE = Font(color='FFFFFF', bold=True, size=12)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
HDR_TXT = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True, size=11)

def hdr(cell, text, fill=BLUE_HDR, font=None, size=12, align='center', colspan=0):
    cell.value = text
    cell.fill = fill
    cell.font = font or Font(color='FFFFFF', bold=True, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = THIN
    if colspan > 1:
        pass

def cell_set(cell, value, fill=None, font=None, align='center', bold=False, size=11):
    cell.value = value
    if fill:
        cell.fill = fill
    cell.font = font or Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = THIN

def money_cell(cell, value, fill=None, good=True):
    cell.value = f'¥{value:.4f}'
    cell.fill = fill or (GREEN if good else RED)
    cell.font = GREEN_TXT if good else RED_TXT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN

def pct_cell(cell, value, fill=None, good=True):
    cell.value = f'{value:.2%}'
    cell.fill = fill or (GREEN if good else RED)
    cell.font = GREEN_TXT if good else RED_TXT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN

# ==================== Sheet 1: 封面 ====================
ws = wb.active
ws.title = '封面'
ws.merge_cells('A3:G3')
ws['A3'] = '工商业储能市场开拓分析报告'
ws['A3'].font = Font(bold=True, size=24, color='1F4E79')
ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[3].height = 50

ws.merge_cells('A5:G5')
ws['A5'] = '2026年3月'
ws['A5'].font = Font(size=16, color='666666')
ws['A5'].alignment = Alignment(horizontal='center')

ws.merge_cells('A7:G7')
ws['A7'] = '编制单位：雪子助手'
ws['A7'].font = Font(size=12)
ws['A7'].alignment = Alignment(horizontal='center')

# 目录
ws.merge_cells('A9:G9')
ws['A9'] = '目 录'
ws['A9'].font = Font(bold=True, size=16, color='FFFFFF')
ws['A9'].fill = BLUE_HDR
ws['A9'].alignment = Alignment(horizontal='center')

dirs = [
    ('Sheet2', '模式一经济账（纯储能）'),
    ('Sheet3', '模式二经济账（光储一体化）'),
    ('Sheet4', '现货市场价格数据'),
    ('Sheet5', '省份梯队划分'),
    ('Sheet6', '全国投资地图'),
    ('Sheet7', '综合结论与建议'),
]
for i, (sheet, name) in enumerate(dirs, 10):
    ws.cell(i, 1, f'{i-8}. {name}').font = Font(size=12)
    ws.cell(i, 1).alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{i}:G{i}')
ws.column_dimensions['A'].width = 40

# ==================== Sheet 2: 模式一经济账 ====================
ws2 = wb.create_sheet('模式一经济账')
ws2.merge_cells('A1:H1')
ws2['A1'] = '模式一：纯储能投资经济账（10MWh）'
ws2['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws2['A1'].fill = BLUE_HDR
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

# 门槛结论
ws2.merge_cells('A3:C3')
ws2['A3'] = '⚠️ 核心门槛结论'
ws2['A3'].font = Font(bold=True, size=14, color='FFFFFF')
ws2['A3'].fill = BLUE_SUB
ws2['A3'].alignment = Alignment(horizontal='center', vertical='center')

data_concl = [
    ['IRR=10% 需要电价差', '≥ 0.6618 元/kWh', '✅ 满足则IRR可达10%'],
    ['回收期≤7年 需要电价差', '≥ 0.5795 元/kWh', '✅ 满足则7年内回本'],
]
for i, row in enumerate(data_concl, 4):
    ws2.merge_cells(f'A{i}:B{i}')
    cell_set(ws2.cell(i,1), row[0], bold=True, fill=YELLOW)
    cell_set(ws2.cell(i,3), row[1], fill=GREEN, bold=True, size=14)
    cell_set(ws2.cell(i,4), row[2])

# 参数表
ws2.merge_cells('A7:H7')
ws2['A7'] = '基础参数'
ws2['A7'].font = Font(bold=True, size=12, color='FFFFFF')
ws2['A7'].fill = BLUE_SUB
ws2['A7'].alignment = Alignment(horizontal='center')

params = [
    ['储能容量', '10 MWh', '放电深度', '95%'],
    ['总投资', '1,000万元', '年充放次数', '330次'],
    ['设备成本', '0.80 元/Wh', '充放电效率', '88%'],
    ['施工成本', '0.20 元/Wh', '年衰减率', '1.5%'],
    ['单位成本', '1.00 元/Wh', '使用年限', '10年'],
    ['运营成本', '10万元/年', '', ''],
]
for i, row in enumerate(params, 8):
    for j, v in enumerate(row):
        col = j*2+1
        c = ws2.cell(i, col)
        cell_set(c, v, fill=PatternFill('solid', fgColor='D9E1F2') if j==0 else None, bold=(j==0))
        if j == 0:
            ws2.merge_cells(start_row=i, start_column=col, end_row=i, end_column=col+1)

# 敏感性分析
ws2.merge_cells('A16:H16')
ws2['A16'] = '敏感性分析 - 各省电价差对比'
ws2['A16'].font = Font(bold=True, size=12, color='FFFFFF')
ws2['A16'].fill = BLUE_SUB
ws2['A16'].alignment = Alignment(horizontal='center')

sensitivity = [
    ['省份', '峰谷价差', 'IRR=10%门槛', '是否满足', '回收期≤7年门槛', '是否满足', '结论'],
    ['山西', '0.40~1.50', '0.6618', '✅', '0.5795', '✅', '⭐第一梯队'],
    ['广东', '0.40~1.29', '0.6618', '✅', '0.5795', '✅', '⭐第一梯队'],
    ['浙江', '0.35~1.27', '0.6618', '✅', '0.5795', '✅', '⭐第一梯队'],
    ['山东', '0.35~0.57', '0.6618', '❌', '0.5795', '⚠️', '⚠️第二梯队'],
    ['甘肃', '0.25~0.46', '0.6618', '❌', '0.5795', '❌', '❌第三梯队'],
]
for i, row in enumerate(sensitivity, 17):
    for j, v in enumerate(row):
        c = ws2.cell(i, j+1)
        fill = PatternFill('solid', fgColor='D9E1F2') if i==17 else None
        good = '✅' in str(v) or '⭐' in str(v)
        bad = '❌' in str(v)
        if j == 3:
            fill = GREEN if good else (RED if bad else YELLOW)
        if j == 5:
            fill = GREEN if good else (RED if bad else (YELLOW if '⚠️' in str(v) else None))
        cell_set(c, v, fill=fill, bold=(i==17), align='center')

for col in range(1, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 16

# ==================== Sheet 3: 模式二经济账 ====================
ws3 = wb.create_sheet('模式二经济账')
ws3.merge_cells('A1:H1')
ws3['A1'] = '模式二：光储一体化投资经济账（1MW+3MWh）'
ws3['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws3['A1'].fill = BLUE_HDR
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

# 门槛结论
ws3.merge_cells('A3:C3')
ws3['A3'] = '⚠️ 核心门槛结论'
ws3['A3'].font = Font(bold=True, size=14, color='FFFFFF')
ws3['A3'].fill = BLUE_SUB

data2 = [
    ['最优配置', '1MW光伏 + 3MWh储能', '光伏利用率98%（最佳配比）'],
    ['IRR=8% 需要储能卖电价', '≥ 0.5260 元/kWh', '✅ 达到则IRR≥8%'],
    ['IRR=10% 需要储能卖电价', '≥ 0.6350 元/kWh', '✅ 达到则IRR≥10%'],
]
for i, row in enumerate(data2, 4):
    ws3.merge_cells(f'A{i}:B{i}')
    cell_set(ws3.cell(i,1), row[0], bold=True, fill=YELLOW)
    cell_set(ws3.cell(i,3), row[1], fill=GREEN, bold=True, size=13)
    cell_set(ws3.cell(i,4), row[2])

# 参数
ws3.merge_cells('A8:H8')
ws3['A8'] = '基础参数'
ws3['A8'].font = Font(bold=True, size=12, color='FFFFFF')
ws3['A8'].fill = BLUE_SUB

p2 = [
    ['光伏装机', '1 MW', '储能容量', '3 MWh'],
    ['光伏成本', '2 元/W (200万)', '储能成本', '1 元/Wh (300万)'],
    ['总投资', '500万元', '日照', '1200h/年 (河南)'],
    ['光伏效率', '80%', '储能衰减', '1.5%/年'],
    ['充放效率', '88%', '运行年限', '20年'],
    ['充放次数', '330次/年', '放电深度', '95%'],
    ['直售电价', '0.25 元/kWh', '换电芯', '不换'],
]
for i, row in enumerate(p2, 9):
    for j, v in enumerate(row):
        c = ws3.cell(i, j*2+1)
        cell_set(c, v, fill=PatternFill('solid', fgColor='D9E1F2') if j==0 else None, bold=(j==0))
        ws3.merge_cells(start_row=i, start_column=j*2+1, end_row=i, end_column=j*2+1) if j==0 else None

# 配置对比
ws3.merge_cells('A18:H18')
ws3['A18'] = '不同储能配置对比（1MW光伏）'
ws3['A18'].font = Font(bold=True, size=12, color='FFFFFF')
ws3['A18'].fill = BLUE_SUB

cfg = [
    ['配置', '储能', '总投资', '光伏利用', 'IRR=8%门槛', 'IRR=10%门槛', '推荐'],
    ['A', '1 MWh', '300万', '32.7%', '0.670', '0.845', '⚡'],
    ['B', '2 MWh', '400万', '65.4%', '0.682', '0.799', '⚡'],
    ['C', '3 MWh', '500万', '98.0%', '0.526', '0.635', '✅最优'],
    ['D', '4 MWh', '600万', '131%', '0.860', '0.981', '❌'],
]
for i, row in enumerate(cfg, 19):
    for j, v in enumerate(row):
        c = ws3.cell(i, j+1)
        fill = PatternFill('solid', fgColor='D9E1F2') if i==19 else None
        if '最优' in str(v):
            fill = GREEN
        if '❌' in str(v):
            fill = RED
        cell_set(c, v, fill=fill, bold=(i==19), align='center')

# 敏感性
ws3.merge_cells('A26:H26')
ws3['A26'] = '敏感性分析 - 储能卖电价IRR对照'
ws3['A26'].font = Font(bold=True, size=12, color='FFFFFF')
ws3['A26'].fill = BLUE_SUB

irr_table = [
    ['储能卖电价', 'IRR', '20年净收益', '结论'],
    ['0.50 元/kWh', '6.2%', '~350万', '❌ 不达标'],
    ['0.55 元/kWh', '7.8%', '~420万', '⚠️ 接近8%'],
    ['0.60 元/kWh', '9.3%', '~500万', '⚠️ 接近10%'],
    ['0.65 元/kWh', '10.8%', '~580万', '✅ 达标'],
    ['0.70 元/kWh', '12.2%', '~650万', '✅ 达标'],
    ['0.80 元/kWh', '15.0%', '~780万', '✅ 优秀'],
]
for i, row in enumerate(irr_table, 27):
    for j, v in enumerate(row):
        c = ws3.cell(i, j+1)
        fill = PatternFill('solid', fgColor='D9E1F2') if i==27 else None
        if '✅' in str(v):
            fill = GREEN
        elif '❌' in str(v):
            fill = RED
        elif '⚠️' in str(v):
            fill = YELLOW
        cell_set(c, v, fill=fill, bold=(i==27), align='center')

for col in range(1, 8):
    ws3.column_dimensions[get_column_letter(col)].width = 16

# ==================== Sheet 4: 现货市场价格 ====================
ws4 = wb.create_sheet('现货市场价格')
ws4.merge_cells('A1:G1')
ws4['A1'] = '各省电力现货市场价格数据（2025-2026年）'
ws4['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws4['A1'].fill = BLUE_HDR
ws4['A1'].alignment = Alignment(horizontal='center')
ws4.row_dimensions[1].height = 30

headers = ['省份', '午间低电价', '晚高峰', '峰谷价差', '负电价', '模式一IRR≥10%?', '模式一回收期≤7年?']
for j, h in enumerate(headers, 1):
    c = ws4.cell(3, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = BLUE_HDR
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN

market_data = [
    ['山东', '0.02~0.05', '0.40~0.50', '0.35~0.57', '频繁(-0.07)', '❌不满足', '⚠️勉强'],
    ['山西', '0~0.01', '0.40~0.80', '0.40~1.50', '频繁零/负', '✅满足', '✅满足'],
    ['广东', '0.02~0.05', '0.45~0.60', '0.40~1.29', '极少', '✅满足', '✅满足'],
    ['甘肃', '0.04', '0.30~0.40', '0.25~0.46', '地板频繁', '❌不满足', '❌不满足'],
    ['浙江', '0.02~0.05', '0.40~0.50', '0.35~1.27', '频繁(-0.20)', '✅满足', '✅满足'],
]
for i, row in enumerate(market_data, 4):
    for j, v in enumerate(row, 1):
        c = ws4.cell(i, j)
        c.value = v
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center')
        if j == 6:
            c.fill = GREEN if '✅' in v else RED
            c.font = GREEN_TXT if '✅' in v else RED_TXT
        elif j == 7:
            if '✅' in v:
                c.fill = GREEN
                c.font = GREEN_TXT
            elif '勉强' in v:
                c.fill = YELLOW
                c.font = YELLOW_TXT
            else:
                c.fill = RED
                c.font = RED_TXT

ws4.merge_cells('A10:H10')
ws4['A10'] = '单位说明：电价单位为 元/kWh'
ws4['A10'].font = Font(size=10, italic=True, color='666666')
ws4.merge_cells('A11:H11')
ws4['A11'] = '数据来源：各省电力交易中心2025-2026年公开数据'
ws4['A11'].font = Font(size=10, italic=True, color='666666')

for col in range(1, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 18

# ==================== Sheet 5: 省份梯队划分 ====================
ws5 = wb.create_sheet('省份梯队划分')
ws5.merge_cells('A1:D1')
ws5['A1'] = '省份梯队划分与开拓建议'
ws5['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws5['A1'].fill = BLUE_HDR
ws5['A1'].alignment = Alignment(horizontal='center')

# 模式一
ws5.merge_cells('A3:D3')
ws5['A3'] = '模式一梯队（纯储能）'
ws5['A3'].font = Font(bold=True, size=12, color='FFFFFF')
ws5['A3'].fill = PatternFill('solid', fgColor='375623')

t1 = [
    ['梯队', '省份', '峰谷价差', '现货市场'],
    ['⭐ 第一梯队', '山西、广东、浙江', '0.40~1.50', '已开通'],
    ['⚠️ 第二梯队', '山东、江苏、河南', '0.35~0.57', '即将全覆盖'],
    ['📌 第三梯队', '甘肃、湖北、蒙西', '0.25~0.46', '已开通'],
]
for i, row in enumerate(t1, 4):
    for j, v in enumerate(row, 1):
        c = ws5.cell(i, j)
        c.value = v
        c.border = THIN
        c.alignment = Alignment(horizontal='center' if j<=3 else 'left', vertical='center')
        if i == 4:
            c.fill = PatternFill('solid', fgColor='D9E1F2')
            c.font = BOLD
        elif '第一梯队' in str(v):
            c.fill = GREEN
            c.font = Font(bold=True, color='006100')
        elif '第二梯队' in str(v):
            c.fill = YELLOW
        elif '第三梯队' in str(v):
            c.fill = RED

# 模式二
ws5.merge_cells('A10:D10')
ws5['A10'] = '模式二梯队（光储一体化）'
ws5['A10'].font = Font(bold=True, size=12, color='FFFFFF')
ws5['A10'].fill = PatternFill('solid', fgColor='375623')

t2 = [
    ['梯队', '省份', '分布式光伏', '现货市场'],
    ['⭐ 第一梯队', '河北、山东、河南、安徽', '全国前五', '已开/即将开'],
    ['⚠️ 第二梯队', '江苏、浙江、广东', '全国前列', '已开通'],
    ['📌 第三梯队', '其他省份', '一般', '待开通'],
]
for i, row in enumerate(t2, 11):
    for j, v in enumerate(row, 1):
        c = ws5.cell(i, j)
        c.value = v
        c.border = THIN
        c.alignment = Alignment(horizontal='center' if j<=3 else 'left', vertical='center')
        if i == 11:
            c.fill = PatternFill('solid', fgColor='D9E1F2')
            c.font = BOLD
        elif '第一梯队' in str(v):
            c.fill = GREEN
            c.font = Font(bold=True, color='006100')
        elif '第二梯队' in str(v):
            c.fill = YELLOW
        elif '第三梯队' in str(v):
            c.fill = RED

for col in 'ABCD':
    ws5.column_dimensions[col].width = 22

# ==================== Sheet 6: 全国投资地图 ====================
ws6 = wb.create_sheet('全国投资地图')
ws6.merge_cells('A1:F1')
ws6['A1'] = '全国投资地图 - 各省梯队分布'
ws6['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws6['A1'].fill = BLUE_HDR
ws6['A1'].alignment = Alignment(horizontal='center')

map_data = [
    ['区域', '第一梯队（优先开拓）', '第二梯队（储备）', '第三梯队（观察）'],
    ['华北', '山西', '山东', '甘肃'],
    ['华东', '浙江', '江苏、上海、安徽', ''],
    ['华南', '广东', '', ''],
    ['华中', '河北、河南', '湖北', ''],
    ['西南', '', '', '待研究'],
    ['东北', '', '', '待研究'],
    ['西北', '', '', '新疆、陕西'],
]
for i, row in enumerate(map_data, 3):
    for j, v in enumerate(row, 1):
        c = ws6.cell(i, j)
        c.value = v
        c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if i == 3:
            c.fill = BLUE_HDR
            c.font = HDR_TXT
        elif j == 2 and v and v != '':
            c.fill = GREEN
        elif j == 3 and v and v != '':
            c.fill = YELLOW
        elif j == 4 and v and v != '':
            c.fill = RED
        if j == 2 and v:
            txt_color = '006100'
        elif j == 3 and v:
            txt_color = '9C6500'
        elif j == 4 and v:
            txt_color = '9C0006'
        else:
            txt_color = '000000'
        c.font = Font(bold=(i==3), color=txt_color)

# 图例
ws6.merge_cells('A12:F12')
ws6['A12'] = '图例'
ws6['A12'].font = Font(bold=True, size=12)
legends = [
    ('✅ 第一梯队（绿色）', GREEN, '山西/广东/浙江/河北/河南/安徽可直接开拓'),
    ('⚠️ 第二梯队（黄色）', YELLOW, '山东/江苏/上海/湖北等可储备观望'),
    ('❌ 第三梯队（红色）', RED, '甘肃/湖北/蒙西等暂不推荐'),
]
for i, (txt, fill, note) in enumerate(legends, 13):
    c = ws6.cell(i, 1)
    c.value = txt
    c.fill = fill
    c.font = Font(bold=True)
    c.border = THIN
    ws6.merge_cells(f'B{i}:F{i}')
    ws6.cell(i, 2).value = note
    ws6.cell(i, 2).border = THIN

for col in 'ABCDEF':
    ws6.column_dimensions[col].width = 20

# ==================== Sheet 7: 综合结论 ====================
ws7 = wb.create_sheet('综合结论')
ws7.merge_cells('A1:B1')
ws7['A1'] = '综合结论与开拓建议'
ws7['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws7['A1'].fill = BLUE_HDR
ws7['A1'].alignment = Alignment(horizontal='center')

conclusions = [
    ['政策背景', '2026年3月起，全国多省取消工商业行政性峰谷电价，改为电力市场化交易'],
    ['模式一定论', '主推山西/广东/浙江，门槛电价差0.6618元/kWh达到IRR=10%'],
    ['模式二结论', '推荐1MW+3MWh配置，门槛0.635元/kWh达IRR=10%（最优配比，光伏利用率98%）'],
    ['优先推荐', '山西（模式一+模式二均为第一梯队），广东/浙江（模式一第一梯队）'],
    ['开拓策略', '以模式一（纯储能）为主，模式二（光储一体化）为辅，选择第一梯队省份优先开拓'],
    ['风险提示', '电价差收窄风险、市场价格波动风险、政策变化风险、业主违约风险'],
    ['数据说明', '现货价格数据来自各省电力交易中心2025-2026年公开数据'],
]
for i, row in enumerate(conclusions, 3):
    ws7.cell(i, 1).value = row[0]
    ws7.cell(i, 2).value = row[1]
    ws7.cell(i, 1).fill = PatternFill('solid', fgColor='D9E1F2')
    ws7.cell(i, 1).font = BOLD
    ws7.cell(i, 1).border = THIN
    ws7.cell(i, 2).border = THIN
    ws7.cell(i, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws7.merge_cells(f'B{i}:B{i}')
    
ws7.column_dimensions['A'].width = 20
ws7.column_dimensions['B'].width = 70

wb.save(OUTPUT)
print(f'✅ Excel已生成: {OUTPUT}')
