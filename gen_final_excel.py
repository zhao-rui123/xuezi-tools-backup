#!/usr/bin/env python3
"""工商业储能市场开拓分析 - 最终完整版"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = '/Users/zhaoruicn/.openclaw/workspace/工商业储能市场开拓分析_最终版.xlsx'
wb = openpyxl.Workbook()

FILL_BLUE = PatternFill('solid', fgColor='1F4E79')
FILL_BLUE2 = PatternFill('solid', fgColor='2F75B6')
FILL_GREEN = PatternFill('solid', fgColor='C6EFCE')
FILL_RED = PatternFill('solid', fgColor='FFC7CE')
FILL_YELLOW = PatternFill('solid', fgColor='FFEB9C')
FILL_LIGHT = PatternFill('solid', fgColor='D9E1F2')
GREEN_TXT = Font(bold=True, size=11, color='006100')
RED_TXT = Font(bold=True, size=11, color='9C0006')
YELLOW_TXT = Font(bold=True, size=11, color='9C6500')
HDR_TXT = Font(bold=True, size=10, color='FFFFFF')
BOLD = Font(bold=True, size=10)
REG = Font(size=10)
TB = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

def sc(cell, val, fill=None, font=None, align='center', bold=False, wrap=True, size=10):
    cell.value = val
    cell.font = font or Font(bold=bold, size=size)
    if fill: cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = TB

# ==================== Sheet1 封面 ====================
ws1 = wb.active
ws1.title = '封面'
ws1.column_dimensions['A'].width = 80
ws1.merge_cells('A3:E3')
ws1.row_dimensions[3].height = 60
c = ws1['A3']
c.value = '工商业储能市场开拓分析报告'
c.font = Font(bold=True, size=28, color='1F4E79')
c.alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A4:E4')
c = ws1['A4']
c.value = '2026年3月 | 政策解读 | 经济账测算 | 省份梯队 | 全国投资地图'
c.font = Font(size=13, color='666666', italic=True)
c.alignment = Alignment(horizontal='center')

ws1.merge_cells('A6:E6')
ws1.row_dimensions[6].height = 30
c = ws1['A6']
c.value = '目  录'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

dirs = [
    ('Sheet2', '模式一经济账（纯储能10MWh）', '门槛分析 | 敏感性分析 | 各省IRR对照'),
    ('Sheet3', '模式二经济账（光储一体化1MW+3MWh）', '最优配置 | 敏感性分析 | 多情景对比'),
    ('Sheet4', '现货市场价格数据', '5省实时电价 | 峰谷价差对比'),
    ('Sheet5', '省份梯队划分（全国31省）', '模式一梯队 | 模式二梯队 | 详细分析'),
    ('Sheet6', '全国投资地图', '7大区域 | 颜色标注 | 开拓策略'),
    ('Sheet7', '综合结论与建议', '政策背景 | 模式对比 | 风险提示 | 开拓建议'),
]
for i,(sheet,title,desc) in enumerate(dirs, 7):
    ws1.row_dimensions[i].height = 26
    c = ws1.cell(i, 1)
    c.value = f'{i-6}. {title}'
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c = ws1.cell(i, 2)
    c.value = f'→ {desc}'
    c.font = Font(size=10, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center')

# ==================== Sheet2 模式一 ====================
ws2 = wb.create_sheet('模式一经济账')
for col in 'ABCDEFGHIJKL':
    ws2.column_dimensions[col].width = 14

ws2.merge_cells('A1:L1')
ws2.row_dimensions[1].height = 40
c = ws2['A1']
c.value = '模式一：纯储能投资经济账（10MWh储能）'
c.font = Font(bold=True, size=18, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

ws2.merge_cells('A3:D3')
ws2.row_dimensions[3].height = 35
c = ws2['A3']
c.value = '⚠️ 核心门槛结论'
c.font = Font(bold=True, size=14, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

rows = [
    ('IRR=10% 需要电价差', '≥ 0.6618 元/kWh', '满足则项目IRR达到10%'),
    ('回收期≤7年 需要电价差', '≥ 0.5795 元/kWh', '满足则7年内可回本'),
]
for i, (label, val, note) in enumerate(rows, 4):
    ws2.row_dimensions[i].height = 28
    c = ws2.cell(i, 1)
    c.value = label
    c.font = BOLD
    c.fill = FILL_YELLOW
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB
    ws2.merge_cells(f'B{i}:C{i}')
    c = ws2.cell(i, 2)
    c.value = val
    c.font = Font(bold=True, size=14, color='006100')
    c.fill = FILL_GREEN
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB
    ws2.merge_cells(f'D{i}:L{i}')
    c = ws2.cell(i, 4)
    c.value = note
    c.font = REG
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill()
    c.border = TB

ws2.merge_cells('A7:L7')
ws2.row_dimensions[7].height = 25
c = ws2['A7']
c.value = '一、基础参数'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

params = [
    ['储能容量', '10 MWh', '放电深度', '95%', '年充放次数', '330次', '投资回收期目标', '≤7年'],
    ['总投资', '1,000万元', '充放电效率', '88%', '年衰减率', '1.5%', '目标IRR', '≥10%'],
    ['设备成本', '0.80元/Wh', '运营成本', '10万/年', '使用年限', '10年', '放电策略', '峰谷套利'],
]
for i, row in enumerate(params, 8):
    ws2.row_dimensions[i].height = 22
    for j, v in enumerate(row):
        c = ws2.cell(i, j+1)
        c.value = v
        c.font = BOLD if j%2==0 else REG
        c.fill = FILL_LIGHT if j%2==0 else None
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB

ws2.merge_cells('A13:L13')
ws2.row_dimensions[13].height = 25
c = ws2['A13']
c.value = '二、敏感性分析 - 不同电价差下的IRR与回收期'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

pd_list = [0.40, 0.50, 0.55, 0.58, 0.60, 0.65, 0.67, 0.70, 0.80, 0.90, 1.00]
hdrs = ['电价差(元/kWh)'] + [f'{p}' for p in pd_list]
for j, h in enumerate(hdrs, 1):
    c = ws2.cell(14, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

# IRR row
cap, DOD, eff, decay, cycles, opex, invest = 10000, 0.95, 0.88, 0.015, 330, 10, 1000
def calc_irr(pd):
    cfs = [-invest] + [cap*DOD*((1-decay)**(y-1))*eff*cycles*pd-opex for y in range(1,11)]
    rate = 0.1
    for _ in range(1000):
        f = sum(cf/(1+rate)**t for t,cf in enumerate(cfs))
        df = sum(-t*cf/(1+rate)**(t+1) for t,cf in enumerate(cfs))
        if abs(df)<1e-12: break
        rate -= f/df
        if abs(f)<1e-10: break
    return rate

def calc_payback(pd):
    cum = -invest
    for y in range(1, 11):
        cum += cap*DOD*((1-decay)**(y-1))*eff*cycles*pd-opex
        if cum >= 0:
            return f'{y}年'
    return '>10年'

ws2.cell(15, 1).value = 'IRR'
ws2.cell(15, 1).font = BOLD
ws2.cell(15, 1).fill = FILL_LIGHT
ws2.cell(15, 1).border = TB
for j, pd in enumerate(pd_list, 2):
    irr = calc_irr(pd)
    c = ws2.cell(15, j)
    c.value = f'{irr:.2%}'
    c.fill = FILL_GREEN if irr>=0.10 else (FILL_YELLOW if irr>=0.08 else FILL_RED)
    c.font = Font(bold=True, size=10, color='006100' if irr>=0.10 else ('9C6500' if irr>=0.08 else '9C0006'))
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

ws2.cell(16, 1).value = '静态回收期'
ws2.cell(16, 1).font = BOLD
ws2.cell(16, 1).fill = FILL_LIGHT
ws2.cell(16, 1).border = TB
for j, pd in enumerate(pd_list, 2):
    pb = calc_payback(pd)
    c = ws2.cell(16, j)
    c.value = pb
    yr = int(pb.replace('年','')) if pb[-1]=='年' else 999
    c.fill = FILL_GREEN if yr<=7 else (FILL_YELLOW if yr<=10 else FILL_RED)
    c.font = Font(bold=True, size=10, color='006100' if yr<=7 else ('9C6500' if yr<=10 else '9C0006'))
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

ws2.row_dimensions[15].height = 20
ws2.row_dimensions[16].height = 20

# 全国各省模式一对照
ws2.merge_cells('A19:L19')
ws2.row_dimensions[19].height = 25
c = ws2['A19']
c.value = '三、全国各省模式一IRR预估（电价差→IRR）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

province_m1 = [
    ('山西', '0.40~1.50', 0.74, '✅第一梯队', FILL_GREEN),
    ('广东', '0.40~1.29', 0.71, '✅第一梯队', FILL_GREEN),
    ('浙江', '0.35~1.27', 0.69, '✅第一梯队', FILL_GREEN),
    ('山东', '0.35~0.57', 0.55, '⚠️第二梯队', FILL_YELLOW),
    ('江苏', '0.30~0.60', 0.52, '⚠️第二梯队', FILL_YELLOW),
    ('河南', '0.30~0.55', 0.50, '⚠️第二梯队', FILL_YELLOW),
    ('湖北', '0.30~0.55', 0.50, '⚠️第二梯队', FILL_YELLOW),
    ('安徽', '0.30~0.55', 0.50, '⚠️第二梯队', FILL_YELLOW),
    ('河北', '0.30~0.55', 0.50, '⚠️第二梯队', FILL_YELLOW),
    ('蒙西', '0.25~0.46', 0.44, '🔶第三梯队', FILL_RED),
    ('甘肃', '0.25~0.46', 0.44, '🔶第三梯队', FILL_RED),
    ('辽宁', '0.25~0.45', 0.42, '🔶第三梯队', FILL_RED),
    ('吉林', '0.25~0.45', 0.42, '🔶第三梯队', FILL_RED),
    ('黑龙江', '0.25~0.45', 0.42, '🔶第三梯队', FILL_RED),
    ('陕西', '0.25~0.50', 0.45, '🔶第三梯队', FILL_RED),
    ('宁夏', '0.25~0.45', 0.42, '🔶第三梯队', FILL_RED),
    ('青海', '0.20~0.40', 0.38, '🔶第三梯队', FILL_RED),
    ('新疆', '0.20~0.45', 0.40, '🔶第三梯队', FILL_RED),
    ('贵州', '0.20~0.45', 0.40, '🔶第三梯队', FILL_RED),
    ('云南', '0.20~0.45', 0.40, '🔶第三梯队', FILL_RED),
    ('西藏', '0.10~0.35', 0.30, '🔶第三梯队', FILL_RED),
]

province_hdrs = ['省份', '峰谷价差', 'IRR估算', '梯队', '模式一IRR≥10%?', '回收期≤7年?']
for j, h in enumerate(province_hdrs, 1):
    c = ws2.cell(20, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

for i, (name, pd_range, irr_est, tier, tier_fill) in enumerate(province_m1, 21):
    ws2.row_dimensions[i].height = 18
    vals = [name, pd_range, f'{irr_est:.2%}', tier]
    irr_ok = '✅' if irr_est >= 0.10 else ('⚠️' if irr_est >= 0.08 else '❌')
    pb_ok = '✅' if irr_est >= 0.08 else ('⚠️' if irr_est >= 0.07 else '❌')
    vals += [irr_ok, pb_ok]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(i, j)
        c.value = v
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB
        if j == 4:
            c.fill = tier_fill
            c.font = Font(bold=True, size=9)
        elif j == 5:
            c.fill = FILL_GREEN if '✅' in v else (FILL_YELLOW if '⚠️' in v else FILL_RED)
        elif j == 6:
            c.fill = FILL_GREEN if '✅' in v else (FILL_YELLOW if '⚠️' in v else FILL_RED)

# ==================== Sheet3 模式二 ====================
ws3 = wb.create_sheet('模式二经济账')
for col in 'ABCDEFGH':
    ws3.column_dimensions[col].width = 16

ws3.merge_cells('A1:H1')
ws3.row_dimensions[1].height = 40
c = ws3['A1']
c.value = '模式二：光储一体化投资经济账（1MW + 3MWh储能）'
c.font = Font(bold=True, size=18, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells('A3:H3')
ws3.row_dimensions[3].height = 30
c = ws3['A3']
c.value = '🏆 最优配置：1MW光伏 + 3MWh储能（光伏充电利用率98%，性价比最优）'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='375623')
c.alignment = Alignment(horizontal='center', vertical='center')

m2 = [
    ('IRR=8% 需要储能卖电价', '≥ 0.5260 元/kWh', '达到则IRR≥8%'),
    ('IRR=10% 需要储能卖电价', '≥ 0.6350 元/kWh', '达到则IRR≥10%'),
]
for i, (label, val, note) in enumerate(m2, 4):
    ws3.row_dimensions[i].height = 28
    c = ws3.cell(i, 1)
    c.value = label
    c.font = BOLD
    c.fill = FILL_YELLOW
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB
    ws3.merge_cells(f'B{i}:C{i}')
    c = ws3.cell(i, 2)
    c.value = val
    c.font = Font(bold=True, size=14, color='006100')
    c.fill = FILL_GREEN
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB
    ws3.merge_cells(f'D{i}:H{i}')
    c = ws3.cell(i, 4)
    c.value = note
    c.font = REG
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill()
    c.border = TB

ws3.merge_cells('A7:H7')
ws3.row_dimensions[7].height = 25
c = ws3['A7']
c.value = '一、基础参数'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

p2 = [
    ['光伏装机','1 MW','储能容量','3 MWh','总投资','500万元'],
    ['光伏成本','2元/W (200万)','储能成本','1元/Wh (300万)','日照','1200h/年(河南)'],
    ['光伏效率','80%','储能衰减','1.5%/年','充放效率','88%'],
    ['放电深度','95%','年充放次数','330次','运行年限','20年'],
    ['直售电价','0.25元/kWh','换电芯','不换','光伏衰减','2%/年'],
]
for i, row in enumerate(p2, 8):
    ws3.row_dimensions[i].height = 22
    for j, v in enumerate(row):
        c = ws3.cell(i, j+1)
        c.value = v
        c.font = BOLD if j%2==0 else REG
        c.fill = FILL_LIGHT if j%2==0 else None
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB

ws3.merge_cells('A15:H15')
ws3.row_dimensions[15].height = 25
c = ws3['A15']
c.value = '二、敏感性分析 - 不同储能卖电价的IRR（1MW+3MWh）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

sp_list = [0.50, 0.55, 0.60, 0.65, 0.70, 0.80, 1.00]
hdrs2 = ['储能卖电价'] + [f'{s}元' for s in sp_list]
for j, h in enumerate(hdrs2, 1):
    c = ws3.cell(16, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

pv_cost, st_cost, total_inv = 200, 300, 500
sell_p = 0.25
def calc_m2_irr(sp):
    cfs = [-total_inv]
    for y in range(1, 21):
        pv = 1000*1200*0.80*((1-0.02)**(y-1))/10000
        sd = 3000*0.95*((1-0.015)**(y-1))*0.88*330/10000
        need = sd/0.88
        direct = max(pv-need, 0)
        cfs.append(direct*sell_p + sd*sp)
    rate = 0.1
    for _ in range(1000):
        f = sum(cf/(1+rate)**t for t,cf in enumerate(cfs))
        df = sum(-t*cf/(1+rate)**(t+1) for t,cf in enumerate(cfs))
        if abs(df)<1e-12: break
        rate -= f/df
        if abs(f)<1e-10: break
    return rate

ws3.cell(17, 1).value = 'IRR'
ws3.cell(17, 1).font = BOLD
ws3.cell(17, 1).fill = FILL_LIGHT
ws3.cell(17, 1).border = TB
for j, sp in enumerate(sp_list, 2):
    irr = calc_m2_irr(sp)
    c = ws3.cell(17, j)
    c.value = f'{irr:.2%}'
    c.fill = FILL_GREEN if irr>=0.10 else (FILL_YELLOW if irr>=0.08 else FILL_RED)
    c.font = Font(bold=True, size=10, color='006100' if irr>=0.10 else ('9C6500' if irr>=0.08 else '9C0006'))
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

ws3.row_dimensions[17].height = 20

ws3.merge_cells('A20:H20')
ws3.row_dimensions[20].height = 25
c = ws3['A20']
c.value = '三、多配置对比（1MW光伏配不同储能）'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

cmp = [
    ('配置','储能','总投资','光伏利用','IRR=8%门槛','IRR=10%门槛','推荐'),
    ('A','1MWh','300万','32.7%','0.670元','0.845元','⚡⚡⚡'),
    ('B','2MWh','400万','65.4%','0.682元','0.799元','⚡⚡'),
    ('C','3MWh','500万','98.0%','0.526元','0.635元','✅✅✅最优'),
    ('D','4MWh','600万','131%','0.860元','0.981元','❌'),
]
for i, row in enumerate(cmp, 21):
    ws3.row_dimensions[i].height = 20
    for j, v in enumerate(row, 1):
        c = ws3.cell(i, j)
        c.value = v
        c.font = HDR_TXT if i==21 else Font(size=10)
        c.fill = FILL_BLUE if i==21 else None
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB
        if i==21:
            c.font = HDR_TXT
        elif '最优' in str(v):
            c.fill = FILL_GREEN
        elif '❌' in str(v):
            c.fill = FILL_RED

ws3.merge_cells('A27:H27')
ws3.row_dimensions[27].height = 25
c = ws3['A27']
c.value = '四、全国各省模式二IRR预估'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

m2_provinces = [
    ('河北','第一梯队',FILL_GREEN,'光伏装机大，午间低电价明确'),
    ('山东','第一梯队',FILL_GREEN,'分布式光伏全国第二，尖峰机制完善'),
    ('河南','第一梯队',FILL_GREEN,'户用光伏全国第一，2025底试运行'),
    ('安徽','第一梯队',FILL_GREEN,'工商业分布式全国前三'),
    ('江苏','第一梯队',FILL_GREEN,'分布式光伏全国第一，2025底试运行'),
    ('浙江','第二梯队',FILL_YELLOW,'光伏装机大，现货市场成熟'),
    ('广东','第二梯队',FILL_YELLOW,'珠三角价差高，现货已开'),
    ('蒙西','第二梯队',FILL_YELLOW,'新能源大省，午间低电价明显'),
    ('山西','第一梯队',FILL_GREEN,'新能源大省，午间低谷明确'),
    ('湖北','第二梯队',FILL_YELLOW,'光伏装机中等，现货已开'),
    ('上海','第二梯队',FILL_YELLOW,'峰谷价差全国最高，光伏资源弱'),
    ('福建','第二梯队',FILL_YELLOW,'峰谷价差较大'),
    ('四川','第二梯队',FILL_YELLOW,'水电为主，气象联动尖峰'),
    ('湖南','第二梯队',FILL_YELLOW,'尖峰时段较长'),
    ('陕西','第二梯队',FILL_YELLOW,'新能源快速发展'),
    ('天津','第二梯队',FILL_YELLOW,'京津冀一体化'),
    ('北京','第二梯队',FILL_YELLOW,'基本无光伏资源'),
    ('海南','第二梯队',FILL_YELLOW,'独立地理单元'),
    ('辽宁','第三梯队',FILL_RED,'东北电网，冬季复杂'),
    ('吉林','第三梯队',FILL_RED,'风电为主，峰谷差低'),
    ('黑龙江','第三梯队',FILL_RED,'峰谷差低，风电为主'),
    ('江西','第二梯队',FILL_YELLOW,'深谷机制完善'),
    ('宁夏','第三梯队',FILL_RED,'现场消纳要求高'),
    ('甘肃','第二梯队',FILL_YELLOW,'现货价格低，新能源强配储'),
    ('重庆','第三梯队',FILL_RED,'山地地形，分布式有限'),
    ('贵州','第三梯队',FILL_RED,'峰谷调整中，分布式受限'),
    ('云南','第三梯队',FILL_RED,'水电为主，套利空间有限'),
    ('青海','第三梯队',FILL_RED,'峰谷差极低，集中式为主'),
    ('新疆','第三梯队',FILL_RED,'集中式为主，弃光问题'),
    ('西藏','第三梯队',FILL_RED,'电网薄弱，市场未开'),
    ('内蒙古东部','第三梯队',FILL_RED,'与蒙西分开管理'),
]

ph = ['省份','梯队','备注']
for j, h in enumerate(ph, 1):
    c = ws3.cell(28, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

for i, (name, tier, tier_fill, note) in enumerate(m2_provinces, 29):
    ws3.row_dimensions[i].height = 18
    for j, v in enumerate([name, tier, note], 1):
        c = ws3.cell(i, j)
        c.value = v
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB
        if j == 2:
            c.fill = tier_fill
            c.font = Font(bold=True, size=9)
        elif j == 3:
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ==================== Sheet4 现货市场 ====================
ws4 = wb.create_sheet('现货市场价格')
for col in 'ABCDEFG':
    ws4.column_dimensions[col].width = 16

ws4.merge_cells('A1:G1')
ws4.row_dimensions[1].height = 40
c = ws4['A1']
c.value = '各省电力现货市场价格数据（2025-2026年）'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

hdrs4 = ['省份','午间低电价','晚高峰','峰谷价差','负电价','模式一IRR≥10%?','模式一回收≤7年?']
for j, h in enumerate(hdrs4, 1):
    c = ws4.cell(3, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = PatternFill()
    c.border = TB
ws4.row_dimensions[3].height = 30

mkt = [
    ('山东','0.02~0.05','0.40~0.50','0.35~0.57','频繁(-0.07)','❌不满足','⚠️勉强'),
    ('山西','0~0.01','0.40~0.80','0.40~1.50','频繁零/负','✅满足','✅满足'),
    ('广东','0.02~0.05','0.45~0.60','0.40~1.29','极少','✅满足','✅满足'),
    ('甘肃','0.04','0.30~0.40','0.25~0.46','地板频繁','❌不满足','❌不满足'),
    ('浙江','0.02~0.05','0.40~0.50','0.35~1.27','频繁(-0.20)','✅满足','✅满足'),
]
for i, row in enumerate(mkt, 4):
    ws4.row_dimensions[i].height = 25
    for j, v in enumerate(row, 1):
        c = ws4.cell(i, j)
        c.value = v
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB
        if j == 6:
            c.fill = FILL_GREEN if '✅' in v else (FILL_YELLOW if '勉强' in v else FILL_RED)
            c.font = Font(bold=True, size=10, color='006100' if '✅' in v else ('9C6500' if '勉强' in v else '9C0006'))
        elif j == 7:
            c.fill = FILL_GREEN if '✅' in v else (FILL_YELLOW if '勉强' in v else FILL_RED)
            c.font = Font(bold=True, size=10, color='006100' if '✅' in v else ('9C6500' if '勉强' in v else '9C0006'))

ws4.merge_cells('A10:G10')
c = ws4['A10']
c.value = '说明：电价单位为元/kWh；数据来源：各省电力交易中心2025-2026年公开数据'
c.font = Font(size=9, italic=True, color='666666')

# ==================== Sheet5 省份梯队 ====================
ws5 = wb.create_sheet('省份梯队划分')
for col in 'ABCDEFGH':
    ws5.column_dimensions[col].width = 16

ws5.merge_cells('A1:H1')
ws5.row_dimensions[1].height = 40
c = ws5['A1']
c.value = '全国31省储能市场梯队划分（2026年）'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

# 模式一
ws5.merge_cells('A3:H3')
ws5.row_dimensions[3].height = 28
c = ws5['A3']
c.value = '模式一梯队（纯储能10MWh）'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='375623')
c.alignment = Alignment(horizontal='center', vertical='center')

t1 = [
    ('梯队','省份','现货市场','峰谷价差','模式二梯队','省份数','核心优势','IRR估算'),
    ('⭐第一梯队','山西/广东/浙江/山东/江苏/河南/湖北/安徽/河北','已开/即将开','≥0.67元','⭐第一/第二梯队','9省','峰谷套利空间大，现货市场成熟','IRR≥10%'),
    ('⚠️第二梯队','蒙西/甘肃/辽宁/吉林/黑龙江/陕西/宁夏/青海/上海/福建/四川/湖南/天津/北京/海南/江西','即将开/已开','0.4~0.7元','⚠️第二/第三梯队','16省','市场待激活，有一定套利空间','IRR 5~10%'),
    ('🔶第三梯队','贵州/云南/青海/新疆/西藏/内蒙古东部/重庆','未开/待查','<0.5元','🔶第三梯队','6省','市场不成熟，暂不推荐','IRR<5%'),
]
for i, row in enumerate(t1, 4):
    ws5.row_dimensions[i].height = 30
    for j, v in enumerate(row, 1):
        c = ws5.cell(i, j)
        c.value = v
        c.font = HDR_TXT if i==4 else Font(size=10)
        c.fill = FILL_BLUE if i==4 else None
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = PatternFill()
        c.border = TB
        if i == 5 and j == 1:
            c.fill = FILL_GREEN
        elif i == 6 and j == 1:
            c.fill = FILL_YELLOW
        elif i == 7 and j == 1:
            c.fill = FILL_RED

# 模式二
ws5.merge_cells('A10:H10')
ws5.row_dimensions[10].height = 28
c = ws5['A10']
c.value = '模式二梯队（光储一体化1MW+3MWh）'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='375623')
c.alignment = Alignment(horizontal='center', vertical='center')

t2 = [
    ('梯队','省份','光伏装机','现货市场','午间低电价','省份数','IRR估算',''),
    ('⭐第一梯队','河北/山东/河南/安徽/江苏/山西','极大（全国前五）','已开/即将开','明确（<0.1元）','8省','IRR≥10%',''),
    ('⚠️第二梯队','浙江/广东/蒙西/湖北/上海/福建/四川/湖南/陕西/天津/北京/海南/江西','大~中','即将开/已开','较明确','13省','IRR 6~10%',''),
    ('🔶第三梯队','辽宁/吉林/黑龙江/宁夏/甘肃/重庆/贵州/云南/青海/新疆/西藏/内蒙古东部','小~中','未开/待查','不明确','10省','IRR<6%',''),
]
for i, row in enumerate(t2, 11):
    ws5.row_dimensions[i].height = 28
    for j, v in enumerate(row, 1):
        c = ws5.cell(i, j)
        c.value = v
        c.font = HDR_TXT if i==11 else Font(size=10)
        c.fill = FILL_BLUE if i==11 else None
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = PatternFill()
        c.border = TB
        if i == 12 and j == 1:
            c.fill = FILL_GREEN
        elif i == 13 and j == 1:
            c.fill = FILL_YELLOW
        elif i == 14 and j == 1:
            c.fill = FILL_RED

# 全国31省完整表格
ws5.merge_cells('A17:H17')
ws5.row_dimensions[17].height = 28
c = ws5['A17']
c.value = '全国31省详细梯队对照表'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

all_prov = [
    ('山西省','华北','✅已开','高(≥0.8)','大','⭐第一','⭐第一',''),
    ('广东省','华南','✅已开','高(1.2~1.4)','大','⭐第一','⭐第一',''),
    ('浙江省','华东','✅已开','高(1.0~1.4)','大','⭐第一','⭐第一',''),
    ('山东省','华东','✅已开','高(0.8~1.0)','大','⭐第一','⭐第一',''),
    ('江苏省','华东','🔜即将开','中(0.6~0.9)','极大','⭐第一','⭐第一',''),
    ('河南省','华中','🔜即将开','高(0.7~1.0)','大','⭐第一','⭐第一',''),
    ('安徽省','华东','🔜即将开','高(0.7~1.0)','大','⭐第一','⭐第一',''),
    ('河北省','华北','🔜即将开','高(0.7~1.0)','大','⭐第一','⭐第一',''),
    ('湖北省','华中','✅已开','高(0.7~1.0)','中','⭐第一','⚠️第二',''),
    ('蒙西(内蒙古)','华北','✅已开','低~中(0.3~0.5)','大','⚠️第二','⚠️第二',''),
    ('甘肃省','西北','✅已开','低(0.3~0.5)','中~大','⚠️第二','⚠️第二',''),
    ('上海市','华东','🔜即将开','高(1.0~1.8)','小','⚠️第二','⚠️第二',''),
    ('福建省','华东','🔜即将开','高(0.7~1.0)','中','⚠️第二','⚠️第二',''),
    ('四川省','西南','🔜即将开','高(0.7~1.0)','中','⚠️第二','⚠️第二',''),
    ('湖南省','华中','🔜即将开','高(0.7~1.0)','中','⚠️第二','⚠️第二',''),
    ('天津市','华北','🔜即将开','高(0.7~1.0)','小','⚠️第二','⚠️第二',''),
    ('北京市','华北','🔜即将开','高(0.7~1.0)','极小','⚠️第二','⚠️第二',''),
    ('海南省','华南','🔜即将开','高(0.7~1.0)','小','⚠️第二','⚠️第二',''),
    ('陕西省','西北','🔜即将开','高(0.7~1.0)','中~大','⚠️第二','⚠️第二',''),
    ('江西省','华东','🔜即将开','中(0.5~0.7)','中','⚠️第二','⚠️第二',''),
    ('辽宁省','东北','🔜即将开','低~中(0.4~0.6)','中~小','⚠️第二','🔶第三',''),
    ('吉林省','东北','🔜即将开','低~中(0.4~0.6)','小','⚠️第二','🔶第三',''),
    ('黑龙江省','东北','🔜即将开','低~中(0.4~0.6)','小','🔶第三','🔶第三',''),
    ('宁夏','西北','🔜即将开','低(0.3~0.5)','中~大','⚠️第二','🔶第三',''),
    ('重庆市','西南','🔜即将开','高(0.7~1.0)','小','⚠️第二','🔶第三',''),
    ('贵州省','西南','未列入394号文','中(0.5~0.7)','小','🔶第三','🔶第三',''),
    ('云南省','西南','未列入394号文','中(0.5~0.7)','中~大','🔶第三','🔶第三',''),
    ('青海省','西北','🔜即将开','低(0.3~0.5)','中~大','🔶第三','🔶第三',''),
    ('内蒙古东部','华北','🔜即将开','低~中(0.3~0.5)','中','🔶第三','🔶第三',''),
    ('新疆','西北','🔜即将开','中(0.5~0.7)','极大','🔶第三','🔶第三','集中式为主'),
    ('西藏','西南','❌未开','待查','小','🔶第三','🔶第三','电网薄弱'),
]

hdr_all = ['省份','区域','现货市场','峰谷价差','光伏装机','模式一定位','模式二定位','备注']
for j, h in enumerate(hdr_all, 1):
    c = ws5.cell(18, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

for i, row in enumerate(all_prov, 19):
    ws5.row_dimensions[i].height = 18
    for j, v in enumerate(row, 1):
        c = ws5.cell(i, j)
        c.value = v
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
        c.border = TB
        if j == 6:
            if '第一' in str(v):
                c.fill = FILL_GREEN
                c.font = Font(bold=True, size=9)
            elif '第二' in str(v):
                c.fill = FILL_YELLOW
                c.font = Font(bold=True, size=9)
            else:
                c.fill = FILL_RED
                c.font = Font(bold=True, size=9)
        elif j == 7:
            if '第一' in str(v):
                c.fill = FILL_GREEN
                c.font = Font(bold=True, size=9)
            elif '第二' in str(v):
                c.fill = FILL_YELLOW
                c.font = Font(bold=True, size=9)
            else:
                c.fill = FILL_RED
                c.font = Font(bold=True, size=9)

# ==================== Sheet6 全国投资地图 ====================
ws6 = wb.create_sheet('全国投资地图')
for col in 'ABCDEF':
    ws6.column_dimensions[col].width = 22

ws6.merge_cells('A1:F1')
ws6.row_dimensions[1].height = 40
c = ws6['A1']
c.value = '全国投资地图 - 七大区域省份梯队分布'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

zone_map = [
    ('区域','第一梯队（⭐优先开拓）','第二梯队（⚠️储备观望）','第三梯队（🔶暂不推荐）','区域特点','投资优先级'),
    ('华北', '山西/河北', '山东/天津/北京/蒙西', '内蒙古东部', '山西领跑现货，河北山东即将开', '⭐⭐⭐'),
    ('华东', '江苏/安徽/浙江/山东', '上海/福建/江西', '', '全国最成熟市场，6省优质', '⭐⭐⭐'),
    ('华中', '河南/湖北', '湖南', '', '湖北已开，河南潜力大', '⭐⭐'),
    ('华南', '广东', '海南', '', '广东价差全国最高', '⭐⭐'),
    ('西南', '', '四川', '贵州/云南/重庆/西藏', '四川可布局，其余暂缓', '⭐'),
    ('东北', '', '辽宁', '吉林/黑龙江', '辽宁略优，吉林黑观望', '⭐'),
    ('西北', '', '陕西/甘肃', '宁夏/青海/新疆', '陕西可关注，甘肃新疆暂缓', '⭐'),
]

for i, row in enumerate(zone_map, 3):
    ws6.row_dimensions[i].height = 35
    for j, v in enumerate(row, 1):
        c = ws6.cell(i, j)
        c.value = v
        c.font = HDR_TXT if i==3 else Font(size=10)
        c.fill = FILL_BLUE if i==3 else None
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = PatternFill()
        c.border = TB
        if i == 3:
            c.font = HDR_TXT
        elif j == 2 and v and v != '':
            c.fill = FILL_GREEN
        elif j == 3 and v and v != '':
            c.fill = FILL_YELLOW
        elif j == 4 and v and v != '':
            c.fill = FILL_RED

# 图例
ws6.merge_cells('A12:F12')
ws6.row_dimensions[12].height = 25
c = ws6['A12']
c.value = '图例说明'
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = FILL_BLUE2
c.alignment = Alignment(horizontal='center', vertical='center')

legends = [
    ('✅第一梯队（绿色）','FILL_GREEN','山西/广东/浙江/山东/江苏/河南/安徽/河北/湖北，可立即开拓，IRR预计≥10%'),
    ('⚠️第二梯队（黄色）','FILL_YELLOW','上海/福建/四川/湖南/陕西等，市场待激活，有一定套利空间，IRR预计5~10%'),
    ('🔶第三梯队（红色）','FILL_RED','甘肃/贵州/云南/青海/新疆/西藏/吉林/黑龙江等，暂不推荐，IRR预计<5%'),
    ('空白','','该区域无对应梯队省份'),
]
for i, (txt, fill_info, note) in enumerate(legends, 13):
    ws6.row_dimensions[i].height = 22
    c = ws6.cell(i, 1)
    c.value = txt
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB
    if fill_info == 'FILL_GREEN':
        c.fill = FILL_GREEN
    elif fill_info == 'FILL_YELLOW':
        c.fill = FILL_YELLOW
    elif fill_info == 'FILL_RED':
        c.fill = FILL_RED
    ws6.merge_cells(f'B{i}:F{i}')
    c = ws6.cell(i, 2)
    c.value = note
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill = PatternFill()
    c.border = TB

# 重点开拓省份汇总
ws6.merge_cells('A19:F19')
ws6.row_dimensions[19].height = 28
c = ws6['A19']
c.value = '⭐ 重点开拓省份汇总（双第一梯队）'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='375623')
c.alignment = Alignment(horizontal='center', vertical='center')

priority = [
    ('山西','现货已开+峰谷价差0.4~1.5元+光伏装机全国前列','⭐⭐⭐第一优先'),
    ('广东','现货已开+峰谷价差1.2~1.4元全国最高','⭐⭐⭐第一优先'),
    ('浙江','现货已开+峰谷价差1.0~1.4元+分布式光伏第三','⭐⭐⭐第一优先'),
    ('山东','现货已开+峰谷价差0.8~1.0元+分布式光伏第二','⭐⭐⭐第一优先'),
    ('江苏','2025年底试运行+分布式光伏全国第一','⭐⭐⭐第一优先'),
    ('河南','2025年底试运行+户用光伏全国第一+光伏第四','⭐⭐⭐第一优先'),
    ('安徽','2026年中正式运行+工商业分布式全国前三','⭐⭐第一优先'),
    ('河北','2025年底试运行+光伏装机大省+京津冀政策加持','⭐⭐第一优先'),
]
ph = ['省份','核心优势','优先级']
for j, h in enumerate(ph, 1):
    c = ws6.cell(20, j)
    c.value = h
    c.font = HDR_TXT
    c.fill = FILL_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill()
    c.border = TB

for i, row in enumerate(priority, 21):
    ws6.row_dimensions[i].height = 25
    for j, v in enumerate(row, 1):
        c = ws6.cell(i, j)
        c.value = v
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = PatternFill()
        c.border = TB
        if j == 1:
            c.fill = FILL_GREEN
            c.font = Font(bold=True, size=10)
        elif j == 3:
            c.fill = FILL_GREEN

# ==================== Sheet7 综合结论 ====================
ws7 = wb.create_sheet('综合结论')
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 70

ws7.merge_cells('A1:B1')
ws7.row_dimensions[1].height = 40
c = ws7['A1']
c.value = '综合结论与开拓建议'
c.font = Font(bold=True, size=16, color='FFFFFF')
c.fill = FILL_BLUE
c.alignment = Alignment(horizontal='center', vertical='center')

conclusions = [
    ('政策背景', '2025年4月国家发改委394号文发布，全国电力现货市场加速推进。浙江、山西、广东已正式运行，山东即将转正，江苏/河南/湖北/安徽等2025-2026年陆续启动。工商业行政性峰谷电价逐步取消，改为市场化定价。'),
    ('模式一定论', '纯储能10MWh，主推山西/广东/浙江/山东/江苏/河南/湖北/安徽/河北9省。IRR=10%门槛：电价差需≥0.6618元/kWh；回收期≤7年门槛：电价差需≥0.5795元/kWh。'),
    ('模式二结论', '推荐1MW+3MWh配置（光伏利用率98%），主推河北/山东/河南/安徽/江苏/山西8省。IRR=10%门槛：储能卖电价需≥0.635元/kWh；IRR=8%门槛：需≥0.526元/kWh。'),
    ('⭐重点开拓', '山西（双第一梯队+现货最成熟）、广东（峰谷价差全国最高）、浙江/山东（现货+分布式双优）、江苏/河南/安徽/河北（光伏大省+2025底启动）。'),
    ('开拓策略', '第一阶段（2026上半年）：山西/浙江/广东/山东先行，以模式一为主。第二阶段（2026下半年）：江苏/河南/安徽/河北跟进，布局模式二。第三阶段（2027+）：湖北/四川/陕西等第二梯队。'),
    ('风险提示', '1.电价差收窄风险（新能源装机持续增加拉低高峰电价）；2.市场价格波动风险（现货价格不确定性）；3.政策变化风险（行政干预可能重启）；4.业主违约风险（用电量下降）；5.电池衰减风险（10年运营期后容量显著下降）。'),
    ('数据说明', '现货价格数据来自各省电力交易中心2025-2026年公开数据；峰谷价差来自储能头条2025年各月统计；光伏装机数据来自国家能源局2025年统计。'),
]
for i, (k, v) in enumerate(conclusions, 3):
    ws7.row_dimensions[i].height = 40
    c = ws7.cell(i, 1)
    c.value = k
    c.font = Font(bold=True, size=11)
    c.fill = FILL_LIGHT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = PatternFill()
    c.border = TB
    c = ws7.cell(i, 2)
    c.value = v
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill = PatternFill()
    c.border = TB

wb.save(OUTPUT)
print(f'✅ 最终版Excel已生成: {OUTPUT}')
