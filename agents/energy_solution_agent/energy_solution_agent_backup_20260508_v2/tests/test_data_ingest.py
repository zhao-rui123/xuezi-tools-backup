from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from energy_solution_agent.data_ingest import _load_numeric_series, ingest_external_series
from energy_solution_agent.material_extract import parse_calb_user_storage_workbook
from energy_solution_agent.province_cycle_rules import parse_province_cycle_rules_workbook
from energy_solution_agent.province_tou_schedule import parse_province_tou_schedule_workbook


class DataIngestExcelSeriesTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = Path(tempfile.mkdtemp(prefix="data_ingest_"))

    def tearDown(self) -> None:
        if self.tmpdir.exists():
            shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_loads_grid_company_day_matrix_excel(self) -> None:
        path = self.tmpdir / "matrix.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "单位名称", "终端地址码", "电能表资产编号", "电能表运行时间", "物联关系序号", "数据日期", "数据类型", "00:00", "00:15", "00:30", "00:45"])
        ws.append([1, "A", "B", "C", "D", "E", "2025-01-01", "总有功功率", 100, 110, 120, 130])
        wb.save(path)

        values = _load_numeric_series(path)
        self.assertEqual(values, [100.0, 110.0, 120.0, 130.0])

    def test_loads_timestamp_value_excel(self) -> None:
        path = self.tmpdir / "timeseries.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["有功功率"])
        ws.append(["15分钟 ; 实际值"])
        ws.append(["时间", "电表名称", "通讯地址", "总有功功率(kW)"])
        ws.append(["2025-10-31 23:45", "X", "Y", 900])
        ws.append(["2025-10-31 23:59", "X", "Y", 1000])
        wb.save(path)

        values = _load_numeric_series(path)
        self.assertEqual(values, [1000.0, 900.0])

    def test_ingests_spot_price_workbook_into_market_data(self) -> None:
        path = self.tmpdir / "spot_prices.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append([None, "星期四", None, "星期五", None])
        ws.append([None, "日前", "实时", "日前", "实时"])
        ws.append(["时段", "2026-01-01", "2026-01-01", "2026-01-02", "2026-01-02"])
        for hour in range(24):
            ws.append(
                [
                    f"{hour:02d}:00-{hour + 1:02d}:00",
                    300 + hour,
                    400 + hour,
                    500 + hour,
                    600 + hour,
                ]
            )
        wb.save(path)

        payload = {"market_data": {"market_price_series_path": str(path)}}
        enriched = ingest_external_series(payload, base_dir=self.tmpdir)
        self.assertEqual(len(enriched["market_data"]["market_price_series"]), 48)
        self.assertEqual(enriched["market_data"]["market_price_series"][:3], [400.0, 401.0, 402.0])
        self.assertEqual(len(enriched["market_data"]["spot_price_daily_profiles"]), 2)
        self.assertEqual(enriched["market_data"]["spot_price_daily_profiles"][0]["date"], "2026-01-01")
        self.assertEqual(enriched["market_data"]["day_ahead_market_price_series"][:2], [300.0, 301.0])

    def test_ingests_spot_price_workbook_with_multiple_month_blocks(self) -> None:
        path = self.tmpdir / "spot_prices_multi.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append([None, "星期四", None])
        ws.append([None, "日前", "实时"])
        ws.append(["时段", "2026-01-01", "2026-01-01"])
        for hour in range(24):
            ws.append([f"{hour:02d}:00-{hour + 1:02d}:00", 100 + hour, 200 + hour])
        ws.append(["均值", 0, 0])
        ws.append([None, None, None])
        ws.append([None, "星期日", None])
        ws.append([None, "日前", "实时"])
        ws.append(["时段", "2026-02-01", "2026-02-01"])
        for hour in range(24):
            ws.append([f"{hour:02d}:00-{hour + 1:02d}:00", 300 + hour, 400 + hour])
        ws.append(["均值", 0, 0])
        wb.save(path)

        payload = {"market_data": {"market_price_series_path": str(path)}}
        enriched = ingest_external_series(payload, base_dir=self.tmpdir)
        self.assertEqual(len(enriched["market_data"]["spot_price_daily_profiles"]), 2)
        self.assertEqual(enriched["market_data"]["spot_price_daily_profiles"][1]["date"], "2026-02-01")
        self.assertEqual(enriched["market_data"]["market_price_series"][0], 200.0)
        self.assertEqual(enriched["market_data"]["market_price_series"][-1], 423.0)

    def test_parse_calb_user_storage_workbook_extracts_arbitrage_plan(self) -> None:
        path = self.tmpdir / "calb.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["瑁呮満瑙勬ā", "鍔熺巼", 125, "KW"])
        ws.append(["瀹归噺", 261, "KWh"])
        ws.append(["灏栧嘲鐢典环", 1.34991333333333, "鍏/KWh"])
        ws.append(["宄扮數浠?", 1.112447, "鍏/KWh"])
        ws.append(["骞崇數浠?", 0.708644, "鍏/KWh"])
        ws.append(["璋风數浠?", 0.345405, "鍏/KWh"])
        ws.append(["璋?灏栧嘲娆℃暟", 82.5, "娆?"])
        ws.append(["璋?宄版鏁?", 440, "娆?"])
        ws.append(["璋?骞虫鏁?", 137.5, "娆?"])
        ws.append(["绗竴娆℃斁鐢靛閲?", 253.17, "kWh"])
        ws.append(["绗簩娆℃斁鐢靛閲?", 253.17, "kWh"])
        ws.append(["绯荤粺鍏呯數鏁堢巼", 0.941695806555])
        ws.append(["绯荤粺鏀剧數鏁堢巼", 0.941695806555])
        wb.save(path)

        parsed = parse_calb_user_storage_workbook(path)
        self.assertEqual(parsed["power_kw"], 125.0)
        self.assertEqual(parsed["energy_kwh"], 261.0)
        self.assertEqual(len(parsed["arbitrage_plan"]["cycles"]), 3)

    def test_parse_province_cycle_rules_workbook_extracts_deep_valley_rule(self) -> None:
        path = self.tmpdir / "province_rules.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["地区", "10kV电价", None, "当地分时电价", None, "电价差"])
        ws.append([None, "（元/kWh）", None, "支持充放策略", None, None])
        ws.append([None, None, None, "（全年330天）", None, None])
        ws.append(["山东", "尖", 1.1058, "谷充峰放次数", 82.5, 0.6925])
        ws.append([None, "峰", 0.9574, "平充峰放次数", 0, None])
        ws.append([None, "平", 0.6110, "谷充尖放次数", 82.5, None])
        ws.append([None, "谷", 0.2649, "平充尖放次数", 0, None])
        ws.append([None, "深谷", 0.1658, "深谷充尖放次数", 247.5, None])
        wb.save(path)

        parsed = parse_province_cycle_rules_workbook(path, "山东")
        self.assertIsNotNone(parsed)
        self.assertEqual(len(parsed["tou_tariff"]), 5)
        self.assertTrue(any(cycle["charge_period"] == "deep_valley" for cycle in parsed["arbitrage_plan"]["cycles"]))

    def test_parse_province_tou_schedule_workbook_extracts_monthly_schedule(self) -> None:
        path = self.tmpdir / "tou_schedule.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "上海"
        ws.append(["上海4月份电价分析"])
        ws.append(["7-8月份"])
        ws.append(["根据上海分时电价确定基础充放电区间控制逻辑，22:00-06:00 谷值时段充电，12:00-14:00 尖值时段放电，15:00-18:00 平值时段充电，18:00-21:00 峰值时段放电。"])
        ws.append(["其他月份"])
        ws.append(["根据上海分时电价确定基础充放电区间控制逻辑，22:00-06:00 谷值时段充电，8:00-11:00 峰值时段放电，11:00-18:00 平值时段充电，18:00-21:00 峰值时段放电。"])
        wb.save(path)

        parsed = parse_province_tou_schedule_workbook(path, "上海")
        self.assertIsNotNone(parsed)
        self.assertTrue(any(item["month"] == 7 for item in parsed))
        self.assertTrue(any("super_peak" in item["periods"] for item in parsed if item["month"] == 7))

    def test_parse_province_tou_schedule_workbook_supports_inline_calb_style_rows(self) -> None:
        path = self.tmpdir / "tou_schedule_inline.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "上海"
        ws.append(["上海充放电策略"])
        ws.append(["7、8月份", "22:00-06:00 谷值时段充电，12:00-14:00 尖值时段放电；15:00-18:00 平值时段充电，18:00-21:00 峰值时段放电"])
        ws.append(["9月份", "22:00-06:00 谷值时段充电，8:00-15:00 尖值时段放电；15:00-18:00 平值时段充电，18:00-21:00 峰值时段放电"])
        ws.append(["1、12月", "22:00-06:00 谷值时段充电，8:00-11:00 峰值时段放电；11:00-18:00 平值时段充电，19:00-21:00 尖值时段放电"])
        ws.append(["其他月份", "22:00-06:00 谷值时段充电，08:00-11:00 峰值时段放电；11:00-18:00 平值时段充电，18:00-21:00 峰值时段放电"])
        wb.save(path)

        parsed = parse_province_tou_schedule_workbook(path, "上海")
        self.assertIsNotNone(parsed)
        july = next(item for item in parsed if item["month"] == 7)
        september = next(item for item in parsed if item["month"] == 9)
        january = next(item for item in parsed if item["month"] == 1)
        march = next(item for item in parsed if item["month"] == 3)
        self.assertIn("super_peak", july["periods"])
        self.assertEqual(september["schedule"]["super_peak"], list(range(8, 15)))
        self.assertEqual(january["schedule"]["peak"], [8, 9, 10])
        self.assertEqual(march["schedule"]["peak"], [8, 9, 10, 18, 19, 20])
