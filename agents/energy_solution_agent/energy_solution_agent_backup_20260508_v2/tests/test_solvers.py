from __future__ import annotations

import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from energy_solution_agent.engine import analyze_project
from energy_solution_agent.live_rules import extract_structured_rule_patch
from energy_solution_agent.network_http import get_proxy_url
from energy_solution_agent.policy_classify import classify_market_policy_mode
from energy_solution_agent.resource_fetch import enrich_with_auto_resource_data
from energy_solution_agent.resource_models import estimate_pv_generation, estimate_wind_generation
from energy_solution_agent.tou_policy_fetch import enrich_with_monthly_tou_policy_history
from energy_solution_agent.annual_series import build_annual_series, extrapolate_sample_period_series
from energy_solution_agent.solvers import apply_storage_product_selection, estimate_storage, infer_cycles_from_monthly_tou_history, infer_daily_cycles_from_schedule, simulate_annual_cycle_value, simulate_commercial_hybrid_value, simulate_rule_based_arbitrage, simulate_spot_intraday_arbitrage, simulate_storage_dispatch_annual
from energy_solution_agent.settlement import build_hourly_price_series, resolve_price_series_start_weekday


# TODO: Extract shared payload patterns (e.g. storage config, tariff data,
# equipment defaults) into reusable fixtures to reduce duplication across
# the test methods below.
class SolverAnnualDispatchTest(unittest.TestCase):
    def test_annual_dispatch_uses_full_series_horizon(self) -> None:
        load_series = [100.0] * 8760
        pv_series = [20.0] * 8760

        dispatch = simulate_storage_dispatch_annual(
            load_series_kw=load_series,
            pv_series_kw=pv_series,
            wind_series_kw=[0.0] * 8760,
            charging_series_kw=[0.0] * 8760,
            thermal_series_kw=[0.0] * 8760,
            storage_power_mw=None,
            storage_energy_mwh=None,
        )

        expected_purchase_mwh = (100.0 - 20.0) * 8760 / 1000
        self.assertEqual(len(dispatch["baseline_grid_series_kw"]), 8760)
        self.assertAlmostEqual(dispatch["annual_grid_purchase_mwh"], expected_purchase_mwh, places=6)

    def test_build_annual_series_accepts_15min_year_series(self) -> None:
        raw = [4.0] * 35040
        series = build_annual_series(raw_series=raw, fallback_daily=[1.0] * 24)
        self.assertEqual(len(series), 8760)
        self.assertTrue(all(value == 4.0 for value in series[:24]))

    def test_build_annual_series_can_pad_near_complete_15min_series(self) -> None:
        raw = [4.0] * 34980
        series = build_annual_series(raw_series=raw, fallback_daily=[1.0] * 24)
        self.assertEqual(len(series), 8760)
        self.assertTrue(all(abs(value - 4.0) < 1e-9 for value in series[:24]))

    def test_extrapolate_sample_period_series_can_expand_sample_months(self) -> None:
        january_points = 31 * 96
        february_points = 28 * 96
        raw = [100.0] * january_points + [200.0] * february_points
        series = extrapolate_sample_period_series(raw, sample_months=[1, 2])
        self.assertIsNotNone(series)
        self.assertEqual(len(series), 365 * 96)
        self.assertEqual(series[0], 100.0)

    def test_extrapolate_sample_period_series_can_use_explicit_cross_year_month_map(self) -> None:
        march_points = 31 * 96
        january_points = 31 * 96
        raw = [300.0] * march_points + [100.0] * january_points
        series = extrapolate_sample_period_series(
            raw,
            sample_months=None,
            sample_month_map=[{"year": 2025, "month": 3, "days": 31}, {"year": 2026, "month": 1, "days": 31}],
        )
        self.assertIsNotNone(series)
        self.assertEqual(len(series), 365 * 96)
        self.assertEqual(series[:96], [100.0] * 96)

    def test_estimate_storage_can_use_sample_day_workbook_for_energy_sizing(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "annual-sample-days"
        ws.cell(3, 1, "充放电百分位")
        ws.cell(3, 3, 0.9)
        ws.cell(1, 1, "一充功率")
        ws.cell(1, 3, 125)
        ws.cell(2, 1, "一放功率")
        ws.cell(2, 3, 125)
        row8 = [None] * 120
        samples = [
            (5277.6, 948.0, 780.0, 4720.8),
            (6074.4, 1375.2, 1260.8, 8947.2),
            (10388.4, 1094.4, 3483.2, 4521.6),
        ]
        starts = [5, 14, 23]
        labels = ["第一日", "第二日", "第三日"]
        for start, label, values in zip(starts, labels, samples):
            row8[start - 1] = label
            row8[start + 2] = values[0]
            row8[start + 3] = values[1]
            row8[start + 4] = values[2]
            row8[start + 5] = values[3]
        for col, value in enumerate(row8, start=1):
            ws.cell(8, col, value)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            payload = {
                "project_info": {"scenario_type": "user_side_storage"},
                "load_data": {"sizing_workbook_path": tmp.name},
                "equipment": {
                    "storage": {
                        "power_candidate_kw": [125],
                        "sizing_target_day_coverage_ratio": 0.9,
                    }
                },
            }
            storage = estimate_storage(payload)
        self.assertEqual(storage["raw_storage_power_mw"], 0.125)
        self.assertEqual(storage["raw_storage_energy_mwh"], 0.219)

    def test_estimate_storage_can_prefer_workbook_coverage_ratio_and_power(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "annual-sample-days"
        ws.cell(3, 1, "充放电百分位")
        ws.cell(3, 3, 0.8)
        ws.cell(1, 1, "一充功率")
        ws.cell(1, 3, 125)
        ws.cell(2, 1, "一放功率")
        ws.cell(2, 3, 150)
        ws.cell(3, 5, "二充功率")
        ws.cell(3, 7, 200)
        ws.cell(4, 5, "二放功率")
        ws.cell(4, 7, 175)
        row8 = [None] * 120
        samples = [
            (5277.6, 948.0, 780.0, 4720.8),
            (6074.4, 1375.2, 1260.8, 8947.2),
            (10388.4, 1094.4, 3483.2, 4521.6),
        ]
        starts = [5, 14, 23]
        labels = ["第一日", "第二日", "第三日"]
        for start, label, values in zip(starts, labels, samples):
            row8[start - 1] = label
            row8[start + 2] = values[0]
            row8[start + 3] = values[1]
            row8[start + 4] = values[2]
            row8[start + 5] = values[3]
        for col, value in enumerate(row8, start=1):
            ws.cell(8, col, value)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            payload = {
                "project_info": {"scenario_type": "user_side_storage"},
                "load_data": {"sizing_workbook_path": tmp.name},
                "equipment": {"storage": {}},
            }
            storage = estimate_storage(payload)
        self.assertEqual(storage["raw_storage_power_mw"], 0.125)
        self.assertEqual(storage["raw_storage_energy_mwh"], 0.243)

    def test_estimate_storage_does_not_treat_load_series_excel_path_as_sample_day_workbook(self) -> None:
        payload = {
            "project_info": {"scenario_type": "user_side_storage"},
            "load_data": {
                "load_series_kw_path": r"C:\tmp\fake-load.xlsx",
                "peak_load_kw": 1000,
            },
            "equipment": {"storage": {"power_candidate_kw": [125]}},
        }
        storage = estimate_storage(payload)
        self.assertEqual(storage["raw_storage_power_mw"], 0.125)
        self.assertEqual(storage["raw_storage_energy_mwh"], 0.25)

    def test_offgrid_price_series_uses_fuel_cost(self) -> None:
        prices = build_hourly_price_series({"market_mode": "offgrid_internal", "fuel_cost_per_kwh": 0.3}, 48)
        self.assertEqual(prices, [0.3] * 48)

    def test_price_series_can_resolve_start_weekday_from_calendar_year(self) -> None:
        market = {
            "tou_tariff": [
                {"period": "peak", "price": 1.0},
                {"period": "flat", "price": 0.2},
            ],
            "tou_schedule": {"peak": [0], "flat": list(range(1, 24))},
            "tou_schedule_weekend": {"flat": list(range(24))},
        }

        start_weekday = resolve_price_series_start_weekday({"project_info": {"calendar_year": 2027}})
        prices = build_hourly_price_series(market, 48, start_weekday=start_weekday)

        self.assertEqual(start_weekday, 4)
        self.assertEqual(prices[0], 1.0)
        self.assertEqual(prices[24], 0.2)

    def test_offgrid_optimizer_can_choose_no_storage(self) -> None:
        payload = {
            "project_info": {
                "project_name": "offgrid-opt-test",
                "scenario_type": "microgrid",
                "province": "Overseas",
                "grid_connection_mode": "microgrid",
                "storage_strategy_mode": "renewable_priority",
            },
            "resource_data": {
                "solar": {
                    "annual_irradiation_kwh_per_m2": 2200,
                    "available_area_m2": 39000,
                    "performance_ratio": 0.82,
                }
            },
            "load_data": {
                "annual_consumption_mwh": 8760,
                "peak_load_kw": 1000,
            },
            "market_data": {
                "market_mode": "offgrid_internal",
            },
            "equipment": {
                "pv": {
                    "candidate_mwp": [0, 6],
                },
                "storage": {
                    "power_candidate_kw": [0, 2000],
                    "energy_candidate_kwh": [0, 8000],
                },
                "conventional_backup": {
                    "enabled": True,
                    "fuel_cost_per_kwh": 0.3,
                },
            },
            "financial": {
                "project_years": 15,
                "discount_rate": 0.08,
                "capex": {
                    "storage_system_cost_per_kwh": 5000,
                    "pv_cost_per_w": 0.5,
                    "storage_replacement_cost_ratio": 0.52,
                },
                "opex": {
                    "annual_om_ratio": 0.01,
                    "annual_opex_escalation_rate": 0.02,
                },
                "degradation": {
                    "storage_capacity_fade_per_year": 0.025,
                    "pv_degradation_per_year": 0.005,
                },
            },
        }

        output, _, _ = analyze_project(payload)
        self.assertEqual(output["project_summary"]["scenario_type"], "microgrid")
        self.assertEqual(output["recommended_solution"]["pv_mwp"], 6.0)
        self.assertIsNone(output["recommended_solution"]["storage_power_mw"])
        self.assertIsNone(output["recommended_solution"]["storage_energy_mwh"])
        expected_diesel_cost = output["simulation_results"]["annual_grid_purchase_mwh"] * 0.3 * 1000
        self.assertAlmostEqual(output["financial_results"]["annual_energy_charge_cost"], expected_diesel_cost, delta=1.5)

    def test_microgrid_can_auto_recommend_pv_tilt_from_coordinates(self) -> None:
        payload = {
            "project_info": {
                "project_name": "microgrid-auto-tilt-test",
                "scenario_type": "microgrid",
                "province": "Overseas",
                "city": "Mauritania Mine",
                "country": "Mauritania",
                "latitude": 23.196941,
                "longitude": -11.959593,
                "grid_connection_mode": "microgrid",
                "storage_strategy_mode": "renewable_priority",
            },
            "resource_data": {
                "solar": {
                    "installed_capacity_mwp": 20,
                    "annual_irradiation_kwh_per_m2": 2263,
                    "tilt_deg": None,
                    "performance_ratio": 0.82,
                },
                "wind": {
                    "installed_capacity_mw": 10,
                    "annual_avg_speed_mps": 6.59,
                    "capacity_factor_assumption": 0.42,
                },
            },
            "load_data": {
                "annual_consumption_mwh": 100000,
                "peak_load_kw": 12000,
                "critical_load_kw": 8000,
                "backup_hours_required": 4,
            },
            "market_data": {
                "market_mode": "offgrid_internal",
            },
            "equipment": {
                "conventional_backup": {
                    "enabled": True,
                    "fuel_cost_per_kwh": 0.32,
                },
                "storage": {
                    "power_candidate_kw": [0, 5000, 10000],
                    "energy_candidate_kwh": [0, 20000, 40000],
                },
            },
        }

        output, _, _ = analyze_project(payload)
        self.assertEqual(output["project_summary"]["scenario_type"], "microgrid")
        self.assertIsNotNone(output["resource_results"]["pv_recommended_tilt_deg"])
        self.assertAlmostEqual(
            output["resource_results"]["pv_effective_tilt_deg"],
            output["resource_results"]["pv_recommended_tilt_deg"],
            places=2,
        )
        self.assertEqual(output["resource_results"]["pv_resource_basis"], "annual_irradiation_kwh_per_m2")

    def test_china_projects_use_domestic_tax_model_while_overseas_defaults_to_zero_tax(self) -> None:
        domestic_payload = {
            "project_info": {
                "project_name": "domestic-tax-test",
                "scenario_type": "zero_carbon_factory",
                "province": "江苏",
                "grid_connection_mode": "user_side",
            },
            "resource_data": {
                "solar": {
                    "annual_irradiation_kwh_per_m2": 1600,
                    "available_area_m2": 39000,
                    "performance_ratio": 0.82,
                }
            },
            "load_data": {
                "annual_consumption_mwh": 8760,
                "peak_load_kw": 1000,
            },
            "market_data": {
                "market_mode": "tou_tariff",
                "tou_tariff": [
                    {"period": "peak", "price": 1.05},
                    {"period": "flat", "price": 0.72},
                    {"period": "valley", "price": 0.38},
                ],
            },
            "equipment": {
                "storage": {
                    "power_candidate_kw": [0],
                    "energy_candidate_kwh": [0],
                }
            },
            "financial": {
                "project_years": 15,
                "discount_rate": 0.08,
                "capex": {
                    "pv_cost_per_w": 0.2,
                },
                "opex": {
                    "annual_om_ratio": 0.01,
                    "annual_opex_escalation_rate": 0.02,
                },
                "degradation": {
                    "pv_degradation_per_year": 0.005,
                },
            },
        }
        overseas_payload = {
            **domestic_payload,
            "project_info": {
                **domestic_payload["project_info"],
                "province": "Overseas",
                "country": "Mauritania",
            },
        }

        domestic_output, _, _ = analyze_project(domestic_payload)
        overseas_output, _, _ = analyze_project(overseas_payload)

        self.assertTrue(domestic_output["financial_results"]["tax_model"].startswith("china_domestic"))
        self.assertGreater(domestic_output["financial_results"]["annual_tax_total"], 0.0)
        self.assertEqual(overseas_output["financial_results"]["tax_model"], "overseas_exempt")
        self.assertEqual(overseas_output["financial_results"]["annual_tax_total"], 0.0)
        self.assertLess(domestic_output["financial_results"]["npv"], overseas_output["financial_results"]["npv"])

    def test_offgrid_optimizer_respects_fixed_wind_and_coverage_constraints(self) -> None:
        payload = {
            "project_info": {
                "project_name": "offgrid-constraint-test",
                "scenario_type": "microgrid",
                "province": "Overseas",
                "country": "Mauritania",
                "grid_connection_mode": "microgrid",
                "storage_strategy_mode": "renewable_priority",
            },
            "resource_data": {
                "solar": {
                    "annual_irradiation_kwh_per_m2": 2263,
                    "available_area_m2": 1200000,
                    "performance_ratio": 0.82,
                },
                "wind": {
                    "annual_avg_speed_mps": 6.59,
                    "capacity_factor_assumption": 0.42,
                },
            },
            "load_data": {
                "annual_consumption_mwh": (18000 * 20 + 44460 * 24) * 365 / 1000,
                "peak_load_kw": 62460,
            },
            "market_data": {
                "market_mode": "offgrid_internal",
                "fuel_cost_per_kwh": 0.3,
            },
            "equipment": {
                "pv": {
                    "candidate_mwp": [110, 120, 130, 180, 190],
                },
                "wind": {
                    "fixed_capacity_mw": 72,
                },
                "storage": {
                    "power_candidate_kw": [0, 60000, 80000, 100000],
                    "energy_candidate_kwh": [0, 360000, 480000, 600000],
                },
                "conventional_backup": {
                    "enabled": True,
                    "fuel_cost_per_kwh": 0.3,
                },
            },
            "financial": {
                "project_years": 15,
                "discount_rate": 0.08,
                "capex": {
                    "storage_system_cost_per_kwh": 850,
                    "pv_cost_per_w": 3.1,
                    "wind_cost_per_w": 6.4,
                },
                "opex": {
                    "annual_om_ratio": 0.018,
                    "annual_opex_escalation_rate": 0.02,
                },
                "degradation": {
                    "storage_capacity_fade_per_year": 0.022,
                    "pv_degradation_per_year": 0.005,
                    "wind_degradation_per_year": 0.003,
                },
                "optimization": {
                    "objective": "max_npv",
                    "fixed_wind_mw": 72,
                    "min_energy_coverage_ratio": 0.99,
                    "min_storage_power_mw": 60,
                    "min_storage_energy_mwh": 360,
                    "min_storage_duration_hours": 4,
                },
            },
        }

        output, _, _ = analyze_project(payload)
        self.assertEqual(output["recommended_solution"]["wind_mw"], 72.0)
        self.assertGreaterEqual(output["simulation_results"]["renewable_energy_coverage_ratio"], 0.99)
        self.assertGreaterEqual(output["recommended_solution"]["storage_power_mw"], 60.0)
        self.assertGreaterEqual(output["recommended_solution"]["storage_energy_mwh"], 360.0)

    def test_pv_annual_irradiation_formula_matches_engineering_specific_yield(self) -> None:
        payload = {
            "resource_data": {
                "solar": {
                    "annual_irradiation_kwh_per_m2": 2263,
                    "available_area_m2": 780000,
                    "performance_ratio": 0.82,
                    "site_latitude_deg": 23.2,
                    "tilt_deg": 20,
                    "azimuth_deg": 180,
                }
            },
            "load_data": {
                "annual_consumption_mwh": 100000,
            },
        }
        result = estimate_pv_generation(payload)
        self.assertEqual(result["pv_mwp"], 120.0)
        self.assertGreater(result["annual_pv_generation_mwh"], 220000)
        self.assertEqual(result["pv_resource_basis"], "annual_irradiation_kwh_per_m2")

    def test_resource_models_accept_8760_high_resolution_series(self) -> None:
        solar_payload = {
            "resource_data": {
                "solar": {
                    "available_area_m2": 6500,
                    "hourly_irradiance_kwh_per_m2": [0.5] * 8760,
                    "hourly_temperature_c": [25.0] * 8760,
                    "performance_ratio": 0.9,
                }
            },
            "load_data": {"annual_consumption_mwh": 1000},
        }
        wind_payload = {
            "resource_data": {
                "wind": {
                    "installed_capacity_mw": 2,
                    "wind_speed_series_mps": [8.0] * 8760,
                    "power_curve": [
                        {"speed_mps": 0, "power_kw": 0},
                        {"speed_mps": 3, "power_kw": 0},
                        {"speed_mps": 6, "power_kw": 1000},
                        {"speed_mps": 8, "power_kw": 1800},
                        {"speed_mps": 10, "power_kw": 2000},
                    ],
                }
            },
            "load_data": {"annual_consumption_mwh": 1000},
        }
        pv_result = estimate_pv_generation(solar_payload)
        wind_result = estimate_wind_generation(wind_payload)
        self.assertEqual(len(pv_result["pv_annual_series_kw"]), 8760)
        self.assertEqual(len(wind_result["wind_annual_series_kw"]), 8760)
        self.assertEqual(pv_result["pv_resource_basis"], "hourly_irradiance_kwh_per_m2_8760")
        self.assertEqual(wind_result["wind_resource_basis"], "wind_speed_series_mps_8760 + power_curve")

    def test_pv_tracking_mode_increases_generation(self) -> None:
        base = {
            "resource_data": {
                "solar": {
                    "annual_irradiation_kwh_per_m2": 2000,
                    "available_area_m2": 6500,
                    "performance_ratio": 0.82,
                    "site_latitude_deg": 25,
                }
            },
            "load_data": {"annual_consumption_mwh": 1000},
        }
        fixed = estimate_pv_generation(base)
        tracking = estimate_pv_generation(
            {
                **base,
                "resource_data": {
                    "solar": {
                        **base["resource_data"]["solar"],
                        "tracking_mode": "single_axis",
                    }
                },
            }
        )
        self.assertGreater(tracking["annual_pv_generation_mwh"], fixed["annual_pv_generation_mwh"])
        self.assertGreater(tracking["pv_tracking_factor"], 1.0)

    def test_pv_dc_ac_clipping_reduces_generation(self) -> None:
        base = {
            "resource_data": {
                "solar": {
                    "installed_capacity_mwp": 1.0,
                    "hourly_irradiance_kwh_per_m2": [1.0] * 8760,
                    "hourly_temperature_c": [25.0] * 8760,
                    "performance_ratio": 0.9,
                }
            },
            "load_data": {"annual_consumption_mwh": 1000},
        }
        unclipped = estimate_pv_generation(base)
        clipped = estimate_pv_generation(
            {
                **base,
                "resource_data": {
                    "solar": {
                        **base["resource_data"]["solar"],
                        "inverter_capacity_mwac": 0.5,
                    }
                },
            }
        )
        self.assertLess(clipped["annual_pv_generation_mwh"], unclipped["annual_pv_generation_mwh"])

    def test_wind_net_factor_reduces_generation(self) -> None:
        payload = {
            "resource_data": {
                "wind": {
                    "installed_capacity_mw": 2,
                    "capacity_factor_assumption": 0.4,
                    "availability_factor": 0.95,
                    "wake_loss_factor": 0.92,
                    "curtailment_factor": 0.97,
                }
            },
            "load_data": {"annual_consumption_mwh": 1000},
        }
        result = estimate_wind_generation(payload)
        expected_net_factor = 0.95 * 0.92 * 0.97
        self.assertAlmostEqual(result["wind_net_factor"], round(expected_net_factor, 4), places=4)
        self.assertAlmostEqual(result["annual_wind_generation_mwh"], round(2 * 8760 * 0.4 * expected_net_factor, 2), places=2)

    def test_wind_air_density_factor_reduces_generation_at_high_temperature(self) -> None:
        base = {
            "resource_data": {
                "wind": {
                    "installed_capacity_mw": 2,
                    "capacity_factor_assumption": 0.4,
                }
            },
            "load_data": {"annual_consumption_mwh": 1000},
        }
        standard = estimate_wind_generation(base)
        hot = estimate_wind_generation(
            {
                **base,
                "resource_data": {
                    "wind": {
                        **base["resource_data"]["wind"],
                        "ambient_temperature_c": 40,
                    }
                },
            }
        )
        self.assertLess(hot["annual_wind_generation_mwh"], standard["annual_wind_generation_mwh"])

    def test_rule_based_arbitrage_uses_days_and_cycle_margin(self) -> None:
        data = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.0277},
                    {"period": "peak", "price": 0.8668},
                    {"period": "flat", "price": 0.4987},
                    {"period": "valley", "price": 0.2289},
                ],
                "arbitrage_plan": {
                    "mode": "rule_based",
                    "cycles": [
                        {"charge_period": "valley", "discharge_period": "super_peak", "days_per_year": 82.5, "charge_energy_kwh": 2135.1, "discharge_energy_kwh": 1841.4},
                        {"charge_period": "valley", "discharge_period": "peak", "days_per_year": 440, "charge_energy_kwh": 2135.1, "discharge_energy_kwh": 1841.4},
                        {"charge_period": "valley", "discharge_period": "flat", "days_per_year": 137.5, "charge_energy_kwh": 2135.1, "discharge_energy_kwh": 1841.4},
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "cycle_life": 6000,
                    "annual_degradation_rate": 0.03,
                }
            },
        }
        storage = {"storage_power_mw": 1.0, "storage_energy_mwh": 2.088}
        result = simulate_rule_based_arbitrage(data, storage)
        self.assertIsNotNone(result)
        # 662127.79 = annual gross margin computed from:
        #   valley price 0.2289 / super_peak 1.0277 × 82.5 days
        # + valley price 0.2289 / peak 0.8668 × 440 days
        # + valley price 0.2289 / flat 0.4987 × 137.5 days
        # with charge/discharge energy 2135.1/1841.4 kWh per cycle.
        self.assertAlmostEqual(result["annual_gross_margin"], 662127.79, delta=500.0)

    def test_rule_based_arbitrage_can_auto_fill_days_from_policy(self) -> None:
        data = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.34991333333333},
                    {"period": "peak", "price": 1.112447},
                    {"period": "flat", "price": 0.708644},
                    {"period": "valley", "price": 0.345405},
                ],
                "arbitrage_plan": {
                    "mode": "rule_based",
                    "auto_days_from_policy": True,
                    "monthly_active_days": 27.5,
                    "default_cycle_day_ratios": {
                        "super_peak": 0.25,
                        "peak": 1.3333333333,
                        "flat": 0.4166666667,
                    },
                    "cycles": [
                        {"charge_period": "valley", "discharge_period": "super_peak", "discharge_energy_kwh": 253.17},
                        {"charge_period": "valley", "discharge_period": "peak", "discharge_energy_kwh": 253.17},
                        {"charge_period": "valley", "discharge_period": "flat", "discharge_energy_kwh": 253.17},
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "cycle_life": 6000,
                    "annual_degradation_rate": 0.03,
                    "battery_charge_efficiency": 0.941695806555,
                    "battery_discharge_efficiency": 0.941695806555,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                    "first_discharge_depth": 0.97,
                }
            },
        }
        storage = {"storage_power_mw": 0.125, "storage_energy_mwh": 0.261}
        result = simulate_rule_based_arbitrage(data, storage)
        self.assertIsNotNone(result)
        self.assertGreater(result["annual_discharge_mwh"], 10.0)
        self.assertGreater(result["annual_gross_margin"], 90000.0)

    def test_rule_based_arbitrage_prefers_monthly_policy_history_when_provided(self) -> None:
        data = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.3},
                    {"period": "peak", "price": 1.0},
                    {"period": "flat", "price": 0.7},
                    {"period": "valley", "price": 0.3},
                ],
                "monthly_tou_policy_history": [
                    {"month": 1, "periods": ["peak", "flat", "valley"], "active_days": 27.5},
                    {"month": 2, "periods": ["super_peak", "peak", "flat", "valley"], "active_days": 27.5},
                ],
                "arbitrage_plan": {
                    "mode": "rule_based",
                    "auto_days_from_policy": True,
                    "monthly_active_days": 27.5,
                    "default_cycle_day_ratios": {
                        "super_peak": 0.25,
                        "peak": 1.0,
                        "flat": 0.5,
                    },
                    "cycles": [
                        {"charge_period": "valley", "discharge_period": "super_peak", "discharge_energy_kwh": 100},
                        {"charge_period": "valley", "discharge_period": "peak", "discharge_energy_kwh": 100},
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.95,
                    "battery_discharge_efficiency": 0.95,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                    "cycle_life": 6000,
                }
            },
        }
        storage = {"storage_power_mw": 0.1, "storage_energy_mwh": 0.2}
        result = simulate_rule_based_arbitrage(data, storage)
        self.assertIsNotNone(result)
        self.assertGreater(result["annual_gross_margin"], 0.0)

    def test_rule_based_arbitrage_supports_month_specific_single_and_double_cycle_templates(self) -> None:
        data = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.3},
                    {"period": "peak", "price": 1.0},
                    {"period": "flat", "price": 0.7},
                    {"period": "valley", "price": 0.3},
                ],
                "arbitrage_plan": {
                    "mode": "rule_based",
                    "monthly_cycle_templates": [
                        {
                            "month": 1,
                            "active_days": 27.5,
                            "cycles": [
                                {"charge_period": "valley", "discharge_period": "peak", "discharge_energy_kwh": 100}
                            ],
                        },
                        {
                            "month": 2,
                            "active_days": 27.5,
                            "cycles": [
                                {"charge_period": "valley", "discharge_period": "super_peak", "discharge_energy_kwh": 100},
                                {"charge_period": "valley", "discharge_period": "peak", "discharge_energy_kwh": 80},
                            ],
                        },
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.95,
                    "battery_discharge_efficiency": 0.95,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                    "cycle_life": 6000,
                }
            },
        }
        storage = {"storage_power_mw": 0.1, "storage_energy_mwh": 0.2}
        result = simulate_rule_based_arbitrage(data, storage)
        self.assertIsNotNone(result)
        self.assertGreater(result["monthly_storage_revenue_breakdown"][1]["gross_margin"], result["monthly_storage_revenue_breakdown"][0]["gross_margin"])

    def test_spot_intraday_arbitrage_selects_multiple_daily_cycles(self) -> None:
        data = {
            "market_data": {
                "arbitrage_plan": {
                    "mode": "spot_intraday",
                    "min_charge_hours": 2,
                    "min_discharge_hours": 2,
                    "min_spread_yuan_per_mwh": 250,
                },
                "spot_price_daily_profiles": [
                    {
                        "date": "2026-01-01",
                        "realtime_prices": [
                            0.10, 0.10, 0.55, 0.55,
                            0.12, 0.12, 0.48, 0.48,
                            0.30, 0.30, 0.32, 0.32,
                            0.31, 0.31, 0.33, 0.33,
                            0.34, 0.34, 0.35, 0.35,
                            0.36, 0.36, 0.37, 0.37,
                        ],
                    }
                ],
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.95,
                    "battery_discharge_efficiency": 0.95,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                    "cycle_life": 6000,
                }
            },
        }
        storage = {"storage_power_mw": 1.0, "storage_energy_mwh": 2.0}
        result = simulate_spot_intraday_arbitrage(data, storage)
        self.assertIsNotNone(result)
        self.assertEqual(result["days_covered"], 1)
        self.assertEqual(result["total_cycles"], 2)
        first_day = result["daily_spot_arbitrage_schedule"][0]
        self.assertEqual(first_day["cycles"][0]["charge_window"], "00:00-02:00")
        self.assertEqual(first_day["cycles"][0]["discharge_window"], "02:00-04:00")
        self.assertEqual(first_day["cycles"][1]["charge_window"], "04:00-06:00")
        self.assertEqual(first_day["cycles"][1]["discharge_window"], "06:00-08:00")

    def test_spot_intraday_defaults_to_daily_independent_optimization(self) -> None:
        data = {
            "market_data": {
                "arbitrage_plan": {
                    "mode": "spot_intraday",
                },
                "spot_price_daily_profiles": [
                    {
                        "date": "2026-01-01",
                        "realtime_prices": [120.0] * 24,
                    }
                ],
            },
            "equipment": {
                "storage": {},
            },
        }
        storage = {"storage_power_mw": 1.0, "storage_energy_mwh": 2.0}

        with patch("energy_solution_agent.solvers.arbitrage._select_best_daily_spot_cycles", return_value=[]) as daily_selector, patch(
            "energy_solution_agent.solvers.arbitrage._select_best_spot_cycles_continuous",
            return_value=[],
        ) as continuous_selector:
            simulate_spot_intraday_arbitrage(data, storage)

        daily_selector.assert_called_once()
        continuous_selector.assert_not_called()

    def test_live_rule_patch_can_emit_monthly_tou_policy_history(self) -> None:
        text = "夏季尖峰时段10:00-12:00、14:00-21:00，平段8:00-10:00，谷段0:00-8:00；冬季峰时段9:00-12:00、16:00-20:00。"
        patch = extract_structured_rule_patch(text)
        self.assertIn("monthly_tou_policy_history", patch)
        self.assertTrue(any(item["month"] == 6 for item in patch["monthly_tou_policy_history"]))

    def test_tou_policy_fetch_can_fill_history_from_profile_sources(self) -> None:
        profile = {"source_links": []}
        data = {"market_data": {"monthly_tou_policy_history": [{"month": 6, "periods": ["peak", "flat", "valley"]}]}}
        enriched, meta = enrich_with_monthly_tou_policy_history(data, profile)
        self.assertEqual(meta["tou_policy_fetch_status"], "not_needed")
        self.assertEqual(enriched["market_data"]["monthly_tou_policy_history"][0]["month"], 6)

    @patch("energy_solution_agent.tou_policy_fetch._fetch_url")
    def test_tou_policy_fetch_can_use_explicit_month_metadata(self, mock_fetch) -> None:
        mock_fetch.return_value = {
            "ok": True,
            "url": "https://example.com/policy",
            "title": "2025年6月分时电价通知",
            "text": "夏季尖峰时段10:00-12:00、14:00-21:00，平段8:00-10:00，谷段0:00-8:00。",
        }
        data = {
            "market_data": {
                "tou_policy_history_links": [
                    {"url": "https://example.com/policy", "month": 6},
                ]
            }
        }
        enriched, meta = enrich_with_monthly_tou_policy_history(data, profile=None)
        self.assertEqual(meta["tou_policy_fetch_status"], "fetched")
        self.assertEqual(enriched["market_data"]["monthly_tou_policy_history"][0]["month"], 6)
        self.assertTrue(enriched["market_data"]["monthly_tou_policy_history"][0]["periods"])

    def test_policy_classifier_can_distinguish_fixed_tou_and_market_based(self) -> None:
        fixed = classify_market_policy_mode(
            {"market_data": {}},
            [{"title": "安徽分时电价通知", "text": "分时电价 尖峰 峰段 平段 谷段 深谷"}],
        )
        market = classify_market_policy_mode(
            {"market_data": {}},
            [{"title": "某省市场化交易政策", "text": "市场化交易 现货交易 中长期交易 代理购电"}],
        )
        self.assertEqual(fixed["market_data"]["market_policy_mode"], "fixed_tou_policy")
        self.assertEqual(market["market_data"]["market_policy_mode"], "market_based")

    def test_proxy_url_can_be_read_from_network_block(self) -> None:
        self.assertEqual(get_proxy_url({"proxy_url": "http://127.0.0.1:7890"}), "http://127.0.0.1:7890")

    def test_commercial_hybrid_mode_a_matches_simple_high_price_sale_logic(self) -> None:
        data = {
            "market_data": {
                "commercial_hybrid_plan": {
                    "mode": "mode_a",
                    "high_price": 0.739537,
                    "sell_discount": 0.85,
                    "high_window_hours_per_day": 2,
                    "operating_days_per_year": 365,
                    "demand_control_kw": 1000,
                    "demand_rate_per_kw_month": 35.2,
                    "vpp_price_per_mwh": 418,
                    "vpp_duration_hours": 2,
                    "vpp_times_per_year": 300,
                    "vpp_effective_ratio": 0.8,
                    "vpp_project_share": 0.5,
                }
            },
            "equipment": {
                "storage": {
                    "soc_max": 0.95,
                    "soc_reserve_ratio": 0.05,
                    "battery_discharge_efficiency": 0.945,
                    "pcs_efficiency": 0.994,
                    "transformer_efficiency": 1.0,
                    "cycle_life": 6000,
                    "annual_degradation_rate": 0.02,
                }
            },
        }
        renewables = {"annual_pv_generation_mwh": 2400}
        storage = {"storage_power_mw": 3.75, "storage_energy_mwh": 7.5}
        result = simulate_commercial_hybrid_value(data, renewables, storage)
        self.assertIsNotNone(result)
        self.assertGreater(result["annual_energy_value"], 1000000)
        self.assertGreater(result["annual_demand_value"], 400000)
        self.assertGreater(result["annual_vpp_value"], 300000)
        self.assertLess(result["deliverable_power_mw"], storage["storage_power_mw"])

    def test_annual_cycle_value_mode_uses_latest_tariff_and_cycle_counts(self) -> None:
        data = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.34991333333333},
                    {"period": "peak", "price": 1.112447},
                    {"period": "flat", "price": 0.708644},
                    {"period": "valley", "price": 0.345405},
                ],
                "arbitrage_plan": {
                    "mode": "annual_cycle_value",
                    "cycles": [
                        {"charge_period": "valley", "discharge_period": "peak", "days_per_year": 440, "discharge_energy_kwh": 253.17},
                        {"charge_period": "valley", "discharge_period": "super_peak", "days_per_year": 82.5, "discharge_energy_kwh": 253.17},
                        {"charge_period": "valley", "discharge_period": "flat", "days_per_year": 137.5, "discharge_energy_kwh": 253.17},
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.941695806555,
                    "battery_discharge_efficiency": 0.941695806555,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                    "cycle_life": 6000,
                    "annual_degradation_rate": 0.03,
                }
            },
        }
        storage = {"storage_power_mw": 0.125, "storage_energy_mwh": 0.261}
        result = simulate_annual_cycle_value(data, storage)
        self.assertIsNotNone(result)
        self.assertGreater(result["annual_gross_margin"], 90000.0)
        self.assertGreater(result["annual_fec"], 200.0)

    def test_annual_cycle_value_respects_revenue_share_ratio(self) -> None:
        base = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.0},
                    {"period": "peak", "price": 0.8},
                    {"period": "flat", "price": 0.6},
                    {"period": "valley", "price": 0.3},
                ],
                "arbitrage_plan": {
                    "mode": "annual_cycle_value",
                    "cycles": [
                        {"charge_period": "valley", "discharge_period": "peak", "days_per_year": 100, "discharge_energy_kwh": 100},
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.95,
                    "battery_discharge_efficiency": 0.95,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                }
            },
        }
        storage = {"storage_power_mw": 0.1, "storage_energy_mwh": 0.2}
        full = simulate_annual_cycle_value(base, storage)
        shared = simulate_annual_cycle_value(
            {
                **base,
                "market_data": {
                    **base["market_data"],
                    "arbitrage_plan": {
                        **base["market_data"]["arbitrage_plan"],
                        "revenue_share_ratio": 0.15,
                    },
                },
            },
            storage,
        )
        self.assertIsNotNone(full)
        self.assertIsNotNone(shared)
        self.assertAlmostEqual(shared["annual_gross_margin"], round(full["annual_gross_margin"] * 0.15, 2), delta=1.0)

    def test_annual_cycle_value_can_use_customer_share_ratio_as_inverse_of_owner_share(self) -> None:
        base = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.0},
                    {"period": "peak", "price": 0.8},
                    {"period": "flat", "price": 0.6},
                    {"period": "valley", "price": 0.3},
                ],
                "arbitrage_plan": {
                    "mode": "annual_cycle_value",
                    "cycles": [
                        {"charge_period": "valley", "discharge_period": "peak", "days_per_year": 100, "discharge_energy_kwh": 100},
                    ],
                },
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.95,
                    "battery_discharge_efficiency": 0.95,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                }
            },
        }
        storage = {"storage_power_mw": 0.1, "storage_energy_mwh": 0.2}
        owner = simulate_annual_cycle_value(
            {
                **base,
                "market_data": {
                    **base["market_data"],
                    "arbitrage_plan": {
                        **base["market_data"]["arbitrage_plan"],
                        "owner_share_ratio": 0.85,
                    },
                },
            },
            storage,
        )
        customer = simulate_annual_cycle_value(
            {
                **base,
                "market_data": {
                    **base["market_data"],
                    "arbitrage_plan": {
                        **base["market_data"]["arbitrage_plan"],
                        "customer_share_ratio": 0.15,
                    },
                },
            },
            storage,
        )
        self.assertIsNotNone(owner)
        self.assertIsNotNone(customer)
        self.assertAlmostEqual(customer["annual_gross_margin"], owner["annual_gross_margin"], delta=1.0)

    def test_infer_cycles_from_monthly_tou_history_can_emit_double_cycle_when_spread_allows(self) -> None:
        data = {
            "market_data": {
                "tou_tariff": [
                    {"period": "super_peak", "price": 1.3},
                    {"period": "peak", "price": 1.0},
                    {"period": "flat", "price": 0.6},
                    {"period": "valley", "price": 0.3},
                ],
                "monthly_tou_policy_history": [
                    {
                        "month": 7,
                        "periods": ["super_peak", "peak", "flat", "valley"],
                        "active_days": 27.5,
                        "schedule": {
                            "valley": [0, 1, 2, 3, 4, 5],
                            "peak": [8, 9],
                            "flat": [10, 11],
                            "super_peak": [12, 13],
                        },
                    }
                ],
                "arbitrage_plan": {
                    "second_cycle_min_spread": 0.0,
                    "first_cycle_min_spread": 0.0,
                },
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.95,
                    "battery_discharge_efficiency": 0.95,
                    "pcs_efficiency": 1.0,
                    "transformer_efficiency": 1.0,
                    "soc_max": 0.95,
                    "soc_reserve_ratio": 0.05,
                }
            },
        }
        storage = {"storage_power_mw": 1.0, "storage_energy_mwh": 2.0}
        cycles = infer_cycles_from_monthly_tou_history(data, storage)
        self.assertEqual(len(cycles), 2)
        self.assertEqual({cycle["discharge_period"] for cycle in cycles}, {"super_peak", "peak"})

    def test_infer_daily_cycles_from_schedule_uses_price_top_points_but_time_order_for_second_cycle(self) -> None:
        schedule = {
            "valley": [0, 1, 2, 3, 4, 5],
            "flat": [6, 7, 10, 11, 15, 16, 17],
            "peak": [8, 9, 18, 19, 20],
            "super_peak": [12, 13],
        }
        prices = {
            "super_peak": 1.4,
            "peak": 1.0,
            "flat": 0.7,
            "valley": 0.3,
        }
        cycles = infer_daily_cycles_from_schedule(
            schedule=schedule,
            prices=prices,
            first_cycle_min_spread=0.0,
            second_cycle_min_spread=0.0,
            discharge_energy_kwh=100.0,
            effective_rte=0.9,
        )
        self.assertEqual(len(cycles), 2)
        self.assertEqual(cycles[0]["discharge_period"], "peak")
        self.assertEqual(cycles[1]["discharge_period"], "super_peak")
        self.assertAlmostEqual(cycles[1]["charge_energy_kwh"], 111.1111111111, places=6)

    def test_infer_daily_cycles_from_schedule_skips_second_cycle_when_first_cycle_not_formed(self) -> None:
        schedule = {
            "flat": [0, 1, 2, 3, 4, 5, 10, 11, 12, 13, 14],
            "peak": [6, 7, 8, 9],
            "super_peak": [15, 16],
        }
        prices = {
            "super_peak": 1.4,
            "peak": 1.0,
            "flat": 0.95,
        }
        cycles = infer_daily_cycles_from_schedule(
            schedule=schedule,
            prices=prices,
            first_cycle_min_spread=0.2,
            second_cycle_min_spread=0.0,
            discharge_energy_kwh=100.0,
            effective_rte=0.9,
        )
        self.assertEqual(cycles, [])

    def test_market_responding_dispatch_respects_daily_cycle_limit(self) -> None:
        dispatch = simulate_storage_dispatch_annual(
            load_series_kw=[1000.0] * 8760,
            pv_series_kw=[0.0] * 8760,
            wind_series_kw=[0.0] * 8760,
            charging_series_kw=[0.0] * 8760,
            thermal_series_kw=[0.0] * 8760,
            storage_power_mw=1.0,
            storage_energy_mwh=2.0,
            strategy_mode="market_responding",
            price_series=[0.2] * 6 + [0.6] * 6 + [1.0] * 6 + [0.5] * 6,
            storage_config={"max_daily_cycles": 1.0},
        )
        self.assertLessEqual(dispatch["storage_equivalent_full_cycles_per_year"], 366.0)

    def test_storage_product_selection_can_preserve_raw_values_and_compute_utilization(self) -> None:
        data = {
            "equipment": {
                "storage": {
                    "selected_product_power_kw": 150,
                    "selected_product_energy_kwh": 261,
                }
            }
        }
        storage = {
            "raw_storage_power_mw": 0.125,
            "raw_storage_energy_mwh": 0.195,
            "storage_power_mw": 0.125,
            "storage_energy_mwh": 0.195,
        }
        selected = apply_storage_product_selection(data, storage)
        self.assertEqual(selected["raw_storage_power_mw"], 0.125)
        self.assertEqual(selected["raw_storage_energy_mwh"], 0.195)
        self.assertEqual(selected["storage_power_mw"], 0.15)
        self.assertEqual(selected["storage_energy_mwh"], 0.261)
        self.assertAlmostEqual(selected["storage_power_utilization_ratio"], 0.8333, places=4)
        self.assertAlmostEqual(selected["storage_energy_utilization_ratio"], 0.7471, places=4)

    def test_storage_product_selection_can_use_explicit_raw_overrides_from_followup_turn(self) -> None:
        data = {
            "equipment": {
                "storage": {
                    "raw_storage_power_kw": 125,
                    "raw_storage_energy_kwh": 195,
                    "selected_product_power_kw": 125,
                    "selected_product_energy_kwh": 261,
                }
            }
        }
        storage = {
            "raw_storage_power_mw": 1.2,
            "raw_storage_energy_mwh": 2.4,
            "storage_power_mw": 1.2,
            "storage_energy_mwh": 2.4,
        }
        selected = apply_storage_product_selection(data, storage)
        self.assertEqual(selected["raw_storage_power_mw"], 0.125)
        self.assertEqual(selected["raw_storage_energy_mwh"], 0.195)
        self.assertEqual(selected["storage_power_mw"], 0.125)
        self.assertEqual(selected["storage_energy_mwh"], 0.261)
        self.assertAlmostEqual(selected["storage_energy_utilization_ratio"], 0.7471, places=4)

    def test_storage_product_selection_can_auto_map_low_voltage_energy_to_261kwh_blocks(self) -> None:
        data = {
            "project_info": {
                "voltage_level_kv": 0.38,
            },
            "equipment": {
                "storage": {}
            },
        }
        storage = {
            "raw_storage_power_mw": 0.125,
            "raw_storage_energy_mwh": 0.195,
            "storage_power_mw": 0.125,
            "storage_energy_mwh": 0.195,
        }
        selected = apply_storage_product_selection(data, storage)
        self.assertEqual(selected["selected_product_energy_mwh"], 0.261)
        self.assertEqual(selected["storage_energy_mwh"], 0.261)
        self.assertAlmostEqual(selected["storage_energy_utilization_ratio"], 0.7471, places=4)

    def test_storage_product_selection_can_keep_raw_energy_for_medium_voltage(self) -> None:
        data = {
            "project_info": {
                "voltage_level_kv": 10,
            },
            "equipment": {
                "storage": {}
            },
        }
        storage = {
            "raw_storage_power_mw": 0.125,
            "raw_storage_energy_mwh": 0.195,
            "storage_power_mw": 0.125,
            "storage_energy_mwh": 0.195,
        }
        selected = apply_storage_product_selection(data, storage)
        self.assertEqual(selected["selected_product_energy_mwh"], 0.195)
        self.assertEqual(selected["storage_energy_mwh"], 0.195)

    def test_estimate_storage_can_use_monthly_load_space_sizing_for_user_side_storage(self) -> None:
        day = []
        for hour in range(24):
            if hour in {0, 1}:
                day.append(80.0)
            elif hour in {18, 19}:
                day.append(25.0)
            elif hour == 12:
                day.append(100.0)
            else:
                day.append(60.0)
        load_series = day * 365
        monthly_history = [
            {
                "month": month,
                "periods": ["valley", "peak"],
                "schedule": {"valley": [0, 1], "peak": [18, 19]},
            }
            for month in range(1, 13)
        ]
        data = {
            "project_info": {"scenario_type": "user_side_storage"},
            "market_data": {
                "tou_tariff": [
                    {"period": "peak", "price": 1.0},
                    {"period": "valley", "price": 0.3},
                ],
                "monthly_tou_policy_history": monthly_history,
            },
            "equipment": {
                "storage": {
                    "sizing_target_day_coverage_ratio": 0.9,
                }
            },
            "load_data": {},
        }
        storage = estimate_storage(data, load_series_kw=load_series)
        self.assertEqual(storage["raw_storage_power_mw"], 0.02)
        self.assertEqual(storage["raw_storage_energy_mwh"], 0.04)

    def test_estimate_storage_can_use_spot_intraday_load_space_sizing_for_power_trading_storage(self) -> None:
        day = []
        for hour in range(24):
            if hour in {10, 11, 12}:
                day.extend([20.0] * 4)
            elif hour in {21, 22}:
                day.extend([120.0] * 4)
            else:
                day.extend([60.0] * 4)
        load_series = day * 31
        daily_profiles = [
            {
                "date": f"2026-01-{day_idx:02d}",
                "realtime_prices": [300.0] * 10 + [-80.0, -80.0, -80.0] + [120.0] * 8 + [280.0, 280.0, 260.0],
            }
            for day_idx in range(1, 32)
        ]
        data = {
            "project_info": {"scenario_type": "user_side_storage"},
            "market_data": {
                "market_mode": "market_price_series",
                "spot_price_daily_profiles": daily_profiles,
                "arbitrage_plan": {
                    "mode": "spot_intraday",
                    "continuous_horizon": False,
                    "min_charge_hours": 2,
                    "min_discharge_hours": 2,
                    "max_charge_hours": 6,
                    "max_discharge_hours": 6,
                    "min_spread_yuan_per_mwh": 250,
                },
                "contract_capacity_kw": 150.0,
                "demand_charge_mode": "contract_capacity",
            },
            "equipment": {
                "storage": {
                    "battery_charge_efficiency": 0.965,
                    "battery_discharge_efficiency": 0.965,
                    "pcs_efficiency": 0.985,
                    "transformer_efficiency": 0.99,
                    "sizing_target_day_coverage_ratio": 0.9,
                }
            },
            "load_data": {},
        }
        storage = estimate_storage(data, load_series_kw=load_series)
        self.assertIsNotNone(storage)
        self.assertGreater(storage["raw_storage_power_mw"], 0.0)
        self.assertGreater(storage["raw_storage_energy_mwh"], 0.0)

    @patch("energy_solution_agent.resource_fetch._fetch_nasa_power_hourly")
    @patch("energy_solution_agent.resource_fetch._fetch_nasa_power_climatology")
    @patch("energy_solution_agent.resource_fetch._fetch_nominatim")
    def test_auto_resource_fetch_can_work_from_place_name(self, mock_geocode, mock_nasa, mock_hourly) -> None:
        mock_geocode.return_value = {
            "latitude": 23.196941,
            "longitude": -11.959593,
            "display_name": "Mauritania Mine",
        }
        mock_hourly.return_value = {
            "hourly_irradiance_kwh_per_m2": [0.3] * 8760,
            "hourly_temperature_c": [28.0] * 8760,
            "hourly_wind_speed_50m_mps": [6.6] * 8760,
        }
        mock_nasa.return_value = {
            "monthly_irradiation_kwh_per_m2": [5.6, 5.9, 6.3, 6.5, 6.7, 6.8, 6.7, 6.5, 6.3, 6.0, 5.7, 5.5],
            "annual_irradiation_kwh_per_m2": 2263,
            "monthly_temperature_c": [21, 22, 24, 27, 30, 33, 35, 35, 33, 30, 26, 22],
            "annual_avg_speed_50m_mps": 6.59,
        }
        payload = {
            "project_info": {
                "place_name": "Mauritania Mine",
            },
            "resource_data": {
                "solar": {},
                "wind": {},
            },
        }

        enriched, meta = enrich_with_auto_resource_data(payload)
        self.assertTrue(meta["resource_fetch_attempted"])
        self.assertEqual(meta["resource_fetch_status"], "fetched_hourly")
        self.assertEqual(enriched["project_info"]["latitude"], 23.196941)
        self.assertEqual(enriched["resource_data"]["solar"]["annual_irradiation_kwh_per_m2"], 2263)
        self.assertEqual(enriched["resource_data"]["wind"]["annual_avg_speed_mps"], 6.59)
        self.assertEqual(len(enriched["resource_data"]["solar"]["hourly_irradiance_kwh_per_m2"]), 8760)
        self.assertEqual(len(enriched["resource_data"]["wind"]["wind_speed_series_mps"]), 8760)
