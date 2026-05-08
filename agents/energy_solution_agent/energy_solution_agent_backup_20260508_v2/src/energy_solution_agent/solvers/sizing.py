from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..settlement import build_hourly_price_series, resolve_price_series_start_weekday
from .utils import (
    _calendar_month_slices,
    _clean_workbook_text,
    _infer_steps_per_hour_for_sizing,
    _integrate_charge_space_kwh,
    _integrate_discharge_space_kwh,
    _max_charge_space_kw,
    _max_discharge_space_kw,
    _percentile_inc,
    _resolve_monthly_billing_baseline_kw,
)
from .cycles import _build_default_schedule_from_periods, infer_daily_cycles_from_schedule
from .arbitrage import (
    _build_daily_spot_profiles_from_series,
    _month_from_profile_or_index,
    _normalize_price_series_to_yuan_per_mwh,
    _normalized_spot_intraday_plan,
    _select_best_daily_spot_cycles,
)

def estimate_storage(
    data: dict[str, Any],
    charging_peak_kw: float = 0.0,
    thermal_coupling_kw: float = 0.0,
    load_series_kw: list[float] | None = None,
    pv_series_kw: list[float] | None = None,
    wind_series_kw: list[float] | None = None,
    operation_mode: str = "",
) -> dict[str, Any]:
    load = data.get("load_data", {})
    market = data.get("market_data", {})
    equipment = data.get("equipment", {}).get("storage", {})
    peak_load_kw = float(load.get("peak_load_kw") or 0.0)
    candidate_powers = [float(v) for v in equipment.get("power_candidate_kw") or []]
    candidate_energies = [float(v) for v in equipment.get("energy_candidate_kwh") or []]
    sample_day_sizing = _estimate_storage_sizing_from_sample_day_workbook(data)
    sample_day_energy_kwh = sample_day_sizing.get("energy_kwh")
    sample_day_power_kw = sample_day_sizing.get("power_kw")

    if sample_day_energy_kwh is None and _supports_series_storage_sizing(data) and load_series_kw:
        if operation_mode in {"renewable_market_cooptimization", "renewable_export_oriented"}:
            sized_from_market = _estimate_storage_from_renewable_market_opportunity(
                data,
                load_series_kw=load_series_kw,
                pv_series_kw=pv_series_kw or [],
                wind_series_kw=wind_series_kw or [],
            )
            if sized_from_market:
                return sized_from_market
        sized_from_spot = _estimate_storage_from_spot_load_spaces(data, load_series_kw)
        if sized_from_spot:
            return sized_from_spot
        sized_from_series = _estimate_storage_from_load_spaces(data, load_series_kw)
        if sized_from_series:
            return sized_from_series

    if sample_day_power_kw is not None and sample_day_power_kw > 0:
        power_kw = sample_day_power_kw
    elif candidate_powers:
        power_kw = candidate_powers[0]
    else:
        base = peak_load_kw * 0.22 + charging_peak_kw * 0.35 + thermal_coupling_kw * 0.2
        if market.get("demand_charge_rule"):
            base *= 1.15
        power_kw = max(500.0, round(base, 0))
    if sample_day_energy_kwh is not None and sample_day_energy_kwh > 0:
        energy_kwh = sample_day_energy_kwh
    elif candidate_energies:
        energy_kwh = candidate_energies[0]
    else:
        energy_kwh = power_kw * 2.0

    rte = float(equipment.get("round_trip_efficiency") or 0.88)
    annual_discharge = power_kw * 0.55 * 365 / 1000
    annual_charge = annual_discharge / max(rte, 0.01)
    return {
        "raw_storage_power_mw": round(power_kw / 1000, 3),
        "raw_storage_energy_mwh": round(energy_kwh / 1000, 3),
        "storage_power_mw": round(power_kw / 1000, 3),
        "storage_energy_mwh": round(energy_kwh / 1000, 3),
        "annual_storage_charge_mwh": round(annual_charge, 2),
        "annual_storage_discharge_mwh": round(annual_discharge, 2),
    }


def _supports_series_storage_sizing(data: dict[str, Any]) -> bool:
    scenario_type = str((data.get("project_info", {}) or {}).get("scenario_type") or "").lower()
    return scenario_type in {"user_side_storage", "source_grid_load_storage", "charging_station"}


def _estimate_storage_from_renewable_market_opportunity(
    data: dict[str, Any],
    load_series_kw: list[float],
    pv_series_kw: list[float],
    wind_series_kw: list[float],
) -> dict[str, Any] | None:
    if not load_series_kw or (not pv_series_kw and not wind_series_kw):
        return None
    market = data.get("market_data", {}) or {}
    prices = build_hourly_price_series(
        market,
        len(load_series_kw),
        start_weekday=resolve_price_series_start_weekday(data),
    )
    if not prices:
        return None
    steps_per_hour = _infer_steps_per_hour_for_sizing(load_series_kw)
    if steps_per_hour is None:
        return None
    day_steps = 24 * steps_per_hour
    if len(load_series_kw) < day_steps:
        return None
    storage_cfg = data.get("equipment", {}).get("storage", {}) or {}
    percentile = min(max(1.0 - float(storage_cfg.get("sizing_target_day_coverage_ratio") or 0.9), 0.0), 1.0)
    spread_margin = float(
        market.get("cooptimization_min_sell_spread_per_kwh")
        or market.get("min_sell_spread_per_kwh")
        or 0.15
    )
    energy_candidates: list[float] = []
    plan = market.get("arbitrage_plan") or {}
    min_charge_hours = max(2, int(plan.get("min_charge_hours") or 2))
    min_discharge_hours = max(2, int(plan.get("min_discharge_hours") or 2))
    horizon = min(len(load_series_kw), len(pv_series_kw) if pv_series_kw else len(load_series_kw), len(wind_series_kw) if wind_series_kw else len(load_series_kw), len(prices))
    for day_start in range(0, horizon - day_steps + 1, day_steps):
        day_load = load_series_kw[day_start : day_start + day_steps]
        day_pv = (pv_series_kw[day_start : day_start + day_steps] if pv_series_kw else [0.0] * day_steps)
        day_wind = (wind_series_kw[day_start : day_start + day_steps] if wind_series_kw else [0.0] * day_steps)
        day_prices = prices[day_start : day_start + day_steps]
        if len(day_load) < day_steps or len(day_prices) < day_steps:
            continue
        daily_renewable_to_storage_kwh = 0.0
        for step in range(day_steps):
            renewable_kw = float(day_pv[step] if step < len(day_pv) else 0.0) + float(day_wind[step] if step < len(day_wind) else 0.0)
            if renewable_kw <= 0:
                continue
            price = float(day_prices[step])
            threshold_price = _renewable_charge_threshold_price_for_step(
                market=market,
                pv_kw=float(day_pv[step] if step < len(day_pv) else 0.0),
                wind_kw=float(day_wind[step] if step < len(day_wind) else 0.0),
            )
            future_prices = day_prices[step:]
            future_peak = max(float(v) for v in future_prices) if future_prices else price
            if price > threshold_price:
                continue
            if future_peak - price < spread_margin:
                continue
            step_hours = 1.0 / steps_per_hour
            daily_renewable_to_storage_kwh += renewable_kw * step_hours
        if daily_renewable_to_storage_kwh > 0:
            energy_candidates.append(daily_renewable_to_storage_kwh)
    if not energy_candidates:
        return None
    energy_kwh = _percentile_inc(energy_candidates, percentile)
    power_kw = max(energy_kwh / max(min_charge_hours, 1), energy_kwh / max(min_discharge_hours, 1))
    if power_kw <= 0 or energy_kwh <= 0:
        return None
    rte = float(storage_cfg.get("battery_charge_efficiency") or 0.96) * float(storage_cfg.get("battery_discharge_efficiency") or 0.96)
    annual_discharge = energy_kwh * 0.78 * 365 / 1000
    annual_charge = annual_discharge / max(rte, 0.01)
    return {
        "raw_storage_power_mw": round(power_kw / 1000, 3),
        "raw_storage_energy_mwh": round(energy_kwh / 1000, 3),
        "storage_power_mw": round(power_kw / 1000, 3),
        "storage_energy_mwh": round(energy_kwh / 1000, 3),
        "annual_storage_charge_mwh": round(annual_charge, 2),
        "annual_storage_discharge_mwh": round(annual_discharge, 2),
    }


def _renewable_charge_threshold_price_for_step(
    market: dict[str, Any],
    pv_kw: float,
    wind_kw: float,
) -> float:
    if market.get("renewable_charge_threshold_price_per_kwh") not in (None, ""):
        return float(market.get("renewable_charge_threshold_price_per_kwh") or 0.0)
    solar_lcoe = float(market.get("solar_lcoe_per_kwh") or 0.18)
    wind_lcoe = float(market.get("wind_lcoe_per_kwh") or 0.35)
    total = max(0.0, pv_kw) + max(0.0, wind_kw)
    if total <= 0:
        return max(solar_lcoe, wind_lcoe)
    weighted = (max(0.0, pv_kw) * solar_lcoe + max(0.0, wind_kw) * wind_lcoe) / total
    return weighted


def _estimate_storage_sizing_from_sample_day_workbook(data: dict[str, Any]) -> dict[str, float | None]:
    load_data = data.get("load_data", {}) or {}
    workbook_path = load_data.get("sizing_workbook_path")
    if not workbook_path or str(workbook_path).lower().endswith(".xlsx") is False:
        return {"energy_kwh": None, "power_kw": None}
    try:
        workbook = load_workbook(Path(str(workbook_path)), data_only=True, read_only=True)
    except Exception:
        return {"energy_kwh": None, "power_kw": None}

    metrics = _extract_sample_day_energy_metrics(workbook)
    if not metrics:
        return {"energy_kwh": None, "power_kw": None}
    storage_cfg = ((data.get("equipment", {}) or {}).get("storage", {}) or {})
    coverage_ratio = float(storage_cfg.get("sizing_target_day_coverage_ratio") or metrics.get("coverage_ratio") or 0.9)
    percentile = min(max(1.0 - coverage_ratio, 0.0), 1.0)
    values = []
    for key in ("charge_1_kwh", "discharge_1_kwh", "charge_2_kwh", "discharge_2_kwh"):
        metric_values = metrics.get(key) or []
        if metric_values:
            values.append(_percentile_inc(metric_values, percentile))
    # Assumes 2 charge/discharge cycles per day (4 events total). The divisor /4.0
    # splits combined charge_1+discharge_1+charge_2+discharge_2 into a per-event
    # energy estimate. Must match the number of aggregated event types in the workbook.
    energy_kwh = round(min(values) / 4.0, 1) if values else None
    power_candidates = []
    for key in ("charge_1_kw", "discharge_1_kw", "charge_2_kw", "discharge_2_kw"):
        metric_values = metrics.get(key) or []
        if metric_values:
            power_candidates.append(_percentile_inc(metric_values, percentile))
    power_kw = round(min(power_candidates), 1) if power_candidates else None
    # Sample-day workbooks store 15-minute quarter-hour totals; convert back to required energy capacity.
    return {"energy_kwh": energy_kwh, "power_kw": power_kw}


def _extract_sample_day_energy_metrics(workbook: Any) -> dict[str, list[float]] | None:
    best_metrics = None
    best_count = 0
    for ws in workbook.worksheets:
        coverage_ratio = _extract_workbook_coverage_ratio(ws)
        power_metrics = _extract_sample_day_power_metrics(ws)
        row_values = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))
        if not row_values:
            continue
        row = list(row_values[0])
        labels = {_clean_workbook_text(value) for value in row if isinstance(value, str)}
        if not any(_looks_like_sample_day_label(label) for label in labels):
            continue
        metrics = {
            "charge_1_kwh": [],
            "discharge_1_kwh": [],
            "charge_2_kwh": [],
            "discharge_2_kwh": [],
            "charge_1_kw": power_metrics.get("charge_1_kw") or [],
            "discharge_1_kw": power_metrics.get("discharge_1_kw") or [],
            "charge_2_kw": power_metrics.get("charge_2_kw") or [],
            "discharge_2_kw": power_metrics.get("discharge_2_kw") or [],
            "coverage_ratio": coverage_ratio,
        }
        for idx, value in enumerate(row):
            text = _clean_workbook_text(value)
            if not _looks_like_sample_day_label(text):
                continue
            numeric = []
            for follower in row[idx + 1 :]:
                if isinstance(follower, (int, float)):
                    numeric.append(float(follower))
                    if len(numeric) == 4:
                        break
            if len(numeric) == 4 and all(value > 0 for value in numeric):
                metrics["charge_1_kwh"].append(numeric[0])
                metrics["discharge_1_kwh"].append(numeric[1])
                metrics["charge_2_kwh"].append(numeric[2])
                metrics["discharge_2_kwh"].append(numeric[3])
        trailing_numeric = [float(value) for value in row if isinstance(value, (int, float))]
        if len(trailing_numeric) >= 4:
            tail = trailing_numeric[-4:]
            if all(value > 0 for value in tail):
                last_existing = (
                    metrics["charge_1_kwh"][-1] if metrics["charge_1_kwh"] else None,
                    metrics["discharge_1_kwh"][-1] if metrics["discharge_1_kwh"] else None,
                    metrics["charge_2_kwh"][-1] if metrics["charge_2_kwh"] else None,
                    metrics["discharge_2_kwh"][-1] if metrics["discharge_2_kwh"] else None,
                )
                if tuple(tail) != last_existing:
                    metrics["charge_1_kwh"].append(tail[0])
                    metrics["discharge_1_kwh"].append(tail[1])
                    metrics["charge_2_kwh"].append(tail[2])
                    metrics["discharge_2_kwh"].append(tail[3])
        if len(metrics["charge_1_kwh"]) > best_count:
            best_metrics = metrics
            best_count = len(metrics["charge_1_kwh"])
    return best_metrics


def _looks_like_sample_day_label(text: str) -> bool:
    return text.startswith("第") and ("日" in text or "天" in text)


def _extract_workbook_coverage_ratio(ws: Any) -> float | None:
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for idx, value in enumerate(row):
            if _clean_workbook_text(value) == "充放电百分位":
                for follower in row[idx + 1 :]:
                    if isinstance(follower, (int, float)):
                        numeric = float(follower)
                        if 0.0 < numeric <= 1.0:
                            return numeric
    return None


def _extract_sample_day_power_metrics(ws: Any) -> dict[str, list[float]]:
    metrics = {
        "charge_1_kw": [],
        "discharge_1_kw": [],
        "charge_2_kw": [],
        "discharge_2_kw": [],
    }
    label_map = {
        "一次充电功率": "charge_1_kw",
        "一充功率": "charge_1_kw",
        "第一次充电功率": "charge_1_kw",
        "一次放电功率": "discharge_1_kw",
        "一放功率": "discharge_1_kw",
        "第一次放电功率": "discharge_1_kw",
        "二次充电功率": "charge_2_kw",
        "二充功率": "charge_2_kw",
        "第二次充电功率": "charge_2_kw",
        "二次放电功率": "discharge_2_kw",
        "二放功率": "discharge_2_kw",
        "第二次放电功率": "discharge_2_kw",
    }
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        clean = [_clean_workbook_text(value) for value in row]
        for idx, text in enumerate(clean):
            metric_key = label_map.get(text)
            if not metric_key:
                continue
            for follower in row[idx + 1 :]:
                if isinstance(follower, (int, float)) and float(follower) > 0:
                    metrics[metric_key].append(float(follower))
                    break
    return metrics


def _estimate_storage_from_load_spaces(data: dict[str, Any], load_series_kw: list[float]) -> dict[str, Any] | None:
    market = data.get("market_data", {}) or {}
    history = market.get("monthly_tou_policy_history") or []
    prices = {str(item.get("period")): float(item.get("price", 0.0)) for item in (market.get("tou_tariff") or [])}
    if not history or not prices or not load_series_kw:
        return None

    steps_per_hour = _infer_steps_per_hour_for_sizing(load_series_kw)
    if steps_per_hour is None:
        return None
    percentile = 1.0 - float((data.get("equipment", {}).get("storage", {}) or {}).get("sizing_target_day_coverage_ratio") or 0.9)
    percentile = min(max(percentile, 0.0), 1.0)
    month_slices = _calendar_month_slices(steps_per_hour)

    power_candidates: list[float] = []
    energy_candidates: list[float] = []

    for item in history:
        month = int(item.get("month") or 0)
        if month < 1 or month > 12 or month not in month_slices:
            continue
        month_start, month_end = month_slices[month]
        month_series = load_series_kw[month_start:month_end]
        if not month_series:
            continue
        schedule = item.get("schedule") or {}
        periods = list(str(period) for period in (item.get("periods") or []))
        if not schedule and periods:
            schedule = _build_default_schedule_from_periods(periods)
        if not schedule:
            continue
        cycles = infer_daily_cycles_from_schedule(
            schedule=schedule,
            prices=prices,
            first_cycle_min_spread=float((market.get("arbitrage_plan") or {}).get("first_cycle_min_spread") or 0.0),
            second_cycle_min_spread=float((market.get("arbitrage_plan") or {}).get("second_cycle_min_spread") or 0.0),
            discharge_energy_kwh=1.0,
            effective_rte=1.0,
        )
        if not cycles:
            continue
        baseline_kw = _resolve_monthly_billing_baseline_kw(data, month_series, month)
        if baseline_kw <= 0:
            continue
        days_in_month = len(month_series) // (24 * steps_per_hour)
        for cycle in cycles:
            charge_hours = [int(hour) for hour in (cycle.get("charge_hours") or [])]
            discharge_hours = [int(hour) for hour in (cycle.get("discharge_hours") or [])]
            if not charge_hours or not discharge_hours:
                continue
            daily_charge_energy: list[float] = []
            daily_discharge_energy: list[float] = []
            daily_charge_power: list[float] = []
            daily_discharge_power: list[float] = []
            for day_idx in range(days_in_month):
                day_start = day_idx * 24 * steps_per_hour
                day_series = month_series[day_start : day_start + 24 * steps_per_hour]
                if len(day_series) < 24 * steps_per_hour:
                    continue
                charge_space = _integrate_charge_space_kwh(day_series, baseline_kw, charge_hours, steps_per_hour)
                discharge_space = _integrate_discharge_space_kwh(day_series, discharge_hours, steps_per_hour)
                charge_power = _max_charge_space_kw(day_series, baseline_kw, charge_hours, steps_per_hour)
                discharge_power = _max_discharge_space_kw(day_series, discharge_hours, steps_per_hour)
                if charge_space > 0:
                    daily_charge_energy.append(charge_space)
                if discharge_space > 0:
                    daily_discharge_energy.append(discharge_space)
                if charge_power > 0:
                    daily_charge_power.append(charge_power)
                if discharge_power > 0:
                    daily_discharge_power.append(discharge_power)
            if daily_charge_energy:
                energy_candidates.append(_percentile_inc(daily_charge_energy, percentile))
            if daily_discharge_energy:
                energy_candidates.append(_percentile_inc(daily_discharge_energy, percentile))
            if daily_charge_power:
                power_candidates.append(_percentile_inc(daily_charge_power, percentile))
            if daily_discharge_power:
                power_candidates.append(_percentile_inc(daily_discharge_power, percentile))

    if not power_candidates or not energy_candidates:
        return None

    power_kw = min(power_candidates)
    energy_kwh = min(energy_candidates)
    return {
        "raw_storage_power_mw": round(power_kw / 1000, 3),
        "raw_storage_energy_mwh": round(energy_kwh / 1000, 3),
        "storage_power_mw": round(power_kw / 1000, 3),
        "storage_energy_mwh": round(energy_kwh / 1000, 3),
        "annual_storage_charge_mwh": None,
        "annual_storage_discharge_mwh": None,
    }


def _estimate_storage_from_spot_load_spaces(data: dict[str, Any], load_series_kw: list[float]) -> dict[str, Any] | None:
    market = data.get("market_data", {}) or {}
    plan = _normalized_spot_intraday_plan(market)
    if str(plan.get("mode") or "").lower() != "spot_intraday":
        return None
    daily_profiles = market.get("spot_price_daily_profiles") or _build_daily_spot_profiles_from_series(market.get("market_price_series") or [])
    if not daily_profiles or not load_series_kw:
        return None

    steps_per_hour = _infer_steps_per_hour_for_sizing(load_series_kw)
    if steps_per_hour is None:
        return None
    day_steps = 24 * steps_per_hour
    profile_count = min(len(daily_profiles), len(load_series_kw) // day_steps)
    if profile_count <= 0:
        return None

    storage_cfg = data.get("equipment", {}).get("storage", {}) or {}
    charge_eff = float(storage_cfg.get("battery_charge_efficiency") or 0.96)
    discharge_eff = float(storage_cfg.get("battery_discharge_efficiency") or 0.96)
    pcs_eff = float(storage_cfg.get("pcs_efficiency") or 0.98)
    transformer_eff = float(storage_cfg.get("transformer_efficiency") or 0.99)
    effective_rte = charge_eff * discharge_eff * pcs_eff * transformer_eff

    min_charge_hours = int(plan["min_charge_hours"])
    min_discharge_hours = int(plan["min_discharge_hours"])
    max_charge_hours = int(plan["max_charge_hours"])
    max_discharge_hours = int(plan["max_discharge_hours"])
    min_spread_yuan_per_mwh = float(plan["min_spread_yuan_per_mwh"])
    percentile = min(max(1.0 - float(storage_cfg.get("sizing_target_day_coverage_ratio") or 0.9), 0.0), 1.0)

    power_candidates: list[float] = []
    energy_candidates: list[float] = []
    for profile_idx in range(profile_count):
        profile = daily_profiles[profile_idx]
        prices = _normalize_price_series_to_yuan_per_mwh(
            [float(v) for v in (profile.get("realtime_prices") or [])[:24]],
            unit_hint=str(plan.get("price_unit") or market.get("spot_price_unit") or market.get("market_price_unit") or ""),
        )
        if len(prices) < 24:
            continue
        cycles = _select_best_daily_spot_cycles(
            prices,
            min_charge_hours=min_charge_hours,
            min_discharge_hours=min_discharge_hours,
            min_spread_yuan_per_mwh=min_spread_yuan_per_mwh,
            power_mw=1.0,
            energy_mwh=float(plan.get("window_energy_probe_mwh") or max_charge_hours),
            usable_depth=1.0,
            charge_path_eff=1.0,
            discharge_path_eff=1.0,
            effective_rte=effective_rte,
            max_charge_hours=max_charge_hours,
            max_discharge_hours=max_discharge_hours,
        )
        if not cycles:
            continue
        day_start = profile_idx * day_steps
        day_series = load_series_kw[day_start : day_start + day_steps]
        if len(day_series) < day_steps:
            continue
        baseline_kw = _resolve_monthly_billing_baseline_kw(data, day_series, _month_from_profile_or_index(profile, profile_idx))
        if baseline_kw <= 0:
            continue
        for cycle in cycles:
            charge_hours = list(range(int(cycle["charge_start"]), int(cycle["charge_end"])))
            discharge_hours = list(range(int(cycle["discharge_start"]), int(cycle["discharge_end"])))
            charge_space = _integrate_charge_space_kwh(day_series, baseline_kw, charge_hours, steps_per_hour)
            discharge_space = _integrate_discharge_space_kwh(day_series, discharge_hours, steps_per_hour)
            charge_power = _max_charge_space_kw(day_series, baseline_kw, charge_hours, steps_per_hour)
            discharge_power = _max_discharge_space_kw(day_series, discharge_hours, steps_per_hour)
            if charge_space > 0 and discharge_space > 0:
                energy_candidates.append(min(charge_space, discharge_space))
            if charge_power > 0 and discharge_power > 0:
                power_candidates.append(min(charge_power, discharge_power))
    if not power_candidates or not energy_candidates:
        return None

    power_kw = _percentile_inc(power_candidates, percentile)
    energy_kwh = _percentile_inc(energy_candidates, percentile)
    if power_kw <= 0 or energy_kwh <= 0:
        return None
    annual_discharge = power_kw * 0.55 * 365 / 1000
    annual_charge = annual_discharge / max(effective_rte, 0.01)
    return {
        "raw_storage_power_mw": round(power_kw / 1000, 3),
        "raw_storage_energy_mwh": round(energy_kwh / 1000, 3),
        "storage_power_mw": round(power_kw / 1000, 3),
        "storage_energy_mwh": round(energy_kwh / 1000, 3),
        "annual_storage_charge_mwh": round(annual_charge, 2),
        "annual_storage_discharge_mwh": round(annual_discharge, 2),
    }


def apply_storage_product_selection(data: dict[str, Any], storage: dict[str, Any]) -> dict[str, Any]:
    storage_cfg = data.get("equipment", {}).get("storage", {}) or {}
    raw_power_kw = storage_cfg.get("raw_storage_power_kw")
    raw_energy_kwh = storage_cfg.get("raw_storage_energy_kwh")
    if raw_power_kw not in (None, "", 0):
        raw_power_mw = float(raw_power_kw) / 1000
    else:
        raw_power_mw = float(storage.get("raw_storage_power_mw") or storage.get("storage_power_mw") or 0.0)
    if raw_energy_kwh not in (None, "", 0):
        raw_energy_mwh = float(raw_energy_kwh) / 1000
    else:
        raw_energy_mwh = float(storage.get("raw_storage_energy_mwh") or storage.get("storage_energy_mwh") or 0.0)

    selected_power_kw = storage_cfg.get("selected_product_power_kw")
    selected_energy_kwh = storage_cfg.get("selected_product_energy_kwh")
    selected_power_mw = float(selected_power_kw) / 1000 if selected_power_kw not in (None, "", 0) else raw_power_mw
    if selected_energy_kwh not in (None, "", 0):
        selected_energy_mwh = float(selected_energy_kwh) / 1000
    else:
        selected_energy_mwh = _auto_selected_product_energy_mwh(data, raw_energy_mwh)

    result = dict(storage)
    result["raw_storage_power_mw"] = round(raw_power_mw, 3) if raw_power_mw > 0 else None
    result["raw_storage_energy_mwh"] = round(raw_energy_mwh, 3) if raw_energy_mwh > 0 else None
    result["storage_power_mw"] = round(selected_power_mw, 3) if selected_power_mw > 0 else None
    result["storage_energy_mwh"] = round(selected_energy_mwh, 3) if selected_energy_mwh > 0 else None
    result["selected_product_power_mw"] = round(selected_power_mw, 3) if selected_power_mw > 0 else None
    result["selected_product_energy_mwh"] = round(selected_energy_mwh, 3) if selected_energy_mwh > 0 else None
    result["storage_power_utilization_ratio"] = round(raw_power_mw / selected_power_mw, 4) if raw_power_mw > 0 and selected_power_mw > 0 else None
    result["storage_energy_utilization_ratio"] = round(raw_energy_mwh / selected_energy_mwh, 4) if raw_energy_mwh > 0 and selected_energy_mwh > 0 else None
    result["storage_power_oversize_ratio"] = round(selected_power_mw / raw_power_mw, 4) if raw_power_mw > 0 and selected_power_mw > 0 else None
    result["storage_energy_oversize_ratio"] = round(selected_energy_mwh / raw_energy_mwh, 4) if raw_energy_mwh > 0 and selected_energy_mwh > 0 else None
    return result


def _auto_selected_product_energy_mwh(data: dict[str, Any], raw_energy_mwh: float) -> float:
    if raw_energy_mwh <= 0:
        return raw_energy_mwh
    if _is_low_voltage_access(data):
        block_mwh = 0.261
        blocks = max(1, math.ceil(raw_energy_mwh / block_mwh))
        return round(blocks * block_mwh, 3)
    return raw_energy_mwh


def _is_low_voltage_access(data: dict[str, Any]) -> bool:
    voltage = (data.get("project_info", {}) or {}).get("voltage_level_kv")
    try:
        level = float(voltage)
    except (TypeError, ValueError):
        return False
    return abs(level - 0.38) < 1e-6


