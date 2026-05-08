from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def enrich_from_material_workbooks(data: dict[str, Any], base_dir: Path | None = None) -> dict[str, Any]:
    market = data.setdefault("market_data", {})
    storage_cfg = data.setdefault("equipment", {}).setdefault("storage", {})
    load_data = data.setdefault("load_data", {})
    base = base_dir or Path.cwd()

    calb_path = market.get("calb_workbook_path")
    if calb_path:
        parsed = parse_calb_user_storage_workbook(base / str(calb_path))
        _merge_if_missing(
            storage_cfg,
            {
                "power_candidate_kw": [parsed["power_kw"]] if parsed.get("power_kw") else None,
                "energy_candidate_kwh": [parsed["energy_kwh"]] if parsed.get("energy_kwh") else None,
                "battery_charge_efficiency": parsed.get("battery_charge_efficiency"),
                "battery_discharge_efficiency": parsed.get("battery_discharge_efficiency"),
                "pcs_charge_efficiency": parsed.get("pcs_charge_efficiency"),
                "pcs_discharge_efficiency": parsed.get("pcs_discharge_efficiency"),
                "max_daily_cycles": parsed.get("max_daily_cycles"),
            },
        )
        if parsed.get("peak_load_kw") and not load_data.get("peak_load_kw"):
            load_data["peak_load_kw"] = parsed["peak_load_kw"]
        if parsed.get("tou_tariff"):
            market.setdefault("tou_tariff", parsed["tou_tariff"])
        if parsed.get("arbitrage_plan"):
            market.setdefault("arbitrage_plan", parsed["arbitrage_plan"])
    return data


def parse_calb_user_storage_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    rows: list[list[Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cleaned = [cell for cell in row if cell not in (None, "")]
            if cleaned:
                rows.append(cleaned)

    def first_numeric_after(label: str) -> float | None:
        for row in rows:
            for idx, cell in enumerate(row):
                if str(cell).strip() == label:
                    for follower in row[idx + 1 :]:
                        if isinstance(follower, (int, float)):
                            return float(follower)
                        if isinstance(follower, str):
                            try:
                                return float(follower.strip())
                            except (ValueError, TypeError):
                                continue
        return None

    def first_numeric_after_any(*labels: str) -> float | None:
        for label in labels:
            value = first_numeric_after(label)
            if value is not None:
                return value
        return None

    # UTF-8 labels first, then GBK-Mojibake fallbacks for legacy Excel files
    # created on older Chinese Windows systems where text was GBK-encoded.
    power_kw = first_numeric_after_any("功率", "鍔熺巼")
    energy_kwh = first_numeric_after_any("容量", "瀹归噺")
    super_peak_price = first_numeric_after_any("尖峰电价", "灏栧嘲鐢典环")
    peak_price = first_numeric_after_any("峰电价", "宄扮數浠?")
    flat_price = first_numeric_after_any("平电价", "骞崇數浠?")
    valley_price = first_numeric_after_any("谷电价", "璋风數浠?")
    valley_super_peak_days = first_numeric_after_any("谷充尖峰次数", "璋?灏栧嘲娆℃暟") or 0.0
    valley_peak_days = first_numeric_after_any("谷充峰次数", "璋?宄版鏁?") or 0.0
    valley_flat_days = first_numeric_after_any("谷充平次数", "璋?骞虫鏁?") or 0.0
    first_discharge_kwh = first_numeric_after_any(
        "第一次放电容量",
        "第一次放电容量（高峰谷）(mWh)",
        "绗竴娆℃斁鐢靛閲?",
        "绗竴娆℃斁鐢靛閲忥紙楂樺嘲璋凤級锛坘Wh锛?",
    )
    second_discharge_kwh = first_numeric_after_any(
        "第二次放电容量",
        "第二次放电容量（高峰平）(mWh)",
        "绗簩娆℃斁鐢靛閲?",
        "绗簩娆℃斁鐢靛閲忥紙楂樺嘲骞筹級锛坘Wh锛?",
    )
    charge_eff = first_numeric_after_any("系统充电效率", "绯荤粺鍏呯數鏁堢巼")
    if charge_eff is None:
        charge_eff = 0.92
        logger.warning("battery charge efficiency not found in workbook, defaulting to 0.92")
    discharge_eff = first_numeric_after_any("系统放电效率", "绯荤粺鏀剧數鏁堢巼")
    if discharge_eff is None:
        discharge_eff = 0.92
        logger.warning("battery discharge efficiency not found in workbook, defaulting to 0.92")
    pcs_charge_eff = first_numeric_after_any("PCS充电效率", "PCS鍏呯數鏁堢巼")
    if pcs_charge_eff is None:
        pcs_charge_eff = 0.97
        logger.warning("PCS charge efficiency not found in workbook, defaulting to 0.97")
    pcs_discharge_eff = first_numeric_after_any("PCS放电效率", "PCS鏀剧數鏁堢巼")
    if pcs_discharge_eff is None:
        pcs_discharge_eff = 0.97
        logger.warning("PCS discharge efficiency not found in workbook, defaulting to 0.97")
    peak_load_kw = first_numeric_after_any("用户需量（变压器容量80%）", "鐢ㄦ埛闇€閲?鍙樺帇鍣ㄥ閲?0%")

    discharge_primary = float(first_discharge_kwh or 0.0)
    discharge_secondary = float(second_discharge_kwh or discharge_primary or 0.0)
    effective_rte = max(1e-9, charge_eff * discharge_eff * pcs_charge_eff * pcs_discharge_eff)

    cycles = []
    if discharge_primary > 0 and valley_super_peak_days > 0 and super_peak_price is not None and valley_price is not None:
        cycles.append(
            {
                "charge_period": "valley",
                "discharge_period": "super_peak",
                "days_per_year": float(valley_super_peak_days),
                "discharge_energy_kwh": discharge_primary,
                "charge_energy_kwh": discharge_primary / effective_rte,
            }
        )
    if discharge_primary > 0 and valley_peak_days > 0 and peak_price is not None and valley_price is not None:
        cycles.append(
            {
                "charge_period": "valley",
                "discharge_period": "peak",
                "days_per_year": float(valley_peak_days),
                "discharge_energy_kwh": discharge_primary,
                "charge_energy_kwh": discharge_primary / effective_rte,
            }
        )
    if discharge_secondary > 0 and valley_flat_days > 0 and flat_price is not None and valley_price is not None:
        cycles.append(
            {
                "charge_period": "valley",
                "discharge_period": "flat",
                "days_per_year": float(valley_flat_days),
                "discharge_energy_kwh": discharge_secondary,
                "charge_energy_kwh": discharge_secondary / effective_rte,
            }
        )

    tou_tariff = []
    if super_peak_price is not None:
        tou_tariff.append({"period": "super_peak", "price": float(super_peak_price)})
    if peak_price is not None:
        tou_tariff.append({"period": "peak", "price": float(peak_price)})
    if flat_price is not None:
        tou_tariff.append({"period": "flat", "price": float(flat_price)})
    if valley_price is not None:
        tou_tariff.append({"period": "valley", "price": float(valley_price)})

    return {
        "power_kw": power_kw,
        "energy_kwh": energy_kwh,
        "peak_load_kw": peak_load_kw,
        "battery_charge_efficiency": charge_eff,
        "battery_discharge_efficiency": discharge_eff,
        "pcs_charge_efficiency": pcs_charge_eff,
        "pcs_discharge_efficiency": pcs_discharge_eff,
        "tou_tariff": tou_tariff,
        "max_daily_cycles": 2.0 if len(cycles) >= 2 else 1.0 if cycles else 0.0,
        "arbitrage_plan": {"mode": "rule_based", "cycles": cycles} if cycles else None,
    }


def _merge_if_missing(target: dict[str, Any], patch: dict[str, Any]) -> None:
    for key, value in patch.items():
        if value is None:
            continue
        if key not in target or target.get(key) in (None, "", [], {}):
            target[key] = value
