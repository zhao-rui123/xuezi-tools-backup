from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


SERIES_FIELDS = [
    ("load_data", "load_series_kw"),
    ("load_data", "cooling_load_series_kw"),
    ("load_data", "heating_load_series_kw"),
    ("resource_data.solar", "hourly_generation_profile_kw"),
    ("resource_data.solar", "hourly_irradiance_kwh_per_m2"),
    ("resource_data.solar", "daily_irradiance_kwh_per_m2"),
    ("resource_data.solar", "hourly_temperature_c"),
    ("resource_data.solar", "daily_temperature_c"),
    ("resource_data.wind", "hourly_generation_profile_kw"),
    ("resource_data.wind", "wind_speed_series_mps"),
    ("resource_data.wind", "hourly_wind_speed_series_mps"),
    ("resource_data.wind", "daily_wind_speed_series_mps"),
    ("charging_data", "arrival_profile"),
]


def _resolve_series_path(raw_path: str, base: Path) -> Path:
    """Resolve a user-provided file path.

    - Relative paths are resolved against ``base`` and checked for directory traversal.
    - Absolute paths are allowed as-is (they represent explicit user intent).
    """
    candidate = Path(raw_path)
    if candidate.is_absolute():
        resolved = candidate.resolve()
        # Reject absolute paths pointing to sensitive system locations while
        # allowing normal user-accessible absolute paths (test fixtures, etc.).
        _FORBIDDEN_PREFIXES = (r"\Windows\System32", r"\Windows\SysWOW64", "/etc/", "/sys/", "/proc/", "/boot/")
        resolved_str = str(resolved)
        for prefix in _FORBIDDEN_PREFIXES:
            if prefix in resolved_str:
                raise ValueError(
                    f"Absolute path references a protected system location: {raw_path}"
                )
        return resolved
    resolved = (base / candidate).resolve()
    if not resolved.is_relative_to(base):
        raise ValueError(
            f"Relative path escapes base directory ({base}): {raw_path}"
        )
    return resolved


def ingest_external_series(data: dict[str, Any], base_dir: Path | None = None) -> dict[str, Any]:
    base = base_dir or Path.cwd()
    for section_path, field in SERIES_FIELDS:
        node = _get_node(data, section_path)
        if not isinstance(node, dict):
            continue
        path_key = f"{field}_path"
        if node.get(path_key):
            node[field] = _load_numeric_series(_resolve_series_path(str(node[path_key]), base))
    market = data.get("market_data")
    if isinstance(market, dict) and market.get("market_price_series_path"):
        _ingest_market_price_series(_resolve_series_path(str(market["market_price_series_path"]), base), market)
    return data


def _get_node(data: dict[str, Any], dotted: str) -> Any:
    current: Any = data
    for part in dotted.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def _load_numeric_series(path: Path) -> list[float]:
    if not path.exists():
        raise FileNotFoundError(f"Series file not found: {path}")
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            return [float(v) for v in payload]
        if isinstance(payload, dict):
            for key in ("values", "series", "data"):
                if key in payload and isinstance(payload[key], list):
                    return [float(v) for v in payload[key]]
        raise ValueError(f"Unsupported JSON series shape: {path}")
    if path.suffix.lower() in {".csv", ".txt"}:
        values: list[float] = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            for row in reader:
                for item in row:
                    item = item.strip()
                    if not item:
                        continue
                    try:
                        values.append(float(item))
                    except ValueError:
                        logger.warning("Non-numeric value skipped in %s: %r", path, item)
                        continue
        return values
    if path.suffix.lower() == ".xlsx":
        return _load_excel_numeric_series(path)
    raise ValueError(f"Unsupported series file type: {path.suffix}")


def _ingest_market_price_series(path: Path, market: dict[str, Any]) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Series file not found: {path}")
    if path.suffix.lower() != ".xlsx":
        market["market_price_series"] = _load_numeric_series(path)
        return
    realtime_series, daily_profiles, day_ahead_series = _load_spot_price_workbook(path)
    if realtime_series:
        market["market_price_series"] = realtime_series
    if daily_profiles:
        market["spot_price_daily_profiles"] = daily_profiles
    if day_ahead_series:
        market["day_ahead_market_price_series"] = day_ahead_series


def _load_excel_numeric_series(path: Path) -> list[float]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            spot_values, _, _ = _load_spot_price_day_pairs(ws)
            if spot_values:
                return spot_values
            matrix_values = _load_day_matrix_series(ws)
            if matrix_values:
                return _sanitize_power_series(matrix_values)
            point_values = _load_timestamp_value_series(ws)
            if point_values:
                return _sanitize_power_series(point_values)
            grid_query_values = _load_grid_query_series(ws)
            if grid_query_values:
                return _sanitize_power_series(grid_query_values)
        raise ValueError(f"Unsupported Excel series shape: {path}")
    finally:
        wb.close()


def _load_spot_price_workbook(path: Path) -> tuple[list[float], list[dict[str, Any]], list[float]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            realtime_values, daily_profiles, day_ahead_values = _load_spot_price_day_pairs(ws)
            if realtime_values:
                return realtime_values, daily_profiles, day_ahead_values
        raise ValueError(f"Unsupported spot price workbook shape: {path}")
    finally:
        wb.close()


def _load_spot_price_day_pairs(ws: Any) -> tuple[list[float], list[dict[str, Any]], list[float]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return [], [], []
    daily_profiles: list[dict[str, Any]] = []
    realtime_values: list[float] = []
    day_ahead_values: list[float] = []
    row_idx = 0
    while row_idx <= len(rows) - 3:
        label_row = list(rows[row_idx + 1]) if row_idx + 1 < len(rows) else []
        date_row = list(rows[row_idx + 2]) if row_idx + 2 < len(rows) else []
        if str(date_row[0] or "").strip() != "时段" or sum(1 for value in label_row if str(value or "").strip() == "实时") < 1:
            row_idx += 1
            continue
        data_rows: list[list[Any]] = []
        cursor = row_idx + 3
        while cursor < len(rows):
            row = list(rows[cursor])
            first = str(row[0] or "").strip() if row else ""
            if not first:
                break
            if first == "均值":
                break
            if "-" not in first:
                break
            data_rows.append(row)
            cursor += 1
        max_col = min(len(label_row), len(date_row))
        for col in range(1, max_col - 1, 2):
            day_ahead_label = str(label_row[col] or "").strip()
            realtime_label = str(label_row[col + 1] or "").strip()
            if day_ahead_label != "日前" or realtime_label != "实时":
                continue
            date_value = date_row[col] or date_row[col + 1]
            if date_value is None:
                continue
            day_ahead_day: list[float] = []
            realtime_day: list[float] = []
            for row in data_rows:
                if col + 1 >= len(row):
                    continue
                day_ahead_value = row[col]
                realtime_value = row[col + 1]
                if day_ahead_value is None or realtime_value is None:
                    continue
                try:
                    day_ahead_day.append(float(day_ahead_value))
                    realtime_day.append(float(realtime_value))
                except (TypeError, ValueError):
                    continue
            if len(realtime_day) < 24:
                continue
            profile = {
                "date": _normalize_excel_date(date_value),
                "day_ahead_prices": day_ahead_day[:24],
                "realtime_prices": realtime_day[:24],
            }
            daily_profiles.append(profile)
            day_ahead_values.extend(profile["day_ahead_prices"])
            realtime_values.extend(profile["realtime_prices"])
        row_idx = cursor + 1
    return realtime_values, daily_profiles, day_ahead_values


def _normalize_excel_date(value: Any) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _load_day_matrix_series(ws: Any) -> list[float]:
    rows = ws.iter_rows(values_only=True)
    label_row = None
    matched_rows = 0
    collected: list[list[float]] = []
    for row in rows:
        values = list(row)
        labels = [str(item).strip() if item is not None else "" for item in values]
        if label_row is None and any("00:00" in label for label in labels):
            label_row = labels
            continue
        if label_row is None:
            continue
        if len(values) < len(label_row):
            values += [None] * (len(label_row) - len(values))
        start_idx = label_row.index("00:00")
        if len(values) <= 7 or values[7] is None:
            logger.warning("Day matrix row missing data_type at column 7, skipping")
            continue
        data_type = str(values[7]).strip()
        if data_type and "总有功功率" not in data_type:
            continue
        numeric_row: list[float] = []
        for item in values[start_idx:]:
            if item in (None, ""):
                continue
            try:
                numeric_row.append(float(item))
            except (TypeError, ValueError):
                break
        if len(numeric_row) >= 4:
            matched_rows += 1
            collected.append(numeric_row[:96])
    if matched_rows > 1:
        series: list[float] = []
        for row in collected:
            series.extend(row)
        return series
    if collected:
        return collected[0]
    return []


def _load_timestamp_value_series(ws: Any) -> list[float]:
    values: list[float] = []
    rows = ws.iter_rows(min_row=1, values_only=True)
    found_header = False
    for row in rows:
        cells = list(row)
        if not found_header:
            labels = [str(item).strip() if item is not None else "" for item in cells]
            if "总有功功率(kW)" in labels:
                found_header = True
            continue
        if len(cells) < 4:
            continue
        try:
            values.append(float(cells[3]))
        except (TypeError, ValueError):
            continue
    if values:
        values.reverse()
    return values


def _load_grid_query_series(ws: Any) -> list[float]:
    rows = ws.iter_rows(values_only=True)
    header = None
    values: list[float] = []
    for row in rows:
        cells = list(row)
        labels = [str(item).strip() if item is not None else "" for item in cells]
        if header is None:
            if "数据日期" in labels and "时间" in labels and "功率(KW)" in labels:
                header = labels
            continue
        if len(cells) < 24:
            continue
        if cells[12] in (None, ""):
            continue
        try:
            values.append(float(cells[23]))
        except (TypeError, ValueError):
            continue
    return values


def _sanitize_power_series(values: list[float]) -> list[float]:
    if not values:
        return values
    positives: list[float] = []
    for v in values:
        if isinstance(v, (int, float)):
            fv = float(v)
            if 1.0 <= fv <= 100000.0:
                positives.append(fv)
    if not positives:
        return list(values)
    sorted_pos = sorted(positives)
    median = sorted_pos[len(sorted_pos) // 2]
    upper_bound = max(median * 8, median + 50000)
    cleaned: list[float] = []
    for value in values:
        numeric = float(value)
        if abs(numeric) > upper_bound:
            continue
        cleaned.append(numeric)
    return cleaned
