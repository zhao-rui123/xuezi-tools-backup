from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


PERIOD_MAP = {
    "尖": "super_peak",
    "尖峰": "super_peak",
    "尖峰时段": "super_peak",
    "深谷": "deep_valley",
    "深谷时段": "deep_valley",
    "峰": "peak",
    "峰时段": "peak",
    "平": "flat",
    "平时段": "flat",
    "谷": "valley",
    "谷时段": "valley",
}

MONTH_RANGE_PATTERNS = (
    re.compile(r"(\d{1,2})\s*[-~至到]\s*(\d{1,2})月"),
    re.compile(r"(\d{1,2})月\s*[-~至到]\s*(?:次年)?(\d{1,2})月"),
)
MONTH_LIST_PATTERN = re.compile(r"((?:\d{1,2}[、,，])+\d{1,2})月")
SINGLE_MONTH_PATTERN = re.compile(r"(\d{1,2})月")
TIME_RANGE_PATTERN = re.compile(
    r"(\d{1,2})[:：]?\d{0,2}\s*[-~至到]\s*(\d{1,2})[:：]?\d{0,2}\s*(尖峰|尖|深谷|峰|平|谷)(?:时段)?"
)


def enrich_from_province_tou_schedule_workbook(data: dict[str, Any], base_dir: Path | None = None) -> dict[str, Any]:
    market = data.setdefault("market_data", {})
    project = data.setdefault("project_info", {})
    province = str(project.get("province") or "").strip()
    workbook_path = market.get("province_tou_schedule_workbook_path")
    if not province or not workbook_path:
        return data

    base = base_dir or Path.cwd()
    parsed = parse_province_tou_schedule_workbook(base / str(workbook_path), province)
    if parsed:
        market.setdefault("monthly_tou_policy_history", parsed)
    return data


def parse_province_tou_schedule_workbook(path: Path, province: str) -> list[dict[str, Any]] | None:
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        logger.warning("Failed to load TOU schedule workbook %s: %s", path, exc)
        return None
    # read_only=False is required for cell fill color access in color grid parser
    try:
        if province not in workbook.sheetnames:
            return None
        sheet = workbook[province]
        return _parse_text_month_groups(sheet) or _parse_calendar_grid_month_groups(sheet)
    finally:
        workbook.close()


def _parse_text_month_groups(sheet: Any) -> list[dict[str, Any]] | None:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [_clean_text(cell) for cell in row if cell not in (None, "")]
        if values:
            rows.append(values)

    # Each entry is (months, raw_schedule_text) where months is a list of ints or the sentinel "other".
    # Uses a "pending months" state to handle rows where month info and schedule text appear on separate lines.
    month_groups: list[tuple[list[int] | str, str]] = []
    pending_months: list[int] | str | None = None
    for values in rows:
        joined = " ".join(values)
        months = next((_parse_month_group(value) for value in values if _parse_month_group(value) is not None), None)
        schedule_text = next((value for value in values if _extract_schedule_from_text(value)), None)
        if months is None:
            months = _parse_month_group(joined)
        if schedule_text is None and _extract_schedule_from_text(joined):
            schedule_text = joined
        if months is not None:
            if schedule_text:
                month_groups.append((months, schedule_text))
                pending_months = None
            else:
                pending_months = months
            continue
        if pending_months is not None and schedule_text:
            month_groups.append((pending_months, schedule_text))
            pending_months = None

    return _build_monthly_history_from_text_groups(month_groups)


def _build_monthly_history_from_text_groups(month_groups: list[tuple[list[int] | str, str]]) -> list[dict[str, Any]] | None:
    if not month_groups:
        return None

    history_by_month: dict[int, dict[str, Any]] = {}
    explicit_month_sets = [months for months, _ in month_groups if isinstance(months, list)]
    covered_months = {month for months in explicit_month_sets for month in months}
    other_months = [month for month in range(1, 13) if month not in covered_months]

    for months, text in month_groups:
        schedule = _extract_schedule_from_text(text)
        if not schedule:
            continue
        target_months = other_months if months == "other" else months
        for month in target_months:
            history_by_month[int(month)] = {
                "month": int(month),
                "periods": list(schedule.keys()),
                "schedule": schedule,
                "source_note": text,
            }

    return [history_by_month[month] for month in sorted(history_by_month)] or None


def _parse_calendar_grid_month_groups(sheet: Any) -> list[dict[str, Any]] | None:
    month_groups: list[tuple[list[int] | str, dict[str, list[int]], str]] = []
    current_months: list[int] | str | None = None
    current_note = ""
    current_schedule: dict[str, list[int]] = {}
    current_score = 0

    for row in sheet.iter_rows():
        values = [_clean_text(cell.value) for cell in row if cell.value not in (None, "")]
        if not values:
            continue

        month_text = next((value for value in values if _parse_month_group(value) is not None), None)
        if month_text is not None:
            if current_months is not None and current_schedule:
                month_groups.append((current_months, current_schedule, current_note))
            current_months = _parse_month_group(month_text)
            current_note = month_text
            current_schedule = {}
            current_score = 0
            continue

        if current_months is None or _looks_like_hour_grid_row(row):
            continue

        schedule = _extract_schedule_from_color_grid_row(row)
        score = sum(len(hours) for hours in schedule.values())
        if score > current_score:
            current_schedule = schedule
            current_score = score

    if current_months is not None and current_schedule:
        month_groups.append((current_months, current_schedule, current_note))

    if not month_groups:
        return None

    history_by_month: dict[int, dict[str, Any]] = {}
    explicit_month_sets = [months for months, _, _ in month_groups if isinstance(months, list)]
    covered_months = {month for months in explicit_month_sets for month in months}
    other_months = [month for month in range(1, 13) if month not in covered_months]

    for months, schedule, note in month_groups:
        target_months = other_months if months == "other" else months
        for month in target_months:
            history_by_month[int(month)] = {
                "month": int(month),
                "periods": list(schedule.keys()),
                "schedule": schedule,
                "source_note": note,
            }

    return [history_by_month[month] for month in sorted(history_by_month)] or None


def _clean_text(value: Any) -> str:
    return str(value).replace("\xa0", "").strip() if value not in (None, "") else ""


def _parse_month_group(text: str) -> list[int] | str | None:
    clean = _clean_text(text).replace("月份", "月").replace(" ", "")
    if not clean:
        return None
    if "其他月份" in clean or "其他月" in clean:
        return "other"
    if "春季" in clean:
        return [3, 4, 5]
    if "夏季" in clean:
        return [6, 7, 8]
    if "秋季" in clean:
        return [9, 10, 11]
    if "冬季" in clean:
        return [12, 1, 2]

    months: set[int] = set()
    for pattern in MONTH_RANGE_PATTERNS:
        for start_text, end_text in pattern.findall(clean):
            _add_month_range(months, int(start_text), int(end_text))

    for group in MONTH_LIST_PATTERN.findall(clean):
        for item in re.split(r"[、,，]", group):
            month = int(item)
            if 1 <= month <= 12:
                months.add(month)

    for item in SINGLE_MONTH_PATTERN.findall(clean):
        month = int(item)
        if 1 <= month <= 12:
            months.add(month)

    return sorted(months) if months else None


def _add_month_range(months: set[int], start: int, end: int) -> None:
    if start <= end:
        months.update(range(start, end + 1))
    else:
        months.update(range(start, 13))
        months.update(range(1, end + 1))


def _extract_schedule_from_text(text: str) -> dict[str, list[int]]:
    schedule: dict[str, set[int]] = {}
    for start, end, label in TIME_RANGE_PATTERN.findall(text):
        period = PERIOD_MAP.get(label)
        if period:
            schedule.setdefault(period, set()).update(_expand_hour_range(int(start), int(end)))
    return {period: sorted(hours) for period, hours in schedule.items()}


def _looks_like_hour_grid_row(row: list[Any]) -> bool:
    hours = [int(cell.value) for cell in row if isinstance(cell.value, (int, float))]
    return len(hours) >= 25 and hours[:4] == [0, 1, 2, 3] and hours[-1] == 24


def _extract_schedule_from_color_grid_row(row: list[Any]) -> dict[str, list[int]]:
    schedule: dict[str, list[int]] = {}
    for hour in range(24):
        start_col = 2 + hour * 2
        if start_col + 1 > len(row):
            break
        period = _period_from_cells(row[start_col - 1 : start_col + 1])
        if period:
            schedule.setdefault(period, []).append(hour)
    return schedule


def _period_from_cells(cells: list[Any]) -> str | None:
    for cell in cells:
        fill = getattr(cell, "fill", None)
        if not fill or getattr(fill, "patternType", None) != "solid":
            continue
        fg = getattr(fill, "fgColor", None)
        if not fg:
            continue
        if getattr(fg, "type", None) == "rgb":
            rgb = str(getattr(fg, "rgb", "") or "").upper()
            if rgb == "FFFF0000":
                return "super_peak"
            if rgb == "FF92D050":
                return "flat"
            if rgb == "FF00B0F0":
                return "valley"
            if rgb in {"", "00000000", "000000", "FFFFFFFF"}:
                continue
        if getattr(fg, "type", None) == "theme":
            return "peak"
    return None


def _expand_hour_range(start: int, end: int) -> list[int]:
    """Expand a time range into a list of hours using exclusive-end convention.

    ``range(start, end)`` produces hours ``[start, start+1, ..., end-1]``.
    For example, ``9:00-12:00`` yields ``[9, 10, 11]``, *not* ``[9, 10, 11, 12]``.
    This matches the standard TOU schedule representation where each hour
    index corresponds to the hour starting at that time.
    """
    if end > start:
        return list(range(start, end))
    if end < start:
        return list(range(start, 24)) + list(range(0, end))
    return [start]
