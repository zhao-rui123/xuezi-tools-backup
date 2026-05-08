"""Design-institute level Excel report for energy solution analysis.

Produces a multi-sheet professional workbook with KPI dashboard,
tables, and embedded charts — suitable for client delivery and further editing."""

from __future__ import annotations

import math
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, BarChart3D, LineChart, PieChart, PieChart3D, Reference,
    Series as ChartSeries,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
# Color System (matching DOCX palette)
# ═══════════════════════════════════════════════════════════════════

NAVY      = "0B1F3F"
DARK_BLUE = "142D5E"
MID_BLUE  = "1F478A"
ACCENT    = "2E6BC4"
LIGHT_BLUE="4A90D9"
PALE_BLUE = "8EBEEB"
ICE       = "D9E8F7"
WHITE     = "FFFFFF"

GREEN       = "1B8C4A"
LIGHT_GREEN = "DFF0E3"
RED         = "C0392B"
LIGHT_RED   = "F9E3E0"
ORANGE      = "D6840B"
LIGHT_ORANGE= "FEF2D6"

GRAY_DARK  = "4A4A4A"
GRAY       = "888888"
GRAY_LIGHT = "BBBBBB"
BG_GRAY    = "F5F6FA"

# ═══════════════════════════════════════════════════════════════════
# Style Definitions
# ═══════════════════════════════════════════════════════════════════

_font_normal = Font(name="Microsoft YaHei", size=10, color=GRAY_DARK)
_font_bold   = Font(name="Microsoft YaHei", size=10, color=GRAY_DARK, bold=True)
_font_title  = Font(name="Microsoft YaHei", size=16, color=NAVY, bold=True)
_font_heading= Font(name="Microsoft YaHei", size=13, color=NAVY, bold=True)
_font_header = Font(name="Microsoft YaHei", size=10, color=WHITE, bold=True)
_font_kpi_val= Font(name="Microsoft YaHei", size=18, color=WHITE, bold=True)
_font_kpi_lbl= Font(name="Microsoft YaHei", size=8, color=WHITE)
_font_small   = Font(name="Microsoft YaHei", size=9, color=GRAY)
_font_white_bold = Font(name="Microsoft YaHei", size=10, color=WHITE, bold=True)

_align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
_align_left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
_align_right  = Alignment(horizontal="right", vertical="center")

_thin_border = Border(
    left=Side(style="thin", color=GRAY_LIGHT),
    right=Side(style="thin", color=GRAY_LIGHT),
    top=Side(style="thin", color=GRAY_LIGHT),
    bottom=Side(style="thin", color=GRAY_LIGHT),
)

_fill_header = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_fill_kpi_green  = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
_fill_kpi_red    = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
_fill_kpi_orange = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
_fill_kpi_blue   = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
_fill_alt   = PatternFill(start_color=BG_GRAY, end_color=BG_GRAY, fill_type="solid")
_fill_ice   = PatternFill(start_color=ICE, end_color=ICE, fill_type="solid")
_fill_green = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
_fill_red   = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
_fill_orange= PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")


def _fmt(v: Any, decimals: int = 2) -> str:
    if v is None: return "—"
    if isinstance(v, bool): return "是" if v else "否"
    try:
        fv = float(v)
        if abs(fv) >= 1e4:
            return f"{fv:,.{decimals}f}"
        if decimals == 0:
            return f"{fv:,.0f}"
        return f"{fv:,.{decimals}f}"
    except: return str(v)

def _pct(v: Any) -> str:
    if v is None: return "—"
    if isinstance(v, bool): return "是" if v else "否"
    try:
        fv = float(v)
        return f"{fv*100:.2f}%" if abs(fv)<1 else f"{fv:.2f}%"
    except: return str(v)

def _style_header_row(ws, row: int, cols: int, fill=_fill_header, font=_font_header):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = _align_center
        cell.border = _thin_border

def _style_data_rows(ws, start_row: int, end_row: int, cols: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _font_normal
            cell.alignment = _align_center if c > 1 else _align_left
            cell.border = _thin_border
            if (r - start_row) % 2 == 1:
                cell.fill = _fill_alt

def _add_title(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font_title

def _add_heading(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font_heading

def _add_kv_table(ws, start_row: int, col_key: int, col_val: int,
                  data: list[tuple[str, str]], key_width: int = 20):
    ws.column_dimensions[get_column_letter(col_key)].width = key_width
    ws.column_dimensions[get_column_letter(col_val)].width = 45
    for i, (k, v) in enumerate(data):
        r = start_row + i
        ck = ws.cell(row=r, column=col_key, value=k)
        ck.font = _font_bold; ck.alignment = _align_left; ck.border = _thin_border
        if i % 2 == 1: ck.fill = _fill_alt
        cv = ws.cell(row=r, column=col_val, value=v)
        cv.font = _font_normal; cv.alignment = _align_left; cv.border = _thin_border
        if i % 2 == 1: cv.fill = _fill_alt

def _sheet_setup(ws, title: str, freeze_row: int = 1):
    ws.sheet_properties.tabColor = NAVY
    ws.freeze_panes = f"A{freeze_row + 1}"


# ═══════════════════════════════════════════════════════════════════
# Chart Builders (openpyxl native charts)
# ═══════════════════════════════════════════════════════════════════

def _add_bar_chart(ws, data_start_row: int, data_end_row: int,
                   data_col: int, label_col: int,
                   anchor: str, title: str, width: int = 18, height: int = 10,
                   color: str = MID_BLUE):
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.width = width; chart.height = height

    cats = Reference(ws, min_col=label_col, min_row=data_start_row,
                     max_row=data_end_row)
    vals = Reference(ws, min_col=data_col, min_row=data_start_row - 1,
                     max_row=data_end_row)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None

    s = chart.series[0]
    s.graphicalProperties.solidFill = color
    s.graphicalProperties.line.solidFill = "FFFFFF"

    ws.add_chart(chart, anchor)

def _add_line_chart(ws, data_start_row: int, data_end_row: int,
                    data_col: int, label_col: int,
                    anchor: str, title: str, width: int = 18, height: int = 10,
                    color: str = ACCENT):
    chart = LineChart()
    chart.title = title
    chart.width = width; chart.height = height
    chart.style = 2

    cats = Reference(ws, min_col=label_col, min_row=data_start_row,
                     max_row=data_end_row)
    vals = Reference(ws, min_col=data_col, min_row=data_start_row - 1,
                     max_row=data_end_row)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    s = chart.series[0]
    s.graphicalProperties.solidFill = color
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = 25000  # EMU

    ws.add_chart(chart, anchor)

def _add_pie_chart(ws, data_start_row: int, data_end_row: int,
                   data_col: int, label_col: int,
                   anchor: str, title: str, width: int = 14, height: int = 10):
    chart = PieChart()
    chart.title = title
    chart.width = width; chart.height = height

    cats = Reference(ws, min_col=label_col, min_row=data_start_row,
                     max_row=data_end_row)
    vals = Reference(ws, min_col=data_col, min_row=data_start_row - 1,
                     max_row=data_end_row)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True

    colors = [MID_BLUE, GREEN, RED, ORANGE, "7D3C98", "148F77", "2471A3", "27AE60"]
    for i, color in enumerate(colors[:data_end_row - data_start_row + 1]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, anchor)


# ═══════════════════════════════════════════════════════════════════
# Main Builder
# ═══════════════════════════════════════════════════════════════════

def build_excel_report(output: dict[str, Any],
                       diagnostics: dict[str, Any],
                       filepath: str | Path) -> str:
    """Generate a professional multi-sheet Excel report."""

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Data ──
    S    = output.get("project_summary", {})
    sol  = output.get("recommended_solution", {})
    sim  = output.get("simulation_results", {})
    disp = output.get("dispatch_results", {})
    res  = output.get("resource_results", {})
    fin  = output.get("financial_results", {})
    carb = output.get("carbon_results", {})
    mkt  = output.get("market_and_settlement", {})
    sens = output.get("sensitivity_results", [])
    des  = output.get("design_and_interconnection", {})
    alt  = output.get("alternative_solutions", [])
    qual = output.get("data_quality_results", {})
    gaps = output.get("data_gaps", [])
    risks = output.get("risks", [])

    proj_name = S.get("project_name") or "未命名"
    province  = S.get("province") or "—"
    detail    = S.get("scenario_detail_label") or S.get("scenario_type") or "—"

    has_pv   = bool(sol.get("pv_mwp") or res.get("pv_p50_generation_mwh"))
    has_wind = bool(sol.get("wind_mw"))
    has_charging = bool(sol.get("charging_capacity_mw") or sim.get("annual_charging_energy_mwh"))
    has_thermal  = bool(sol.get("cooling_capacity_rt") or sol.get("heating_capacity_mwth"))
    has_market   = mkt.get("market_mode") not in (None, "", "none")

    # ══════════════════════════════════════════════════════════════
    # Sheet 1: 项目摘要 (Summary Dashboard)
    # ══════════════════════════════════════════════════════════════
    ws0 = wb.create_sheet("项目摘要", 0)
    _sheet_setup(ws0, "摘要", freeze_row=17)
    ws0.sheet_properties.tabColor = NAVY

    # Title block
    ws0.merge_cells("A1:G1")
    _add_title(ws0, 1, 1, f"工商业储能可行性研究报告 - {proj_name}")
    ws0.merge_cells("A2:G2")
    ws0.cell(row=2, column=1, value=f"方案类型：{detail}    |    项目地点：{province}    |    报告日期：{datetime.now().strftime('%Y-%m-%d')}").font = _font_small

    # KPI Row
    irr = fin.get("irr"); npv = fin.get("npv"); payback = fin.get("payback_years")
    abate = carb.get("annual_reduction_tco2e")
    stg_mwh = sol.get("storage_energy_mwh") or 0

    irr_color = GREEN if (irr and irr >= 0.08) else (RED if irr is not None and irr < 0.06 else ORANGE)
    npv_color = GREEN if (npv and npv > 0) else RED

    kpis = [
        ("IRR", f"{irr*100:.2f}%" if irr is not None else "N/A", irr_color),
        ("NPV", f"{npv/1e4:,.1f}万" if npv is not None else "N/A", npv_color),
        ("回收期", f"{payback:.1f}年" if payback is not None else "N/A", MID_BLUE),
        ("年减排", f"{abate:,.0f}t" if abate else "—", GREEN),
        ("储能规模", f"{stg_mwh:,.0f} MWh", MID_BLUE),
    ]
    for i, (label, value, color) in enumerate(kpis):
        col = i * 2 + 1
        ws0.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws0.cell(row=4, column=col, value=label).font = Font(name="Microsoft YaHei", size=9, color=WHITE, bold=True)
        ws0.cell(row=4, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws0.cell(row=4, column=col).alignment = _align_center
        ws0.cell(row=4, column=col+1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws0.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        ws0.cell(row=5, column=col, value=value).font = _font_kpi_val
        ws0.cell(row=5, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws0.cell(row=5, column=col).alignment = _align_center
        ws0.cell(row=5, column=col+1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws0.column_dimensions[get_column_letter(col)].width = 10
        ws0.column_dimensions[get_column_letter(col+1)].width = 14

    # Project info
    _add_heading(ws0, 7, 1, "项目概况")
    pv_mw = sol.get("pv_mwp") or 0; wind_mw = sol.get("wind_mw") or 0
    stg_mw = sol.get("storage_power_mw") or 0
    info_data = [
        ("项目名称", proj_name),
        ("场景类型", detail),
        ("省份/区域", province),
        ("运行模式", S.get("operation_mode") or "—"),
        ("分析模式", S.get("analysis_mode") or "—"),
        ("数据完整度", diagnostics.get("data_completeness_grade") or "—"),
        ("数据质量", f"{qual.get('score', '—')} ({qual.get('level', '—')})"),
        ("光伏容量", f"{pv_mw} MWp"),
        ("风电容量", f"{wind_mw} MW"),
        ("储能容量", f"{stg_mw} MW / {stg_mwh} MWh"),
        ("充电容量", f"{sol.get('charging_capacity_mw') or 0} MW"),
    ]
    _add_kv_table(ws0, 8, 1, 2, info_data, key_width=18)

    # Key metrics
    _add_heading(ws0, 21, 1, "核心测算指标")
    calc_data = [
        ("年光伏发电量", f"{_fmt(sim.get('annual_pv_generation_mwh'))} MWh"),
        ("年风电发电量", f"{_fmt(sim.get('annual_wind_generation_mwh'))} MWh"),
        ("年储能放电量", f"{_fmt(sim.get('annual_storage_discharge_mwh'))} MWh"),
        ("年储能充电量", f"{_fmt(sim.get('annual_storage_charge_mwh'))} MWh"),
        ("年电网购电量", f"{_fmt(sim.get('annual_grid_purchase_mwh'))} MWh"),
        ("年外送电量", f"{_fmt(sim.get('annual_export_mwh'))} MWh"),
        ("年弃电量", f"{_fmt(sim.get('annual_curtailment_mwh'))} MWh"),
        ("年等效满循环", f"{_fmt(disp.get('storage_equivalent_full_cycles_per_year'))} 次"),
        ("估算储能寿命", f"{_fmt(disp.get('storage_life_years_estimate'))} 年"),
        ("日循环次数", _fmt(disp.get("daily_storage_cycles"))),
        ("削峰量", f"{_fmt(disp.get('estimated_peak_reduction_kw'))} kW"),
    ]
    _add_kv_table(ws0, 22, 1, 2, calc_data, key_width=20)

    # Risk callouts
    if risks:
        row = 35
        _add_heading(ws0, row, 1, "风险提示")
        for r in risks:
            row += 1
            c = ws0.cell(row=row, column=1, value=f"⚠ {r}")
            c.font = Font(name="Microsoft YaHei", size=10, color=RED)
            c.fill = _fill_red
            ws0.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)


    # ══════════════════════════════════════════════════════════════
    # Sheet 2: 资源评估
    # ══════════════════════════════════════════════════════════════
    if has_pv or has_wind:
        ws1 = wb.create_sheet("资源评估")
        _sheet_setup(ws1, "资源")
        ws1.sheet_properties.tabColor = ORANGE

        r = 1
        _add_heading(ws1, r, 1, "光伏资源评估"); r += 1
        pv_headers = ["参数", "数值"]
        for i, h in enumerate(pv_headers, 1):
            ws1.cell(row=r, column=i, value=h)
        _style_header_row(ws1, r, 2)
        r += 1

        pv_data = [
            ("资源口径", f"{res.get('pv_resource_basis') or '—'} / {res.get('pv_resource_accuracy') or '—'}"),
            ("P50 年发电量 (MWh)", _fmt(res.get("pv_p50_generation_mwh"))),
            ("P90 年发电量 (MWh)", _fmt(res.get("pv_p90_generation_mwh"))),
            ("综合 PR", _fmt(res.get("pv_pr_effective"))),
            ("有效倾角 (°)", _fmt(res.get("pv_effective_tilt_deg"), 0)),
            ("推荐倾角 (°)", _fmt(res.get("pv_recommended_tilt_deg"), 0)),
            ("倾角修正系数", _fmt(res.get("pv_tilt_factor"))),
            ("方位角修正系数", _fmt(res.get("pv_azimuth_factor"))),
            ("跟踪支架系数", _fmt(res.get("pv_tracking_factor"))),
            ("温度修正系数", _fmt(res.get("pv_temperature_factor"))),
            ("双面增益系数", _fmt(res.get("pv_bifacial_gain"))),
            ("遮挡修正系数", _fmt(res.get("pv_shading_factor"))),
        ]
        for k, v in pv_data:
            ws1.cell(row=r, column=1, value=k)
            ws1.cell(row=r, column=2, value=v)
            r += 1
        _style_data_rows(ws1, 3, r - 1, 2)
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 22

        # PV 24h data for chart
        pv24 = (sim.get("pv_hourly_profile_kw") or res.get("pv_hourly_profile_kw") or [])[:24]
        if pv24 and max(pv24) > 0:
            r += 2
            ws1.cell(row=r, column=1, value="小时").font = _font_bold
            ws1.cell(row=r, column=2, value="光伏功率 (kW)").font = _font_bold
            _style_header_row(ws1, r, 2, fill=_fill_kpi_orange)
            chart_start = r + 1
            for h, v in enumerate(pv24):
                ws1.cell(row=chart_start + h, column=1, value=h)
                ws1.cell(row=chart_start + h, column=2, value=round(v, 1))
            _style_data_rows(ws1, chart_start, chart_start + 23, 2)
            _add_line_chart(ws1, chart_start, chart_start + 23, 2, 1,
                           f"E{chart_start}", "光伏典型日发电曲线", color="D6840B")

        # Wind
        if has_wind:
            r += 27
            _add_heading(ws1, r, 1, "风能资源评估"); r += 1
            for i, h in enumerate(["参数", "数值"], 1):
                ws1.cell(row=r, column=i, value=h)
            _style_header_row(ws1, r, 2)
            r += 1
            wind_data = [
                ("资源口径", f"{res.get('wind_resource_basis') or '—'} / {res.get('wind_resource_accuracy') or '—'}"),
                ("P50 年发电量 (MWh)", _fmt(res.get("wind_p50_generation_mwh"))),
                ("P90 年发电量 (MWh)", _fmt(res.get("wind_p90_generation_mwh"))),
                ("功率曲线", _fmt(res.get("wind_power_curve_used"))),
                ("净折减系数", _fmt(res.get("wind_net_factor"))),
            ]
            for k, v in wind_data:
                ws1.cell(row=r, column=1, value=k)
                ws1.cell(row=r, column=2, value=v)
                r += 1
            _style_data_rows(ws1, r - len(wind_data), r - 1, 2)


    # ══════════════════════════════════════════════════════════════
    # Sheet 3: 储能系统
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("储能系统")
    _sheet_setup(ws2, "储能")
    ws2.sheet_properties.tabColor = MID_BLUE

    r = 1
    _add_heading(ws2, r, 1, "储能容量配置"); r += 1
    for i, h in enumerate(["参数", "数值"], 1):
        ws2.cell(row=r, column=i, value=h)
    _style_header_row(ws2, r, 2)
    r += 1

    stg_mw_sol = sol.get("storage_power_mw") or 0
    stg_mwh_sol = sol.get("storage_energy_mwh") or 0
    raw_mw = sol.get("raw_storage_power_mw") or 0
    raw_mwh = sol.get("raw_storage_energy_mwh") or 0

    stg_data = [
        ("推荐容量", f"{stg_mw_sol} MW / {stg_mwh_sol} MWh"),
        ("计算容量", f"{raw_mw} MW / {raw_mwh} MWh" if raw_mwh else "—"),
        ("额定充放电时长", f"{stg_mwh_sol / max(stg_mw_sol, 0.001):.1f} 小时"),
        ("策略模式", disp.get("storage_strategy_mode") or "—"),
        ("测算口径", disp.get("storage_sizing_basis") or "—"),
    ]
    for k, v in stg_data:
        ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
        r += 1
    _style_data_rows(ws2, 3, r - 1, 2)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 25

    r += 1
    _add_heading(ws2, r, 1, "运行指标"); r += 1
    for i, h in enumerate(["指标", "数值"], 1):
        ws2.cell(row=r, column=i, value=h)
    _style_header_row(ws2, r, 2)
    hr_start = r + 1
    r += 1

    op_data = [
        ("日循环次数", _fmt(disp.get("daily_storage_cycles"))),
        ("年等效满循环", _fmt(disp.get("storage_equivalent_full_cycles_per_year"))),
        ("年储能吞吐量 (MWh)", _fmt(disp.get("storage_annual_throughput_mwh"))),
        ("估算寿命 (年)", _fmt(disp.get("storage_life_years_estimate"))),
        ("有效往返效率", _pct(disp.get("storage_effective_round_trip_efficiency"))),
        ("年衰减率", _pct(disp.get("storage_degradation_per_year"))),
        ("寿命末容量比", _pct(disp.get("storage_end_of_life_capacity_ratio"))),
        ("绿电充电比", _pct(disp.get("storage_charge_from_renewables_ratio"))),
        ("预留 SOC", _pct(disp.get("storage_reserved_soc_ratio"))),
        ("保供 SOC", _pct(disp.get("storage_backup_soc_ratio"))),
    ]
    for k, v in op_data:
        ws2.cell(row=r, column=1, value=k)
        ws2.cell(row=r, column=2, value=v)
        r += 1
    _style_data_rows(ws2, hr_start, r - 1, 2)

    # Dispatch data
    dp = disp.get("dispatch_series_kw") or []
    if dp and len(dp) >= 24:
        r += 2
        ws2.cell(row=r, column=1, value="小时").font = _font_bold
        ws2.cell(row=r, column=2, value="充电 (kW)").font = _font_bold
        ws2.cell(row=r, column=3, value="放电 (kW)").font = _font_bold
        ws2.cell(row=r, column=4, value="SOC (%)").font = _font_bold
        _style_header_row(ws2, r, 4)
        dp_start = r + 1
        soc_raw = disp.get("soc_series") or []
        soc_m = max(max(soc_raw[:24], key=lambda x: max(x, 0)), 0.01) if soc_raw else 1
        for h in range(24):
            v = dp[h] if h < len(dp) else 0
            r2 = dp_start + h
            ws2.cell(row=r2, column=1, value=h)
            ws2.cell(row=r2, column=2, value=round(max(v, 0), 1))
            ws2.cell(row=r2, column=3, value=round(abs(min(v, 0)), 1))
            ws2.cell(row=r2, column=4, value=round((soc_raw[h]/soc_m*100) if soc_raw and h < len(soc_raw) else 50, 1))
        _style_data_rows(ws2, dp_start, dp_start + 23, 4)
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 15

        # Charge/Discharge chart
        _add_bar_chart(ws2, dp_start, dp_start + 23, 2, 1,
                      f"F{dp_start - 1}", "储能典型日充放电曲线", color=MID_BLUE)

    # Monthly table
    monthly = disp.get("monthly_storage_revenue_breakdown") or []
    if monthly:
        r = dp_start + 26 if dp else r + 2
        _add_heading(ws2, r, 1, "月度运行数据")
        r += 1
        mh = ["月份", "充电 (MWh)", "放电 (MWh)", "毛收益"]
        for i, h in enumerate(mh, 1):
            ws2.cell(row=r, column=i, value=h)
        _style_header_row(ws2, r, 4)
        m_start = r + 1
        for j, m in enumerate(monthly):
            r2 = m_start + j
            ws2.cell(row=r2, column=1, value=f"{m.get('month','')}月")
            ws2.cell(row=r2, column=2, value=round(m.get("charge_mwh") or 0, 1))
            ws2.cell(row=r2, column=3, value=round(m.get("discharge_mwh") or 0, 1))
            ws2.cell(row=r2, column=4, value=round(m.get("gross_margin") or 0, 0))
        _style_data_rows(ws2, m_start, m_start + len(monthly) - 1, 4)


    # ══════════════════════════════════════════════════════════════
    # Sheet 4: 财务分析
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("财务分析")
    _sheet_setup(ws3, "财务")
    ws3.sheet_properties.tabColor = GREEN

    r = 1
    _add_heading(ws3, r, 1, "投资估算"); r += 1
    fin_headers = ["项目", "金额"]
    for i, h in enumerate(fin_headers, 1):
        ws3.cell(row=r, column=i, value=h)
    _style_header_row(ws3, r, 2)
    r += 1

    capex_data = [
        ("系统总投资 (CAPEX)", f"{_fmt(fin.get('capex_total'))} 元"),
        ("年运维费用 (OPEX)", f"{_fmt(fin.get('opex_annual'))} 元/年"),
        ("运维递增率", _pct(fin.get("opex_escalation_rate"))),
        ("储能更换年份", _fmt(fin.get("storage_replacement_year"), 0) if fin.get("storage_replacement_year") else "—"),
        ("储能更换成本", f"{_fmt(fin.get('storage_replacement_cost'))} 元" if fin.get("storage_replacement_cost") else "—"),
    ]
    for k, v in capex_data:
        ws3.cell(row=r, column=1, value=k)
        ws3.cell(row=r, column=2, value=v)
        r += 1
    _style_data_rows(ws3, 3, r - 1, 2)
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 22

    r += 1
    _add_heading(ws3, r, 1, "收益结构"); r += 1
    for i, h in enumerate(["项目", "金额"], 1):
        ws3.cell(row=r, column=i, value=h)
    _style_header_row(ws3, r, 2)
    rev_start = r + 1
    r += 1
    rev_data = [
        ("年收益/节费", f"{_fmt(fin.get('annual_savings_or_revenue'))} 元"),
        ("年电量电费", f"{_fmt(fin.get('annual_energy_charge_cost'))} 元"),
        ("年需量电费", f"{_fmt(fin.get('annual_demand_charge_cost'))} 元"),
        ("年辅助服务收益", f"{_fmt(fin.get('annual_ancillary_service_revenue'))} 元"),
        ("年需求响应收益", f"{_fmt(fin.get('annual_demand_response_revenue'))} 元"),
        ("年外送收益", f"{_fmt(fin.get('annual_export_revenue'))} 元"),
    ]
    for k, v in rev_data:
        ws3.cell(row=r, column=1, value=k)
        ws3.cell(row=r, column=2, value=v)
        r += 1
    _style_data_rows(ws3, rev_start, r - 1, 2)

    # Revenue pie chart data
    rev_items = [
        ("辅助服务", fin.get("annual_ancillary_service_revenue")),
        ("需求响应", fin.get("annual_demand_response_revenue")),
        ("外送收益", fin.get("annual_export_revenue")),
    ]
    pie_data = [(l, v) for l, v in rev_items if v and v > 0]
    savings = fin.get("annual_savings_or_revenue")
    if savings and savings > 0:
        pie_data.insert(0, ("综合节费", savings))
    if len(pie_data) >= 2:
        r += 1
        ws3.cell(row=r, column=1, value="收益类别").font = _font_bold
        ws3.cell(row=r, column=2, value="金额 (元)").font = _font_bold
        _style_header_row(ws3, r, 2, fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"))
        pie_start = r + 1
        for label, val in pie_data:
            r += 1
            ws3.cell(row=r, column=1, value=label)
            ws3.cell(row=r, column=2, value=round(val, 0))
        _style_data_rows(ws3, pie_start, r, 2)
        _add_pie_chart(ws3, pie_start, r, 2, 1, f"D{pie_start - 1}", "年收益构成")

    r += 1
    _add_heading(ws3, r, 1, "核心财务指标"); r += 1
    for i, h in enumerate(["指标", "数值", "评价"], 1):
        ws3.cell(row=r, column=i, value=h)
    _style_header_row(ws3, r, 3)
    kpi_start = r + 1
    r += 1

    kpi_check = lambda v: "✓ 达标" if v else ("—" if v is None else "✗ 未达标")
    kpi_data = [
        ("IRR", f"{_pct(irr)}" if irr is not None else "N/A",
         kpi_check(irr and irr >= 0.08) if irr is not None else "—"),
        ("NPV", f"{_fmt(npv)} 元" if npv is not None else "N/A",
         kpi_check(npv and npv > 0) if npv is not None else "—"),
        ("静态回收期", f"{_fmt(payback)} 年" if payback is not None else "N/A", "≤8年为良好"),
        ("LCOE", f"{_fmt(fin.get('lcoe'))} 元/kWh" if fin.get("lcoe") else "—", "—"),
        ("LCOS", f"{_fmt(fin.get('lcos'))} 元/kWh" if fin.get("lcos") else "—", "—"),
        ("单位减碳成本", f"{_fmt(fin.get('abatement_cost_per_tco2e'))} 元/tCO₂" if fin.get("abatement_cost_per_tco2e") else "—", "—"),
    ]
    for k, v, e in kpi_data:
        ws3.cell(row=r, column=1, value=k)
        ws3.cell(row=r, column=2, value=v)
        ws3.cell(row=r, column=3, value=e)
        r += 1
    _style_data_rows(ws3, kpi_start, r - 1, 3)
    ws3.column_dimensions["C"].width = 18

    # Cashflow data
    cf = fin.get("cashflow_series") or []
    if cf and len(cf) > 1:
        r += 2
        ws3.cell(row=r, column=1, value="年份").font = _font_bold
        ws3.cell(row=r, column=2, value="净现金流 (元)").font = _font_bold
        ws3.cell(row=r, column=3, value="累计净现金流 (元)").font = _font_bold
        _style_header_row(ws3, r, 3)
        cf_start = r + 1
        cumsum = 0
        for y, val in enumerate(cf):
            cumsum += val
            r2 = cf_start + y
            ws3.cell(row=r2, column=1, value=f"第{y+1}年")
            ws3.cell(row=r2, column=2, value=round(val, 0))
            ws3.cell(row=r2, column=3, value=round(cumsum, 0))
        _style_data_rows(ws3, cf_start, cf_start + len(cf) - 1, 3)
        _add_bar_chart(ws3, cf_start, cf_start + len(cf) - 1, 2, 1,
                      f"E{cf_start - 1}", "全生命周期现金流", color=GREEN)

    # Alternative solutions
    if alt:
        r = cf_start + len(cf) + 2 if cf else r + 2
        _add_heading(ws3, r, 1, "方案对比"); r += 1
        for i, h in enumerate(["方案", "容量", "IRR", "回收期", "NPV"], 1):
            ws3.cell(row=r, column=i, value=h)
        _style_header_row(ws3, r, 5)
        alt_start = r + 1
        for j, a in enumerate(alt):
            r2 = alt_start + j
            ws3.cell(row=r2, column=1, value=a.get("label", "—"))
            ws3.cell(row=r2, column=2, value=f"{_fmt(a.get('storage_power_mw'))}MW/{_fmt(a.get('storage_energy_mwh'))}MWh")
            ws3.cell(row=r2, column=3, value=_pct(a.get("irr")))
            ws3.cell(row=r2, column=4, value=_fmt(a.get("payback_years")))
            ws3.cell(row=r2, column=5, value=_fmt(a.get("npv")))
        _style_data_rows(ws3, alt_start, alt_start + len(alt) - 1, 5)


    # ══════════════════════════════════════════════════════════════
    # Sheet 5: 碳减排
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("碳减排分析")
    _sheet_setup(ws4, "碳")
    ws4.sheet_properties.tabColor = GREEN

    r = 1
    if abate and float(abate) > 0:
        ws4.merge_cells("A1:C1")
        c = ws4.cell(row=1, column=1,
                     value=f"✓ 项目年减排 CO₂ 约 {_fmt(abate)} 吨，等效植树 {_fmt(float(abate) * 50)} 棵/年")
        c.font = Font(name="Microsoft YaHei", size=11, color=GREEN, bold=True)
        c.fill = _fill_green
        r = 3

    _add_heading(ws4, r, 1, "碳排放指标"); r += 1
    for i, h in enumerate(["指标", "数值"], 1):
        ws4.cell(row=r, column=i, value=h)
    _style_header_row(ws4, r, 2)
    cb_start = r + 1
    r += 1

    cb_data = [
        ("基线排放 (tCO₂e)", _fmt(carb.get("baseline_emissions_tco2e"))),
        ("项目后排放 (tCO₂e)", _fmt(carb.get("post_project_emissions_tco2e"))),
        ("年减排量 (tCO₂e)", _fmt(abate)),
        ("范围1减排 (tCO₂e)", _fmt(carb.get("scope1_reduction_tco2e"))),
        ("范围2减排 (tCO₂e)", _fmt(carb.get("scope2_reduction_tco2e"))),
        ("范围3减排 (tCO₂e)", _fmt(carb.get("scope3_reduction_tco2e"))),
        ("绿电覆盖率", _pct(carb.get("green_power_coverage_ratio"))),
        ("声明边界", carb.get("claim_boundary_summary") or "—"),
        ("减碳成本 (元/tCO₂)", _fmt(carb.get("abatement_cost_per_tco2e"))),
    ]
    for k, v in cb_data:
        ws4.cell(row=r, column=1, value=k)
        ws4.cell(row=r, column=2, value=v)
        r += 1
    _style_data_rows(ws4, cb_start, r - 1, 2)
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 22

    # Carbon path breakdown
    pbd = carb.get("carbon_path_breakdown") or []
    if pbd:
        r += 2
        _add_heading(ws4, r, 1, "减排路径拆分"); r += 1
        for i, h in enumerate(["路径", "减排量 (tCO₂e)", "占比"], 1):
            ws4.cell(row=r, column=i, value=h)
        _style_header_row(ws4, r, 3)
        pb_start = r + 1
        r += 1
        for p in pbd:
            ws4.cell(row=r, column=1, value=p.get("path", "—"))
            ws4.cell(row=r, column=2, value=_fmt(p.get("reduction_tco2e")))
            ws4.cell(row=r, column=3, value=_pct(p.get("share")))
            r += 1
        _style_data_rows(ws4, pb_start, r - 1, 3)
        ws4.column_dimensions["C"].width = 16


    # ══════════════════════════════════════════════════════════════
    # Sheet 6: 敏感性 (if data)
    # ══════════════════════════════════════════════════════════════
    if sens:
        ws5 = wb.create_sheet("敏感性分析")
        _sheet_setup(ws5, "敏感性")
        ws5.sheet_properties.tabColor = ORANGE

        _add_heading(ws5, 1, 1, "IRR 敏感性分析")
        for i, h in enumerate(["因素", "年收益影响", "IRR 敏感度"], 1):
            ws5.cell(row=2, column=i, value=h)
        _style_header_row(ws5, 2, 3)
        for j, s in enumerate(sens):
            r = 3 + j
            ws5.cell(row=r, column=1, value=s.get("factor", "—"))
            ws5.cell(row=r, column=2, value=_fmt(s.get("impact_on_annual_revenue")))
            ws5.cell(row=r, column=3, value=_fmt(s.get("impact_on_irr")))
        _style_data_rows(ws5, 3, 2 + len(sens), 3)
        ws5.column_dimensions["A"].width = 22
        ws5.column_dimensions["B"].width = 18
        ws5.column_dimensions["C"].width = 16

        # Sensitivity chart
        impacts = []
        for s in sens:
            try: impacts.append(float(s.get("impact_on_irr") or 0))
            except: impacts.append(0)
        if any(abs(v) > 0.001 for v in impacts):
            # Write data in a chart-friendly way: labels + values
            ws5.cell(row=2 + len(sens) + 2, column=1, value="因素").font = _font_bold
            ws5.cell(row=2 + len(sens) + 2, column=2, value="IRR变化(%)").font = _font_bold
            _style_header_row(ws5, 2 + len(sens) + 2, 2, fill=_fill_kpi_orange)
            s2_start = 2 + len(sens) + 3
            for j, (f, v) in enumerate(zip([s.get("factor","") for s in sens], impacts)):
                ws5.cell(row=s2_start + j, column=1, value=f)
                ws5.cell(row=s2_start + j, column=2, value=v)
            _style_data_rows(ws5, s2_start, s2_start + len(sens) - 1, 2)
            _add_bar_chart(ws5, s2_start, s2_start + len(sens) - 1, 2, 1,
                          f"D{s2_start}", "IRR 敏感性分析")


    # ══════════════════════════════════════════════════════════════
    # Sheet 7: 充电/空调/市场 (conditional)
    # ══════════════════════════════════════════════════════════════
    if has_charging:
        ws_ch = wb.create_sheet("充电设施")
        _sheet_setup(ws_ch, "充电")
        ws_ch.sheet_properties.tabColor = ACCENT
        r = 1
        _add_heading(ws_ch, r, 1, "充电设施分析"); r += 1
        ch_data = [
            ("年充电量 (MWh)", _fmt(sim.get("annual_charging_energy_mwh"))),
            ("充电峰值 (kW)", _fmt(disp.get("charging_peak_kw"))),
            ("利用率", _pct(disp.get("charging_utilization_ratio"))),
            ("排队风险", disp.get("charging_queue_risk") or "—"),
            ("排队指数", _fmt(disp.get("charging_queue_index"))),
            ("多样性系数", _fmt(disp.get("charging_diversity_factor"))),
        ]
        _add_kv_table(ws_ch, r, 1, 2, ch_data, key_width=20)
        seg = disp.get("charging_segment_summary") or []
        if seg:
            r += len(ch_data) + 2
            ws_ch.cell(row=r, column=1, value="车型").font = _font_bold
            ws_ch.cell(row=r, column=2, value="日充电量").font = _font_bold
            ws_ch.cell(row=r, column=3, value="峰值 (kW)").font = _font_bold
            _style_header_row(ws_ch, r, 3)
            for j, s in enumerate(seg):
                r2 = r + 1 + j
                ws_ch.cell(row=r2, column=1, value=s.get("vehicle_type", "—"))
                ws_ch.cell(row=r2, column=2, value=f"{_fmt(s.get('daily_energy_kwh'))} kWh")
                ws_ch.cell(row=r2, column=3, value=_fmt(s.get("peak_kw")))
            _style_data_rows(ws_ch, r + 1, r + len(seg), 3)

    if has_thermal:
        ws_th = wb.create_sheet("暖通空调")
        _sheet_setup(ws_th, "暖通")
        ws_th.sheet_properties.tabColor = ORANGE
        th_data = [
            ("年供冷量 (MWh)", _fmt(sim.get("annual_cooling_energy_mwh"))),
            ("年供热量 (MWh)", _fmt(sim.get("annual_heating_energy_mwh"))),
            ("冷峰值 (kWth)", _fmt(disp.get("thermal_cooling_peak_kwth"))),
            ("热峰值 (kWth)", _fmt(disp.get("thermal_heating_peak_kwth"))),
            ("制冷电耗峰值 (kW)", _fmt(disp.get("thermal_electric_peak_kw"))),
            ("年锅炉燃料当量 (MWh)", _fmt(disp.get("thermal_annual_boiler_fuel_equivalent_mwh"))),
        ]
        _add_kv_table(ws_th, 2, 1, 2, th_data, key_width=22)

    if has_market:
        ws_mk = wb.create_sheet("电力市场")
        _sheet_setup(ws_mk, "市场")
        ws_mk.sheet_properties.tabColor = RED
        mk_data = [
            ("市场模式", mkt.get("market_mode") or "—"),
            ("政策模式", mkt.get("market_policy_mode") or "—"),
            ("充电基准价 (元/kWh)", _fmt(mkt.get("trading_charge_benchmark_price_per_kwh"))),
            ("放电基准价 (元/kWh)", _fmt(mkt.get("trading_discharge_benchmark_price_per_kwh"))),
            ("交易价差 (元/kWh)", _fmt(mkt.get("trading_price_spread_per_kwh"))),
            ("波动指数", _fmt(mkt.get("trading_volatility_index"))),
            ("最优月份", mkt.get("trading_best_month") or "—"),
            ("最弱月份", mkt.get("trading_worst_month") or "—"),
        ]
        _add_kv_table(ws_mk, 2, 1, 2, mk_data, key_width=22)

    # ── Save ──
    out_path = Path(filepath)
    wb.save(str(out_path))
    return str(out_path.resolve())
