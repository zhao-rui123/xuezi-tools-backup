from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PRICE_LABELS = {
    "尖": "super_peak",
    "尖峰": "super_peak",
    "峰": "peak",
    "平": "flat",
    "谷": "valley",
    "深谷": "deep_valley",
    # GBK-Mojibake fallbacks for legacy Excel files from older Chinese Windows systems
    "灏?": "super_peak",
    "宄?": "peak",
    "骞?": "flat",
    "璋?": "valley",
    "娣辫胺": "deep_valley",
}

STRATEGY_MAP = {
    "谷充峰放": ("valley", "peak"),
    "平充峰放": ("flat", "peak"),
    "谷充尖放": ("valley", "super_peak"),
    "平充尖放": ("flat", "super_peak"),
    "深谷充尖放": ("deep_valley", "super_peak"),
    "深谷充峰放": ("deep_valley", "peak"),
    "谷充平放": ("valley", "flat"),
    "深谷充平放": ("deep_valley", "flat"),
    # GBK-Mojibake fallbacks for legacy Excel files
    "璋峰厖宄版斁": ("valley", "peak"),
    "骞冲厖宄版斁": ("flat", "peak"),
    "璋峰厖灏栨斁": ("valley", "super_peak"),
    "骞冲厖灏栨斁": ("flat", "super_peak"),
    "娣辫胺鍏呭皷鏀?": ("deep_valley", "super_peak"),
    "娣辫胺鍏呭嘲鏀?": ("deep_valley", "peak"),
    "璋峰厖骞虫斁": ("valley", "flat"),
    "娣辫胺鍏呭钩鏀?": ("deep_valley", "flat"),
}


def enrich_from_province_cycle_rules(data: dict[str, Any], base_dir: Path | None = None) -> dict[str, Any]:
    market = data.setdefault("market_data", {})
    project = data.setdefault("project_info", {})
    province = str(project.get("province") or "").strip()
    workbook_path = market.get("province_cycle_rules_workbook_path")
    if not province or not workbook_path:
        return data

    base = base_dir or Path.cwd()
    parsed = parse_province_cycle_rules_workbook(base / str(workbook_path), province)
    if not parsed:
        return data

    if parsed.get("tou_tariff") and not market.get("tou_tariff"):
        market["tou_tariff"] = parsed["tou_tariff"]
    if parsed.get("arbitrage_plan"):
        merged = dict(parsed["arbitrage_plan"])
        merged.update(market.get("arbitrage_plan") or {})
        market["arbitrage_plan"] = merged
    market.setdefault("province_cycle_rule_source", str(workbook_path))
    return data


def parse_province_cycle_rules_workbook(path: Path, province: str) -> dict[str, Any] | None:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        # Intentional: first sheet contains all province data in a merged layout,
        # so we always use sheetnames[0] regardless of province value.
        sheet = workbook[workbook.sheetnames[0]]
        rows = [tuple(cell for cell in row[:6]) for row in sheet.iter_rows(values_only=True)]
        province = province.strip()

        for index, row in enumerate(rows):
            if _clean_text(row[0]) != province:
                continue
            block = rows[index : index + 6]
            tou_tariff = []
            cycles = []
            for item in block:
                period = PRICE_LABELS.get(_clean_text(item[1]))
                price = _to_float(item[2])
                strategy = _clean_text(item[3]).replace("次数", "").replace("娆℃暟", "")
                count = _to_float(item[4])
                if period and price is not None:
                    tou_tariff.append({"period": period, "price": price})
                cycle = _strategy_to_cycle(strategy, count)
                if cycle:
                    cycles.append(cycle)
            if not tou_tariff:
                return None
            return {
                "tou_tariff": tou_tariff,
                "arbitrage_plan": {
                    "mode": "rule_based",
                    "auto_days_from_policy": False,
                    "cycles": cycles,
                },
            }
        return None
    finally:
        workbook.close()


def _strategy_to_cycle(strategy: str, count: float | None) -> dict[str, Any] | None:
    if not strategy or count is None or count <= 0:
        return None
    if strategy not in STRATEGY_MAP:
        return None
    charge_period, discharge_period = STRATEGY_MAP[strategy]
    return {
        "charge_period": charge_period,
        "discharge_period": discharge_period,
        "days_per_year": int(float(count)),
    }


def _clean_text(value: Any) -> str:
    return str(value).replace("\xa0", "").strip() if value not in (None, "") else ""


def _to_float(value: Any) -> float | None:
    try:
        if value in (None, ""):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None
