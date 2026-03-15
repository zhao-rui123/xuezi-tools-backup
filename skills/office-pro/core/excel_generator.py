#!/usr/bin/env python3
"""
Excel 模板生成器
支持多种专业样式的Excel模板
"""

import os
import json
from typing import Dict, List, Optional, Union, Any
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, RadarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.reference import Reference
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np


class ExcelStyle:
    STYLES = {
        "professional": {
            "header_fill": "2C3E50",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "ECF0F1",
            "border_color": "BDC3C7",
            "accent_color": "3498DB"
        },
        "blue": {
            "header_fill": "4472C4",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "D6EAF8",
            "border_color": "85C1E9",
            "accent_color": "2E86C1"
        },
        "green": {
            "header_fill": "27AE60",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "D5F5E3",
            "border_color": "82E0AA",
            "accent_color": "1E8449"
        },
        "red": {
            "header_fill": "C0392B",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "FADBD8",
            "border_color": "F1948A",
            "accent_color": "922B21"
        },
        "purple": {
            "header_fill": "8E44AD",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "EDDAF3",
            "border_color": "D2B4DE",
            "accent_color": "6C3483"
        },
        "dark": {
            "header_fill": "1A1A1A",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "F2F2F2",
            "border_color": "666666",
            "accent_color": "333333"
        },
        "orange": {
            "header_fill": "E67E22",
            "header_font": "FFFFFF",
            "header_font_bold": True,
            "alternating_row": "FCF3CF",
            "border_color": "F5B041",
            "accent_color": "CA6F1E"
        }
    }

    @classmethod
    def get_style(cls, style_name: str = "professional") -> Dict:
        return cls.STYLES.get(style_name, cls.STYLES["professional"])


class ExcelTemplateGenerator:
    def __init__(self):
        self.workbook = None
        self.current_sheet = None

    def create_workbook(self, style: str = "professional"):
        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.style = ExcelStyle.get_style(style)
        return self

    def add_sheet(self, sheet_name: str):
        ws = self.workbook.create_sheet(title=sheet_name)
        self.current_sheet = ws
        return self

    def write_data(self, data: Union[pd.DataFrame, List[List], Dict],
                   headers: List[str] = None, start_row: int = 1, start_col: int = 1):
        if isinstance(data, pd.DataFrame):
            for r_idx, row in enumerate(data_to_rows(data), start_row):
                for c_idx, value in enumerate(row, start_col):
                    cell = self.current_sheet.cell(row=r_idx, column=c_idx, value=value)
        elif isinstance(data, dict):
            if headers is None:
                headers = list(data.keys())
            self.write_data([headers], start_row=start_row)
            values = list(data.values()) if isinstance(list(data.values())[0], (int, float, str)) else list(data.values())[0]
            for i, row in enumerate(values, start_row + 1):
                for j, val in enumerate(row, start_col):
                    self.current_sheet.cell(row=i, column=j, value=val)
        else:
            for r_idx, row in enumerate(data, start_row):
                for c_idx, value in enumerate(row, start_col):
                    self.current_sheet.cell(row=r_idx, column=c_idx, value=value)
        return self

    def apply_header_style(self, start_row: int = 1, num_cols: int = None):
        if num_cols is None:
            num_cols = self.current_sheet.max_column

        header_fill = PatternFill(start_color=self.style["header_fill"],
                                  end_color=self.style["header_fill"],
                                  fill_type="solid")
        header_font = Font(color=self.style["header_font"],
                          bold=self.style["header_font_bold"],
                          size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin', color=self.style["border_color"]),
            right=Side(style='thin', color=self.style["border_color"]),
            top=Side(style='thin', color=self.style["border_color"]),
            bottom=Side(style='thin', color=self.style["border_color"])
        )

        for col in range(1, num_cols + 1):
            cell = self.current_sheet.cell(row=start_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        return self

    def apply_alternating_rows(self, start_row: int = 2):
        alt_fill = PatternFill(start_color=self.style["alternating_row"],
                              end_color=self.style["alternating_row"],
                              fill_type="solid")

        border = Border(
            left=Side(style='thin', color=self.style["border_color"]),
            right=Side(style='thin', color=self.style["border_color"]),
            top=Side(style='thin', color=self.style["border_color"]),
            bottom=Side(style='thin', color=self.style["border_color"])
        )

        for row in range(start_row, self.current_sheet.max_row + 1):
            if row % 2 == 0:
                for col in range(1, self.current_sheet.max_column + 1):
                    cell = self.current_sheet.cell(row=row, column=col)
                    cell.fill = alt_fill
                    cell.border = border

        return self

    def auto_adjust_columns(self):
        for column in self.current_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.current_sheet.column_dimensions[column_letter].width = adjusted_width

        return self

    def add_chart(self, chart_type: str, data_range: str, title: str,
                  position: str = "E5", categories_range: str = None):
        chart_type = chart_type.lower()

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "doughnut":
            chart = DoughnutChart()
        elif chart_type == "radar":
            chart = RadarChart()
        else:
            chart = BarChart()

        chart.title = title
        chart.style = 10
        chart.height = 10
        chart.width = 16

        if chart_type in ["pie", "doughnut"]:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True

        if data_range:
            data_ref = Reference(self.current_sheet, range_string=data_range)
            chart.add_data(data_ref, titles_from_data=True)

        if categories_range:
            cats_ref = Reference(self.current_sheet, range_string=categories_range)
            chart.set_categories(cats_ref)

        self.current_sheet.add_chart(chart, position)
        return self

    def add_title(self, title: str, row: int = 1, font_size: int = 16):
        cell = self.current_sheet.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(size=font_size, bold=True, color=self.style["accent_color"])
        self.current_sheet.merge_cells(start_row=row, start_column=1,
                                       end_row=row, end_column=self.current_sheet.max_column)
        return self

    def add_summary(self, data: Dict, start_row: int = None):
        if start_row is None:
            start_row = self.current_sheet.max_row + 2

        cell = self.current_sheet.cell(row=start_row, column=1)
        cell.value = "统计摘要"
        cell.font = Font(bold=True, size=12, color=self.style["accent_color"])

        row = start_row + 1
        for key, value in data.items():
            self.current_sheet.cell(row=row, column=1, value=key)
            self.current_sheet.cell(row=row, column=2, value=value)
            row += 1

        return self

    def freeze_panes(self, cell: str = "A2"):
        self.current_sheet.freeze_panes = cell
        return self

    def apply_filter(self):
        self.current_sheet.auto_filter.ref = self.current_sheet.dimensions
        return self

    def save(self, output_path: str):
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        self.workbook.save(output_path)
        print(f"Excel 已生成: {output_path}")
        return output_path


class ExcelTemplateLibrary:
    TEMPLATES = {
        "sales_monthly": {
            "name": "月度销售报表",
            "style": "blue",
            "description": "专业的月度销售数据报表模板",
            "sheets": ["销售概览", "明细数据", "统计分析"],
            "charts": ["销售额趋势", "产品占比", "地区对比"]
        },
        "financial_report": {
            "name": "财务报表",
            "style": "professional",
            "description": "专业财务报表模板",
            "sheets": ["资产负债表", "利润表", "现金流量表"],
            "charts": ["收入趋势", "支出分析", "利润率"]
        },
        "project_tracker": {
            "name": "项目跟踪表",
            "style": "green",
            "description": "项目管理与进度跟踪模板",
            "sheets": ["项目概览", "任务列表", "里程碑"],
            "charts": ["进度跟踪", "资源分配", "风险评估"]
        },
        "inventory": {
            "name": "库存管理表",
            "style": "orange",
            "description": "库存管理与分析模板",
            "sheets": ["库存概览", "入库记录", "出库记录"],
            "charts": ["库存趋势", "周转率", "缺货预警"]
        },
        "hr_employee": {
            "name": "员工信息表",
            "style": "purple",
            "description": "人力资源员工信息管理模板",
            "sheets": ["员工列表", "部门分布", "离职分析"],
            "charts": ["部门人数", "入职趋势", "离职率"]
        },
        "kpi_dashboard": {
            "name": "KPI仪表盘",
            "style": "dark",
            "description": "关键绩效指标仪表盘模板",
            "sheets": ["KPI概览", "目标vs实际", "趋势分析"],
            "charts": ["KPI完成率", "目标差距", "月度趋势"]
        }
    }

    @classmethod
    def list_templates(cls) -> List[Dict]:
        return [
            {"id": k, "name": v["name"], "description": v["description"], "style": v["style"]}
            for k, v in cls.TEMPLATES.items()
        ]

    @classmethod
    def create_template(cls, template_id: str, data: pd.DataFrame = None,
                       output_path: str = None) -> str:
        if template_id not in cls.TEMPLATES:
            raise ValueError(f"Template not found: {template_id}")

        template = cls.TEMPLATES[template_id]
        generator = ExcelTemplateGenerator()
        generator.create_workbook(style=template["style"])

        for sheet_name in template["sheets"]:
            generator.add_sheet(sheet_name)

        if data is not None:
            for i, sheet_name in enumerate(template["sheets"]):
                generator.current_sheet = generator.workbook[sheet_name]

                if i == 0:
                    generator.add_title(f"{template['name']} - {datetime.now().strftime('%Y-%m')}")
                    generator.write_data(data.head(20), headers=data.columns.tolist())
                    generator.apply_header_style()
                    generator.apply_alternating_rows()
                    generator.auto_adjust_columns()
                    generator.freeze_panes()
                    generator.apply_filter()
                else:
                    generator.write_data(data.sample(min(10, len(data))))
                    generator.apply_header_style()
                    generator.auto_adjust_columns()

        if output_path is None:
            output_path = f"{template_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        generator.save(output_path)
        return output_path


def create_excel_with_template(template_id: str, data: Union[pd.DataFrame, Dict, str],
                               output_path: str = None) -> str:
    if isinstance(data, str):
        if data.endswith('.csv'):
            data = pd.read_csv(data)
        elif data.endswith(('.xlsx', '.xls')):
            data = pd.read_excel(data)
        elif data.endswith('.json'):
            data = pd.read_json(data)

    return ExcelTemplateLibrary.create_template(template_id, data, output_path)


def create_professional_excel(data: Union[pd.DataFrame, Dict],
                             output_path: str,
                             style: str = "professional",
                             title: str = "数据报告",
                             include_charts: bool = True) -> str:
    generator = ExcelTemplateGenerator()
    generator.create_workbook(style=style)

    generator.current_sheet.title = "数据"
    generator.add_title(title)

    if isinstance(data, pd.DataFrame):
        generator.write_data(data, headers=data.columns.tolist())
        generator.apply_header_style()
        generator.apply_alternating_rows()
        generator.auto_adjust_columns()
        generator.freeze_panes()

        if include_charts and len(data.columns) >= 2:
            numeric_cols = data.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) >= 1:
                data_ref = f"B2:{get_column_letter(min(4, len(data.columns)))}{len(data)+1}"
                cats_ref = f"A2:A{len(data)+1}"
                generator.add_chart("bar", data_ref, "数据对比", "H5", cats_ref)

    generator.save(output_path)
    return output_path


if __name__ == "__main__":
    sample_data = pd.DataFrame({
        "产品": ["产品A", "产品B", "产品C", "产品D", "产品E"],
        "销售额": [12000, 15000, 18000, 16000, 20000],
        "利润": [3000, 4500, 5500, 4800, 6500],
        "数量": [100, 120, 150, 140, 180]
    })

    print("创建专业Excel模板...")
    create_professional_excel(sample_data, "professional_report.xlsx", style="blue")

    print("\n使用模板库...")
    for tmpl in ExcelTemplateLibrary.list_templates():
        print(f"  - {tmpl['name']}: {tmpl['description']}")

    ExcelTemplateLibrary.create_template("sales_monthly", sample_data, "sales_template.xlsx")
    print("\n模板生成完成!")
