#!/usr/bin/env python3
"""
Excel 数据分析和图表生成器
"""

import argparse
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json


def style_header(ws):
    """设置表头样式"""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def add_chart(ws, chart_type, data_ref, cats_ref, title, position):
    """添加图表"""
    if chart_type == 'bar':
        chart = BarChart()
    elif chart_type == 'line':
        chart = LineChart()
    elif chart_type == 'pie':
        chart = PieChart()
    else:
        chart = BarChart()
    
    chart.title = title
    chart.style = 10
    
    if chart_type == 'pie':
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, position)


def analyze_data(input_file, output_file, chart_type='bar', sheet_name='分析结果'):
    """分析数据并生成 Excel 报告"""
    # 读取数据
    if input_file.endswith('.csv'):
        df = pd.read_csv(input_file)
    elif input_file.endswith('.json'):
        with open(input_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        df = pd.DataFrame(data)
    else:
        df = pd.read_excel(input_file)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 写入数据
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 设置表头样式
    style_header(ws)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 添加图表
    if len(df.columns) >= 2:
        # 使用正确的 Reference 格式
        from openpyxl.chart.reference import Reference
        data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1, max_col=min(3, len(df.columns)))
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
        add_chart(ws, chart_type, data, cats, f'{df.columns[1]} 分析', 'E5')
    
    # 添加统计信息
    stats_row = len(df) + 3
    ws.cell(stats_row, 1, '统计信息').font = Font(bold=True, size=12)
    
    numeric_cols = df.select_dtypes(include=['number']).columns
    for i, col in enumerate(numeric_cols[:3], 1):  # 最多3个数值列
        ws.cell(stats_row + i, 1, f'{col} 平均值:')
        ws.cell(stats_row + i, 2, df[col].mean())
        ws.cell(stats_row + i, 3, f'{col} 总和:')
        ws.cell(stats_row + i, 4, df[col].sum())
    
    wb.save(output_file)
    print(f"分析完成: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Excel 数据分析器')
    parser.add_argument('--input', '-i', required=True, help='输入文件 (CSV/JSON/Excel)')
    parser.add_argument('--output', '-o', required=True, help='输出 Excel 文件')
    parser.add_argument('--chart-type', '-c', default='bar', 
                       choices=['bar', 'line', 'pie'], help='图表类型')
    parser.add_argument('--sheet-name', '-s', default='分析结果', help='工作表名称')
    
    args = parser.parse_args()
    analyze_data(args.input, args.output, args.chart_type, args.sheet_name)


if __name__ == '__main__':
    main()
