#!/usr/bin/env python3
"""
项目投资财务报告生成器 - Excel版
生成带图表的专业财务分析报告
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import sys

sys.path.insert(0, '/Users/zhaoruicn/.openclaw/workspace/skills/project-finance-model')
from project_finance_model import ProjectFinancialModel, ProjectInputs

class ExcelReportGenerator:
    """Excel财务报告生成器"""
    
    def __init__(self, model: ProjectFinancialModel):
        self.model = model
        self.inputs = model.inputs
        self.metrics = model.metrics if model.metrics else model.calculate_financial_metrics()
        self.wb = openpyxl.Workbook()
        
        # 定义样式
        self.setup_styles()
    
    def setup_styles(self):
        """设置Excel样式"""
        # 标题样式
        self.title_style = NamedStyle(name='title_style')
        self.title_style.font = Font(name='Microsoft YaHei', size=18, bold=True, color='FFFFFF')
        self.title_style.fill = PatternFill(start_color='1a3a6c', end_color='1a3a6c', fill_type='solid')
        self.title_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # 副标题样式
        self.subtitle_style = NamedStyle(name='subtitle_style')
        self.subtitle_style.font = Font(name='Microsoft YaHei', size=12, bold=True, color='1a3a6c')
        self.subtitle_style.fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
        self.subtitle_style.alignment = Alignment(horizontal='left', vertical='center')
        
        # 表头样式
        self.header_style = NamedStyle(name='header_style')
        self.header_style.font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill(start_color='2c5282', end_color='2c5282', fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据样式
        self.data_style = NamedStyle(name='data_style')
        self.data_style.font = Font(name='Microsoft YaHei', size=10)
        self.data_style.alignment = Alignment(horizontal='right', vertical='center')
        
        # 金额样式
        self.currency_style = NamedStyle(name='currency_style')
        self.currency_style.font = Font(name='Microsoft YaHei', size=10)
        self.currency_style.number_format = '#,##0.00'
        self.currency_style.alignment = Alignment(horizontal='right', vertical='center')
        
        # 百分比样式
        self.percent_style = NamedStyle(name='percent_style')
        self.percent_style.font = Font(name='Microsoft YaHei', size=10)
        self.percent_style.number_format = '0.00%'
        self.percent_style.alignment = Alignment(horizontal='right', vertical='center')
        
        # 高亮样式（绿色-好）
        self.good_style = NamedStyle(name='good_style')
        self.good_style.font = Font(name='Microsoft YaHei', size=10, bold=True, color='2e7d32')
        self.good_style.fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
        
        # 警告样式（橙色-临界）
        self.warning_style = NamedStyle(name='warning_style')
        self.warning_style.font = Font(name='Microsoft YaHei', size=10, bold=True, color='e65100')
        self.warning_style.fill = PatternFill(start_color='fff3e0', end_color='fff3e0', fill_type='solid')
        
        # 危险样式（红色-差）
        self.danger_style = NamedStyle(name='danger_style')
        self.danger_style.font = Font(name='Microsoft YaHei', size=10, bold=True, color='c62828')
        self.danger_style.fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')
        
        # 添加到工作簿
        for style in [self.title_style, self.subtitle_style, self.header_style, 
                     self.data_style, self.currency_style, self.percent_style,
                     self.good_style, self.warning_style, self.danger_style]:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
    
    def set_column_width(self, ws, col, width):
        """设置列宽"""
        ws.column_dimensions[get_column_letter(col)].width = width
    
    def create_summary_sheet(self):
        """创建摘要工作表"""
        ws = self.wb.active
        ws.title = '财务摘要'
        
        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = '项目投资财务分析报告'
        ws['A1'].style = 'title_style'
        ws.row_dimensions[1].height = 35
        
        # 项目信息
        ws['A3'] = '项目基本信息'
        ws['A3'].style = 'subtitle_style'
        ws.merge_cells('A3:F3')
        
        info_data = [
            ['项目名称', self.inputs.project_name, '', '项目类型', self.inputs.project_type, ''],
            ['运营期', f'{self.inputs.operation_years}年', '', '企业类型', self.inputs.enterprise_type, ''],
            ['分析日期', datetime.now().strftime('%Y-%m-%d'), '', '所在地区', self.inputs.location, ''],
        ]
        
        for i, row in enumerate(info_data, start=4):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                if j in [1, 4]:
                    cell.font = Font(bold=True)
        
        # 财务指标（核心）
        ws['A8'] = '核心财务指标'
        ws['A8'].style = 'subtitle_style'
        ws.merge_cells('A8:F8')
        
        metrics_data = [
            ['指标', '数值', '判断', '指标', '数值', '判断'],
            ['净现值(NPV)', self.metrics.npv, self._judge_npv(), 
             '内部收益率(IRR)', self.metrics.irr, self._judge_irr()],
            ['静态回收期', self.metrics.payback_period, self._judge_payback(),
             '动态回收期', self.metrics.dynamic_payback, self._judge_payback_dynamic()],
            ['投资利润率(ROI)', self.metrics.roi, '',
             '资本金利润率(ROE)', self.metrics.roe, ''],
        ]
        
        for i, row in enumerate(metrics_data, start=9):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 9:
                    cell.style = 'header_style'
                else:
                    if j in [2, 5]:  # 数值列
                        if '率' in str(row[j-2]):
                            cell.number_format = '0.00%'
                        elif '期' in str(row[j-2]):
                            cell.number_format = '0.00"年"'
                        else:
                            cell.number_format = '#,##0.00'
                    if j == 3 and val:  # 判断列
                        if '✓' in str(val):
                            cell.style = 'good_style'
                        elif '⚠' in str(val):
                            cell.style = 'warning_style'
                        elif '✗' in str(val):
                            cell.style = 'danger_style'
        
        # 设置列宽
        self.set_column_width(ws, 1, 18)
        self.set_column_width(ws, 2, 18)
        self.set_column_width(ws, 3, 15)
        self.set_column_width(ws, 4, 18)
        self.set_column_width(ws, 5, 18)
        self.set_column_width(ws, 6, 15)
        
        # 添加图表 - 投资构成饼图
        ws['A15'] = '投资构成分析'
        ws['A15'].style = 'subtitle_style'
        ws.merge_cells('A15:D15')
        
        # 投资数据
        investment_data = [
            ['投资项', '金额(万元)'],
            ['土地购置', self.inputs.land_cost/10000],
            ['建设工程', self.inputs.construction_cost/10000],
            ['设备购置', self.inputs.equipment_cost/10000],
            ['安装工程', self.inputs.installation_cost/10000],
            ['其他费用', self.inputs.other_cost/10000],
            ['流动资金', self.inputs.working_capital/10000],
        ]
        
        for i, row in enumerate(investment_data, start=16):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 16:
                    cell.style = 'header_style'
                else:
                    if j == 2:
                        cell.number_format = '#,##0.00'
        
        # 创建饼图
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=17, max_row=22)
        data = Reference(ws, min_col=2, min_row=16, max_row=22)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = '投资构成'
        pie.style = 10
        ws.add_chart(pie, 'E16')
        
        # 结论
        ws['A25'] = '分析结论'
        ws['A25'].style = 'subtitle_style'
        ws.merge_cells('A25:F25')
        
        conclusion = self._generate_conclusion()
        ws['A26'] = conclusion
        ws['A26'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A26:F30')
        ws.row_dimensions[26].height = 80
    
    def create_cashflow_sheet(self):
        """创建现金流明细表"""
        ws = self.wb.create_sheet('现金流明细')
        
        # 标题
        ws.merge_cells('A1:I1')
        ws['A1'] = '项目现金流明细表'
        ws['A1'].style = 'title_style'
        ws.row_dimensions[1].height = 30
        
        # 表头
        headers = ['年份', '营业收入', '营业成本', '折旧', '利息', '所得税', 
                   '净利润', '现金流', '累计现金流']
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.style = 'header_style'
        
        # 数据
        for i, year_data in enumerate(self.model.yearly_data, start=4):
            row_data = [
                year_data.year,
                year_data.revenue/10000,
                year_data.cost/10000,
                year_data.depreciation/10000,
                year_data.interest/10000,
                year_data.tax/10000,
                year_data.net_profit/10000,
                year_data.cash_flow/10000,
                year_data.cumulative_cf/10000,
            ]
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                if j == 1:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.number_format = '#,##0.00'
                    if val < 0:
                        cell.font = Font(color='c62828')
        
        # 设置列宽
        for i in range(1, 10):
            self.set_column_width(ws, i, 14)
        
        # 添加折线图 - 现金流趋势
        line = LineChart()
        line.title = '现金流趋势'
        line.style = 10
        line.y_axis.title = '金额(万元)'
        line.x_axis.title = '年份'
        
        # 现金流数据
        data = Reference(ws, min_col=8, min_row=3, max_row=3+len(self.model.yearly_data))
        categories = Reference(ws, min_col=1, min_row=4, max_row=3+len(self.model.yearly_data))
        line.add_data(data, titles_from_data=True)
        line.set_categories(categories)
        
        ws.add_chart(line, 'K3')
    
    def create_investment_sheet(self):
        """创建投资分析表"""
        ws = self.wb.create_sheet('投资分析')
        
        # 标题
        ws.merge_cells('A1:E1')
        ws['A1'] = '项目投资分析'
        ws['A1'].style = 'title_style'
        
        # 投资构成
        ws['A3'] = '投资构成明细'
        ws['A3'].style = 'subtitle_style'
        ws.merge_cells('A3:E3')
        
        inv = self.model.calculate_total_investment()
        
        investment_items = [
            ['投资项', '金额(万元)', '占比', '说明'],
            ['土地购置', self.inputs.land_cost/10000, 
             self.inputs.land_cost/inv['total_investment'], ''],
            ['建设工程', self.inputs.construction_cost/10000,
             self.inputs.construction_cost/inv['total_investment'], ''],
            ['设备购置', self.inputs.equipment_cost/10000,
             self.inputs.equipment_cost/inv['total_investment'], '核心投资'],
            ['安装工程', self.inputs.installation_cost/10000,
             self.inputs.installation_cost/inv['total_investment'], ''],
            ['其他费用', self.inputs.other_cost/10000,
             self.inputs.other_cost/inv['total_investment'], ''],
            ['流动资金', self.inputs.working_capital/10000,
             self.inputs.working_capital/inv['total_investment'], ''],
            ['合计', inv['total_investment']/10000, 1.0, ''],
        ]
        
        for i, row in enumerate(investment_items, start=4):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 4:
                    cell.style = 'header_style'
                else:
                    if j == 2:
                        cell.number_format = '#,##0.00'
                    elif j == 3:
                        cell.number_format = '0.00%'
                    if i == 11:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='e8f4f8', end_color='e8f4f8', fill_type='solid')
        
        # 资金来源
        ws['A14'] = '资金来源'
        ws['A14'].style = 'subtitle_style'
        ws.merge_cells('A14:E14')
        
        funding_data = [
            ['资金来源', '金额(万元)', '占比', '说明'],
            ['资本金', inv['equity']/10000, self.inputs.equity_ratio, '自有资金'],
            ['银行贷款', inv['loan']/10000, self.inputs.loan_ratio, f'利率{self.inputs.loan_rate*100}%，{self.inputs.loan_term}年'],
            ['合计', inv['total_investment']/10000, 1.0, ''],
        ]
        
        for i, row in enumerate(funding_data, start=15):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                if i == 15:
                    cell.style = 'header_style'
                else:
                    if j == 2:
                        cell.number_format = '#,##0.00'
                    elif j == 3:
                        cell.number_format = '0.00%'
        
        # 设置列宽
        self.set_column_width(ws, 1, 15)
        self.set_column_width(ws, 2, 18)
        self.set_column_width(ws, 3, 12)
        self.set_column_width(ws, 4, 25)
    
    def generate_report(self, output_path: str = None):
        """生成完整报告"""
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'/Users/zhaoruicn/.openclaw/workspace/项目财务分析报告_{timestamp}.xlsx'
        
        # 创建各工作表
        self.create_summary_sheet()
        self.create_cashflow_sheet()
        self.create_investment_sheet()
        
        # 保存
        self.wb.save(output_path)
        return output_path
    
    def _judge_npv(self) -> str:
        """判断NPV"""
        if self.metrics.npv > 0:
            return '✓ 可行'
        elif self.metrics.npv > -1000000:
            return '⚠ 临界'
        else:
            return '✗ 不可行'
    
    def _judge_irr(self) -> str:
        """判断IRR"""
        if self.metrics.irr > 0.15:
            return '✓ 优秀'
        elif self.metrics.irr > 0.08:
            return '✓ 良好'
        elif self.metrics.irr > 0.05:
            return '⚠ 一般'
        else:
            return '✗ 较差'
    
    def _judge_payback(self) -> str:
        """判断回收期"""
        if self.metrics.payback_period < 8:
            return '✓ 优秀'
        elif self.metrics.payback_period < 10:
            return '✓ 良好'
        elif self.metrics.payback_period < 12:
            return '⚠ 一般'
        else:
            return '✗ 较长'
    
    def _judge_payback_dynamic(self) -> str:
        """判断动态回收期"""
        if self.metrics.dynamic_payback < 10:
            return '✓ 优秀'
        elif self.metrics.dynamic_payback < 12:
            return '✓ 良好'
        elif self.metrics.dynamic_payback < 15:
            return '⚠ 一般'
        else:
            return '✗ 较长'
    
    def _generate_conclusion(self) -> str:
        """生成分析结论"""
        conclusions = []
        
        # NPV判断
        if self.metrics.npv > 0:
            conclusions.append(f"✓ 项目净现值NPV为¥{self.metrics.npv/10000:.2f}万元，大于0，项目财务上可行。")
        else:
            conclusions.append(f"✗ 项目净现值NPV为¥{self.metrics.npv/10000:.2f}万元，小于0，项目财务上不可行。")
        
        # IRR判断
        if self.metrics.irr > 0.08:
            conclusions.append(f"✓ 内部收益率IRR为{self.metrics.irr*100:.2f}%，高于折现率8%，投资回报良好。")
        else:
            conclusions.append(f"⚠ 内部收益率IRR为{self.metrics.irr*100:.2f}%，低于折现率8%，投资回报一般。")
        
        # 回收期
        conclusions.append(f"• 静态回收期{self.metrics.payback_period:.2f}年，动态回收期{self.metrics.dynamic_payback:.2f}年。")
        
        # 投资利润
        conclusions.append(f"• 投资利润率ROI为{self.metrics.roi*100:.2f}%，资本金利润率ROE为{self.metrics.roe*100:.2f}%。")
        
        # 综合建议
        if self.metrics.npv > 0 and self.metrics.irr > 0.08:
            conclusions.append(f"\n【综合建议】该项目财务指标良好，建议投资。")
        elif self.metrics.npv > 0:
            conclusions.append(f"\n【综合建议】该项目基本可行，但收益率偏低，建议优化方案或谨慎投资。")
        else:
            conclusions.append(f"\n【综合建议】该项目财务指标不佳，建议重新评估或放弃投资。")
        
        return '\n'.join(conclusions)

# 示例使用
def example():
    """生成示例报告"""
    inputs = ProjectInputs(
        project_name="100MWh储能电站",
        project_type="储能电站",
        operation_years=15,
        land_cost=5000000,
        construction_cost=15000000,
        equipment_cost=30000000,
        installation_cost=5000000,
        other_cost=3000000,
        working_capital=2000000,
        annual_revenue=12000000,
        annual_cost=2000000,
        maintenance_cost=800000,
        labor_cost=600000,
        enterprise_type="高新技术企业",
        location="市区",
        discount_rate=0.08,
    )
    
    model = ProjectFinancialModel(inputs)
    model.calculate_financial_metrics()
    
    generator = ExcelReportGenerator(model)
    output_path = generator.generate_report()
    
    print(f"✅ 财务报告已生成: {output_path}")
    return output_path

if __name__ == "__main__":
    example()
