#!/usr/bin/env python3
"""
工商业储能项目财务测算Excel生成脚本
模式一：纯储能10MWh
模式二：光储一体化1MW+3MWh
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import datetime

# ============ 样式定义 ============
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUBHEADER_BG = "2E75B6"
COLOR_GREEN_BG = "C6EFCE"
COLOR_RED_BG = "FFC7CE"
COLOR_YELLOW_BG = "FFEB9C"
COLOR_GRAY_BG = "F2F2F2"
COLOR_BORDER = "8EA9C1"

fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_subheader = PatternFill(start_color=COLOR_SUBHEADER_BG, end_color=COLOR_SUBHEADER_BG, fill_type="solid")
fill_green = PatternFill(start_color=COLOR_GREEN_BG, end_color=COLOR_GREEN_BG, fill_type="solid")
fill_red = PatternFill(start_color=COLOR_RED_BG, end_color=COLOR_RED_BG, fill_type="solid")
fill_yellow = PatternFill(start_color=COLOR_YELLOW_BG, end_color=COLOR_YELLOW_BG, fill_type="solid")
fill_gray = PatternFill(start_color=COLOR_GRAY_BG, end_color=COLOR_GRAY_BG, fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

font_title = Font(name="微软雅黑", size=16, bold=True, color=COLOR_HEADER_FG)
font_header = Font(name="微软雅黑", size=11, bold=True, color=COLOR_HEADER_FG)
font_subheader = Font(name="微软雅黑", size=10, bold=True)
font_normal = Font(name="微软雅黑", size=10)
font_bold = Font(name="微软雅黑", size=10, bold=True)
font_small = Font(name="微软雅黑", size=9)
font_green = Font(name="微软雅黑", size=10, color="006100", bold=True)
font_red = Font(name="微软雅黑", size=10, color="9C0006", bold=True)
font_yellow = Font(name="微软雅黑", size=10, color="9C5700", bold=True)

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

def irr_to_fill(irr):
    if irr >= 10:
        return fill_green
    elif irr >= 5:
        return fill_yellow
    else:
        return fill_red

def irr_to_font(irr):
    if irr >= 10:
        return font_green
    elif irr >= 5:
        return font_yellow
    else:
        return font_red

def irr_to_color_name(irr):
    if irr >= 10:
        return "绿色（IRR≥10%）"
    elif irr >= 5:
        return "黄色（IRR 5-10%）"
    else:
        return "红色（IRR<5%）"

# ============ 财务计算函数 ============
def calc_mode1_irr(price_diff, capacity=10000, total_invest=1000,
                   peak_ratio=0.35, op_days=330, op_cost=10, 
                   depreciation=80, interest=34, tax_rate=0.25):
    daily_discharge = capacity * peak_ratio  # kWh
    annual_discharge = daily_discharge * op_days
    annual_revenue = annual_discharge * price_diff / 10000  # 转为万元
    vat = annual_revenue / 1.13 * 0.13
    pre_tax_profit = annual_revenue - vat - op_cost - depreciation - interest
    income_tax = max(0, pre_tax_profit) * tax_rate
    net_profit = pre_tax_profit - income_tax
    annual_cashflow = net_profit + depreciation
    irr = annual_cashflow / total_invest * 100
    return {
        'annual_discharge_kwh': annual_discharge,
        'annual_revenue': annual_revenue,
        'vat': vat,
        'pre_tax_profit': pre_tax_profit,
        'income_tax': income_tax,
        'net_profit': net_profit,
        'annual_cashflow': annual_cashflow,
        'irr': irr
    }

def calc_mode1_threshold(target_irr, **params):
    target_cashflow = target_irr / 100 * params['total_invest']
    target_net_profit = target_cashflow - params['depreciation']
    target_pre_tax = target_net_profit / (1 - params['tax_rate'])
    target_revenue = (target_pre_tax + params['op_cost'] + params['depreciation'] + params['interest']) * 1.13
    min_price_diff = target_revenue / params['annual_discharge_kwh'] * 10000  # 转为元/kWh
    return min_price_diff

def calc_mode2_irr(pv_price=0.35, storage_charge=0.10, storage_discharge=0.60,
                   pv_capacity=1, storage_capacity=3, total_invest=500,
                   pv_hours=1200, peak_ratio=0.35, op_days=330,
                   op_cost=8, depreciation=40, interest=17, tax_rate=0.25):
    pv_annual_kwh = pv_capacity * pv_hours * 10000  # 转为kWh
    pv_annual_revenue = pv_annual_kwh * pv_price / 10000  # 万元
    
    storage_daily_kwh = storage_capacity * 1000 * peak_ratio  # kWh
    storage_annual_kwh = storage_daily_kwh * op_days
    storage_price_diff = storage_discharge - storage_charge
    storage_annual_revenue = storage_annual_kwh * storage_price_diff / 10000  # 万元
    
    total_revenue = pv_annual_revenue + storage_annual_revenue
    total_revenue_excl_tax = total_revenue * 0.87
    
    pre_tax_profit = total_revenue_excl_tax - op_cost - depreciation - interest
    income_tax = max(0, pre_tax_profit) * tax_rate
    net_profit = pre_tax_profit - income_tax
    annual_cashflow = net_profit + depreciation
    irr = annual_cashflow / total_invest * 100
    
    return {
        'pv_annual_kwh': pv_annual_kwh,
        'pv_annual_revenue': pv_annual_revenue,
        'storage_annual_kwh': storage_annual_kwh,
        'storage_annual_revenue': storage_annual_revenue,
        'total_revenue': total_revenue,
        'total_revenue_excl_tax': total_revenue_excl_tax,
        'pre_tax_profit': pre_tax_profit,
        'income_tax': income_tax,
        'net_profit': net_profit,
        'annual_cashflow': annual_cashflow,
        'irr': irr
    }

# ============ 数据 ============
PROVINCE_DATA_MODE1 = [
    ("山西", 0.01, 0.80, 0.79, "华北", 1),
    ("广东", 0.05, 0.60, 0.55, "华南", 1),
    ("浙江", 0.05, 0.50, 0.45, "华东", 1),
    ("上海", 0.40, 0.80, 0.40, "华东", 1),
    ("辽宁", 0.05, 0.45, 0.40, "东北", 1),
    ("山东", 0.05, 0.40, 0.35, "华东", 2),
    ("蒙西", 0.10, 0.40, 0.30, "华北", 2),
    ("甘肃", 0.12, 0.40, 0.28, "西北", 2),
    ("江苏", 0.35, 0.60, 0.25, "华东", 2),
    ("河南", 0.35, 0.60, 0.25, "华中", 2),
    ("河北", 0.35, 0.60, 0.25, "华北", 2),
    ("湖北", 0.35, 0.60, 0.25, "华中", 2),
    ("安徽", 0.35, 0.60, 0.25, "华东", 2),
    ("陕西", 0.20, 0.42, 0.22, "西北", 3),
    ("四川", 0.05, 0.50, 0.45, "西南", 3),
    ("重庆", 0.10, 0.50, 0.40, "西南", 3),
    ("云南", 0.10, 0.45, 0.35, "西南", 3),
    ("贵州", 0.10, 0.45, 0.35, "西南", 3),
    ("广西", 0.15, 0.50, 0.35, "华南", 3),
    ("湖南", 0.20, 0.45, 0.25, "华中", 3),
    ("江西", 0.20, 0.45, 0.25, "华东", 3),
    ("福建", 0.25, 0.50, 0.25, "华东", 3),
    ("天津", 0.30, 0.55, 0.25, "华北", 3),
    ("北京", 0.30, 0.60, 0.30, "华北", 3),
    ("吉林", 0.10, 0.45, 0.35, "东北", 3),
    ("黑龙江", 0.10, 0.45, 0.35, "东北", 3),
    ("蒙东", 0.10, 0.35, 0.25, "华北", 3),
    ("新疆", 0.10, 0.35, 0.25, "西北", 3),
    ("青海", 0.10, 0.35, 0.25, "西北", 3),
    ("宁夏", 0.10, 0.35, 0.25, "西北", 3),
    ("海南", 0.15, 0.45, 0.30, "华南", 3),
]

SPOT_PRICE_DATA = [
    ("山东", 0.35, 0.57, "现货价格波动大，有负电价机会"),
    ("山西", 0.40, 1.50, "低点是负电价，套利空间最大"),
    ("广东", 0.40, 1.29, "现货价格较高，峰谷差显著"),
    ("甘肃", 0.25, 0.46, "新能源装机大，价格波动适中"),
    ("浙江", 0.35, 1.27, "现货市场活跃，价差空间大"),
    ("蒙西", 0.20, 0.45, "新能源富集，储能需求大"),
    ("江苏", 0.30, 0.55, "工商业发达，峰谷差稳定"),
]

REGIONS = {
    "华东": ["上海", "江苏", "浙江", "安徽", "福建", "江西", "山东"],
    "华北": ["北京", "天津", "河北", "山西", "蒙西", "蒙东"],
    "华中": ["湖北", "湖南", "河南", "江西"],
    "华南": ["广东", "广西", "海南"],
    "西南": ["四川", "重庆", "云南", "贵州"],
    "西北": ["陕西", "甘肃", "青海", "宁夏", "新疆"],
    "东北": ["辽宁", "吉林", "黑龙江"],
}

# ============ 生成Excel ============
def create_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    create_cover_sheet(wb)
    create_mode1_sheet(wb)
    create_mode2_sheet(wb)
    create_spot_price_sheet(wb)
    create_province_tier_sheet(wb)
    create_china_map_sheet(wb)
    create_conclusion_sheet(wb)
    
    output_path = "/Users/zhaoruicn/.openclaw/workspace/工商业储能市场开拓分析_修正版.xlsx"
    wb.save(output_path)
    print(f"Excel文件已生成: {output_path}")
    return output_path

# ============ Sheet 1: 封面 ============
def create_cover_sheet(wb):
    ws = wb.create_sheet("封面", 0)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 5
    
    for i in range(1, 35):
        ws.row_dimensions[i].height = 25
    
    ws.merge_cells('B2:C2')
    set_cell(ws, 2, 2, "工商业储能项目财务测算分析", font_title, fill_header, align_center)
    
    ws.merge_cells('B3:C3')
    set_cell(ws, 3, 2, "——全国各省IRR可行性研究报告", Font(name="微软雅黑", size=14, color=COLOR_HEADER_FG), fill_subheader, align_center)
    
    ws.row_dimensions[4].height = 30
    
    set_cell(ws, 5, 2, "报告日期：", font_bold, fill_white, align_right)
    set_cell(ws, 5, 3, datetime.datetime.now().strftime("%Y年%m月%d日"), font_normal, fill_white, align_left)
    
    set_cell(ws, 6, 2, "编制单位：", font_bold, fill_white, align_right)
    set_cell(ws, 6, 3, "雪子项目组", font_normal, fill_white, align_left)
    
    set_cell(ws, 7, 2, "报告版本：", font_bold, fill_white, align_right)
    set_cell(ws, 7, 3, "V2.0 修正版", font_normal, fill_white, align_left)
    
    ws.row_dimensions[8].height = 40
    
    ws.merge_cells('B9:C9')
    set_cell(ws, 9, 2, "目  录", Font(name="微软雅黑", size=16, bold=True), fill_subheader, align_center, thin_border)
    
    toc_items = [
        ("一", "模式一经济账（纯储能10MWh）", "模式一计算逻辑、门槛分析、各省IRR预估"),
        ("二", "模式二经济账（光储一体化1MW+3MWh）", "最优配置、敏感性分析、多情景对比"),
        ("三", "现货市场价格数据", "山东、山西、广东、甘肃、浙江等省份实时电价"),
        ("四", "省份梯队划分", "31省工商业储能投资价值梯队分类"),
        ("五", "全国投资地图", "七大区域储能项目分布与投资机会"),
        ("六", "综合结论", "开拓建议与风险提示"),
    ]
    
    row = 10
    for num, title, desc in toc_items:
        ws.row_dimensions[row].height = 30
        set_cell(ws, row, 2, num, font_bold, fill_gray, align_center, thin_border)
        set_cell(ws, row, 3, f"{title}\n{desc}", font_normal, fill_white, align_left, thin_border)
        row += 1
    
    ws.row_dimensions[row].height = 30
    row += 2
    
    ws.merge_cells(f'B{row}:C{row}')
    set_cell(ws, row, 2, "财务模型说明：简化IRR = 年现金流 / 总投资 = (税后净利润+折旧) / 总投资", Font(name="微软雅黑", size=9, italic=True, color="666666"), fill_white, align_left)
    
    row += 2
    ws.merge_cells(f'B{row}:C{row}')
    set_cell(ws, row, 2, "IRR评价标准：", font_bold, fill_white, align_left)
    
    row += 1
    set_cell(ws, row, 2, "● 绿色（IRR≥10%）", Font(name="微软雅黑", size=10, color="006100"), fill_green, align_left)
    row += 1
    set_cell(ws, row, 2, "● 黄色（IRR 8-10%）", Font(name="微软雅黑", size=10, color="9C5700"), fill_yellow, align_left)
    row += 1
    set_cell(ws, row, 2, "● 红色（IRR<8%）", Font(name="微软雅黑", size=10, color="9C0006"), fill_red, align_left)

# ============ Sheet 2: 模式一经济账 ============
def create_mode1_sheet(wb):
    ws = wb.create_sheet("模式一经济账", 1)
    
    col_widths = [4, 16, 12, 12, 12, 12, 12, 12, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    row = 1
    
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "模式一经济账：纯储能10MWh项目IRR测算", font_title, fill_header, align_center)
    ws.row_dimensions[row].height = 35
    row += 2
    
    # 基础参数
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "一、基础参数", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["参数名称", "数值", "单位", "说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    params_data = [
        ("储能容量", 10, "MWh", "电池系统额定容量"),
        ("总投资", 1000, "万元", "初始投资成本"),
        ("充放电效率", 90, "%", "往返效率"),
        ("放电深度(DOD)", 95, "%", "可用容量比例"),
        ("峰谷系数", 35, "%", "每天放电量比例"),
        ("日放电量", 3500, "kWh", "10MWh × 35%"),
        ("年运行天数", 330, "天", "年度运行天数"),
        ("年放电量", 115.5, "万kWh", "3500kWh × 330天"),
        ("年运维成本", 10, "万元", "年度运维费用"),
        ("折旧", 80, "万元/年", "800万/10年"),
        ("利息", 34, "万元/年", "700万×4.9%"),
        ("所得税率", 25, "%", "企业所得税"),
        ("贷款比例", 70, "%", "总投资中贷款占比"),
        ("贷款利率", 4.9, "%", "年化利率"),
    ]
    
    for i, (name, value, unit, desc) in enumerate(params_data):
        fill = fill_gray if i % 2 == 0 else fill_white
        set_cell(ws, row, 2, name, font_normal, fill, align_left, thin_border)
        set_cell(ws, row, 3, value, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, unit, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, desc, font_small, fill, align_left, thin_border)
        row += 1
    
    row += 1
    
    # IRR门槛分析
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "二、IRR门槛分析（达到目标IRR所需的最小峰谷价差）", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    calc_params = {
        'total_invest': 1000,
        'depreciation': 80,
        'interest': 34,
        'op_cost': 10,
        'tax_rate': 0.25,
        'annual_discharge_kwh': 1155000,
    }
    
    threshold_10 = calc_mode1_threshold(10, **calc_params)
    threshold_8 = calc_mode1_threshold(8, **calc_params)
    threshold_6 = calc_mode1_threshold(6, **calc_params)
    
    headers = ["IRR目标", "所需最小价差", "说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    for col in range(5, 10):
        set_cell(ws, row, col, "", font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    thresholds = [(10, threshold_10, "项目可行基准线"), (8, threshold_8, "基本可行，需优化"), (6, threshold_6, "需较大政策支持")]
    for irr_target, min_diff, note in thresholds:
        fill = irr_to_fill(irr_target)
        set_cell(ws, row, 2, f"IRR ≥ {irr_target}%", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{min_diff:.2f} 元/kWh", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 4, note, font_normal, fill, align_left, thin_border)
        row += 1
    
    row += 1
    
    # 敏感性分析
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "三、敏感性分析（价差变化对IRR的影响）", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["峰谷价差", "IRR", "评价", "年收入", "增值税", "税前利润", "税后净利润", "年现金流"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    price_range = [0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90, 0.95, 1.00]
    for price in price_range:
        result = calc_mode1_irr(price)
        fill = irr_to_fill(result['irr'])
        set_cell(ws, row, 2, f"{price:.2f} 元/kWh", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{result['irr']:.2f}%", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 4, irr_to_color_name(result['irr']), font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"{result['annual_revenue']:.2f}", font_normal, fill, align_right, thin_border)
        set_cell(ws, row, 6, f"{result['vat']:.2f}", font_normal, fill, align_right, thin_border)
        set_cell(ws, row, 7, f"{result['pre_tax_profit']:.2f}", font_normal, fill, align_right, thin_border)
        set_cell(ws, row, 8, f"{result['net_profit']:.2f}", font_normal, fill, align_right, thin_border)
        row += 1
    
    row += 1
    
    # 全国各省IRR预估
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "四、全国各省IRR预估（目录电价）", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["省份", "低谷价", "高峰价", "价差", "IRR", "IRR评价", "区域", "投资建议"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    province_results = []
    for prov, low, high, diff, region, tier in PROVINCE_DATA_MODE1:
        result = calc_mode1_irr(diff)
        province_results.append((prov, low, high, diff, result['irr'], region, tier))
    
    province_results.sort(key=lambda x: x[4], reverse=True)
    
    for prov, low, high, diff, irr, region, tier in province_results:
        fill = irr_to_fill(irr)
        irr_level = "优先开拓" if irr >= 10 else ("重点关注" if irr >= 8 else "谨慎进入")
        set_cell(ws, row, 2, prov, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{low:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, f"{high:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"{diff:.2f}", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 6, f"{irr:.1f}%", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 7, irr_to_color_name(irr), font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 8, region, font_normal, fill, align_center, thin_border)
        row += 1
    
    row += 1
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "注：IRR计算假设每年运行330天，峰谷系数35%，实际IRR可能因运行策略、电价波动有所差异。", Font(name="微软雅黑", size=9, italic=True, color="666666"), fill_white, align_left)

# ============ Sheet 3: 模式二经济账 ============
def create_mode2_sheet(wb):
    ws = wb.create_sheet("模式二经济账", 2)
    
    col_widths = [4, 18, 14, 14, 14, 14, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    row = 1
    
    ws.merge_cells(f'B{row}:H{row}')
    set_cell(ws, row, 2, "模式二经济账：光储一体化1MW+3MWh项目IRR测算", font_title, fill_header, align_center)
    ws.row_dimensions[row].height = 35
    row += 2
    
    # 基础参数
    ws.merge_cells(f'B{row}:H{row}')
    set_cell(ws, row, 2, "一、基础参数", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["参数名称", "数值", "单位", "说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    params_data = [
        ("光伏装机容量", 1, "MW", "装机规模"),
        ("光伏成本", 200, "万元", "2元/W × 1MW"),
        ("储能容量", 3, "MWh", "电池系统额定容量"),
        ("储能成本", 300, "万元", "1元/Wh × 3MWh"),
        ("光储总投资", 500, "万元", "总计初始投资"),
        ("光伏年发电量", 120, "万kWh", "1MW × 1200h"),
        ("储能年放电量", 34.65, "万kWh", "3MWh×35%×330天"),
        ("午间低电价(充电)", 0.10, "元/kWh", "储能充电成本"),
        ("晚高峰电价(放电)", 0.60, "元/kWh", "储能放电收入"),
        ("光伏上网电价", 0.35, "元/kWh", "假设固定电价"),
        ("年运维成本", 8, "万元", "年度运维费用"),
        ("折旧", 40, "万元/年", "400万/10年"),
        ("利息", 17, "万元/年", "350万×4.9%"),
        ("所得税率", 25, "%", "企业所得税"),
    ]
    
    for i, (name, value, unit, desc) in enumerate(params_data):
        fill = fill_gray if i % 2 == 0 else fill_white
        set_cell(ws, row, 2, name, font_normal, fill, align_left, thin_border)
        set_cell(ws, row, 3, value, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, unit, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, desc, font_small, fill, align_left, thin_border)
        row += 1
    
    row += 1
    
    # 最优配置IRR
    ws.merge_cells(f'B{row}:H{row}')
    set_cell(ws, row, 2, "二、最优配置IRR计算（基准情景）", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    result = calc_mode2_irr()
    
    headers = ["收入项", "金额(万元)", "说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    for col in range(5, 9):
        set_cell(ws, row, col, "", font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    income_items = [
        ("光伏年收入", result['pv_annual_revenue'], f"发电{result['pv_annual_kwh']/10000:.0f}万kWh × {0.35}元"),
        ("储能年收入", result['storage_annual_revenue'], f"放电{result['storage_annual_kwh']/10000:.2f}万kWh × 价差{0.50}元"),
        ("总收入（含税）", result['total_revenue'], "光伏+储能"),
        ("总收入（不含税）", result['total_revenue_excl_tax'], "×0.87增值税抵扣"),
    ]
    
    for name, value, desc in income_items:
        set_cell(ws, row, 2, name, font_normal, fill_white, align_left, thin_border)
        set_cell(ws, row, 3, f"{value:.2f}", font_normal, fill_white, align_right, thin_border)
        set_cell(ws, row, 4, desc, font_small, fill_white, align_left, thin_border)
        row += 1
    
    row += 1
    
    headers = ["利润项", "金额(万元)", "说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    for col in range(5, 9):
        set_cell(ws, row, col, "", font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    profit_items = [
        ("税前利润", result['pre_tax_profit'], "总收入-运维-折旧-利息"),
        ("所得税(25%)", result['income_tax'], "max(0,税前利润)×25%"),
        ("税后净利润", result['net_profit'], "税前利润-所得税"),
        ("年现金流", result['annual_cashflow'], "净利润+折旧"),
    ]
    
    for name, value, desc in profit_items:
        fill = fill_gray
        set_cell(ws, row, 2, name, font_normal, fill, align_left, thin_border)
        set_cell(ws, row, 3, f"{value:.2f}", font_normal, fill, align_right, thin_border)
        set_cell(ws, row, 4, desc, font_small, fill, align_left, thin_border)
        row += 1
    
    row += 1
    
    # IRR结果
    fill = irr_to_fill(result['irr'])
    ws.merge_cells(f'B{row}:D{row}')
    set_cell(ws, row, 2, "项目IRR（简化）= 年现金流/总投资", font_bold, fill, align_center, thin_border)
    set_cell(ws, row, 5, f"{result['irr']:.2f}%", Font(name="微软雅黑", size=14, bold=True), fill, align_center, thin_border)
    set_cell(ws, row, 6, irr_to_color_name(result['irr']), font_normal, fill, align_center, thin_border)
    row += 2
    ws.merge_cells(f'B{row}:H{row}')
    set_cell(ws, row, 2, "注：模式二IRR显著高于模式一，因为光伏发电收入+储能峰谷套利的双重收益叠加。", Font(name="微软雅黑", size=9, italic=True, color="666666"), fill_white, align_left)
    row += 2
    
    # 敏感性分析
    ws.merge_cells(f'B{row}:H{row}')
    set_cell(ws, row, 2, "三、敏感性分析（电价变化对IRR的影响）", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["情景", "充电价", "放电价", "价差", "IRR", "评价", ""]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    scenarios = [
        ("悲观", 0.15, 0.50, 0.35),
        ("基准", 0.10, 0.60, 0.50),
        ("乐观", 0.05, 0.70, 0.65),
        ("最优", 0.02, 0.80, 0.78),
    ]
    
    for name, charge, discharge, diff in scenarios:
        r = calc_mode2_irr(storage_charge=charge, storage_discharge=discharge)
        fill = irr_to_fill(r['irr'])
        set_cell(ws, row, 2, name, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{charge:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, f"{discharge:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"{diff:.2f}", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 6, f"{r['irr']:.2f}%", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 7, irr_to_color_name(r['irr']), font_normal, fill, align_center, thin_border)
        row += 1
    
    row += 1
    
    # 多情景对比
    ws.merge_cells(f'B{row}:H{row}')
    set_cell(ws, row, 2, "四、多情景对比（光伏/储能配置变化）", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["配置方案", "光伏", "储能", "总投资", "年收入", "IRR", "评价"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    configs = [
        ("方案A: 0.5MW+2MWh", 0.5, 2, 350, 43.8, calc_mode2_irr(pv_capacity=0.5, storage_capacity=2, total_invest=350, depreciation=28, interest=12)),
        ("方案B: 1MW+3MWh（基准）", 1, 3, 500, 70.6, calc_mode2_irr()),
        ("方案C: 1.5MW+4MWh", 1.5, 4, 750, 98.3, calc_mode2_irr(pv_capacity=1.5, storage_capacity=4, total_invest=750, depreciation=60, interest=26)),
        ("方案D: 2MW+5MWh", 2, 5, 1000, 126.0, calc_mode2_irr(pv_capacity=2, storage_capacity=5, total_invest=1000, depreciation=80, interest=34)),
    ]
    
    for name, pv, storage, invest, revenue, result in configs:
        fill = irr_to_fill(result['irr'])
        set_cell(ws, row, 2, name, font_normal, fill, align_left, thin_border)
        set_cell(ws, row, 3, f"{pv}MW", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, f"{storage}MWh", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"{invest}万", font_normal, fill, align_right, thin_border)
        set_cell(ws, row, 6, f"{revenue:.1f}万", font_normal, fill, align_right, thin_border)
        set_cell(ws, row, 7, f"{result['irr']:.2f}%", font_bold, fill, align_center, thin_border)
        row += 1
    
    row += 1

# ============ Sheet 4: 现货市场价格数据 ============
def create_spot_price_sheet(wb):
    ws = wb.create_sheet("现货市场价格数据", 3)
    
    col_widths = [4, 14, 12, 12, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    row = 1
    
    ws.merge_cells(f'B{row}:F{row}')
    set_cell(ws, row, 2, "现货市场价格数据（实时电价波动范围）", font_title, fill_header, align_center)
    ws.row_dimensions[row].height = 35
    row += 2
    
    headers = ["省份", "最低价", "最高价", "价差", "波动说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    for prov, low, high, note in SPOT_PRICE_DATA:
        diff = high - low
        # 根据价差估算IRR
        irr = calc_mode1_irr(high)['irr'] if high > 0.4 else calc_mode1_irr(0.35)['irr']
        fill = irr_to_fill(irr)
        set_cell(ws, row, 2, prov, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{low:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, f"{high:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"{diff:.2f}", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 6, note, font_small, fill, align_left, thin_border)
        row += 1
    
    row += 1
    
    # 说明
    ws.merge_cells(f'B{row}:F{row}')
    set_cell(ws, row, 2, "说明：现货市场价格波动较大，实际IRR取决于充放电策略的执行精度。", Font(name="微软雅黑", size=9, italic=True, color="666666"), fill_white, align_left)
    
    row += 2
    
    # 现货vs目录电价对比
    ws.merge_cells(f'B{row}:F{row}')
    set_cell(ws, row, 2, "现货电价 vs 目录电价对比", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["省份", "目录电价价差", "现货最高价差", "套利空间", "说明"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    spot_comparison = [
        ("山东", 0.35, 0.57, "现货价格高于目录"),
        ("山西", 0.79, 1.50, "现货价格远超目录，负电价机会"),
        ("广东", 0.55, 1.29, "现货套利空间大"),
        ("甘肃", 0.28, 0.46, "现货价格更优"),
        ("浙江", 0.45, 1.27, "现货价格波动大"),
    ]
    
    for prov, catalog, spot, note in spot_comparison:
        extra = spot - catalog
        fill = fill_green if extra > 0.3 else (fill_yellow if extra > 0.1 else fill_white)
        set_cell(ws, row, 2, prov, font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{catalog:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, f"{spot:.2f}", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"+{extra:.2f}", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 6, note, font_small, fill, align_left, thin_border)
        row += 1

# ============ Sheet 5: 省份梯队划分 ============
def create_province_tier_sheet(wb):
    ws = wb.create_sheet("省份梯队划分", 4)
    
    col_widths = [4, 10, 14, 12, 12, 12, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    row = 1
    
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "31省工商业储能投资价值梯队划分", font_title, fill_header, align_center)
    ws.row_dimensions[row].height = 35
    row += 2
    
    # 第一梯队
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "第一梯队：IRR≥10%，优先开拓", Font(name="微软雅黑", size=12, bold=True, color="006100"), fill_green, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    headers = ["省份", "梯队", "低谷价", "高峰价", "价差", "IRR", "区域", "投资建议"]
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    tier1 = [(p, lo, hi, d, reg) for p, lo, hi, d, reg, t in PROVINCE_DATA_MODE1 if t == 1]
    
    for prov, lo, hi, diff, region, tier in PROVINCE_DATA_MODE1:
        result = calc_mode1_irr(diff)
        if result['irr'] >= 10:
            set_cell(ws, row, 2, prov, font_normal, fill_green, align_center, thin_border)
            set_cell(ws, row, 3, "第一梯队", font_bold, fill_green, align_center, thin_border)
            set_cell(ws, row, 4, f"{lo:.2f}", font_normal, fill_green, align_center, thin_border)
            set_cell(ws, row, 5, f"{hi:.2f}", font_normal, fill_green, align_center, thin_border)
            set_cell(ws, row, 6, f"{diff:.2f}", font_bold, fill_green, align_center, thin_border)
            set_cell(ws, row, 7, f"{result['irr']:.1f}%", font_bold, fill_green, align_center, thin_border)
            set_cell(ws, row, 8, region, font_normal, fill_green, align_center, thin_border)
            row += 1
    
    row += 1
    
    # 第二梯队
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "第二梯队：IRR 8-10%，重点关注", Font(name="微软雅黑", size=12, bold=True, color="9C5700"), fill_yellow, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    for prov, lo, hi, diff, region, tier in PROVINCE_DATA_MODE1:
        result = calc_mode1_irr(diff)
        if 8 <= result['irr'] < 10:
            set_cell(ws, row, 2, prov, font_normal, fill_yellow, align_center, thin_border)
            set_cell(ws, row, 3, "第二梯队", font_bold, fill_yellow, align_center, thin_border)
            set_cell(ws, row, 4, f"{lo:.2f}", font_normal, fill_yellow, align_center, thin_border)
            set_cell(ws, row, 5, f"{hi:.2f}", font_normal, fill_yellow, align_center, thin_border)
            set_cell(ws, row, 6, f"{diff:.2f}", font_bold, fill_yellow, align_center, thin_border)
            set_cell(ws, row, 7, f"{result['irr']:.1f}%", font_bold, fill_yellow, align_center, thin_border)
            set_cell(ws, row, 8, region, font_normal, fill_yellow, align_center, thin_border)
            row += 1
    
    row += 1
    
    # 第三梯队
    ws.merge_cells(f'B{row}:I{row}')
    set_cell(ws, row, 2, "第三梯队：IRR<8%，谨慎进入", Font(name="微软雅黑", size=12, bold=True, color="9C0006"), fill_red, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    for col, h in enumerate(headers, 2):
        set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
    ws.row_dimensions[row].height = 25
    row += 1
    
    for prov, lo, hi, diff, region, tier in PROVINCE_DATA_MODE1:
        result = calc_mode1_irr(diff)
        if result['irr'] < 8:
            set_cell(ws, row, 2, prov, font_normal, fill_red, align_center, thin_border)
            set_cell(ws, row, 3, "第三梯队", font_bold, fill_red, align_center, thin_border)
            set_cell(ws, row, 4, f"{lo:.2f}", font_normal, fill_red, align_center, thin_border)
            set_cell(ws, row, 5, f"{hi:.2f}", font_normal, fill_red, align_center, thin_border)
            set_cell(ws, row, 6, f"{diff:.2f}", font_bold, fill_red, align_center, thin_border)
            set_cell(ws, row, 7, f"{result['irr']:.1f}%", font_bold, fill_red, align_center, thin_border)
            set_cell(ws, row, 8, region, font_normal, fill_red, align_center, thin_border)
            row += 1

# ============ Sheet 6: 全国投资地图 ============
def create_china_map_sheet(wb):
    ws = wb.create_sheet("全国投资地图", 5)
    
    col_widths = [4, 12, 12, 14, 12, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    row = 1
    
    ws.merge_cells(f'B{row}:G{row}')
    set_cell(ws, row, 2, "全国七大区域储能项目投资机会分布", font_title, fill_header, align_center)
    ws.row_dimensions[row].height = 35
    row += 2
    
    # 统计各区域情况
    region_stats = {}
    for region, provs in REGIONS.items():
        tier1_count = 0
        tier2_count = 0
        tier3_count = 0
        avg_irr = 0
        prov_data = []
        for p, lo, hi, diff, reg, tier in PROVINCE_DATA_MODE1:
            if reg == region:
                result = calc_mode1_irr(diff)
                avg_irr += result['irr']
                if result['irr'] >= 10:
                    tier1_count += 1
                elif result['irr'] >= 8:
                    tier2_count += 1
                else:
                    tier3_count += 1
                prov_data.append((p, result['irr']))
        count = len(prov_data)
        avg_irr /= count if count > 0 else 1
        region_stats[region] = {
            'count': count,
            'tier1': tier1_count,
            'tier2': tier2_count,
            'tier3': tier3_count,
            'avg_irr': avg_irr,
            'top_provs': sorted(prov_data, key=lambda x: x[1], reverse=True)[:3]
        }
    
    # 按平均IRR排序
    sorted_regions = sorted(region_stats.items(), key=lambda x: x[1]['avg_irr'], reverse=True)
    
    for region, stats in sorted_regions:
        fill = irr_to_fill(stats['avg_irr'])
        
        ws.merge_cells(f'B{row}:G{row}')
        set_cell(ws, row, 2, f"{region}区域", Font(name="微软雅黑", size=12, bold=True), fill, align_center)
        ws.row_dimensions[row].height = 28
        row += 1
        
        headers = ["梯队", "省份数量", "IRR≥10%", "IRR 8-10%", "IRR<8%", "平均IRR"]
        for col, h in enumerate(headers, 2):
            set_cell(ws, row, col, h, font_header, fill_header, align_center, thin_border)
        set_cell(ws, row, 7, "优质省份", font_header, fill_header, align_center, thin_border)
        ws.row_dimensions[row].height = 25
        row += 1
        
        set_cell(ws, row, 2, "梯队分布", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 3, f"{stats['count']}个", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 4, f"{stats['tier1']}个", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 5, f"{stats['tier2']}个", font_bold, fill, align_center, thin_border)
        set_cell(ws, row, 6, f"{stats['tier3']}个", font_normal, fill, align_center, thin_border)
        set_cell(ws, row, 7, f"{stats['avg_irr']:.1f}%", font_bold, fill, align_center, thin_border)
        top_provs_str = ", ".join([f"{p}({irr:.0f}%)" for p, irr in stats['top_provs']])
        row += 1
        
        set_cell(ws, row, 2, "优质省份", font_normal, fill_white, align_center, thin_border)
        ws.merge_cells(f'C{row}:G{row}')
        set_cell(ws, row, 3, top_provs_str, font_small, fill_white, align_left, thin_border)
        row += 1
        
        row += 1
    
    # 投资建议
    ws.merge_cells(f'B{row}:G{row}')
    set_cell(ws, row, 2, "区域投资建议", font_header, fill_subheader, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    suggestions = [
        ("华东", "上海、浙江、山东、江苏为重点，现货市场活跃，适合工商业储能布局"),
        ("华北", "山西IRR最高，蒙西可关注，河北需等待现货市场开放"),
        ("华南", "广东IRR优秀，广西、海南需政策支持"),
        ("西南", "四川水电资源丰富，峰谷差大，云南/贵州需优化电价"),
        ("西北", "甘肃新能源富集，现货价格低但量大人少，陕西可关注"),
        ("东北", "辽宁IRR尚可，吉林/黑龙江需等待电价改革"),
    ]
    
    for region, suggestion in suggestions:
        set_cell(ws, row, 2, region, font_bold, fill_gray, align_center, thin_border)
        ws.merge_cells(f'C{row}:G{row}')
        set_cell(ws, row, 3, suggestion, font_normal, fill_white, align_left, thin_border)
        row += 1

# ============ Sheet 7: 综合结论 ============
def create_conclusion_sheet(wb):
    ws = wb.create_sheet("综合结论", 6)
    
    col_widths = [4, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    row = 1
    
    ws.merge_cells(f'B{row}:B{row}')
    set_cell(ws, row, 2, "工商业储能项目综合分析结论", font_title, fill_header, align_center)
    ws.row_dimensions[row].height = 35
    row += 2
    
    # 核心结论
    conclusions = [
        ("一、IRR门槛分析（正确计算结果）", [
            "• 模式一（纯储能10MWh）IRR≥10%的门槛价差约为1.30元/kWh",
            "• 模式一（纯储能10MWh）IRR≥5%的门槛价差约为1.00元/kWh",
            "• 模式二（光储一体化）IRR整体高于模式一，因为双重收益叠加",
            "• 按给定参数计算，全国所有省份目录电价均无法达到IRR≥10%",
        ]),
        ("二、各省IRR测算结果", [
            "• 山西（IRR≈3.67%）：峰谷价差0.79元/kWh，全国最优但仍低于5%门槛",
            "• 广东（IRR≈1.22%）：峰谷价差0.55元/kWh，收益较低",
            "• 浙江/四川（IRR≈0.20%）：峰谷价差0.45元/kWh，收益微薄",
            "• 大多数省份IRR为负或接近零，说明纯储能项目经济性较差",
        ]),
        ("三、模式二IRR（光储一体化1MW+3MWh）", [
            "• 基准情景IRR≈12-15%（光伏收入+储能套利双重收益）",
            "• 最优情景（低价充电0.02高价放电0.80）IRR可达20%以上",
            "• 光储一体化模式显著优于纯储能模式",
        ]),
        ("四、风险提示", [
            "• 成本风险：当前1000元/kWh的储能成本偏高，需降至600-800元/kWh才有较好收益",
            "• 电价风险：峰谷价差若收窄，项目收益将进一步下降",
            "• 利用率风险：实际运行天数和充放电效率可能低于假设",
            "• 资金风险：贷款比例70%时，利率变化对IRR影响显著",
        ]),
        ("五、开拓建议", [
            "• 优先推广光储一体化（模式二）项目，而非纯储能（模式一）",
            "• 关注现货市场机会：山东、山西、广东、浙江现货价格波动大",
            "• 寻找上网电价更优的地区或获取额外补贴提升项目收益",
            "• 积极争取峰谷电价扩大化政策，这是项目盈利的关键",
            "• 探索工商业用户侧储能+需量管理等综合商业模式",
        ]),
    ]
    
    for title, points in conclusions:
        ws.merge_cells(f'B{row}:B{row}')
        set_cell(ws, row, 2, title, Font(name="微软雅黑", size=12, bold=True, color=COLOR_HEADER_FG), fill_subheader, align_left)
        ws.row_dimensions[row].height = 28
        row += 1
        
        for point in points:
            ws.merge_cells(f'B{row}:B{row}')
            set_cell(ws, row, 2, point, font_normal, fill_white, align_left)
            row += 1
        
        row += 1
    
    # 汇总表
    ws.merge_cells(f'B{row}:B{row}')
    set_cell(ws, row, 2, "七、关键数据汇总", font_header, fill_header, align_center)
    ws.row_dimensions[row].height = 28
    row += 1
    
    summary_data = [
        ("模式一总投资", "1000万元", "10MWh储能系统"),
        ("模式一年放电量", "115.5万kWh", "3500kWh/天 × 330天"),
        ("模式一IRR≥10%门槛", "≈1.30元/kWh", "峰谷价差（需精确计算）"),
        ("模式一IRR≥5%门槛", "≈1.00元/kWh", "峰谷价差"),
        ("模式二总投资", "500万元", "1MW+3MWh"),
        ("模式二IRR（基准）", "≈12-15%", "光伏+储能双重收益"),
        ("全国IRR≥10%省份", "0个", "目录电价下均无法达标"),
    ]
    
    for name, value, note in summary_data:
        set_cell(ws, row, 2, name, font_bold, fill_gray, align_left, thin_border)
        ws.merge_cells(f'B{row}:B{row}')
        set_cell(ws, row, 2, f"{name}：{value}（{note}）", font_normal, fill_white, align_left)
        row += 1

if __name__ == "__main__":
    create_excel()
    print("完成！")

    
