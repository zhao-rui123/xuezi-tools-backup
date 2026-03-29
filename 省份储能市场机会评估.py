import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "省份储能市场机会评估"

# 标题
headers = ["省份名", "优先级", "市场化价差(元/kWh)", "光伏红线", "电力外送型", "综合评分", "备注"]

# 31省份数据
data = [
    ["山东", "A", 0.72, "否", "否", 95, "峰谷价差大,用户侧储能盈利强"],
    ["广东", "A", 0.68, "否", "否", 92, "工商业发达,储能需求旺盛"],
    ["浙江", "A", 0.65, "否", "否", 90, "峰谷电价差显著,补贴政策好"],
    ["上海", "A", 0.62, "否", "否", 88, "峰谷价差大,电网侧储能空间大"],
    ["江苏", "A", 0.58, "否", "否", 86, "工商业储能装机量大"],
    ["河南", "A", 0.55, "否", "否", 84, "新能源配储需求高"],
    ["安徽", "A", 0.52, "否", "否", 82, "峰谷电价差持续拉大"],
    ["江西", "A", 0.50, "否", "否", 80, "锂电产业基础好"],
    ["青海", "B", 0.42, "是", "是", 75, "光伏红线省份,外送为主"],
    ["甘肃", "B", 0.40, "是", "是", 72, "风电光伏装机大,弃电严重"],
    ["宁夏", "B", 0.38, "是", "是", 70, "新能源外送基地"],
    ["山西", "B", 0.45, "是", "是", 68, "火电调峰需求大"],
    ["内蒙古", "B", 0.35, "是", "是", 65, "风光储一体化项目多"],
    ["新疆", "B", 0.30, "是", "是", 62, "新能源外送通道建设"],
    ["河北", "B", 0.48, "是", "否", 60, "光伏装机量大,消纳压力大"],
    ["吉林", "B", 0.32, "是", "是", 58, "风电消纳问题突出"],
    ["黑龙江", "B", 0.28, "是", "是", 55, "冬季调峰需求大"],
    ["辽宁", "B", 0.35, "是", "否", 52, "工业转型期"],
    ["北京", "C", 0.55, "否", "否", 78, "虚拟电厂政策支持"],
    ["天津", "C", 0.50, "否", "否", 75, "港口储能示范项目"],
    ["福建", "C", 0.48, "否", "否", 72, "海风资源好"],
    ["湖北", "C", 0.45, "否", "否", 70, "水电储能潜力大"],
    ["湖南", "C", 0.42, "否", "否", 68, "电力辅助服务市场"],
    ["四川", "C", 0.40, "否", "否", 65, "水电调节需求"],
    ["重庆", "C", 0.38, "否", "否", 62, "山地城市特性"],
    ["陕西", "C", 0.36, "否", "否", 60, "新能源快速发展"],
    ["贵州", "C", 0.34, "否", "否", 58, "煤炭转型期"],
    ["云南", "C", 0.32, "否", "否", 55, "水电为主"],
    ["广西", "C", 0.30, "否", "否", 52, "新能源增长快"],
    ["海南", "C", 0.38, "否", "否", 50, "岛式储能示范"],
    ["西藏", "C", 0.25, "否", "否", 45, "成本高,规模小"],
]

# 样式
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
odd_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 写标题行
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# 写数据行
for row_idx, row_data in enumerate(data, 2):
    fill = odd_fill if row_idx % 2 == 1 else even_fill
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center_align

# 列宽自适应
for col in range(1, len(headers) + 1):
    max_length = len(headers[col - 1])
    for row in range(2, len(data) + 2):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    adjusted_width = min(max_length + 4, 30)
    ws.column_dimensions[get_column_letter(col)].width = adjusted_width

# 自动筛选
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

# 冻结首行
ws.freeze_panes = "A2"

wb.save("/Users/zhaoruicn/.openclaw/workspace/省份储能市场机会评估.xlsx")
print("Excel文件已生成: 省份储能市场机会评估.xlsx")
