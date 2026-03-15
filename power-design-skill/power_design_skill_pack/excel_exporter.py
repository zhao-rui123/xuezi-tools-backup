
# ==================== Excel导出模块 ====================
# 用于导出计算结果到Excel文件

import os
from datetime import datetime
from typing import Dict, List, Any, Optional

class ExcelExporter:
    """
    Excel导出器
    用于将计算结果导出为Excel文件
    """

    def __init__(self, output_dir: str = None):
        """
        初始化Excel导出器

        Args:
            output_dir: 输出目录，默认为当前目录
        """
        self.output_dir = output_dir or os.getcwd()
        os.makedirs(self.output_dir, exist_ok=True)

    def export_load_calculation(
        self,
        results: List[Dict[str, Any]],
        total_result: Any = None,
        filename: str = None
    ) -> str:
        """
        导出负荷计算结果

        Args:
            results: 各组负荷计算结果
            total_result: 总负荷结果
            filename: 输出文件名

        Returns:
            输出文件路径
        """
        if filename is None:
            filename = f"负荷计算结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "负荷计算"

            # 设置标题
            ws['A1'] = '负荷计算结果'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            ws['A1'].alignment = Alignment(horizontal='center')

            # 设置表头
            headers = ['设备名称', '有功功率(kW)', '无功功率(kvar)', '视在功率(kVA)', '计算电流(A)', '功率因数']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center')

            # 填充数据
            row = 4
            for result in results:
                ws.cell(row=row, column=1, value=result.get('name', ''))
                ws.cell(row=row, column=2, value=round(result['result'].active_power, 2))
                ws.cell(row=row, column=3, value=round(result['result'].reactive_power, 2))
                ws.cell(row=row, column=4, value=round(result['result'].apparent_power, 2))
                ws.cell(row=row, column=5, value=round(result['result'].calculation_current, 2))
                ws.cell(row=row, column=6, value=round(result['result'].power_factor, 2))
                row += 1

            # 添加总计行
            if total_result:
                ws.cell(row=row, column=1, value='总计')
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2, value=round(total_result.active_power, 2))
                ws.cell(row=row, column=3, value=round(total_result.reactive_power, 2))
                ws.cell(row=row, column=4, value=round(total_result.apparent_power, 2))
                ws.cell(row=row, column=5, value=round(total_result.calculation_current, 2))
                ws.cell(row=row, column=6, value=round(total_result.power_factor, 2))

                for col in range(1, 7):
                    ws.cell(row=row, column=col).font = Font(bold=True)
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

            # 调整列宽
            column_widths = [20, 15, 15, 15, 15, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            wb.save(filepath)
            return filepath

        except ImportError:
            # 如果没有openpyxl，使用CSV格式
            return self._export_to_csv(results, total_result, filename.replace('.xlsx', '.csv'))

    def export_short_circuit(
        self,
        result: Any,
        filename: str = None
    ) -> str:
        """
        导出短路电流计算结果

        Args:
            result: 短路电流计算结果
            filename: 输出文件名

        Returns:
            输出文件路径
        """
        if filename is None:
            filename = f"短路电流计算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "短路电流"

            # 标题
            ws['A1'] = '短路电流计算结果'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = Alignment(horizontal='center')

            # 数据
            data = [
                ['三相短路电流(kA)', round(result.three_phase_current, 2)],
                ['两相短路电流(kA)', round(result.two_phase_current, 2)],
                ['单相短路电流(kA)', round(result.single_phase_current, 2)],
                ['短路容量(MVA)', round(result.short_circuit_capacity, 2)],
            ]

            for row_idx, (label, value) in enumerate(data, 3):
                ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row_idx, column=2, value=value)
                ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                ws.cell(row=row_idx, column=1).font = Font(bold=True, color='FFFFFF')

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20

            wb.save(filepath)
            return filepath

        except ImportError:
            return self._export_simple_csv(data, filename.replace('.xlsx', '.csv'))

    def export_cable_selection(
        self,
        results: List[Dict[str, Any]],
        filename: str = None
    ) -> str:
        """
        导出电缆选择结果

        Args:
            results: 电缆选择结果列表
            filename: 输出文件名

        Returns:
            输出文件路径
        """
        if filename is None:
            filename = f"电缆选择结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "电缆选择"

            # 标题
            ws['A1'] = '电缆截面选择结果'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center')

            # 表头
            headers = ['设备名称', '电缆截面(mm²)', '载流量(A)', '电压降(%)', '热稳定校验']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # 数据
            for row_idx, result in enumerate(results, 4):
                cable = result['cable']
                ws.cell(row=row_idx, column=1, value=result.get('name', ''))
                ws.cell(row=row_idx, column=2, value=cable.cross_section)
                ws.cell(row=row_idx, column=3, value=round(cable.current_carrying_capacity, 1))
                ws.cell(row=row_idx, column=4, value=round(cable.voltage_drop, 2))
                ws.cell(row=row_idx, column=5, value='通过' if cable.thermal_stability else '不通过')

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15

            wb.save(filepath)
            return filepath

        except ImportError:
            return self._export_cable_csv(results, filename.replace('.xlsx', '.csv'))

    def export_protection_settings(
        self,
        results: List[Dict[str, Any]],
        filename: str = None
    ) -> str:
        """
        导出保护整定结果

        Args:
            results: 保护整定结果列表
            filename: 输出文件名

        Returns:
            输出文件路径
        """
        if filename is None:
            filename = f"保护整定结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "保护整定"

            # 标题
            ws['A1'] = '继电保护整定结果'
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center')

            # 表头
            headers = ['设备名称', '速断电流(A)', '过流定值(A)', '延时时间(s)', '灵敏度']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # 数据
            for row_idx, result in enumerate(results, 4):
                protection = result['protection']
                ws.cell(row=row_idx, column=1, value=result.get('name', ''))
                ws.cell(row=row_idx, column=2, value=round(protection.instantaneous_current, 1))
                ws.cell(row=row_idx, column=3, value=round(protection.time_delay_current, 1))
                ws.cell(row=row_idx, column=4, value=round(protection.time_delay, 1))
                ws.cell(row=row_idx, column=5, value=round(protection.sensitivity, 2))

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12

            wb.save(filepath)
            return filepath

        except ImportError:
            return self._export_protection_csv(results, filename.replace('.xlsx', '.csv'))

    def export_full_report(
        self,
        calculator_results: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        导出完整计算报告

        Args:
            calculator_results: 综合计算器结果
            filename: 输出文件名

        Returns:
            输出文件路径
        """
        if filename is None:
            filename = f"电力设计计算报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()

            # 1. 负荷计算工作表
            ws1 = wb.active
            ws1.title = "负荷计算"

            ws1['A1'] = '电力系统设计计算报告'
            ws1['A1'].font = Font(size=18, bold=True)
            ws1.merge_cells('A1:F1')
            ws1['A1'].alignment = Alignment(horizontal='center')

            ws1['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws1.merge_cells('A2:F2')
            ws1['A2'].alignment = Alignment(horizontal='center')

            # 负荷计算表头
            headers = ['设备名称', '有功功率(kW)', '无功功率(kvar)', '视在功率(kVA)', '计算电流(A)', '功率因数']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # 负荷数据
            row = 5
            for load in calculator_results.get('loads', []):
                ws1.cell(row=row, column=1, value=load.get('name', ''))
                ws1.cell(row=row, column=2, value=round(load['result'].active_power, 2))
                ws1.cell(row=row, column=3, value=round(load['result'].reactive_power, 2))
                ws1.cell(row=row, column=4, value=round(load['result'].apparent_power, 2))
                ws1.cell(row=row, column=5, value=round(load['result'].calculation_current, 2))
                ws1.cell(row=row, column=6, value=round(load['result'].power_factor, 2))
                row += 1

            # 总负荷
            total = calculator_results.get('total_load')
            if total:
                ws1.cell(row=row, column=1, value='总计')
                ws1.cell(row=row, column=1).font = Font(bold=True)
                ws1.cell(row=row, column=2, value=round(total.active_power, 2))
                ws1.cell(row=row, column=3, value=round(total.reactive_power, 2))
                ws1.cell(row=row, column=4, value=round(total.apparent_power, 2))
                ws1.cell(row=row, column=5, value=round(total.calculation_current, 2))
                ws1.cell(row=row, column=6, value=round(total.power_factor, 2))
                for col in range(1, 7):
                    ws1.cell(row=row, column=col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    ws1.cell(row=row, column=col).font = Font(bold=True)

            # 2. 短路电流工作表
            ws2 = wb.create_sheet("短路电流")
            ws2['A1'] = '短路电流计算结果'
            ws2['A1'].font = Font(size=16, bold=True)
            ws2.merge_cells('A1:B1')
            ws2['A1'].alignment = Alignment(horizontal='center')

            short = calculator_results.get('short_circuit')
            if short:
                data = [
                    ['三相短路电流(kA)', round(short.three_phase_current, 2)],
                    ['两相短路电流(kA)', round(short.two_phase_current, 2)],
                    ['单相短路电流(kA)', round(short.single_phase_current, 2)],
                    ['短路容量(MVA)', round(short.short_circuit_capacity, 2)],
                ]
                for row_idx, (label, value) in enumerate(data, 3):
                    ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                    ws2.cell(row=row_idx, column=2, value=value)
                    ws2.cell(row=row_idx, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    ws2.cell(row=row_idx, column=1).font = Font(bold=True, color='FFFFFF')

            # 3. 变压器校验工作表
            ws3 = wb.create_sheet("变压器校验")
            ws3['A1'] = '变压器校验结果'
            ws3['A1'].font = Font(size=16, bold=True)
            ws3.merge_cells('A1:B1')
            ws3['A1'].alignment = Alignment(horizontal='center')

            trans_check = calculator_results.get('transformer_check')
            if trans_check:
                data = [
                    ['负载率(%)', round(trans_check['load_rate'], 1)],
                    ['是否满足要求', '是' if trans_check['is_suitable'] else '否'],
                ]
                for row_idx, (label, value) in enumerate(data, 3):
                    ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                    ws3.cell(row=row_idx, column=2, value=value)
                    ws3.cell(row=row_idx, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    ws3.cell(row=row_idx, column=1).font = Font(bold=True, color='FFFFFF')

            # 4. 电缆选择工作表
            ws4 = wb.create_sheet("电缆选择")
            ws4['A1'] = '电缆截面选择结果'
            ws4['A1'].font = Font(size=16, bold=True)
            ws4.merge_cells('A1:E1')
            ws4['A1'].alignment = Alignment(horizontal='center')

            headers = ['设备名称', '电缆截面(mm²)', '载流量(A)', '电压降(%)', '热稳定校验']
            for col, header in enumerate(headers, 1):
                cell = ws4.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            for row_idx, cable in enumerate(calculator_results.get('cables', []), 4):
                ws4.cell(row=row_idx, column=1, value=cable.get('name', ''))
                ws4.cell(row=row_idx, column=2, value=cable['cable'].cross_section)
                ws4.cell(row=row_idx, column=3, value=round(cable['cable'].current_carrying_capacity, 1))
                ws4.cell(row=row_idx, column=4, value=round(cable['cable'].voltage_drop, 2))
                ws4.cell(row=row_idx, column=5, value='通过' if cable['cable'].thermal_stability else '不通过')

            # 5. 保护整定工作表
            ws5 = wb.create_sheet("保护整定")
            ws5['A1'] = '继电保护整定结果'
            ws5['A1'].font = Font(size=16, bold=True)
            ws5.merge_cells('A1:E1')
            ws5['A1'].alignment = Alignment(horizontal='center')

            headers = ['设备名称', '速断电流(A)', '过流定值(A)', '延时时间(s)', '灵敏度']
            for col, header in enumerate(headers, 1):
                cell = ws5.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            for row_idx, protection in enumerate(calculator_results.get('protections', []), 4):
                ws5.cell(row=row_idx, column=1, value=protection.get('name', ''))
                ws5.cell(row=row_idx, column=2, value=round(protection['protection'].instantaneous_current, 1))
                ws5.cell(row=row_idx, column=3, value=round(protection['protection'].time_delay_current, 1))
                ws5.cell(row=row_idx, column=4, value=round(protection['protection'].time_delay, 1))
                ws5.cell(row=row_idx, column=5, value=round(protection['protection'].sensitivity, 2))

            # 调整所有工作表的列宽
            for ws in [ws1, ws2, ws3, ws4, ws5]:
                for col in range(1, 7):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

            wb.save(filepath)
            return filepath

        except ImportError:
            # 如果没有openpyxl，返回CSV格式
            return self._export_full_csv(calculator_results, filename.replace('.xlsx', '.csv'))

    def _export_to_csv(self, results, total_result, filename):
        """导出到CSV格式"""
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            import csv
            writer = csv.writer(f)
            writer.writerow(['负荷计算结果'])
            writer.writerow([])
            writer.writerow(['设备名称', '有功功率(kW)', '无功功率(kvar)', '视在功率(kVA)', '计算电流(A)', '功率因数'])

            for result in results:
                writer.writerow([
                    result.get('name', ''),
                    round(result['result'].active_power, 2),
                    round(result['result'].reactive_power, 2),
                    round(result['result'].apparent_power, 2),
                    round(result['result'].calculation_current, 2),
                    round(result['result'].power_factor, 2)
                ])

            if total_result:
                writer.writerow([
                    '总计',
                    round(total_result.active_power, 2),
                    round(total_result.reactive_power, 2),
                    round(total_result.apparent_power, 2),
                    round(total_result.calculation_current, 2),
                    round(total_result.power_factor, 2)
                ])

        return filepath

    def _export_simple_csv(self, data, filename):
        """导出简单CSV"""
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            import csv
            writer = csv.writer(f)
            writer.writerow(['短路电流计算结果'])
            writer.writerow([])
            for label, value in data:
                writer.writerow([label, value])
        return filepath

    def _export_cable_csv(self, results, filename):
        """导出电缆CSV"""
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            import csv
            writer = csv.writer(f)
            writer.writerow(['电缆截面选择结果'])
            writer.writerow([])
            writer.writerow(['设备名称', '电缆截面(mm²)', '载流量(A)', '电压降(%)', '热稳定校验'])

            for result in results:
                cable = result['cable']
                writer.writerow([
                    result.get('name', ''),
                    cable.cross_section,
                    round(cable.current_carrying_capacity, 1),
                    round(cable.voltage_drop, 2),
                    '通过' if cable.thermal_stability else '不通过'
                ])
        return filepath

    def _export_protection_csv(self, results, filename):
        """导出保护CSV"""
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            import csv
            writer = csv.writer(f)
            writer.writerow(['继电保护整定结果'])
            writer.writerow([])
            writer.writerow(['设备名称', '速断电流(A)', '过流定值(A)', '延时时间(s)', '灵敏度'])

            for result in results:
                protection = result['protection']
                writer.writerow([
                    result.get('name', ''),
                    round(protection.instantaneous_current, 1),
                    round(protection.time_delay_current, 1),
                    round(protection.time_delay, 1),
                    round(protection.sensitivity, 2)
                ])
        return filepath

    def _export_full_csv(self, calculator_results, filename):
        """导出完整CSV"""
        filepath = os.path.join(self.output_dir, filename)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            import csv
            writer = csv.writer(f)

            writer.writerow(['电力系统设计计算报告'])
            writer.writerow([f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
            writer.writerow([])

            # 负荷计算
            writer.writerow(['一、负荷计算结果'])
            writer.writerow(['设备名称', '有功功率(kW)', '无功功率(kvar)', '视在功率(kVA)', '计算电流(A)', '功率因数'])

            for load in calculator_results.get('loads', []):
                writer.writerow([
                    load.get('name', ''),
                    round(load['result'].active_power, 2),
                    round(load['result'].reactive_power, 2),
                    round(load['result'].apparent_power, 2),
                    round(load['result'].calculation_current, 2),
                    round(load['result'].power_factor, 2)
                ])

            total = calculator_results.get('total_load')
            if total:
                writer.writerow([
                    '总计',
                    round(total.active_power, 2),
                    round(total.reactive_power, 2),
                    round(total.apparent_power, 2),
                    round(total.calculation_current, 2),
                    round(total.power_factor, 2)
                ])

            writer.writerow([])

            # 短路电流
            writer.writerow(['二、短路电流计算结果'])
            short = calculator_results.get('short_circuit')
            if short:
                writer.writerow(['三相短路电流(kA)', round(short.three_phase_current, 2)])
                writer.writerow(['两相短路电流(kA)', round(short.two_phase_current, 2)])
                writer.writerow(['单相短路电流(kA)', round(short.single_phase_current, 2)])
                writer.writerow(['短路容量(MVA)', round(short.short_circuit_capacity, 2)])

            writer.writerow([])

            # 变压器校验
            writer.writerow(['三、变压器校验'])
            trans_check = calculator_results.get('transformer_check')
            if trans_check:
                writer.writerow(['负载率(%)', round(trans_check['load_rate'], 1)])
                writer.writerow(['是否满足要求', '是' if trans_check['is_suitable'] else '否'])

        return filepath


# 便捷函数
def export_to_excel(
    calculator_results: Dict[str, Any],
    output_dir: str = None,
    filename: str = None
) -> str:
    """
    便捷函数：导出计算结果到Excel

    Args:
        calculator_results: 综合计算器结果
        output_dir: 输出目录
        filename: 文件名

    Returns:
        输出文件路径
    """
    exporter = ExcelExporter(output_dir)
    return exporter.export_full_report(calculator_results, filename)

