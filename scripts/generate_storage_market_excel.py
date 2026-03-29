#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
储能市场机会评估Excel生成器
基于Claude深度思考架构设计
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ==================== 样式定义 ====================
title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
normal_font = Font(name='微软雅黑', size=9)
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
grade1_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')  # 金色-第一梯队
grade2_fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')  # 绿色-第二梯队
grade3_fill = PatternFill(start_color='9DC3E6', end_color='9DC3E6', fill_type='solid')  # 蓝色-第三梯队
grade4_fill = PatternFill(start_color='B4A7D6', end_color='B4A7D6', fill_type='solid')  # 紫色-第四梯队
watch_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # 灰色-观望
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def get_grade_fill(grade):
    grade_map = {
        '第一梯队': grade1_fill,
        '第二梯队': grade2_fill,
        '第三梯队': grade3_fill,
        '第四梯队': grade4_fill,
        '观望': watch_fill
    }
    return grade_map.get(grade, None)

# ==================== 省份完整数据 ====================
# 省份数据：(省份, 区域, 梯队, 价差, 光伏红线, 光伏装机MW, 外送, 综合评分, 机会分析)
provinces_data = [
    # 第一梯队
    ('山东', '华东', '第一梯队', 0.82, '部分红区', 7613, '否', 95, '价差大+光伏红线严格+分布式最强省'),
    ('河南', '华中', '第一梯队', 0.72, '部分红区', 4100, '否', 88, '分布式光伏强+工商业价差大'),
    ('河北', '华北', '第一梯队', 0.65, '部分红区', 7202, '是', 85, '光伏装机第二大省+部分红区'),
    ('山西', '华北', '第一梯队', 0.68, '是', 4100, '是', 82, '现货市场已运行+外送为主但价差可观'),
    # 第二梯队
    ('广东', '华南', '第二梯队', 0.75, '部分红区', 5200, '否', 80, '价差大+工商业发达+市场化程度高'),
    ('江苏', '华东', '第二梯队', 0.75, '否', 6165, '否', 78, '分布式光伏强省+工业用电量大'),
    ('浙江', '华东', '第二梯队', 0.73, '否', 4800, '否', 76, '工商业分布式发达+市场化推进中'),
    ('安徽', '华东', '第二梯队', 0.70, '否', 3200, '否', 72, '价差达标+工业基础好'),
    ('四川', '西南', '第二梯队', 0.72, '否', 890, '是', 70, '水电为主但价差较大+新能源快速发展'),
    # 第三梯队
    ('甘肃', '西北', '第三梯队', 0.58, '是', 3520, '是', 62, '现货已运行+外送为主价差受限'),
    ('福建', '华东', '第三梯队', 0.52, '部分暂停', 1600, '否', 60, '2025现货启动+价差达标'),
    ('湖北', '华中', '第三梯队', 0.62, '全省受限', 2100, '是', 58, '光伏受限需配储+外送为主'),
    ('江西', '华东', '第三梯队', 0.68, '否', 1800, '否', 65, '价差较大+新能源发展快'),
    # 第四梯队
    ('新疆', '西北', '第四梯队', 0.45, '是', 5675, '是', 45, '光伏装机大省+红线严格+外送为主'),
    ('青海', '西北', '第四梯队', 0.55, '是', 2530, '是', 42, '清洁能源94%+红线严格+外送为主'),
    ('宁夏', '西北', '第四梯队', 0.52, '是', 1560, '是', 40, '光伏红线+外送为主价差受限'),
    # 观望
    ('内蒙古', '华北', '观望', 0.50, '部分取消', 4811, '是', 35, '已取消强制配储+外送为主'),
    ('云南', '西南', '观望', 0.45, '是', 2850, '是', 30, '西电东送+强制配储10%'),
    ('陕西', '西北', '观望', 0.52, '否', 2800, '是', 38, '价差达标但外送为主'),
    ('贵州', '西南', '观望', 0.55, '否', 1100, '是', 35, '价差达标但外送为主'),
    ('广西', '华南', '观望', 0.48, '否', 890, '是', 32, '南电北送+外送为主'),
    ('海南', '华南', '观望', 0.62, '否', 380, '否', 55, '独立电网+峰谷价差大'),
    ('北京', '华北', '观望', 0.55, '否', 80, '否', 40, '光伏规模小+城市级市场'),
    ('天津', '华北', '观望', 0.60, '否', 280, '否', 50, '价差达标+港口工业'),
    ('上海', '华东', '观望', 1.00, '否', 350, '否', 60, '价差高+工商业电价高'),
    ('重庆', '西南', '观望', 0.60, '否', 520, '是', 45, '价差达标但外送为主'),
    ('黑龙江', '东北', '观望', 0.52, '已缓解', 1100, '是', 38, '红区已减少+东北市场'),
    ('吉林', '东北', '观望', 0.48, '是', 1200, '是', 35, '光伏红线+东北市场'),
    ('辽宁', '东北', '观望', 0.55, '部分红区', 1500, '是', 40, '现货已运行+外送为主'),
    ('西藏', '西南', '观望', '-', '否', 120, '是', 20, '规模小+外送为主'),
]

# 市场化数据
market_data = {
    '山东': ('高度', 0.52, 0.48, '已运行', 95),
    '河南': ('中度', 0.35, 0.30, '2025启动', 70),
    '河北': ('中度', 0.32, 0.28, '2025南网启动', 68),
    '山西': ('高度', 0.48, 0.42, '已运行', 92),
    '广东': ('高度', 0.55, 0.50, '已运行', 95),
    '江苏': ('中度', 0.38, 0.32, '2025启动', 72),
    '浙江': ('中度', 0.40, 0.35, '2025启动', 75),
    '安徽': ('低度', 0.30, 0.25, '否', 65),
    '四川': ('中度', 0.35, 0.30, '2025启动', 68),
    '甘肃': ('高度', 0.42, 0.38, '已运行', 88),
    '福建': ('中度', 0.28, 0.24, '2025启动', 62),
    '湖北': ('中度', 0.32, 0.28, '2025启动', 65),
    '江西': ('中度', 0.30, 0.26, '2025启动', 62),
    '新疆': ('中度', 0.25, 0.22, '2025启动', 55),
    '青海': ('中度', 0.28, 0.24, '2025启动', 52),
    '宁夏': ('中度', 0.26, 0.22, '2025启动', 50),
}

# 光伏红线政策数据
pv_policy_data = {
    '山东': ('是', '15-30%', '2-4h', '严格', 90),
    '河南': ('部分', '10-20%', '2h', '部分', 60),
    '河北': ('部分', '10-15%', '2h', '部分', 55),
    '山西': ('是', '10-20%', '2h', '严格', 75),
    '广东': ('部分', '无强制', '-', '已取消', 40),
    '江苏': ('否', '无强制', '-', '无要求', 30),
    '浙江': ('否', '无强制', '-', '无要求', 30),
    '安徽': ('否', '无强制', '-', '无要求', 30),
    '四川': ('否', '无强制', '-', '无要求', 30),
    '甘肃': ('是', '15-20%', '2h', '严格', 70),
    '福建': ('是', '部分暂停', '-', '调整中', 50),
    '湖北': ('是', '10-15%', '2h', '严格', 65),
    '江西': ('否', '无强制', '-', '无要求', 30),
    '新疆': ('是', '15-25%', '2h', '严格', 80),
    '青海': ('是', '20-25%', '2h', '严格', 78),
    '宁夏': ('是', '10-15%', '2h', '严格', 72),
}

# ==================== Sheet1: 省份总览 ====================
ws1 = wb.active
ws1.title = '省份总览'

# 标题
ws1.merge_cells('A1:I1')
set_cell(ws1, 1, 1, '省级储能市场机会评估总览（2025年）', 
         font=title_font, fill=title_fill, 
         alignment=Alignment(horizontal='center', vertical='center'))
ws1.row_dimensions[1].height = 35

# 表头
headers = ['省份', '区域', '梯队', '市场化价差\n(元/kWh)', '光伏红线', 
           '分布式光伏\n装机(MW)', '电力外送', '综合评分', '机会分析']
for col, h in enumerate(headers, 1):
    set_cell(ws1, 2, col, h, font=header_font, fill=header_fill,
             alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), 
             border=thin_border)
ws1.row_dimensions[2].height = 35

# 数据
for i, row_data in enumerate(provinces_data, 3):
    province, region, grade, spread, pv_limit, pv_install, export_type, score, analysis = row_data
    row_fill = get_grade_fill(grade)
    
    values = [province, region, grade, spread, pv_limit, pv_install, export_type, score, analysis]
    for col, val in enumerate(values, 1):
        cell = set_cell(ws1, i, col, val, font=normal_font, fill=row_fill, border=thin_border,
                       alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))
        if col == 4 and isinstance(val, (int, float)):
            cell.number_format = '0.00'
        elif col == 8:
            cell.number_format = '0'

# 列宽
col_widths = [10, 8, 10, 12, 12, 12, 10, 10, 45]
for i, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

ws1.freeze_panes = 'A3'
ws1.auto_filter.ref = f'A2:I{len(provinces_data)+2}'

# ==================== Sheet2: 市场化分析 ====================
ws2 = wb.create_sheet('市场化分析')

ws2.merge_cells('A1:F1')
set_cell(ws2, 1, 1, '各省电力市场化程度分析（2025年）', 
         font=title_font, fill=title_fill, 
         alignment=Alignment(horizontal='center', vertical='center'))
ws2.row_dimensions[1].height = 35

headers2 = ['省份', '市场化程度', '日前市场价差\n(元/kWh)', '实时市场价差\n(元/kWh)', '市场参与情况', '成熟度评分']
for col, h in enumerate(headers2, 1):
    set_cell(ws2, 2, col, h, font=header_font, fill=header_fill,
             alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), 
             border=thin_border)
ws2.row_dimensions[2].height = 35

row = 3
for province, data in market_data.items():
    market_level, day_ahead, real_time, status, maturity = data
    set_cell(ws2, row, 1, province, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws2, row, 2, market_level, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws2, row, 3, day_ahead, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'), number_format='0.00')
    set_cell(ws2, row, 4, real_time, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'), number_format='0.00')
    set_cell(ws2, row, 5, status, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws2, row, 6, maturity, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'), number_format='0')
    row += 1

col_widths2 = [12, 12, 15, 15, 20, 12]
for i, width in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

ws2.freeze_panes = 'A3'
ws2.auto_filter.ref = f'A2:F{row}'

# ==================== Sheet3: 光伏红线政策 ====================
ws3 = wb.create_sheet('光伏红线政策')

ws3.merge_cells('A1:E1')
set_cell(ws3, 1, 1, '各省光伏红线与配储政策（2025年）', 
         font=title_font, fill=title_fill, 
         alignment=Alignment(horizontal='center', vertical='center'))
ws3.row_dimensions[1].height = 35

headers3 = ['省份', '光伏红线', '配储比例要求', '配储时长要求', '政策状态']
for col, h in enumerate(headers3, 1):
    set_cell(ws3, 2, col, h, font=header_font, fill=header_fill,
             alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), 
             border=thin_border)
ws3.row_dimensions[2].height = 35

row = 3
for province, data in pv_policy_data.items():
    is_red, ratio, duration, status, policy_score = data
    set_cell(ws3, row, 1, province, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws3, row, 2, is_red, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws3, row, 3, ratio, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws3, row, 4, duration, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    set_cell(ws3, row, 5, status, font=normal_font, border=thin_border,
            alignment=Alignment(horizontal='center'))
    row += 1

col_widths3 = [12, 12, 15, 15, 15]
for i, width in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

ws3.freeze_panes = 'A3'
ws3.auto_filter.ref = f'A2:E{row}'

# ==================== Sheet4: 综合评分 ====================
ws4 = wb.create_sheet('综合评分')

ws4.merge_cells('A1:H1')
set_cell(ws4, 1, 1, '省级储能市场综合评分（满分100分）', 
         font=title_font, fill=title_fill, 
         alignment=Alignment(horizontal='center', vertical='center'))
ws4.row_dimensions[1].height = 35

headers4 = ['省份', '梯队', '价差得分\n(满分30)', '光伏红线得分\n(满分25)', 
            '政策支持得分\n(满分15)', '市场容量得分\n(满分10)', '综合得分', '排名']
for col, h in enumerate(headers4, 1):
    set_cell(ws4, 2, col, h, font=header_font, fill=header_fill,
             alignment=Alignment(horizontal='center', vertical='center', wrap_text=True), 
             border=thin_border)
ws4.row_dimensions[2].height = 40

# 计算评分
def calc_scores(province, data):
    """计算各维度得分"""
    _, region, grade, spread, pv_limit, pv_install, export_type, total_score, _ = data
    
    # 价差得分（满分30）
    if isinstance(spread, (int, float)):
        if spread >= 0.5: spread_score = 30
        elif spread >= 0.3: spread_score = 24
        elif spread >= 0.2: spread_score = 18
        else: spread_score = 12
    else:
        spread_score = 10
    
    # 光伏红线得分（满分25）
    if pv_limit == '是' or pv_limit == '严格':
        pv_score = 25
    elif pv_limit == '部分' or pv_limit == '部分红区':
        pv_score = 15
    elif pv_limit == '部分暂停':
        pv_score = 18
    else:
        pv_score = 5
    
    # 政策支持得分（满分15）- 简化处理
    if grade == '第一梯队':
        policy_score = 15
    elif grade == '第二梯队':
        policy_score = 12
    elif grade == '第三梯队':
        policy_score = 8
    else:
        policy_score = 5
    
    # 市场容量得分（满分10）
    if pv_install >= 5000: cap_score = 10
    elif pv_install >= 3000: cap_score = 8
    elif pv_install >= 1500: cap_score = 6
    elif pv_install >= 500: cap_score = 4
    else: cap_score = 2
    
    return spread_score, pv_score, policy_score, cap_score

# 按综合得分排序
sorted_provinces = sorted(provinces_data, key=lambda x: x[7], reverse=True)

row = 3
for rank, data in enumerate(sorted_provinces, 1):
    province = data[0]
    grade = data[2]
    row_fill = get_grade_fill(grade)
    spread_score, pv_score, policy_score, cap_score = calc_scores(province, data)
    total = spread_score + pv_score + policy_score + cap_score
    
    values = [province, grade, spread_score, pv_score, policy_score, cap_score, total, rank]
    for col, val in enumerate(values, 1):
        cell = set_cell(ws4, row, col, val, font=normal_font, fill=row_fill, border=thin_border,
                       alignment=Alignment(horizontal='center', vertical='center'))
        if col == 7:
            cell.number_format = '0'
    row += 1

col_widths4 = [12, 10, 12, 15, 15, 15, 12, 8]
for i, width in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = width

ws4.freeze_panes = 'A3'
ws4.auto_filter.ref = f'A2:H{row}'

# ==================== 保存 ====================
output_path = '/Users/zhaoruicn/.openclaw/workspace/储能市场机会评估_完整版.xlsx'
wb.save(output_path)
print(f"Excel已生成: {output_path}")
print(f"共 {len(provinces_data)} 个省份数据")
print(f"4个工作表：省份总览、市场经济分析、光伏红线政策、综合评分")
