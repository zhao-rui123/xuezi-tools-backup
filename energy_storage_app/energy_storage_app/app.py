# -*- coding: utf-8 -*-
"""
CALB工商业储能收资清单 - Web应用
支持在线查看、填写、下载Excel模板、导出填写数据
"""

from flask import Flask, render_template, render_template_string, request, send_file, redirect, url_for, flash, session
import pandas as pd
import openpyxl
from notify import send_feishu_notification  # 导入飞书通知
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
from datetime import datetime
import uuid

app = Flask(__name__)
app.secret_key = 'energy-storage-secret-key-2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('submissions', exist_ok=True)

# Excel模板路径
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'CALB-工商业储能收资清单.xlsx')
SUBMISSIONS_FILE = 'submissions/submissions.json'
TEMPLATE_OPTIONS = {
    'energy_storage': {
        'name': 'CALB工商业储能收资清单',
        'download_name': 'CALB-工商业储能收资清单.xlsx',
        'export_name': '工商业储能收资清单_导出数据.xlsx',
        'template_output': 'submissions/工商业储能收资清单_模板.xlsx',
        'template_title': 'CALB工商业储能收资清单 - 填写说明',
        'icon': '⚡',
    },
    'zero_carbon_park': {
        'name': '零碳园区收资清单',
        'download_name': '零碳园区收资清单.xlsx',
        'export_name': '零碳园区收资清单_导出数据.xlsx',
        'template_output': 'submissions/零碳园区收资清单_模板.xlsx',
        'template_title': '零碳园区收资清单 - 填写说明',
        'icon': '🌿',
    }
}


def load_submissions():
    """加载已提交的数据"""
    if os.path.exists(SUBMISSIONS_FILE):
        with open(SUBMISSIONS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_submission(data):
    """保存提交的数据"""
    submissions = load_submissions()
    data['id'] = str(uuid.uuid4())[:8]
    data['submit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    submissions.append(data)
    with open(SUBMISSIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(submissions, f, ensure_ascii=False, indent=2)
    return data['id']


def get_form_data(template_type='energy_storage'):
    """获取表单字段配置"""
    forms = {
        'energy_storage': {
        # 1、基本信息
        'company_name': {'label': '*公司名称', 'type': 'text', 'required': True, 'section': '1、基本信息'},
        'project_name': {'label': '*项目名称', 'type': 'text', 'required': True, 'section': '1、基本信息'},
        'project_location': {'label': '*项目地点', 'type': 'text', 'required': True, 'section': '1、基本信息'},
        'contact_person': {'label': '*对接人', 'type': 'text', 'required': True, 'section': '1、基本信息'},
        'contact_phone': {'label': '*联系方式', 'type': 'tel', 'required': True, 'section': '1、基本信息'},
        'industry': {'label': '项目客户行业', 'type': 'select', 'options': ['冶金', '有色', '机械', '轻工', '纺织', '烟草', '商贸', '其他'], 'section': '1、基本信息'},
        'vehicle_access': {'label': '*车辆吊装路径情况', 'type': 'textarea', 'required': True, 'section': '1、基本信息'},

        # 2、配电信息
        'power_type': {'label': '*企业用电性质', 'type': 'select', 'options': ['大工业', '一般工商业'], 'required': True, 'section': '2、配电信息'},
        'voltage_level': {'label': '*计量电压等级', 'type': 'select', 'options': ['35KV', '10KV', '380V', '220V'], 'required': True, 'section': '2、配电信息'},
        'transformer_capacity': {'label': '*主变容量(kVA)', 'type': 'text', 'required': True, 'section': '2、配电信息'},
        'transformer_count': {'label': '台数', 'type': 'number', 'section': '2、配电信息'},
        'basic_price': {'label': '*基本电价', 'type': 'select', 'options': ['需量 元/kVA×月', '容量 元/kVA×月'], 'required': True, 'section': '2、配电信息'},
        'transformer_status': {'label': '*变压器开备情况', 'type': 'textarea', 'required': True, 'section': '2、配电信息'},
        'usage_days': {'label': '*正常用电天数/年', 'type': 'number', 'required': True, 'section': '2、配电信息'},
        'storage_voltage': {'label': '*储能系统接入电网电压等级', 'type': 'select', 'options': ['10KV', '380V'], 'required': True, 'section': '2、配电信息'},
        'load_max': {'label': '*用电负荷(KW)最大', 'type': 'number', 'required': True, 'section': '2、配电信息'},
        'load_min': {'label': '最小', 'type': 'number', 'section': '2、配电信息'},
        'load_avg': {'label': '平均', 'type': 'number', 'section': '2、配电信息'},
        'usage_uniform': {'label': '全年白天和晚上用电是否均匀', 'type': 'textarea', 'section': '2、配电信息'},
        'maintenance_time': {'label': '设备大修每年停工时间', 'type': 'textarea', 'section': '2、配电信息'},
        'backup_interval': {'label': '有无备用间隔供储能系统接入', 'type': 'textarea', 'section': '2、配电信息'},
        'storage_distance': {'label': '*储能及配电房之间的距离(米)', 'type': 'number', 'required': True, 'section': '2、配电信息'},

        # 3、其他信息
        'other_info': {'label': '其他信息', 'type': 'textarea', 'section': '3、其他信息'},

        # 4、收资清单
        'location_photo': {'label': '*储能系统安装位置照片', 'type': 'file', 'required': True, 'section': '4、项目收资清单', 'note': '图片格式'},
        'location_layout': {'label': '*储能系统安装位置平面布局图', 'type': 'file', 'required': True, 'section': '4、项目收资清单', 'note': 'CAD图纸'},
        'underground_pipeline': {'label': '储能系统安装位置地下管线图', 'type': 'file', 'section': '4、项目收资清单', 'note': 'CAD图纸'},
        'transformer_params': {'label': '*储能系统接入的变压器设备参数', 'type': 'textarea', 'required': True, 'section': '4、项目收资清单'},
        'load_data': {'label': '*以日为单位负载功率数据（最近6个月）', 'type': 'file', 'required': True, 'section': '4、项目收资清单', 'note': '国网app可查看'},
        'electricity_bill': {'label': '*近12个月电费单（基本电费+电度电费）', 'type': 'file', 'required': True, 'section': '4、项目收资清单', 'note': '电业局打印'},
        'electrical_drawing': {'label': '*企业用电电气一次图', 'type': 'file', 'required': True, 'section': '4、项目收资清单', 'note': 'CAD图纸'},
        'distribution_drawing': {'label': '*配电室电气一次接线图', 'type': 'file', 'required': True, 'section': '4、项目收资清单', 'note': 'CAD图纸'},
        },
        'zero_carbon_park': {
            # 1、基本信息
            'park_name': {'label': '*园区名称', 'type': 'text', 'required': True, 'section': '1、基本信息'},
            'park_location': {'label': '*园区所在地', 'type': 'text', 'required': True, 'section': '1、基本信息'},
            'park_area': {'label': '*园区规划面积（km²）', 'type': 'number', 'required': True, 'section': '1、基本信息'},
            'construction_stage': {'label': '*建设阶段', 'type': 'select', 'options': ['规划中', '建设中', '运营中', '改造升级'], 'required': True, 'section': '1、基本信息'},
            'management_entity': {'label': '*运营/管理主体', 'type': 'text', 'required': True, 'section': '1、基本信息'},
            'contact_person': {'label': '*对接人', 'type': 'text', 'required': True, 'section': '1、基本信息'},
            'contact_phone': {'label': '*联系方式', 'type': 'tel', 'required': True, 'section': '1、基本信息'},
            'park_type': {'label': '园区类型', 'type': 'select', 'options': ['工业园区', '高新区', '经开区', '物流园区', '综合园区', '其他'], 'section': '1、基本信息'},
            'leading_industries': {'label': '*主导产业', 'type': 'textarea', 'required': True, 'section': '1、基本信息'},

            # 2、规划指标
            'planned_output_value': {'label': '规划产值（亿元/年）', 'type': 'number', 'section': '2、规划指标'},
            'enterprise_count': {'label': '入园企业数量', 'type': 'number', 'section': '2、规划指标'},
            'building_area': {'label': '建筑面积（万㎡）', 'type': 'number', 'section': '2、规划指标'},
            'green_building_ratio': {'label': '*绿色建筑占比目标（%）', 'type': 'number', 'required': True, 'section': '2、规划指标'},
            'renewable_energy_target': {'label': '*可再生能源占比目标（%）', 'type': 'number', 'required': True, 'section': '2、规划指标'},
            'carbon_intensity_target': {'label': '*单位产值碳排放目标（tCO2e/万元）', 'type': 'text', 'required': True, 'section': '2、规划指标'},
            'energy_intensity_target': {'label': '单位产值能耗目标（tce/万元）', 'type': 'text', 'section': '2、规划指标'},
            'zero_carbon_target_year': {'label': '*零碳达成年份', 'type': 'number', 'required': True, 'section': '2、规划指标'},

            # 3、产业准入
            'admission_policy': {'label': '*产业准入政策', 'type': 'textarea', 'required': True, 'section': '3、产业准入'},
            'restricted_industries': {'label': '限制/淘汰产业清单', 'type': 'textarea', 'section': '3、产业准入'},
            'key_enterprises': {'label': '重点用能/排放企业情况', 'type': 'textarea', 'section': '3、产业准入'},
            'low_carbon_requirements': {'label': '*低碳准入要求', 'type': 'textarea', 'required': True, 'section': '3、产业准入'},
            'circular_economy_plan': {'label': '循环经济协同规划', 'type': 'textarea', 'section': '3、产业准入'},

            # 4、基础设施
            'power_supply': {'label': '*供配电系统现状', 'type': 'textarea', 'required': True, 'section': '4、基础设施'},
            'water_supply': {'label': '给排水系统现状', 'type': 'textarea', 'section': '4、基础设施'},
            'heating_cooling': {'label': '供热/供冷系统现状', 'type': 'textarea', 'section': '4、基础设施'},
            'wastewater_treatment': {'label': '污水处理设施情况', 'type': 'textarea', 'section': '4、基础设施'},
            'solid_waste_disposal': {'label': '固废与危废处置体系', 'type': 'textarea', 'section': '4、基础设施'},
            'green_transport': {'label': '绿色交通设施情况', 'type': 'textarea', 'section': '4、基础设施'},
            'smart_platform': {'label': '智慧能碳管理平台情况', 'type': 'textarea', 'section': '4、基础设施'},

            # 5、绿色能源
            'pv_capacity': {'label': '光伏装机容量（MW）', 'type': 'number', 'section': '5、绿色能源'},
            'wind_capacity': {'label': '风电装机容量（MW）', 'type': 'number', 'section': '5、绿色能源'},
            'storage_capacity': {'label': '储能规模（MWh）', 'type': 'number', 'section': '5、绿色能源'},
            'green_power_ratio_current': {'label': '当前绿电占比（%）', 'type': 'number', 'section': '5、绿色能源'},
            'green_power_procurement': {'label': '绿电/绿证采购机制', 'type': 'textarea', 'section': '5、绿色能源'},
            'energy_station': {'label': '综合能源站建设情况', 'type': 'textarea', 'section': '5、绿色能源'},
            'microgrid_plan': {'label': '源网荷储/微电网规划', 'type': 'textarea', 'section': '5、绿色能源'},
            'charging_facilities': {'label': '充换电设施规模', 'type': 'textarea', 'section': '5、绿色能源'},

            # 6、碳排放核算
            'accounting_boundary': {'label': '*核算边界', 'type': 'textarea', 'required': True, 'section': '6、碳排放核算'},
            'baseline_year': {'label': '*基准年', 'type': 'number', 'required': True, 'section': '6、碳排放核算'},
            'total_emissions': {'label': '年度碳排放总量（tCO2e）', 'type': 'number', 'section': '6、碳排放核算'},
            'scope1_emissions': {'label': '范围1排放（tCO2e）', 'type': 'number', 'section': '6、碳排放核算'},
            'scope2_emissions': {'label': '范围2排放（tCO2e）', 'type': 'number', 'section': '6、碳排放核算'},
            'scope3_emissions': {'label': '范围3排放（tCO2e）', 'type': 'number', 'section': '6、碳排放核算'},
            'accounting_standard': {'label': '*采用核算标准', 'type': 'textarea', 'required': True, 'section': '6、碳排放核算'},
            'carbon_reduction_plan': {'label': '*减排路径与项目库', 'type': 'textarea', 'required': True, 'section': '6、碳排放核算'},

            # 7、项目收资清单
            'overall_plan_file': {'label': '*园区总体规划/控制性详细规划', 'type': 'file', 'required': True, 'section': '7、项目收资清单', 'note': 'PDF/CAD'},
            'industry_catalog_file': {'label': '*产业准入与企业清单', 'type': 'file', 'required': True, 'section': '7、项目收资清单', 'note': 'Excel/PDF'},
            'infrastructure_layout_file': {'label': '*基础设施总平图', 'type': 'file', 'required': True, 'section': '7、项目收资清单', 'note': 'CAD/PDF'},
            'energy_data_file': {'label': '*近12个月能源消费数据', 'type': 'file', 'required': True, 'section': '7、项目收资清单', 'note': 'Excel/CSV'},
            'renewable_energy_file': {'label': '绿电、光伏、储能项目资料', 'type': 'file', 'section': '7、项目收资清单', 'note': 'PDF/Excel'},
            'carbon_inventory_file': {'label': '*碳排放核算基础数据/盘查报告', 'type': 'file', 'required': True, 'section': '7、项目收资清单', 'note': 'PDF/Excel'},
            'policy_support_file': {'label': '政策文件与奖补材料', 'type': 'file', 'section': '7、项目收资清单', 'note': 'PDF'},
            'other_attachments': {'label': '其他补充附件', 'type': 'file', 'section': '7、项目收资清单', 'note': '压缩包/PDF'},
        }
    }
    return forms.get(template_type, forms['energy_storage'])


def get_template_meta(template_type='energy_storage'):
    return TEMPLATE_OPTIONS.get(template_type, TEMPLATE_OPTIONS['energy_storage'])


def get_field_value(request_obj, key, config):
    if config['type'] == 'checkbox':
        return ', '.join(request_obj.form.getlist(key))
    return request_obj.form.get(key, '')


def get_submission_name(submission):
    return submission.get('project_name') or submission.get('park_name') or submission.get('company_name') or 'unknown'


def grouped_sections(form_data):
    sections = {}
    for key, config in form_data.items():
        sections.setdefault(config['section'], []).append((key, config))
    return sections


def render_template_selector(target_endpoint):
    return render_template_string("""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>模板选择</title>
        <style>
            body { font-family: 'Microsoft YaHei', sans-serif; background: linear-gradient(135deg, #eef6ff, #f7fff8); margin: 0; color: #1a1a2e; }
            .wrap { max-width: 960px; margin: 0 auto; padding: 48px 20px; }
            h1 { margin: 0 0 12px; font-size: 32px; }
            p { color: #5b6475; margin-bottom: 28px; }
            .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 24px; }
            .card { background: white; border-radius: 18px; padding: 28px; box-shadow: 0 14px 40px rgba(27, 46, 94, 0.12); }
            .icon { font-size: 36px; }
            h2 { margin: 12px 0 10px; font-size: 24px; }
            .desc { min-height: 44px; }
            .btn { display: inline-block; margin-top: 18px; padding: 12px 24px; border-radius: 999px; background: #1f7aec; color: white; text-decoration: none; font-weight: bold; }
            .btn.alt { background: #18a058; }
        </style>
    </head>
    <body>
        <div class="wrap">
            <h1>选择收资模板</h1>
            <p>请选择需要使用的收资清单模板。</p>
            <div class="grid">
                <div class="card">
                    <div class="icon">⚡</div>
                    <h2>CALB工商业储能</h2>
                    <div class="desc">适用于工商业储能项目的现场、配电与附件收资。</div>
                    <a class="btn" href="{{ url_for(target_endpoint, template_type='energy_storage') }}">使用该模板</a>
                </div>
                <div class="card">
                    <div class="icon">🌿</div>
                    <h2>零碳园区</h2>
                    <div class="desc">覆盖园区基本信息、规划指标、产业准入、基础设施、绿色能源与碳核算收资。</div>
                    <a class="btn alt" href="{{ url_for(target_endpoint, template_type='zero_carbon_park') }}">使用该模板</a>
                </div>
            </div>
        </div>
    </body>
    </html>
    """, target_endpoint=target_endpoint)


def render_dynamic_form(template_type, form_data):
    template_meta = get_template_meta(template_type)
    return render_template_string("""
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{{ template_meta.name }} - 在线填写</title>
        <style>
            * { box-sizing: border-box; }
            body { margin: 0; font-family: 'Microsoft YaHei', sans-serif; background: #f5f7fa; color: #263238; }
            .header { background: linear-gradient(135deg, #17324d, #276749); color: white; padding: 22px 28px; display: flex; justify-content: space-between; align-items: center; }
            .header h1 { margin: 0; font-size: 24px; }
            .header a { color: white; text-decoration: none; border: 1px solid rgba(255,255,255,0.35); padding: 8px 16px; border-radius: 999px; }
            .container { max-width: 980px; margin: 32px auto; padding: 0 20px 40px; }
            .note { background: #fff8e6; border-left: 4px solid #f5b700; padding: 16px 18px; border-radius: 8px; margin-bottom: 24px; }
            .section { background: white; border-radius: 16px; padding: 28px; margin-bottom: 22px; box-shadow: 0 8px 24px rgba(17, 24, 39, 0.08); }
            .section h2 { margin: 0 0 20px; font-size: 22px; border-bottom: 2px solid #dbe7f3; padding-bottom: 12px; }
            .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 18px; }
            .field { margin-bottom: 18px; }
            .field.full { grid-column: 1 / -1; }
            label { display: block; font-weight: 600; margin-bottom: 8px; }
            input[type=text], input[type=number], input[type=tel], select, textarea, input[type=file] { width: 100%; padding: 12px 14px; border: 1px solid #cfd8e3; border-radius: 10px; font-size: 15px; font-family: inherit; background: white; }
            textarea { min-height: 110px; resize: vertical; }
            .note-text { display: block; margin-top: 6px; color: #6b7280; font-size: 13px; }
            .actions { text-align: center; margin-top: 26px; }
            button { border: none; background: linear-gradient(90deg, #2f80ed, #27ae60); color: white; border-radius: 999px; padding: 14px 36px; font-size: 16px; font-weight: 700; cursor: pointer; }
            .flash-messages { position: fixed; top: 84px; right: 18px; z-index: 1000; }
            .flash { padding: 14px 20px; border-radius: 10px; margin-bottom: 10px; color: white; }
            .flash.success { background: #18a058; }
            .flash.error { background: #d14343; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>{{ template_meta.icon }} {{ template_meta.name }} - 在线填写</h1>
            <a href="/">返回首页</a>
        </div>
        <div class="flash-messages">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="flash {{ category }}">{{ message }}</div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
        </div>
        <div class="container">
            <div class="note">标 * 为必填项；收资附件建议上传 PDF、Excel、CAD 或压缩包等可追溯资料。</div>
            <form method="POST" enctype="multipart/form-data">
                <input type="hidden" name="template_type" value="{{ template_type }}">
                {% for section, fields in sections.items() %}
                    <div class="section">
                        <h2>{{ section }}</h2>
                        <div class="grid">
                            {% for key, config in fields %}
                                <div class="field {% if config.type in ['textarea', 'file'] %}full{% endif %}">
                                    <label>{{ config.label }}</label>
                                    {% if config.type == 'select' %}
                                        <select name="{{ key }}" {% if config.required %}required{% endif %}>
                                            <option value="">请选择</option>
                                            {% for option in config.options %}
                                                <option value="{{ option }}">{{ option }}</option>
                                            {% endfor %}
                                        </select>
                                    {% elif config.type == 'textarea' %}
                                        <textarea name="{{ key }}" {% if config.required %}required{% endif %}></textarea>
                                    {% elif config.type == 'file' %}
                                        <input type="file" name="{{ key }}" {% if config.required %}required{% endif %}>
                                    {% else %}
                                        <input type="{{ config.type }}" name="{{ key }}" {% if config.required %}required{% endif %}>
                                    {% endif %}
                                    {% if config.note %}
                                        <span class="note-text">{{ config.note }}</span>
                                    {% endif %}
                                </div>
                            {% endfor %}
                        </div>
                    </div>
                {% endfor %}
                <div class="actions">
                    <button type="submit">提交收资清单</button>
                </div>
            </form>
        </div>
    </body>
    </html>
    """, template_meta=template_meta, template_type=template_type, sections=grouped_sections(form_data))


def build_excel_workbook(template_type, include_instructions=False):
    form_data = get_form_data(template_type)
    template_meta = get_template_meta(template_type)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '收资清单'

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, (_, config) in enumerate(form_data.items(), 1):
        cell = ws.cell(row=1, column=col_idx, value=config['label'].lstrip('*'))
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        width = max(14, min(34, len(config['label']) + 6))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in range(2, 102):
            body_cell = ws.cell(row=row, column=col_idx)
            body_cell.border = border
            body_cell.alignment = Alignment(vertical='top', wrap_text=True)

    if include_instructions:
        ws2 = wb.create_sheet('填写说明')
        ws2['A1'] = template_meta['template_title']
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A3'] = '1. 标 * 的字段为必填项'
        ws2['A4'] = '2. 每行代表一个项目，请保持字段名称与模板一致'
        ws2['A5'] = '3. 附件类资料请通过网页表单上传，Excel 中填写资料名称或说明'
        ws2['A6'] = '4. 推荐上传零碳园区相关规划、能耗、碳排放、绿电与基础设施资料'
        ws2.column_dimensions['A'].width = 72

    return wb


def export_submissions_to_workbook(submissions_list, template_type):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '提交数据'
    form_data = get_form_data(template_type)

    headers = ['提交ID', '提交时间', '模板类型'] + [config['label'].lstrip('*') for config in form_data.values()]
    key_map = ['id', 'submit_time', 'template_type'] + list(form_data.keys())

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(32, len(header) + 6))

    for row_idx, sub in enumerate(submissions_list, 2):
        for col_idx, key in enumerate(key_map, 1):
            value = sub.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    return wb


@app.route('/')
def index():
    """首页"""
    return render_template('index.html')


@app.route('/view')
def view():
    """在线查看表格"""
    return render_template('view.html')


@app.route('/form', methods=['GET', 'POST'])
def form():
    """在线填写表单"""
    template_type = request.args.get('template_type') or request.form.get('template_type') or 'energy_storage'
    if request.method == 'GET' and 'template_type' not in request.args:
        return render_template_selector('form')

    form_data = get_form_data(template_type)

    if request.method == 'POST':
        # 处理表单提交
        submission = {'template_type': template_type, 'template_name': get_template_meta(template_type)['name']}

        for key, config in form_data.items():
            if config['type'] == 'file':
                file = request.files.get(key)
                if file and file.filename:
                    # 保存文件
                    filename = f"{get_submission_name(submission)}_{key}_{file.filename}"
                    filename = "".join(c for c in filename if c.isalnum() or c in '._-')
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    submission[key] = filename
                else:
                    submission[key] = ''
            else:
                submission[key] = get_field_value(request, key, config)

        # 保存提交
        submit_id = save_submission(submission)
        flash(f'提交成功！提交ID: {submit_id}', 'success')
        
        # 发送飞书通知
        try:
            send_feishu_notification(submission)
        except Exception as e:
            print(f"[WARN] 飞书通知发送失败: {e}")
        
        return redirect(url_for('form', template_type=template_type))

    if template_type == 'energy_storage':
        return render_template('form.html', form_data=form_data)
    return render_dynamic_form(template_type, form_data)


@app.route('/submissions')
def submissions():
    """查看已提交的数据"""
    submissions_list = load_submissions()
    return render_template('submissions.html', submissions=submissions_list)


@app.route('/submissions/export')
def export_submissions():
    """导出所有提交数据为Excel"""
    template_type = request.args.get('template_type')
    if not template_type:
        return render_template_selector('export_submissions')

    submissions_list = [
        sub for sub in load_submissions()
        if sub.get('template_type', 'energy_storage') == template_type
    ]

    if not submissions_list:
        flash('没有可导出的数据', 'warning')
        return redirect(url_for('submissions'))

    wb = export_submissions_to_workbook(submissions_list, template_type)
    output_path = os.path.join('submissions', get_template_meta(template_type)['export_name'])
    wb.save(output_path)

    return send_file(output_path, as_attachment=True, download_name=get_template_meta(template_type)['export_name'])


@app.route('/download/template')
def download_template():
    """下载Excel模板"""
    template_type = request.args.get('template_type')
    if not template_type:
        return render_template_selector('download_template')

    # 如果模板存在则下载，否则自动生成
    if template_type == 'energy_storage' and os.path.exists(TEMPLATE_PATH):
        return send_file(TEMPLATE_PATH, as_attachment=True, download_name='CALB-工商业储能收资清单.xlsx')
    else:
        wb = build_excel_workbook(template_type, include_instructions=True)
        output_path = get_template_meta(template_type)['template_output']
        wb.save(output_path)

        return send_file(output_path, as_attachment=True, download_name=get_template_meta(template_type)['download_name'])


@app.route('/import', methods=['GET', 'POST'])
def import_excel():
    """导入Excel文件提交"""
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or not file.filename.endswith('.xlsx'):
            flash('请上传 .xlsx 格式的Excel文件', 'error')
            return redirect(url_for('import_excel'))
        
        try:
            # 读取Excel
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # 读取表头和数据
            headers = [cell.value for cell in ws[1]]
            data_row = [cell.value for cell in ws[2]]  # 取第一行数据
            
            if not data_row or not any(data_row):
                flash('Excel文件中没有数据', 'error')
                return redirect(url_for('import_excel'))
            
            # 映射字段
            header_map = {
                '公司名称': 'company_name', '项目名称': 'project_name',
                '项目地点': 'project_location', '对接人': 'contact_person',
                '联系方式': 'contact_phone', '企业用电性质': 'power_type',
                '计量电压等级': 'voltage_level', '主变容量(kVA)': 'transformer_capacity',
                '变压器台数': 'transformer_count', '基本电价方式': 'basic_price',
                '正常用电天数/年': 'usage_days', '储能接入电压等级': 'storage_voltage',
                '最大负荷(KW)': 'load_max', '储能距配电房距离(米)': 'storage_distance',
                '其他信息': 'other_info'
            }
            
            # 构建提交数据
            submission = {'source': 'excel_import', 'template_type': 'energy_storage', 'template_name': get_template_meta('energy_storage')['name']}
            for header, value in zip(headers, data_row):
                if header in header_map:
                    submission[header_map[header]] = str(value) if value else ''
            
            # 保存并通知
            submit_id = save_submission(submission)
            flash(f'Excel导入成功！提交ID: {submit_id}', 'success')
            
            try:
                send_feishu_notification(submission)
            except Exception as e:
                print(f"[WARN] 飞书通知发送失败: {e}")
            
            return redirect(url_for('submissions'))
            
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'error')
            return redirect(url_for('import_excel'))
    
    return """
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Excel批量导入 - CALB储能收资清单</title>
        <style>
            body { font-family: 'Microsoft YaHei', sans-serif; background: #f5f7fa; padding: 50px; }
            .container { max-width: 600px; margin: 0 auto; background: white; padding: 40px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            h1 { color: #1a1a2e; margin-bottom: 30px; }
            .info { background: #e8f4fd; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
            .info h3 { margin-bottom: 10px; color: #366092; }
            .info ol { margin-left: 20px; }
            .form-group { margin-bottom: 20px; }
            input[type=file] { padding: 10px; border: 1px solid #ddd; border-radius: 8px; width: 100%; }
            .btn { padding: 12px 30px; background: linear-gradient(135deg, #00d4ff, #00ff88); color: white; border: none; border-radius: 25px; cursor: pointer; font-size: 1em; }
            .btn:hover { transform: translateY(-2px); box-shadow: 0 5px 20px rgba(0,212,255,0.3); }
            a { color: #00d4ff; text-decoration: none; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 Excel批量导入</h1>
            <div class="info">
                <h3>操作步骤：</h3>
                <ol>
                    <li>先 <a href="/download/template">下载Excel模板</a></li>
                    <li>按照模板格式填写数据</li>
                    <li>保存为 .xlsx 格式</li>
                    <li>上传填写好的文件</li>
                </ol>
            </div>
            <form method="POST" enctype="multipart/form-data">
                <div class="form-group">
                    <input type="file" name="excel_file" accept=".xlsx" required>
                </div>
                <button type="submit" class="btn">📤 导入提交</button>
            </form>
            <p style="margin-top: 20px;"><a href="/">返回首页</a></p>
        </div>
    </body>
    </html>
    """


@app.route('/submissions/<submit_id>')
def submission_detail(submit_id):
    """查看单条提交详情"""
    submissions_list = load_submissions()
    submission = None
    for sub in submissions_list:
        if sub.get('id') == submit_id:
            submission = sub
            break

    if not submission:
        flash('未找到该提交记录', 'error')
        return redirect(url_for('submissions'))

    form_data = get_form_data(submission.get('template_type', 'energy_storage'))
    return render_template('detail.html', submission=submission, form_data=form_data)


if __name__ == '__main__':
    # 运行服务器
    app.run(host='0.0.0.0', port=5000, debug=True)
